#!/usr/bin/env python3
"""Build the master build tracker: one prioritised Excel list of everything
still to build for BroadbandPicker.co.uk — new pages AND product/UX features.

Combines two sources:
1. The Content Gap Roadmap tab of docs/broadbandpicker-keyword-mapping.xlsx
   (page builds, each with an SEO Build Priority Score and Revenue Priority
   Score already computed by scripts/build_keyword_mapping.py).
2. A curated list of feature/UX/content-strategy builds from the growth
   playbook (docs/competitor-landscape-scan.json research + the sequencing
   already agreed), each hand-scored on the same rough 0-100 scale so the
   two sources can be ranked together honestly rather than presented as two
   disconnected lists.

Both feed into one Master Tracker tab, ranked highest priority first. A separate
Pending Build Priority tab is generated from the same records so the next work
queue is explicit and contains no completed items.

Re-running is safe: Status, Owner, Target Date and Notes are preserved from
the previous docs/master-build-tracker.xlsx if one exists (matched by a
stable Item ID), so manual progress updates are never overwritten by a
regeneration.

Usage:
    python3 scripts/build_master_tracker.py
    python3 scripts/build_master_tracker.py --source docs/broadbandpicker-keyword-mapping.xlsx
"""

from __future__ import annotations

import argparse
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "docs" / "master-build-tracker.xlsx"
DEFAULT_SOURCE = ROOT / "docs" / "broadbandpicker-keyword-mapping.xlsx"
RUN_DATE = datetime.now(timezone.utc).date().isoformat()

# Preserved across regenerations by Item ID.
MANUAL_HEADERS = ("Status", "Owner", "Target Date", "Notes")
DEFAULT_STATUS = "Not started"

# ---------------------------------------------------------------------------
# Feature / UX / content-strategy builds — curated from the growth playbook.
# Hand-scored on the same rough 0-100 scale as the page-build priority score
# so both sources can be ranked in one list without pretending the numbers
# came from the same formula.
# ---------------------------------------------------------------------------

FEATURE_BUILDS: list[dict[str, Any]] = [
    {
        "item_id": "ops-conversion-reporting-loop",
        "type": "Measurement",
        "pillar": "Analytics",
        "title": "Validate conversion events and establish the GSC/analytics reporting loop",
        "description": (
            "Confirm postcode submits, speed-test completions, filter use, outbound provider "
            "clicks and assisted conversions are recorded reliably, then use query and "
            "engagement evidence to reprioritise future work."
        ),
        "priority_score": 63,
        "impact_score": 72,
        "effort": "Medium",
        "target": "GSC, analytics and conversion reporting",
        "dependencies": "Sufficient live traffic and event data",
        "source": "BroadbandPicker SEO & Content Plan — Live: Build Status / Technical SEO",
        "verified_status": "Done",
        "completion_notes": (
            "Shipped to Vercel production 2026-08-23: consent-aware GA4 events for postcode "
            "submits, speed-test starts/completions/failures, deal filter/sort use and outbound "
            "provider clicks, with session first-touch attribution and no full postcode sent. "
            "Measurement contract and weekly GSC/GA4 loop documented in docs/analytics-measurement-plan.md."
        ),
    },
    {
        "item_id": "growth-awin-advertiser-outreach",
        "type": "Partnerships",
        "pillar": "Monetisation",
        "title": "Continue selective Awin advertiser outreach",
        "description": (
            "Apply to best-fit provider programmes with tailored pitches, relevant live URLs, "
            "audience evidence, promotional method and compliance controls; log decisions and feedback. "
            "Real status pulled from the Awin Publisher API 2026-08-23 (see "
            "ops-awin-publisher-api-integration): joined -- TalkTalk, Broadband Genie, Zzoomm, "
            "Highland Broadband. Pending -- Sky ROI, Community Fibre, National Broadband, Trooli, "
            "Pine Media, Cuckoo, Zen Internet. Rejected -- BT (consumer + business), Vodafone, "
            "Plusnet, EE, Virgin Media, Hyperoptic, toob, giffgaff. The 8 rejected mainstream "
            "providers are the priority re-application targets; Zzoomm and Highland Broadband are "
            "approved but not yet added as providers on the site at all."
        ),
        "priority_score": 60,
        "impact_score": 70,
        "effort": "Medium",
        "target": "Awin publisher 2942019 and partner application log",
        "dependencies": "Current audience evidence and compliant commercial pages",
        "source": "BroadbandPicker SEO & Content Plan — Live: 90 Day Roadmap",
        "initial_status": "In progress",
    },
    {
        "item_id": "ops-adsense-site-review",
        "type": "Monetisation",
        "pillar": "Infrastructure",
        "title": "Complete AdSense site review and authorisation monitoring",
        "description": (
            "Monitor the existing AdSense review, consent implementation and ads.txt authorisation; "
            "resolve any concrete policy or crawler issue reported by AdSense."
        ),
        "priority_score": 58,
        "impact_score": 55,
        "effort": "Low",
        "target": "AdSense status and /ads.txt",
        "dependencies": "Google site-review processing",
        "source": "BroadbandPicker SEO & Content Plan — Live: Build Status",
        "initial_status": "In progress",
    },
    {
        "item_id": "growth-original-research-outreach",
        "type": "Growth",
        "pillar": "Content",
        "title": "Run original-research outreach and digital PR",
        "description": (
            "Build a relevant outreach list and earn citations and links to the live Ofcom-backed "
            "customer-satisfaction research."
        ),
        "priority_score": 54,
        "impact_score": 68,
        "effort": "Medium",
        "target": "/research/uk-broadband-customer-satisfaction",
        "dependencies": "Research page and source methodology are live",
        "source": "BroadbandPicker SEO & Content Plan — Live: Build Status / 90 Day Roadmap",
        "initial_status": "In progress",
    },
    {
        "item_id": "growth-quarterly-data-reprioritisation",
        "type": "Strategy",
        "pillar": "Growth",
        "title": "Reprioritise quarter two using GSC, engagement and affiliate evidence",
        "description": (
            "Review query/page performance, assisted conversions, Awin decisions and engagement "
            "data, then record the next-quarter build decisions."
        ),
        "priority_score": 53,
        "impact_score": 63,
        "effort": "Medium",
        "target": "Next-quarter decision log",
        "dependencies": "ops-conversion-reporting-loop and sufficient observation period",
        "source": "BroadbandPicker SEO & Content Plan — Live: 90 Day Roadmap",
    },
    {
        "item_id": "audit-core-web-vitals-reporting",
        "type": "Technical SEO",
        "pillar": "Performance",
        "title": "Establish monthly Core Web Vitals reporting by template",
        "description": "Track LCP, INP and CLS by page template and prioritise JS or image-weight fixes from evidence.",
        "priority_score": 47,
        "impact_score": 50,
        "effort": "Medium",
        "target": "Sitewide template performance report",
        "dependencies": "Field data or representative lab baselines",
        "source": "BroadbandPicker SEO & Content Plan — Live: Technical SEO",
    },
    {
        "item_id": "audit-broken-links-redirects",
        "type": "Technical SEO",
        "pillar": "Crawlability",
        "title": "Add routine broken-link and redirect reporting",
        "description": "Monitor broken inbound and internal URLs and maintain relevant one-hop redirects.",
        "priority_score": 44,
        "impact_score": 48,
        "effort": "Low",
        "target": "Sitewide crawl and redirect report",
        "dependencies": "None",
        "source": "BroadbandPicker SEO & Content Plan — Live: Technical SEO",
    },
    {
        "item_id": "ops-git-vercel-reconcile",
        "type": "Ops / risk",
        "pillar": "Infrastructure",
        "title": "Reconcile git history with the live Vercel deployment",
        "description": (
            "Production currently serves pages (the 16-page pipeline batch) that were "
            "deployed straight from local disk via `vercel --prod` and were never "
            "committed or pushed to git. Git and production have diverged — the next "
            "`git push` triggering a Vercel git-integration deploy, or any reset of the "
            "local working tree, risks rolling the live site back. Commit and push the "
            "current working tree to close the gap."
        ),
        "priority_score": 96,
        "impact_score": 90,
        "effort": "Low",
        "target": "Repository + Vercel project",
        "dependencies": "None — blocks everything else being deployed safely",
        "source": "Discovered while preparing this tracker",
    },
    {
        "item_id": "feat-product-offer-schema",
        "type": "Feature",
        "pillar": "Functionality",
        "title": "Product/Offer schema on deal cards",
        "description": (
            "Uswitch marks up individual deals as Product + Quotation JSON-LD — the "
            "basis for price/rating rich snippets in search results. BroadbandPicker's "
            "deal cards don't carry this yet. Contained, well-scoped, clear SEO payoff."
        ),
        "priority_score": 74,
        "impact_score": 70,
        "effort": "Low",
        "target": "components/DealCard and deal-listing pages",
        "dependencies": "None",
        "source": "docs/competitor-landscape-scan.json",
    },
    {
        "item_id": "feat-persistent-postcode",
        "type": "Feature",
        "pillar": "UX",
        "title": "Persistent postcode across the site",
        "description": (
            "Once a postcode is entered on /postcode, carry it (localStorage, no "
            "account needed) into /deals, /compare and every provider page so the "
            "journey feels personal rather than resetting the question each time."
        ),
        "priority_score": 71,
        "impact_score": 68,
        "effort": "Medium",
        "target": "PostcodeChecker + shared client state",
        "dependencies": "None",
        "source": "Growth playbook — UX pillar",
    },
    {
        "item_id": "feat-shortlist-compare-basket",
        "type": "Feature",
        "pillar": "UX",
        "title": "Shortlist / compare-basket pattern",
        "description": (
            "Let users tick 2–3 deals and see them side by side, matching Uswitch and "
            "broadband.co.uk. Current /compare is one full table — good for browsing, "
            "weaker for a final decision between finalists."
        ),
        "priority_score": 61,
        "impact_score": 64,
        "effort": "Medium",
        "target": "app/compare",
        "dependencies": "None",
        "source": "Growth playbook — UX pillar",
        "verified_status": "Done",
        "completion_notes": (
            "Shipped to Vercel production 2026-08-23: visitors can shortlist two or three "
            "providers on /compare, retain the shortlist in local storage, compare price, "
            "maximum speed, contract, setup fee, coverage and customer rating in responsive "
            "finalist cards, then continue to a provider review or tracked deal link. Includes "
            "accessible controls, a three-provider limit and GA4 shortlist/basket events."
        ),
    },
    {
        "item_id": "feat-review-aggregation-module",
        "type": "Feature",
        "pillar": "Functionality",
        "title": "Real review-aggregation trust module",
        "description": (
            "Blend the existing Ofcom-complaints research methodology with live "
            "Trustpilot scores into one visible trust module reused across provider "
            "and comparison pages, instead of a single quoted score per page."
        ),
        "priority_score": 59,
        "impact_score": 62,
        "effort": "Medium",
        "target": "New shared component, used on provider + comparison pages",
        "dependencies": "research/uk-broadband-customer-satisfaction methodology",
        "source": "Growth playbook — Functionality pillar",
    },
    {
        "item_id": "audit-city-hub-thinness",
        "type": "Content audit",
        "pillar": "Content",
        "title": "Verify each new city hub has a genuinely local fact",
        "description": (
            "8 new city hubs (Birmingham, Bristol, Edinburgh, Glasgow, Leeds, "
            "Liverpool, Manchester, Sheffield) plus 50 existing postcode-prefix pages "
            "is a lot of near-identical templates. Each needs at least one real local "
            "fact (a named local altnet, a real coverage figure) or it reads as "
            "duplicate content — audit before building the next batch of cities."
        ),
        "priority_score": 56,
        "impact_score": 50,
        "effort": "Medium",
        "target": "data/priority-pages.ts city entries",
        "dependencies": "None",
        "source": "Growth playbook — Content pillar",
    },
    {
        "item_id": "feat-price-drop-alerts",
        "type": "Feature",
        "pillar": "Functionality",
        "title": "Price-drop alerts",
        "description": (
            "data/provider-live-deals.json already tracks live pricing. An email/"
            "notification layer on top (\"tell me if broadband in my postcode drops "
            "below £X\") is mostly wiring, and it's a retention mechanic none of the "
            "readable competitors emphasised as much as their comparison tables."
        ),
        "priority_score": 49,
        "impact_score": 55,
        "effort": "Medium",
        "target": "New subscription flow + data/provider-live-deals.json",
        "dependencies": "Email delivery provider",
        "source": "Growth playbook — Functionality pillar",
    },
    {
        "item_id": "audit-mobile-checkout-flow",
        "type": "UX audit",
        "pillar": "UX",
        "title": "Mobile checkout-flow audit",
        "description": (
            "The competitor evidence (15k–23k word desktop hubs) is desktop-shaped. "
            "On mobile that content needs a different information architecture — "
            "sticky summary bar, collapsed sections, thumb-reachable CTAs. Worth a "
            "dedicated pass once analytics show where mobile users drop off."
        ),
        "priority_score": 45,
        "impact_score": 58,
        "effort": "Medium",
        "target": "Sitewide mobile layout",
        "dependencies": "Mobile analytics/drop-off data",
        "source": "Growth playbook — UX pillar",
    },
    {
        "item_id": "bet-annual-research-report",
        "type": "Bigger bet",
        "pillar": "Content",
        "title": "Annual 'State of UK Broadband' research report",
        "description": (
            "No competitor scanned publishes original research. Built on the "
            "existing /research page's methodology, an annual report is a natural "
            "press/backlink asset a comparison-table page can never be."
        ),
        "priority_score": 41,
        "impact_score": 65,
        "effort": "High",
        "target": "New annual research page + methodology extension",
        "dependencies": "research/uk-broadband-customer-satisfaction methodology",
        "source": "Growth playbook — Content pillar",
    },
    {
        "item_id": "bet-componentise-interactive-layer",
        "type": "Bigger bet",
        "pillar": "Functionality",
        "title": "Componentise the interactive layer on accessible primitives",
        "description": (
            "PostcodeChecker, SpeedTest, StickyFilterBar and SocialProofCounter "
            "already exist as separate components. As the shortlist and price-alert "
            "features land, adopt accessible primitives (Radix-style) underneath "
            "rather than hand-rolling focus/keyboard behaviour again each time."
        ),
        "priority_score": 38,
        "impact_score": 45,
        "effort": "High",
        "target": "components/",
        "dependencies": "feat-shortlist-compare-basket, feat-price-drop-alerts",
        "source": "Growth playbook — Functionality pillar",
    },
    {
        "item_id": "feat-uk-wide-postcode-coverage",
        "type": "Feature",
        "pillar": "Content/Trust",
        "title": "UK-wide real postcode coverage data (Ofcom)",
        "description": (
            "Only ~50 curated postcode prefixes had real local data; every other UK "
            "postcode showed a generic 'we're expanding coverage' fallback with no "
            "area-specific information. Built from Ofcom's open Connected Nations "
            "postcode-level dataset (Open Government Licence), aggregated to "
            "district level, covering all 2,818 UK postcode districts with real "
            "gigabit/superfast/ultrafast availability percentages."
        ),
        "priority_score": 80,
        "impact_score": 85,
        "effort": "Medium",
        "target": "data/postcode-district-coverage.json, data/postcodeDistrictCoverage.ts, app/postcode/[area]/page.tsx",
        "dependencies": "scripts/build_postcode_coverage.py (re-run when Ofcom publishes a newer postcode-level edition)",
        "source": "User request — real postcode-level availability",
    },
    {
        "item_id": "feat-broadband-match-quiz",
        "type": "Feature",
        "pillar": "Functionality",
        "title": "Broadband Match: personalised recommendation quiz",
        "description": (
            "Researched 2026-08-22 against Uswitch, Compare the Market, "
            "MoneySuperMarket, Which?, broadbandchoices, choose.co.uk and "
            "broadband.co.uk: none has a multi-question quiz that outputs a "
            "ranked provider/package recommendation — confirmed genuine gap, "
            "not a crowded feature. Closest analog, RightSpeed UK, only "
            "outputs a speed-tier range and hands off to a third-party site "
            "rather than affiliate-linking directly. Plan: 6-8 questions "
            "(reason for looking, household size, use cases — WFH/gaming/"
            "streaming/uploads, budget, contract-length preference, postcode), "
            "scored client-side against the existing provider dataset (speed, "
            "price, contract, Trustpilot, coverage — no new data pipeline "
            "needed), producing a ranked top 3 with per-pick reasoning and "
            "direct affiliate CTAs — going further than RightSpeed's speed-"
            "only output. Since AI Overviews cite static pages, not live quiz "
            "results (confirmed in research), pair the tool with internal "
            "links to/from the existing best-broadband-for-gaming/WFH/"
            "students/streaming guides so the underlying logic stays citable "
            "even though the live tool isn't. Directly serves the standing "
            "goal of making BroadbandPicker a site providers want in their "
            "affiliate programme: a completed-quiz click is a much stronger "
            "purchase-intent signal than a generic comparison-table click."
        ),
        "priority_score": 78,
        "impact_score": 83,
        "effort": "Medium-High",
        "target": "New route (e.g. /tools/broadband-match), new scoring component, reuses data/providers.ts + data/postcodes.ts + data/postcodeDistrictCoverage.ts",
        "dependencies": "None — reuses existing provider/coverage datasets",
        "source": "User request 2026-08-22, researched before scoring per the Strategic Lens process",
    },
    {
        "item_id": "feat-nav-header-footer-ia",
        "type": "Feature",
        "pillar": "UX",
        "title": "Navigation/header/footer IA improvements",
        "description": (
            "Researched via scripts/scrape_navigation_patterns.py against 5 best-in-class "
            "sites. broadband.co.uk (closest single-vertical comparable) stood out on sticky "
            "header, breadcrumbs and a trust badge near the top; BroadbandPicker already "
            "matches on the first two. Header nav link count split cleanly into supersite "
            "(127-556 links) vs specialist (21 links) patterns, confirming the growth "
            "playbook's 'specialist, not supersite' positioning. Added Postcode to primary "
            "nav; restructured the footer from 4 to 6 columns (added Tools, Find Broadband "
            "by Area, Research) to fix a crawl/link-equity gap for the postcode-district and "
            "tools pages built this session."
        ),
        "priority_score": 55,
        "impact_score": 60,
        "effort": "Low",
        "target": "app/layout.tsx",
        "dependencies": "None",
        "source": "User request 2026-08-22 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "feat-homepage-visual-redesign",
        "type": "Feature",
        "pillar": "UX",
        "title": "Homepage visual redesign: illustrations + interactivity",
        "description": (
            "Researched via scripts/analyze_homepage_visual_design.py against 5 UK broadband "
            "comparison homepages: none uses stock-photo hero banners; the dominant visual "
            "language is heavy inline SVG (Uswitch ships 195) and provider logos, no carousel/"
            "animation library detected anywhere. Generated 7 original on-brand SVG "
            "illustrations (scripts/generate_homepage_illustrations.py: hero network graphic, "
            "3 colour-filled step icons, 2 gradient blobs, a quiz illustration) rather than "
            "stock photography, matching what's actually proven in this vertical. Added a "
            "scroll-reveal component, hover interactions, and restructured the Broadband Match "
            "promo to a two-column illustrated layout. Content/copy unchanged, visual/UX only."
        ),
        "priority_score": 52,
        "impact_score": 55,
        "effort": "Medium",
        "target": "app/page.tsx, components/ScrollReveal.tsx, public/illustrations/",
        "dependencies": "None",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "feat-footer-logo-interactivity",
        "type": "Feature",
        "pillar": "UX",
        "title": "Footer logo placement and interactive elements",
        "description": (
            "Extended scripts/analyze_homepage_visual_design.py with footer-specific "
            "detection (logo, social icons, accordion/back-to-top, hover density). Confirmed "
            "directly in app/layout.tsx (not just inferred from the scan) that the footer had "
            "no logo at all. Added components/Logo.tsx (extracted from the header, reused "
            "larger in the footer with a tagline), made the 6 footer link columns "
            "<details open> elements with a mobile-only collapse toggle (desktop behaves "
            "exactly as static headings did, via lg:pointer-events-none, so nothing is at "
            "risk of accidental collapse), upgraded the social link to a hover-interactive "
            "pill button, and added a 'Back to top' link using the existing global "
            "smooth-scroll behaviour."
        ),
        "priority_score": 48,
        "impact_score": 40,
        "effort": "Low",
        "target": "app/layout.tsx, components/Logo.tsx",
        "dependencies": "None",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "feat-mega-menu-navigation",
        "type": "Feature",
        "pillar": "UX",
        "title": "Detailed mega-menu navigation",
        "description": (
            "Extended scripts/scrape_navigation_patterns.py with mega-menu structural "
            "detection (top-level item count, icons-in-dropdown, descriptive subtext, "
            "nesting depth). Found Uswitch uses icons in dropdown items and broadband.co.uk "
            "uses one-line descriptions under category groups — both implemented. Built "
            "components/MainNav.tsx: Providers, Postcode, Guides and Tools are now dropdowns "
            "with icons and descriptions; Compare and Deals stay direct links as primary "
            "single-destination actions. The Guides dropdown is a real mega-menu pulling the "
            "site's actual 6 guide categories and their real guides live from data/guides.ts "
            "— stays correct as the pipeline adds guides, no hardcoding. CSS-only "
            "(group-hover/group-focus-within), keyboard-reachable via Tab, no JS dependency."
        ),
        "priority_score": 50,
        "impact_score": 58,
        "effort": "Medium",
        "target": "components/MainNav.tsx, app/layout.tsx",
        "dependencies": "None",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "feat-page-type-ux-redesign",
        "type": "Feature",
        "pillar": "UX",
        "title": "Page-type UX redesign: deals, providers, guides, postcode hub",
        "description": (
            "New scripts/analyze_page_type_ux.py scans competitor pages by theme (deals "
            "listing, provider review, guide article, postcode/checker hub) rather than "
            "homepage-only, against Uswitch, broadband.co.uk, choose.co.uk and "
            "MoneySavingExpert (403, reported not bypassed). Confirmed signals: comparison "
            "checkboxes near-universal; badge/chip labels on deals listings + review index; "
            "TOC on the guide article + one deals listing; pros/cons on provider review + "
            "deals; related-content on provider review + deals; rating widget on one deals "
            "listing. No dedicated postcode-hub competitor page exists in this vertical. "
            "Shipped strictly against confirmed signals: computed Best Value/Fastest/Editor's "
            "Pick badges in components/DealTable.tsx; new components/RatingStars.tsx replacing "
            "plain-text Trustpilot scores plus a 'How {provider} compares' related-comparisons "
            "module on provider pages; a jump-to-section table of contents on guide pages via "
            "new lib/extractHeadings.tsx (JSX-tree heading-ID injection, avoids hand-editing "
            "40 existing guide definitions), gated behind 3+ headings; and the same "
            "cheapest/fastest badges plus rating stars carried onto postcode-hub provider "
            "cards since deals and hub serve the same comparison job."
        ),
        "priority_score": 49,
        "impact_score": 52,
        "effort": "Medium",
        "target": (
            "components/DealTable.tsx, components/RatingStars.tsx, lib/extractHeadings.tsx, "
            "app/providers/[slug]/page.tsx, app/guides/[slug]/page.tsx, "
            "app/postcode/[area]/page.tsx"
        ),
        "dependencies": "None",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "feat-mobile-cross-device-navigation",
        "type": "Feature",
        "pillar": "UX",
        "title": "Mobile and cross-device navigation overhaul",
        "description": (
            "New scripts/analyze_mobile_ux.py scanned Uswitch, broadband.co.uk and "
            "choose.co.uk (MoneySavingExpert 403, reported not bypassed) with a mobile "
            "Safari user agent for markup/CSS signals: viewport meta and hamburger menu "
            "markers were universal but already correct on our own site; tel: click-to-call "
            "was absent everywhere (0/4, skipped); sticky bottom CTA (1/4) and "
            "apple-touch-icon (2/4) were too weak/incomplete to act on. The real gap was "
            "found by direct inspection, not the scan: the mobile nav was a small anchored "
            "dropdown of 10 flat links with ~36px touch targets, while the desktop nav beside "
            "it was a full mega-menu, and the header postcode checker was desktop-only so "
            "phone users had no quick postcode access outside the homepage/deals page. "
            "Replaced it with new components/MobileNav.tsx: a full-width slide-down panel "
            "reusing the desktop mega-menu's own icon/link data (exported from MainNav.tsx, "
            "so the two can't drift apart), every link at a 44px (min-h-11) touch target, "
            "and the postcode checker embedded at the top so it's reachable from any page. "
            "Added anchor ids to /guides category sections so the new mobile Guides menu "
            "jumps to a real target instead of an unread query param."
        ),
        "priority_score": 51,
        "impact_score": 56,
        "effort": "Medium",
        "target": "components/MobileNav.tsx, components/MainNav.tsx, app/layout.tsx, app/guides/page.tsx",
        "dependencies": "None",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "feat-animated-interactive-mobile-menu",
        "type": "Feature",
        "pillar": "UX",
        "title": "Animated, interactive mobile menu with click-outside-to-close",
        "description": (
            "Direct UX/interaction-design request, not a scraping pass. Rewrote "
            "components/MobileNav.tsx from a native <details> element (no animation, no "
            "click-outside-to-close) to a controlled client component: hamburger icon morphs "
            "into an X on open; a blurred backdrop closes the menu on click or Escape; the "
            "panel and its sections slide/fade in with a staggered delay reusing the same "
            "transition idiom as components/ScrollReveal.tsx; Providers/Postcode/Guides/Tools "
            "became single-open accordions using a CSS-only grid-template-rows 0fr->1fr "
            "transition (no JS height measurement); the menu auto-closes on route change via "
            "usePathname(); background scroll is locked while open; focus moves into the panel "
            "on open and returns to the trigger on Escape. Global prefers-reduced-motion "
            "handling already in app/globals.css applies automatically. Not visually verified "
            "on a real device — no browser automation tool available this session, validated "
            "via build output, generated CSS and server-rendered markup only."
        ),
        "priority_score": 50,
        "impact_score": 54,
        "effort": "Medium",
        "target": "components/MobileNav.tsx",
        "dependencies": "None",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "feat-contact-form-email-delivery",
        "type": "Feature",
        "pillar": "Infrastructure",
        "title": "Working contact form delivering to a real inbox, no third-party service",
        "description": (
            "Previously /contact only offered mailto: links to unverified custom-domain "
            "addresses (editorial@/partnerships@/hello@broadbandpicker.co.uk). Built a real "
            "contact form (components/ContactForm.tsx) posting to a new "
            "app/api/contact route, which sends via Gmail SMTP using nodemailer — no "
            "Supabase, no database, no third-party form/CRM service, per explicit "
            "instruction. All submissions land in sayedsahil.elt@gmail.com (CONTACT_TO_EMAIL "
            "env var), sent from a Gmail account authenticated with an App Password "
            "(CONTACT_SMTP_USER / CONTACT_SMTP_PASS). Reply-To is set to the sender's own "
            "address so replying goes straight back to them. Spam mitigation without any "
            "external service: a hidden honeypot field (bots that fill every input get "
            "silently dropped) and a submit-time check rejecting anything sent under 1.5s "
            "after the form rendered. Server-side validates name/email/reason/message "
            "independently of the client. The original mailto: cards are kept underneath as "
            "a secondary direct-email option. Requires CONTACT_SMTP_USER/PASS to be set in "
            "Vercel — the API route fails gracefully with a friendly error and a server log "
            "line if they're absent, rather than silently dropping messages."
        ),
        "priority_score": 55,
        "impact_score": 50,
        "effort": "Low",
        "target": "app/api/contact/route.ts, components/ContactForm.tsx, app/contact/page.tsx",
        "dependencies": "CONTACT_SMTP_USER / CONTACT_SMTP_PASS Gmail App Password set in Vercel env vars",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
        "initial_status": "In progress",
    },
    {
        "item_id": "feat-postcode-journey-personalisation",
        "type": "Feature",
        "pillar": "UX",
        "title": "Postcode-aware personalisation across the customer journey, free tier only",
        "description": (
            "User asked for a postcode -> select-your-house-number address picker with "
            "customised content, and explicitly asked whether it could be built free. "
            "Researched live (WebSearch, 2026-08-23): full UK address-level lookup (the data "
            "behind 'pick your house from a list') is Royal Mail's licensed PAF data, not open "
            "-- getAddress.io, a popular free-tier provider, was shut down in Oct 2025 after a "
            "Royal Mail/IDDQD IP claim; Ideal Postcodes gives a 1-month/50-credit trial then "
            "pay-as-you-go; no provider offers a genuine unlimited free tier at real traffic. "
            "Separately, even a working address picker wouldn't improve accuracy without a paid "
            "per-address line-checker API (Openreach or provider-specific), which is a "
            "commercial relationship, not a script. Presented both findings and 3 options to the "
            "user via AskUserQuestion; they chose the free path: deepen personalisation using "
            "data already held, no new API, no new cost. Shipped: extended the existing "
            "components/PostcodeContextBar.tsx (previously only on deals/compare/provider "
            "pages) to guide pages and all 3 tool pages (speed test, broadband match, cost "
            "calculator) so a stored postcode follows the visitor through the full journey; new "
            "components/ProviderAvailabilityBadge.tsx shows a real yes/no ('BT is/isn't listed "
            "as available in {town}') on provider review pages, but only for the 51 postcode "
            "areas with real availableProviders data in data/postcodes.ts -- silently renders "
            "nothing outside that set rather than guessing; new "
            "components/ReturningVisitorBanner.tsx greets a returning visitor on the homepage "
            "with a direct link back to their own postcode's deals."
        ),
        "priority_score": 53,
        "impact_score": 55,
        "effort": "Medium",
        "target": (
            "components/ProviderAvailabilityBadge.tsx, components/ReturningVisitorBanner.tsx, "
            "components/PostcodeContextBar.tsx, app/page.tsx, app/providers/[slug]/page.tsx, "
            "app/guides/[slug]/page.tsx, app/speed-test/page.tsx, app/tools/broadband-match/page.tsx, "
            "app/tools/broadband-cost-calculator/page.tsx"
        ),
        "dependencies": "None -- deliberately built on data already held, no new API/cost",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "ops-awin-publisher-api-integration",
        "type": "Tooling",
        "pillar": "Monetisation",
        "title": "Direct Awin Publisher API access — programme status and link generation",
        "description": (
            "User asked to connect Claude to their Awin account and provided a real API "
            "token. There's no one-click connector for Awin (unlike Google Drive/Gmail), "
            "so this uses Awin's own Publisher API directly. First live run corrected two "
            "wrong assumptions from documentation research: 'programmedetails' requires a "
            "specific advertiserId (not a bulk listing endpoint) -- the actual bulk-status "
            "endpoint is GET /publishers/{id}/programmes?relationship=X, and 'any' is not "
            "a valid value for it (loops joined/pending/suspended/rejected instead; "
            "'notjoined' returns the ~21k whole-platform catalogue so it's excluded from "
            "the default and only fetched on request). POST /publishers/{id}/"
            "linkbuilder/generate confirmed working, verified against TalkTalk (advertiser "
            "3674). REAL FINDING from the first run, both surfaced to the user and worth "
            "acting on: only 4 programmes are joined (TalkTalk, Broadband Genie, Zzoomm, "
            "Highland Broadband), 7 pending (Sky ROI, Community Fibre, National Broadband, "
            "Trooli, Pine Media, Cuckoo, Zen Internet), and 9 REJECTED including BT "
            "(both consumer and business), Vodafone, Plusnet, EE, Virgin Media, Hyperoptic "
            "and toob. Cross-checked data/providers.ts: none of the currently-live "
            "affiliateUrl values (including TalkTalk, which IS approved) use Awin's "
            "awin1.com tracking format -- they're all plain provider URLs, so the site is "
            "very likely not earning commission on any current outbound click, joined or "
            "not, despite the on-page commercial disclosures implying it might."
        ),
        "priority_score": 47,
        "impact_score": 45,
        "effort": "Low",
        "target": "scripts/awin_sync.py",
        "dependencies": "AWIN_API_TOKEN from the user's Awin account (user menu -> API Credentials)",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
        "initial_status": "In progress",
    },
    {
        "item_id": "content-awin-approved-provider-deep-content",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep, researched content for Awin-approved providers (TalkTalk, Zzoomm, Highland Broadband)",
        "description": (
            "Following the Awin programme-status pull (ops-awin-publisher-api-integration), "
            "built full researched content for the 3 real providers among the 4 joined "
            "programmes (Broadband Genie is a comparison site, not an ISP, and was excluded). "
            "Per docs/page-build-pipeline-brief.md Stage 3/4/4a: for each provider, fetched "
            "the provider's own site for current pricing/speeds/contract terms, WebSearched "
            "Trustpilot and, where relevant, Ofcom's own complaints data, and wrote an 8-section "
            "contentSections block plus FAQs matching the site's deepest existing template "
            "(Gigaclear). All copy hand-checked for zero em dashes and zero banned AI-tell "
            "vocabulary/phrases per the Stage 4a list. TalkTalk's existing entry was materially "
            "stale (Trustpilot claimed 2.8, real current figure 1.5 from 50k+ reviews in "
            "Trustpilot's 'Bad' band; Ofcom's Q1 2026 report names TalkTalk the UK's most "
            "complained-about broadband provider for the third consecutive quarter; the old "
            "'price-lock guarantee' highlight was false, two scheduled April 2027/2028 rises "
            "are now built into every contract) -- fully rewritten, not just extended. Zzoomm "
            "and Highland Broadband were brand-new Provider entries (real Awin tracking links, "
            "placeholder text-wordmark logos per the brief's guardrail against fabricating brand "
            "marks). In passing, corrected Plusnet's stale trustpilotScore field (3.9 -> 2.0, "
            "the real current figure) since new comparison copy was about to cite it -- Plusnet's "
            "own deep content rewrite is out of scope (not an Awin-approved programme). Two new "
            "comparison pages: talktalk-vs-plusnet (a genuinely stark, well-evidenced contrast -- "
            "same price bracket, opposite ends of Ofcom's complaints table) and "
            "zzoomm-vs-hyperoptic (two symmetrical full-fibre altnets with almost non-overlapping "
            "coverage footprints). Internal linking relies on the existing 'How {provider} "
            "compares' module built earlier this session; both new comparisons are discoverable "
            "via /providers/compare and the sitemap even where they don't make a crowded "
            "provider's top-3 related slice."
        ),
        "priority_score": 58,
        "impact_score": 62,
        "effort": "High",
        "target": (
            "data/providers.ts (talktalk, zzoomm, highland-broadband, plusnet trustpilotScore), "
            "data/provider-comparisons.ts (talktalk-vs-plusnet, zzoomm-vs-hyperoptic), "
            "public/logos/zzoomm.svg, public/logos/highland-broadband.svg"
        ),
        "dependencies": "ops-awin-publisher-api-integration",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-awin-pending-provider-deep-content",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep, researched content for Awin-pending providers (Community Fibre, Cuckoo, Zen Internet, National Broadband, Trooli, Pine Media)",
        "description": (
            "Follow-up to content-awin-approved-provider-deep-content, extending the same "
            "research process to the 6 real UK-relevant advertisers pending Awin approval "
            "(Sky ROI excluded -- Republic of Ireland Sky, not UK, out of scope for a UK site, "
            "confirmed with the user). Community Fibre and Zen Internet were existing but thin "
            "entries, fully rewritten. Cuckoo was already deep and dated the previous day; "
            "verified rather than rewritten. National Broadband, Trooli and Pine Media were "
            "brand-new Provider entries. National Broadband is the site's first 4G/5G "
            "fixed-wireless provider (a genuinely new content category, not another fibre "
            "altnet); Pine Media is a hyperlocal single-city (Sheffield-only) provider with a "
            "two-tier product split (its own symmetrical GIG network vs an Openreach-based GLO "
            "product) that needed explaining clearly to avoid conflating the two. Key design "
            "decision: because these programmes are pending, not joined, affiliateUrl stays as "
            "each provider's own plain site URL rather than an Awin tracking link -- confirmed "
            "via a live API test that Awin's linkbuilder will generate a structurally valid "
            "tracking URL even for a pending programme, which would be misleading to publish "
            "since it likely would not earn commission (or could breach Awin's terms) before "
            "approval. Each reviewSources array carries an explicit 'pending, not yet approved' "
            "note on the Awin source, matching the existing Gigaclear convention. All copy "
            "hand-checked for zero em dashes and zero banned AI-tell vocabulary/phrases. Three "
            "new comparison pages, chosen for genuine evidentiary value over one-per-provider "
            "completeness: community-fibre-vs-hyperoptic (London's two biggest symmetrical "
            "altnets, direct competitors), trooli-vs-zzoomm (two multi-region altnets with "
            "near-zero coverage overlap and matching no-price-rise policies), and "
            "national-broadband-vs-highland-broadband (5G available now vs fibre still being "
            "built -- a genuine buy-now-or-wait decision for the same rural Scottish "
            "customer). Cuckoo, Zen Internet and Pine Media were deliberately left without a "
            "new comparison this round -- no natural, non-forced head-to-head partner existed "
            "for any of them without inventing artificial relevance."
        ),
        "priority_score": 55,
        "impact_score": 58,
        "effort": "High",
        "target": (
            "data/providers.ts (community-fibre, cuckoo, zen-internet, national-broadband, "
            "trooli, pine-media), data/provider-comparisons.ts (community-fibre-vs-hyperoptic, "
            "trooli-vs-zzoomm, national-broadband-vs-highland-broadband), "
            "public/logos/national-broadband.svg, public/logos/trooli.svg, "
            "public/logos/pine-media.svg"
        ),
        "dependencies": "content-awin-approved-provider-deep-content, ops-awin-publisher-api-integration",
        "source": "User request 2026-08-23 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-awin-rejected-provider-deep-content",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep, researched content for Awin-rejected providers (BT, Virgin Media, EE, Vodafone, Plusnet, Hyperoptic, toob, giffgaff)",
        "description": (
            "Third and final batch in the Awin content series (after approved and pending), "
            "covering all 8 real providers rejected on Awin: BT (both consumer 3041 and "
            "business 3042 programmes), Virgin Media, EE, Vodafone, Plusnet, Hyperoptic, toob "
            "and giffgaff. All 7 existing entries were thin (no contentSections) and fully "
            "rewritten; giffgaff did not exist on the site at all and was added new. "
            "IMPORTANT CORRECTION made mid-batch: while researching BT's price rise, "
            "discovered Ofcom banned inflation-linked/percentage-based mid-contract price rise "
            "terms in all new contracts from 17 January 2025 (providers must now disclose a "
            "flat pounds-and-pence figure instead). This retroactively invalidated 5 sentences "
            "written in the two earlier Awin batches (Community Fibre, Zen Internet and Zzoomm "
            "entries in data/providers.ts, plus the Community Fibre vs Hyperoptic comparison) "
            "that described CPI-linked percentage rises as a current practice -- all 5 were "
            "corrected in this same batch before continuing, and this correction is itself the "
            "kind of thing worth surfacing: a fact later research revealed to be wrong wasn't "
            "left standing. Real findings surfaced across this batch: BT and EE both maintain "
            "two separate Trustpilot pages (a mobile/brand-dominated headline score and a much "
            "lower broadband-specific one, roughly 4.0 vs 1.5 for BT, 4.2 vs 1.3 for EE) -- the "
            "broadband-specific figure was used as the primary trustpilotScore field in both "
            "cases, with the discrepancy explained in the copy since it's a genuinely useful, "
            "non-obvious insight. Plusnet's low Trustpilot score sits alongside the *best* "
            "Ofcom complaints record of any major UK provider (4 per 100,000 in Q1 2026, "
            "against TalkTalk's 10, Vodafone's 8 and BT's 7) -- a real, evidenced gap between "
            "self-selected review-platform sentiment and regulatory complaint-volume data, "
            "explained rather than picking one source and ignoring the other. Virgin Media "
            "shows the same pattern in reverse: an extremely low Trustpilot score (~1.4) next "
            "to an Ofcom complaints record at or below the industry average. giffgaff is a "
            "genuinely new product (broadband launched September 2025 on Nexfibre, the Virgin "
            "Media O2 wholesale network, not Openreach) with no separate broadband Trustpilot "
            "page yet -- flagged honestly as thin evidence rather than borrowing giffgaff's "
            "much larger mobile-customer score uncritically. All affiliateUrl values stay as "
            "each provider's own plain site URL, since Awin rejected these applications outright "
            "(not merely pending) -- every reviewSources array states this explicitly. Three new "
            "comparison pages targeting genuine content gaps rather than one-per-provider "
            "completeness (this batch's providers already had extensive existing head-to-head "
            "coverage from earlier site work): bt-vs-hyperoptic (safe wide-coverage default vs "
            "better-value niche upgrade), plusnet-vs-ee (same BT Group parent, opposite ends of "
            "Ofcom's complaints table), toob-vs-giffgaff (two newer no-price-rise challenger "
            "brands on completely different networks). All copy hand-checked for zero em dashes "
            "and zero banned AI-tell vocabulary/phrases."
        ),
        "priority_score": 60,
        "impact_score": 65,
        "effort": "High",
        "target": (
            "data/providers.ts (bt, virgin-media, ee, vodafone, plusnet, hyperoptic, toob, "
            "giffgaff, plus corrections to community-fibre, zen-internet, zzoomm), "
            "data/provider-comparisons.ts (bt-vs-hyperoptic, plusnet-vs-ee, toob-vs-giffgaff, "
            "plus a correction to community-fibre-vs-hyperoptic), public/logos/giffgaff.svg"
        ),
        "dependencies": "ops-awin-publisher-api-integration, content-awin-approved-provider-deep-content, content-awin-pending-provider-deep-content",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-broadband-tv-deals-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Research-led editorial refresh: Best Broadband and TV Deals guide",
        "description": (
            "User-requested application of the same Stage 3/4/4a research process from the "
            "Awin content batches to a single existing guide page, "
            "/guides/best-broadband-and-tv-deals. Unlike the thin provider stubs rewritten in "
            "the Awin batches, this guide was already well-built (tables, methodology, 7 FAQs, "
            "real Ofcom citations, last checked 29 July 2026) -- not a stub needing a ground-up "
            "rewrite, so this was a verify-and-deepen pass grounded in fresh research rather "
            "than a full replacement. Scraped 2 top-ranking competitors (MoneySuperMarket, "
            "Uswitch) per Stage 3.0: found MoneySupermarket runs a 15-question FAQ (ours had "
            "7) and leads with an expert-quoted bundling-savings statistic. Traced that claim "
            "to its actual primary source rather than citing the secondary restatement: Ofcom's "
            "own February 2026 research, savings of £26-£48/month from bundling and pay-TV "
            "averaging £12/month within a bundle (23% real-terms fall) -- a stronger, more "
            "current, primary-sourced figure than the competitor's 2023-dated stat, added as a "
            "new section. Also found and added two genuinely new, specific facts the page "
            "didn't have: Virgin Media's Max Volt bundle (real current pricing including its "
            "two scheduled rises to April 2028) as a concrete worked example, and Virgin "
            "Media's 'Most Reliable Broadband Provider' win at the 2026 Uswitch Telecoms "
            "Awards (Opensignal-verified, a genuine third-party measurement). Named the "
            "specific Sky bundle example (Sky Stream, Sky TV, Netflix + Full Fibre 300, "
            "£35/month) that the existing copy referenced only vaguely. Added 2 new FAQs tied "
            "to the new content. Fixed 2 pre-existing em dashes found during the pass (a "
            "legitimate en-dash numeric range, '50-100Mbps', was left alone -- not an AI-tell, "
            "different typographic use). Updated the page's stated fact-check date and "
            "guides.ts updatedDate to 2026-08-24."
        ),
        "priority_score": 48,
        "impact_score": 45,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-broadband-and-tv-deals), data/guides.ts",
        "dependencies": "None",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "ops-content-priority-analysis-tooling",
        "type": "Tooling",
        "pillar": "Content",
        "title": "Content priority analysis: real SEO + GEO signals, build vs. update on one list",
        "description": (
            "User asked for a script that does detailed SEO analysis of the site's own data "
            "and prioritises what to build vs. update, weighting generative-AI/GEO reports "
            "specifically. Built scripts/analyze_content_priority.py: combines real mapped "
            "search volume already researched in docs/broadbandpicker-keyword-mapping.xlsx "
            "with a live crawl of every provider/comparison/guide page (word count against "
            "this session's own depth floors, FAQPage schema, a detectable answer-first "
            "paragraph, a parsed checked/reviewed date). Mid-build discovered the user's "
            "actual request: a real Google Search Console 'Performance on Search Generative "
            "AI Features' export had been dropped into data/GSC/, which is ground truth for "
            "AI Overview surfacing, not a proxy -- rebuilt the script to load and heavily "
            "weight it. Caught and fixed two real bugs before reporting any findings: word "
            "counts of 15k-17k+ caused by the React RSC hydration payload inside <script> tags "
            "being counted as visible text, and every page falsely flagging 'no answer-first "
            "paragraph' because the <p> search matched header mega-menu labels before reaching "
            "the main content -- both fixed by parsing with lxml, stripping script/style, and "
            "scoping to <main>. Real finding once the GSC data was wired in: "
            "/guides/best-full-fibre-broadband-uk earned 512 real AI-feature impressions in the "
            "last 28 days while sitting at 636 words, the clearest 'already working, just needs "
            "depth' signal on the site -- now the top-ranked refresh target, above even "
            "/providers/sky's 172,237 monthly search volume, since Sky currently earns almost "
            "no AI-feature traction (4 impressions) and represents a different kind of "
            "opportunity. Sky also does not appear in any of the 3 Awin relationship lists at "
            "all (joined/pending/rejected), confirmed via a direct re-check -- no record of "
            "ever applying. Every row in the Content Gap Roadmap is already marked built "
            "(33/33), so there are currently no new-page gaps on the existing curated keyword "
            "list. Findings written up in docs/content-priority-analysis.md."
        ),
        "priority_score": 56,
        "impact_score": 60,
        "effort": "Medium",
        "target": "scripts/analyze_content_priority.py",
        "dependencies": "docs/broadbandpicker-keyword-mapping.xlsx (existing keyword research)",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-full-fibre-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Full Fibre Broadband UK guide (top content-priority-analysis target)",
        "description": (
            "Top-ranked target from scripts/analyze_content_priority.py's first real run: "
            "512 AI-feature impressions in the last 28 days (real GSC data) at only 636 words. "
            "Fully rewritten to 1,596 words, drawing on already-verified provider facts from "
            "the three Awin content batches (BT, EE, Vodafone, Plusnet, Community Fibre, "
            "Hyperoptic, toob, Zen Internet) plus new research for Sky and a genuinely "
            "counter-intuitive finding worth surfacing on its own: every major national "
            "provider has a low Trustpilot score, but Ofcom's Q1 2026 complaints data "
            "separates them clearly (Plusnet best at 4/100k, TalkTalk worst at 10/100k) -- "
            "explained as two measures of different things, not a contradiction. Also "
            "verified and included Vodafone's real 2026 award wins (Expert Reviews Broadband "
            "Awards, Uswitch Most Popular Provider) alongside its poor Ofcom complaints "
            "position, rather than picking one narrative. Caught and fixed a real bug before "
            "shipping: wrote several apostrophes as literal backslash-escapes ( \\' ) copying "
            "the .ts string-literal habit from data/providers.ts into JSX text content, where "
            "they would have rendered as a visible backslash -- fixed to the file's existing "
            "&apos; convention and verified against the live rendered HTML, not just the build "
            "passing, before considering it done."
        ),
        "priority_score": 54,
        "impact_score": 58,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-full-fibre-broadband-uk), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-cheapest-broadband-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Cheapest Broadband Deals UK guide (2nd content-priority-analysis target)",
        "description": (
            "Second target from scripts/analyze_content_priority.py's ranked list: 130 real "
            "GSC AI-feature impressions at 481 words, and materially stale (last touched "
            "2026-06-01, still citing NOW Broadband at a since-changed £17.99). Rewritten to "
            "1,101 words, drawing on session-verified pricing across 11 providers rather than "
            "fresh research for most of them. Real correction made: NOW Broadband's usable "
            "Full Fibre 75 now starts at £23/mo, not the stale £17.99 the old copy cited -- no "
            "longer reliably the cheapest option, now similar to or pricier than Plusnet "
            "(£21.99) and EE (£22.99). Real new finding: Community Fibre's Essential 35 at "
            "£12.50/mo is the genuinely cheapest full-fibre deal on the market (London/Surrey/"
            "Sussex only). Verified current social tariff pricing (Virgin Media £12.50, BT "
            "Home Essentials £15-20, Sky Broadband Basics £20 existing-customers-only) plus a "
            "real, citable Ofcom stat: ~4.2 million eligible households, only ~532,000 "
            "actually claiming one. Caught the same JSX-vs-.ts apostrophe-escaping bug found "
            "on the previous guide (one instance this time, in the NOW Broadband paragraph) "
            "before shipping -- confirmed via the file's own reviewSources-style checked-date "
            "convention. 1,101 words sits just under this session's own 1,200-word guide "
            "depth floor; left as is rather than padding for its own sake, since the content "
            "is already comprehensive (2 real data tables, 5 FAQs, all newly sourced)."
        ),
        "priority_score": 52,
        "impact_score": 55,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (cheapest-broadband-uk), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-business-broadband-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Business Broadband Providers UK guide (3rd content-priority-analysis target)",
        "description": (
            "Third target from scripts/analyze_content_priority.py's ranked list: 142 real "
            "GSC AI-feature impressions at 596 words. This page lives on a different template "
            "than the [slug]-based guides (app/guides/best-business-broadband-providers-uk/ "
            "is a standalone PrioritySeoPage route backed by data/priority-pages.ts), and had "
            "2 pre-existing em dashes fixed as part of the pass. Rewritten to 1,050 words with "
            "real, specific per-provider data: Vodafone Business has the most competitive "
            "entry-level price (~£20-22/mo excl. VAT), Zen Business leads on satisfaction "
            "(Which? Recommended since 2021, 84% score) -- consistent with, and corroborating, "
            "the consumer Zen Internet findings from the Awin-rejected batch, Virgin Media "
            "Business has the fastest raw speed (442 Mbps average), and a genuinely useful, "
            "specific stat: out-of-contract business customers overpay 24.86% more per month "
            "on average. Added a real contended-vs-uncontended leased-line explainer with "
            "current pricing (from ~£69/mo entry-level). Also fixed a real bug in the analysis "
            "script itself found while validating this page: DATE_RE didn't recognise the "
            "PrioritySeoPage template's 'Last researched and reviewed:' phrasing (colon, "
            "different verb), so every page on that template was incorrectly scored as having "
            "no checked date -- fixed and confirmed it also corrected the score for "
            "/guides/satellite-broadband-uk on the same template."
        ),
        "priority_score": 50,
        "impact_score": 52,
        "effort": "Medium",
        "target": "data/priority-pages.ts (business), scripts/analyze_content_priority.py",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-5g-home-broadband-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best 5G Home Broadband UK guide (4th content-priority-analysis target)",
        "description": (
            "Fourth target from scripts/analyze_content_priority.py's ranked list: 30 real "
            "GSC AI-feature impressions at 601 words, and the old copy named no real "
            "providers or prices at all, just generic 5G-vs-fibre advice. Rewritten to 1,036 "
            "words with real current data for all 4 major UK 5G home broadband options: Three "
            "(best value, from GBP29/mo, ~150 Mbps), Vodafone GigaCube (cheapest fixed term at "
            "GBP21/mo, or a genuine GBP60/mo no-contract option), EE Smart 5G Hub (broadest UK "
            "5G coverage, GBP30-50/mo), and National Broadband (already deeply covered in the "
            "Awin-pending batch) framed correctly here as a multi-network specialist rather "
            "than a single-network alternative -- it connects to whichever of the 4 UK "
            "networks is strongest at a given address, a genuinely different value "
            "proposition worth explaining rather than just listing as a fourth option. Found "
            "and included a real Trustpilot-vs-Ofcom divergence for Three specifically (4.5 "
            "Trustpilot, worst-quartile Ofcom complaints alongside EE and Vodafone) and linked "
            "to the full-fibre guide's deeper explanation of that pattern rather than "
            "repeating it in full."
        ),
        "priority_score": 47,
        "impact_score": 46,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-5g-home-broadband-uk), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-phone-and-broadband-deals-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Phone and Broadband Deals guide (5th content-priority-analysis target)",
        "description": (
            "Fifth target from scripts/analyze_content_priority.py's ranked list: the old "
            "copy was generic Digital Voice/line-rental advice with no dates, no stats and "
            "one pre-existing em dash, at 457 words. Rewritten with the real PSTN "
            "switch-off timeline: full shutdown now locked in for 31 January 2027 (moved "
            "from the original December 2025 target after Openreach confirmed the "
            "technical barriers were resolved), with PSTN customer numbers falling from "
            "5.2 million (July 2024) to 3.2 million (July 2025). Added Ofcom's minimum "
            "one-hour battery backup requirement for emergency-services access during a "
            "power cut, and the industry PSTN Charter's commitment not to migrate "
            "telecare/vulnerable users without confirmed-compatible equipment. Added a "
            "real, specific call-plan cost example (TalkTalk's Anytime Calls add-on at "
            "GBP12/mo on top of base broadband) to replace the previous fully generic "
            "'compare the whole contract' advice. FAQs expanded from 3 to 5. Fixed the "
            "pre-existing em dash in the dek field."
        ),
        "priority_score": 45,
        "impact_score": 44,
        "effort": "Medium",
        "target": "data/priority-pages.ts (phone)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-sky-provider-page",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Sky provider page (6th content-priority-analysis target, first provider-type page)",
        "description": (
            "Sixth target from scripts/analyze_content_priority.py's ranked list, and the "
            "first provider-page (not guide-page) target: /providers/sky carries the "
            "single highest mapped search volume of any page on the site (172,237/mo) but "
            "was only 387 words with no excerpt or contentSections at all -- the sky entry "
            "in data/providers.ts was missing those two fields entirely, unlike the "
            "already-rewritten bt and virgin-media entries. Rewritten to ~1,700 words with "
            "6 contentSections (deals, Openreach network/speeds, the April 2026 flat GBP3 "
            "price rise -- Sky's first year using pounds-and-pence instead of a percentage, "
            "down from 6.2% in 2025 -- Sky Broadband Shield/TV bundles, a "
            "Trustpilot-vs-Ofcom section, and a final verdict), plus 4 FAQs and full "
            "reviewSources. Corrected 3 stale top-level fields that were actively wrong: "
            "trustpilotScore 3.8 -> 2.7 (cross-checked against this site's own existing "
            "bt-vs-sky and sky-vs-vodafone comparison pages, which already used 2.7 as of "
            "21 August 2026 -- caught and fixed a first-pass error of my own where a single "
            "WebFetch summary suggested 1.4, before corroborating against two independent "
            "sources and this site's own prior content), contractLengths [18] -> [24] "
            "(Sky has not offered an 18-month contract for some time), and monthlyPriceFrom "
            "25.00 -> 23.00 to match the real current cheapest tier (Superfast, 67 Mbps)."
        ),
        "priority_score": 66,
        "impact_score": 60,
        "effort": "Medium",
        "target": "data/providers.ts (sky)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-now-broadband-provider-page",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: NOW Broadband provider page (7th content-priority-analysis target)",
        "description": (
            "Seventh target from scripts/analyze_content_priority.py's ranked list: "
            "/providers/now-broadband was 386 words with no excerpt or contentSections, "
            "and its top-level fields were stale on two counts that research corrected -- "
            "contractLengths was still [12] (NOW dropped its flexible 12-month contracts; "
            "every current package is now a standard 24-month term, per thinkbroadband's "
            "dedicated news item on the change) and monthlyPriceFrom was a stale GBP17.99 "
            "against a real current cheapest tier of GBP23 (Full Fibre 100). Rewritten to "
            "~1,450 words with 5 contentSections. The genuinely distinctive finding here, "
            "different from every other Trustpilot-vs-Ofcom section written this session: "
            "NOW is a real exception to the usual pattern, scoring poorly on BOTH Trustpilot "
            "(1.2/5, corrected from a stale 3.2) AND Ofcom's Q4 2025 complaints data (11 per "
            "100k vs the 8 industry average), noticeably worse than sister brand Sky on the "
            "same underlying Openreach network -- written up as a genuine data point about "
            "support quality diverging from infrastructure quality within the same corporate "
            "group, not just review-platform noise. Also corrected trustpilotScore 3.2 -> 1.2."
        ),
        "priority_score": 44,
        "impact_score": 40,
        "effort": "Medium",
        "target": "data/providers.ts (now-broadband)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-youfibre-provider-page",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: YouFibre provider page (8th content-priority-analysis target)",
        "description": (
            "Eighth target from scripts/analyze_content_priority.py's ranked list: "
            "/providers/youfibre was 397 words with no excerpt or contentSections. "
            "Rewritten to ~1,575 words with 6 contentSections. The standout finding, "
            "verified against a primary Virgin Media O2 press release after a first-pass "
            "WebSearch AI summary over-stated it as a done deal: YouFibre's parent company, "
            "Substantial Group (Netomnia), is being acquired for GBP2bn by nexfibre (a joint "
            "venture of InfraVia, Liberty Global and Telefonica), which then plans to resell "
            "the YouFibre/Brsk retail brands to Virgin Media O2 for GBP150m -- announced Feb "
            "2026, referred to the CMA for a Phase 2 investigation on 1 July 2026, and NOT "
            "yet completed (expected ~Q3 2026 at the earliest). Also wrote up the real, "
            "documented fallout from the March 2026 Brsk-to-YouFibre customer migration "
            "(billing errors, login/password-reset failures, reported speed drops, "
            "overwhelmed support chat), sourced from ISPreview rather than the vague "
            "pre-existing 'Trustpilot sentiment has been mixed' line. Updated the speeds "
            "array to the real current tiers (200/1000/2000/8000 Mbps symmetrical, "
            "replacing stale 200/900/1800) and contractLengths from [1,12,24] to [1,24] "
            "to match the two options confirmed by current pricing research."
        ),
        "priority_score": 43,
        "impact_score": 39,
        "effort": "Medium",
        "target": "data/providers.ts (youfibre)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-broadband-providers-uk-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Broadband Providers UK ranking guide (9th content-priority-analysis target)",
        "description": (
            "Ninth target from scripts/analyze_content_priority.py's ranked list. This "
            "ranking page's data was badly stale: it cited Q4 2025 Ofcom complaints "
            "figures and June-2026 Trustpilot scores that had already been superseded by "
            "this site's own freshly-researched provider pages (Sky, NOW Broadband, BT, "
            "Virgin Media, EE, Plusnet, Vodafone, TalkTalk from earlier this session). "
            "Rather than re-researching each provider from scratch, cross-referenced this "
            "page's ranking table against data/providers.ts's own already-vetted current "
            "fields (price, Trustpilot, reviewedDate) and fetched a fresh single-source "
            "Ofcom Q1 2026 complaints table (published 23 July 2026, record-low industry "
            "average of 6 per 100k) to replace the mixed-quarter data. Corrected several "
            "materially wrong claims: TalkTalk was still framed as 'budget only' at "
            "GBP19.99 when it is now GBP25 and no longer the cheapest big-name option; "
            "Community Fibre was described as 'London-only' when coverage has expanded "
            "into Surrey and Sussex; EE's 'best for reliability' halo no longer holds "
            "under Q1 2026 data (6 per 100k, tied with Virgin Media, not uniquely best). "
            "Added a new section contrasting sister brands Sky (5 per 100k) and NOW "
            "Broadband (11 per 100k, different quarter, footnoted) on the identical "
            "Openreach network, a genuine support-quality data point. Word count "
            "872 -> 1,333."
        ),
        "priority_score": 42,
        "impact_score": 38,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-broadband-providers-uk), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-broadband-without-phone-line-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Broadband Without a Phone Line guide (10th content-priority-analysis target)",
        "description": (
            "Tenth target from scripts/analyze_content_priority.py's ranked list: "
            "770 words, stale pricing table (every one of 8 providers priced against "
            "months-old figures already corrected on this site's own provider pages), and "
            "a wrong PSTN switch-off date ('by the end of 2027' -- the real, now-locked-in "
            "date is 31 January 2027). Rewrote using the same PSTN research gathered for "
            "the best-phone-and-broadband-deals refresh earlier this session (5.2m -> 3.2m "
            "migration stat, Ofcom's 1-hour power-cut battery backup minimum, the PSTN "
            "Charter's telecare protections), added a genuinely missing section -- 'What "
            "happens to a home phone during a power cut?' -- since the old version covered "
            "whether a line is needed at all but never addressed the actual practical "
            "question once Digital Voice is the default. Refreshed all 8 provider prices "
            "against data/providers.ts's own current fields, updated Virgin Media coverage "
            "52% (was 53%) and added Ofcom's Spring 2026 Connected Nations split (89% "
            "gigabit-capable, 82% full fibre, 93% urban vs 66% rural). Word count 770 -> "
            "1,196; left just under the self-imposed 1,200-word floor rather than pad."
        ),
        "priority_score": 41,
        "impact_score": 37,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (broadband-without-phone-line), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-broadband-moving-house-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Broadband When Moving House checklist guide (11th content-priority-analysis target)",
        "description": (
            "Eleventh target from scripts/analyze_content_priority.py's ranked list: 948 "
            "words, structurally sound (evergreen checklist format) but missing a genuinely "
            "newsworthy development -- Ofcom closed its One Touch Switch enforcement case "
            "on 11 June 2026 after finding more than 2 million customers had already used "
            "the process successfully, meaning it is now the industry's permanent standard "
            "rather than a monitored pilot, which the old copy still implied. Added this, "
            "plus a genuinely missing section on what a house move means for a home phone "
            "line given the 31 January 2027 PSTN switch-off (reusing research from the "
            "phone-and-broadband and broadband-without-phone-line refreshes earlier this "
            "session), Ofcom's Spring 2026 coverage split for the 'check availability' "
            "step, and corrected the ETC explanation to reflect Ofcom's actual cap rule "
            "(remaining payments, ex-VAT) rather than a flat multiplication. Re-verified "
            "the pre-existing GBP183.60/year switching-saving stat against fresh 2026 "
            "sources (found independently corroborated) and added the 4.2m/GBP100-240 "
            "social tariff stat already used elsewhere on this site. Word count 948 -> "
            "1,446."
        ),
        "priority_score": 40,
        "impact_score": 36,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (broadband-moving-house), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-satellite-broadband-uk-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Satellite Broadband UK guide (12th content-priority-analysis target)",
        "description": (
            "Twelfth target from scripts/analyze_content_priority.py's ranked list, and "
            "the page whose DATE_RE regex bug was found and fixed earlier this session "
            "(58.2 -> 53.2) without ever getting a content rewrite. The old copy was "
            "generic to the point of naming no real prices at all -- 'Starlink is the "
            "most visible option' with zero named plans, figures or alternatives. "
            "Rewritten with real, current Starlink UK pricing (Residential 100/200/Max "
            "at GBP40/60/80 per month, Roam at GBP55-100, Standard Kit GBP449 commonly "
            "discounted to ~GBP299 or free on Max), and a genuinely current news item: "
            "Starlink's Global Roam plan stopped taking new customers 15 July 2026 and "
            "was withdrawn entirely for existing customers on 17 August 2026, one week "
            "before this rewrite. Named 2 real smaller alternatives (SkyDSL, Bigblu "
            "Broadband) and clarified that Eutelsat/OneWeb, despite frequent mention "
            "alongside Starlink, sells no UK consumer home-broadband plans at all. Word "
            "count 437 -> 1,038."
        ),
        "priority_score": 39,
        "impact_score": 35,
        "effort": "Medium",
        "target": "data/priority-pages.ts (satellite)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-broadband-for-gaming-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Broadband for Gaming UK guide (13th content-priority-analysis target)",
        "description": (
            "Thirteenth target from scripts/analyze_content_priority.py's ranked list: "
            "836 words with stale provider prices (EE GBP26.99, Community Fibre GBP21.99, "
            "Hyperoptic GBP22, BT GBP30.99, all already corrected on this site's own "
            "provider pages) and a wrong Community Fibre top-tier speed claim (920Mbps "
            "symmetrical, against the real current 3,000Mbps top tier confirmed in "
            "data/providers.ts). Refreshed all 5 existing provider picks against already-"
            "vetted current pricing, and added a genuinely new 6th pick, Zen Internet, for "
            "a use case the old page didn't cover at all: hosting a private game server, "
            "where Zen's free static IP (already documented on its own provider page) is a "
            "real, specific advantage over every other provider on the list. Added a short "
            "section on how 24-month contracts and scheduled price rises affect the real "
            "two-year cost of a gaming-suitable package. Word count 836 -> 1,184; left "
            "just under the self-imposed 1,200-word floor rather than pad."
        ),
        "priority_score": 38,
        "impact_score": 34,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-broadband-for-gaming-uk), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-broadband-social-tariffs-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Broadband Social Tariffs UK guide (14th content-priority-analysis target)",
        "description": (
            "Fourteenth target from scripts/analyze_content_priority.py's ranked list: "
            "777 words with a materially wrong 'cheapest tariff' claim -- Vodafone "
            "Together Social was listed at GBP12.50/mo for 38Mbps and named the cheapest "
            "option, but the real current Vodafone Essentials Broadband tariff is "
            "GBP20/mo for 73Mbps (verified directly against Vodafone's own tariff page), "
            "meaning Virgin Media Essential Broadband and Community Fibre Essential "
            "(both genuinely GBP12.50/mo) are actually jointly cheapest. Also corrected "
            "Virgin Media's own tariff, which the old page listed at GBP20/mo. Added a "
            "genuinely newsworthy find: a thinkbroadband report published in August 2026 "
            "(days before this rewrite) tested how easy major providers make it to find "
            "their own social tariffs and scored BT, Sky, Virgin Media and Vodafone "
            "poorly, with a pointed BT quote, against smaller altnets scoring better. "
            "Corrected the eligibility table (Sky accepts UC/PC only, existing customers "
            "only, a real restriction the old page omitted) and added KCOM's Hull-only "
            "tariff. Refreshed the awareness stat with the newer, more specific "
            "thinkbroadband figures (34% aware, 8.6% UC take-up). Word count 777 -> "
            "1,161."
        ),
        "priority_score": 37,
        "impact_score": 34,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (broadband-social-tariffs-uk), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-broadband-working-from-home-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Broadband for Working From Home guide (15th content-priority-analysis target)",
        "description": (
            "Fifteenth target from scripts/analyze_content_priority.py's ranked list: "
            "565 words, no real prices anywhere, and a stale 'EE is consistently rated "
            "the most reliable major UK broadband provider' claim that no longer holds "
            "under this session's own Q1 2026 Ofcom complaints research. Rewrote with "
            "real current prices for Community Fibre, Hyperoptic, Zen Internet and BT, "
            "and added a genuinely new section this page never had: when a business-grade "
            "line (static IP, faster fault-fix SLA) is actually worth it for a freelancer "
            "or remote worker over a residential package, using real Vodafone Business/"
            "Zen Business/Sky Business pricing already researched for this site's business "
            "broadband guide. Added concrete video-call upload-bandwidth guidance (3-4 "
            "Mbps per HD call) rather than vague 'upload matters' framing. Word count "
            "565 -> 986."
        ),
        "priority_score": 36,
        "impact_score": 33,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-broadband-for-working-from-home), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-broadband-deals-under-20-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Broadband Deals Under GBP20 guide (16th content-priority-analysis target)",
        "description": (
            "Sixteenth target from scripts/analyze_content_priority.py's ranked list, "
            "and the one with the largest factual break of this whole refresh sequence: "
            "the page's entire premise named NOW Broadband and TalkTalk as 'the most "
            "common names in this price bracket,' but both now start above GBP20/mo "
            "(NOW at GBP23, TalkTalk at GBP25, both already corrected on their own "
            "provider pages this session). Rebuilt the page around who is actually under "
            "GBP20 today by querying data/providers.ts directly for every provider with "
            "monthlyPriceFrom < 20: Community Fibre (GBP12.50), Onestream (GBP18.50, 94% "
            "Openreach coverage -- the most realistic pick for most addresses), Gigaclear "
            "(GBP19.00, rural symmetrical), toob (GBP19.50) and Trooli (GBP19.99). Also "
            "caught and excluded Shell Energy (GBP19.99 in the data file) after finding "
            "its own provider-page highlights note it closed to new customers and "
            "migrated to TalkTalk in 2024 -- an old, non-orderable plan that would have "
            "been a real error to list as a current deal. Word count 608 -> 890."
        ),
        "priority_score": 35,
        "impact_score": 42,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (broadband-deals-under-20), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-bt-vs-sky-comparison",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: BT vs Sky comparison page (17th content-priority-analysis target)",
        "description": (
            "Seventeenth target from scripts/analyze_content_priority.py's ranked list, "
            "and the first comparison-type (not guide or provider) page in this refresh "
            "sequence: 496 words of entirely generic, number-free copy ('Sky starts "
            "cheaper,' 'BT usually charges more,' 'Sky is typically shorter at 18 "
            "months') that was also factually stale -- Sky moved to 24-month contracts "
            "some time ago, the same length as BT, a real premise this page's whole "
            "'contract flexibility' differentiator was built on. Rewrote using the "
            "already-vetted current data from this session's own BT and Sky provider-page "
            "rewrites: prices have converged to within a pound (BT GBP23.99 vs Sky "
            "GBP23.00), Sky has the clearly better Ofcom Q1 2026 complaints record (5 vs "
            "7 per 100k), BT has wider coverage (98% vs 95%), and BT's flat GBP4/year "
            "March price rise is more transparently disclosed for future years than "
            "Sky's 'prices may rise' contract wording. Populated the previously-empty "
            "factSnapshot field. Word count 496 -> 717."
        ),
        "priority_score": 34,
        "impact_score": 32,
        "effort": "Medium",
        "target": "data/provider-comparisons.ts (bt-vs-sky)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-rolling-monthly-broadband-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Rolling Monthly Broadband Deals guide (18th content-priority-analysis target)",
        "description": (
            "Eighteenth target from scripts/analyze_content_priority.py's ranked list: "
            "522 words repeating the same stale claim caught twice already this session "
            "-- 'NOW Broadband is often relevant because of its shorter 12-month "
            "positioning,' when NOW dropped flexible contracts entirely and now sells "
            "only a standard 24-month term. This page's whole premise rested on that "
            "wrong claim, so it needed a real replacement for 'who actually offers "
            "flexible broadband now,' not just a correction. Rewrote around YouFibre, "
            "confirmed via its own already-rewritten provider page as one of the only "
            "genuine rolling-monthly full-fibre options left (from GBP33.99/mo, 30 days' "
            "notice, vs GBP20/mo on its own 24-month term), plus Hyperoptic/Community "
            "Fibre/Onestream's 12-month middle-ground option, and a note on social "
            "tariffs' no-exit-fee terms. Word count 522 -> 907."
        ),
        "priority_score": 33,
        "impact_score": 31,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-rolling-monthly-broadband-deals), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-broadband-price-rises-2026-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Broadband Price Rises 2026 guide (19th and final content-priority-analysis target this session)",
        "description": (
            "Nineteenth target from scripts/analyze_content_priority.py's ranked list. "
            "This page had the most individually-wrong data points found in a single "
            "page this session: TalkTalk was listed at +GBP2.50/mo (real current figure, "
            "verified via ISPreview, is +GBP4.00/mo for contracts from 16 November 2025 "
            "-- TalkTalk raised its own rise from GBP3 to GBP4 partway through the year); "
            "Hyperoptic and Community Fibre were both listed as 'No rise' when both "
            "apply a real scheduled increase (~GBP4/mo and a capped GBP2/mo respectively, "
            "confirmed against their own already-researched provider pages). Also caught "
            "and fixed a bug introduced earlier THIS session: the NOW Broadband provider "
            "page wrongly claimed NOW mirrors Sky's price rise at a flat GBP4/mo, "
            "corrected first in data/providers.ts (5 instances) before writing this page, "
            "to GBP0 in April 2026 with GBP3 scheduled for April 2027 and 2028. Added a "
            "dedicated section directly naming the 'no-rise altnet' assumption as false "
            "for 2 of the 4 providers commonly assumed to qualify. Word count 758 -> "
            "1,100."
        ),
        "priority_score": 32,
        "impact_score": 40,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (broadband-price-rises-2026), data/guides.ts, data/providers.ts (now-broadband fix)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-24 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-ee-provider-page",
        "type": "Content",
        "pillar": "Content",
        "title": "Update: EE provider page Ofcom complaints figure (20th content-priority-analysis target)",
        "description": (
            "Twentieth target from scripts/analyze_content_priority.py's wider ranked "
            "list (top 30): /providers/ee carries the single highest mapped search volume "
            "of any page on the site (202,237/mo) and was already well-built (1,649 "
            "words, no thinness flag), so this was a targeted correction rather than a "
            "full rewrite. The page's own reviewSources explicitly noted it could not "
            "find EE's exact Q1 2026 Ofcom complaints figure and was still citing Q4 "
            "2025's 'worst-three bracket, 10 per 100k' framing as the operative claim. "
            "Filled that gap with the exact figure from the same broadbandswitch 'The "
            "Complaints Floor' Q1 2026 table already used for the price-rises and "
            "broadband-providers-ranking guides: EE recorded 6 per 100,000, exactly the "
            "record-low industry average, a real, measurable improvement no longer in "
            "the worst bracket. Rewrote the Ofcom section, cons, excerpt, FAQ and 'Is EE "
            "Worth It' verdict to reflect this, reframing EE as a new example of the "
            "low-Trustpilot-vs-strong-Ofcom pattern (joining Sky, Plusnet, Virgin Media) "
            "rather than a provider genuinely weak on both measures."
        ),
        "priority_score": 60,
        "impact_score": 55,
        "effort": "Low",
        "target": "data/providers.ts (ee)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-bt-vs-virgin-media-comparison",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: BT vs Virgin Media comparison page (21st content-priority-analysis target)",
        "description": (
            "Twenty-first target from the wider top-30 content-priority-analysis list: "
            "526 words of entirely generic, number-free copy ('BT has no setup fee in "
            "the current dataset' -- itself wrong, BT's setup fee is GBP30). Rewrote "
            "using already-vetted current data from this session's own BT and Virgin "
            "Media provider-page work: BT from GBP23.99/mo (24-month, 98% coverage, "
            "Trustpilot 4.0/1.5 two-page split) vs Virgin Media from GBP33.00/mo "
            "(18-month, 52% coverage, 1,130Mbps fastest widely available speed, "
            "Trustpilot 1.4 but a genuinely better Q1 2026 Ofcom complaints record than "
            "BT, 6 vs 7 per 100k). Added the genuinely important practical point this "
            "page never covered: switching to Virgin Media never qualifies for One Touch "
            "Switch since it runs on a separate network from Openreach. Populated the "
            "previously-empty factSnapshot field. Word count 526 -> 718."
        ),
        "priority_score": 41,
        "impact_score": 30,
        "effort": "Medium",
        "target": "data/provider-comparisons.ts (bt-vs-virgin-media)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-full-fibre-broadband-explained-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Full Fibre Broadband Explained guide (22nd content-priority-analysis target)",
        "description": (
            "Twenty-second target from the wider content-priority-analysis list: 538 "
            "words with a stale 'speeds of up to 1,000 Mbps (1 Gbps) are achievable' "
            "ceiling claim and a vague 'Openreach targeting 25 million premises by 2026' "
            "coverage stat with no real current figure. Rewrote with Ofcom's actual "
            "Spring 2026 Connected Nations data (82% full-fibre coverage, 89% "
            "gigabit-capable, 93% urban vs 66% rural split) and updated the speed "
            "ceiling to reflect real current top tiers already verified on this site's "
            "own provider pages -- Community Fibre up to 3,000 Mbps and YouFibre up to "
            "8,000 Mbps, both genuinely symmetrical, far beyond the old 1 Gbps "
            "benchmark. Added a new section connecting full fibre to the 31 January 2027 "
            "PSTN switch-off, reusing this session's own Digital Voice/power-cut "
            "research. Word count 538 -> 922."
        ),
        "priority_score": 31,
        "impact_score": 29,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (full-fibre-broadband-explained), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-broadband-for-students-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Broadband for Students guide (23rd content-priority-analysis target)",
        "description": (
            "Twenty-third target from the wider content-priority-analysis list, and the "
            "third page this session found repeating the exact same stale claim: 'NOW "
            "Broadband is often relevant because its shorter-term positioning can suit "
            "student tenancies' (also wrongly caught and fixed on best-rolling-monthly-"
            "broadband-deals and broadband-deals-under-20 earlier this session). Also "
            "corrected 'TalkTalk can work well for low-cost shared houses', no longer "
            "true now that TalkTalk starts at GBP25/mo. Replaced both with the real "
            "current options: YouFibre's genuine rolling monthly contract (from "
            "GBP33.99/mo) for uncertain tenancies, and Onestream (GBP18.50/mo, 94% "
            "Openreach coverage) as the realistic default for most student addresses, "
            "alongside Hyperoptic/Community Fibre's 12-month middle ground. Added a new "
            "common-mistake bullet: assuming a provider still offers the contract type "
            "it was known for in previous years. Word count 668 -> 887."
        ),
        "priority_score": 30,
        "impact_score": 28,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-broadband-for-students), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-virgin-media-provider-page",
        "type": "Content",
        "pillar": "Content",
        "title": "Update: Virgin Media provider page Ofcom complaints figure (24th content-priority-analysis target)",
        "description": (
            "Twenty-fourth target, found via the same spot-check pattern that caught "
            "EE's stale figure: /providers/virgin-media (63,112/mo volume, already "
            "well-built at 1,798 words) was still citing its Q4 2025 Ofcom complaints "
            "figure (7 per 100k) with vague framing ('has not appeared among the top "
            "complained-about providers in Q1 2026 either'). Replaced with the exact "
            "Q1 2026 figure from the same broadbandswitch table used for EE and the "
            "price-rises guide: Virgin Media recorded 6 per 100,000, exactly the "
            "industry average, a further improvement on its already-decent Q4 2025 "
            "position. Added a dedicated source citation and updated reviewedDate."
        ),
        "priority_score": 45,
        "impact_score": 40,
        "effort": "Low",
        "target": "data/providers.ts (virgin-media)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-how-to-switch-broadband-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Update: How to Switch Broadband Provider guide (25th content-priority-analysis target)",
        "description": (
            "Twenty-fifth target: /guides/how-to-switch-broadband-uk (6,000/mo volume) "
            "was already excellent -- 1,913 words, dated citations, real Automatic "
            "Compensation Scheme figures (GBP6.46/day, GBP32.31 per missed appointment), "
            "no thinness or staleness flags. This was a targeted enhancement rather than "
            "a rewrite: added the one genuinely relevant, on-topic development the page "
            "was missing -- Ofcom closed its dedicated One Touch Switch enforcement "
            "programme on 11 June 2026 after finding more than 2 million customers had "
            "already used it successfully, confirming it as the industry's permanent "
            "standard rather than a still-bedding-in process. Added this to the intro, "
            "a new FAQ, and the guide's keyTakeaways/sources. Word count 1,913 -> 2,043."
        ),
        "priority_score": 44,
        "impact_score": 38,
        "effort": "Low",
        "target": "app/guides/[slug]/page.tsx (how-to-switch-broadband-uk), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-broadband-for-streaming-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Broadband for Streaming guide (26th content-priority-analysis target)",
        "description": (
            "Twenty-sixth target: 699 words naming providers (Sky, Virgin Media, "
            "Community Fibre, Hyperoptic) with zero real prices attached. Added current "
            "prices for all 4, plus Virgin Media's real 1,130 Mbps top speed. Added a "
            "genuinely new, technically accurate explanation this page never covered: "
            "why evening buffering specifically affects FTTC and cable connections "
            "(shared capacity per street cabinet/node) but not full fibre (dedicated "
            "line per property), with a matching new FAQ. Word count 699 -> 893."
        ),
        "priority_score": 29,
        "impact_score": 27,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-broadband-for-streaming), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-can-i-leave-broadband-early-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Can I Leave Broadband Early After a Price Rise guide (27th content-priority-analysis target)",
        "description": (
            "Twenty-seventh target: 783 words of correct but entirely abstract legal "
            "explanation with zero real examples of what 'clearly disclosed' vs 'vague' "
            "price-rise wording actually looks like in practice. Added the concrete "
            "contrast this session's own research had already surfaced: BT and "
            "TalkTalk both disclose a flat, dated pounds-and-pence rise upfront (the "
            "clear case), while Sky's contract terms for years after the first only say "
            "prices 'may rise' without a fixed figure (the genuinely weaker disclosure "
            "case worth flagging). Added the specific 30-day notice/exit-window rule "
            "with a matching new FAQ. Caught and fixed 2 instances of the recurring "
            "backslash-apostrophe-in-JSX bug before shipping. Word count 783 -> 1,068."
        ),
        "priority_score": 28,
        "impact_score": 26,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (can-i-leave-broadband-early-after-price-rise), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-best-broadband-for-rural-areas-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Best Broadband for Rural Areas guide (28th content-priority-analysis target)",
        "description": (
            "Twenty-eighth target: 635 words naming only BT and EE, with a real content "
            "gap for a rural-specific guide -- no mention of satellite broadband at all, "
            "despite Starlink being one of the classic genuine rural use cases already "
            "covered in depth on this site's own satellite-broadband-uk guide earlier "
            "this session. Added real Ofcom Spring 2026 rural coverage data (66% rural "
            "vs 93% urban gigabit-capable), real 5G pricing (National Broadband from "
            "GBP34.99/mo as the multi-network specialist genuinely suited to rural "
            "addresses, Three from GBP29/mo, EE Smart 5G Hub), a full Starlink section "
            "with real pricing, and the Gigabit Broadband Voucher Scheme as a "
            "cost-effective alternative to satellite worth checking first. Word count "
            "635 -> 908."
        ),
        "priority_score": 27,
        "impact_score": 26,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (best-broadband-for-rural-areas-uk), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-4-thin-comparison-pages",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: 4 thin comparison pages (29th-32nd content-priority-analysis targets)",
        "description": (
            "Final 4 targets in this batch, all comparison pages with the exact same "
            "problem: entirely generic, number-free copy repeating a 'TalkTalk is "
            "usually cheaper' claim that is now factually wrong on 2 of the 4 pages -- "
            "bt-vs-talktalk and sky-vs-talktalk both still claimed TalkTalk was the "
            "cheaper option, when TalkTalk now starts at GBP25/mo, more expensive than "
            "both BT (GBP23.99) and Sky (GBP23.00). Rewrote all 4 with real current "
            "prices, Ofcom Q1 2026 complaints figures and populated factSnapshot "
            "fields: bt-vs-talktalk (BT now cheaper AND better complaints record), "
            "sky-vs-talktalk (Sky now cheaper AND far better complaints record), "
            "sky-vs-virgin-media (both genuinely decent on Ofcom despite very different "
            "Trustpilot scores -- a real, useful pattern this page never mentioned), "
            "ee-vs-bt (both BT Group brands, EE's complaints record improved to the "
            "industry average, EE's real automatic mobile-backup differentiator). Word "
            "counts: 515->670, 514->655, 505->621, 481->642."
        ),
        "priority_score": 45,
        "impact_score": 42,
        "effort": "High",
        "target": "data/provider-comparisons.ts (bt-vs-talktalk, sky-vs-talktalk, sky-vs-virgin-media, ee-vs-bt)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-ee-vs-talktalk-comparison",
        "type": "Content",
        "pillar": "Content",
        "title": "Update: EE vs TalkTalk comparison page (33rd content-priority-analysis target)",
        "description": (
            "First target in the second 13-page batch, and notable for having 38 real "
            "GSC AI-feature impressions despite only 390/mo mapped volume -- the highest "
            "AI-feature signal of any page examined this batch. The page itself was "
            "already well-built (already followed the site's fixed comparison-page H2 "
            "structure and dated sourcing, likely produced by the separate autonomous "
            "content pipeline on 16 August), so this was a targeted correction: EE's "
            "price was stale at GBP26.99/mo (real current, already fixed on EE's own "
            "provider page this session, is GBP22.99) and its speed was overstated at "
            "'1.6 Gbps' against EE's real current 900 Mbps flagship tier. Corrected both "
            "throughout (factSnapshot, verdict, keyDifferences, all 4 FAQs, source "
            "dates), and added EE's genuinely improved Q1 2026 Ofcom complaints figure "
            "(6 per 100k, industry average) which the page previously omitted."
        ),
        "priority_score": 50,
        "impact_score": 45,
        "effort": "Low",
        "target": "data/provider-comparisons.ts (ee-vs-talktalk)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-broadband-speeds-explained-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Broadband Speeds Explained guide (34th content-priority-analysis target)",
        "description": (
            "Second target in the second 13-page batch: 476 words, the thinnest guide "
            "examined all session, with entirely generic, dateless content. Added a real "
            "Ofcom stat this page never had: average maximum available download speed is "
            "285 Mbps (up ~30% from 223 Mbps), with the important nuance that median "
            "speed actually delivered on FTTC is closer to 80-100 Mbps -- the average and "
            "the typical experience are genuinely different things. Reused the FTTC/cable "
            "evening-congestion explanation (shared capacity per street cabinet/node) "
            "already written for best-broadband-for-streaming earlier this session, with "
            "a matching new FAQ. Added real current symmetrical-speed examples (Community "
            "Fibre 3,000Mbps, YouFibre 8,000Mbps). Word count 476 -> 787."
        ),
        "priority_score": 26,
        "impact_score": 25,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (broadband-speeds-explained), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-broadband-deals-with-cashback-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Broadband Deals With Cashback guide (35th content-priority-analysis target)",
        "description": (
            "Third target in the second 13-page batch: 622 words with zero real "
            "numbers -- entirely abstract advice about 'gift cards' and 'cashback' with "
            "no concrete example of what a real reward is actually worth. Added BT's "
            "real reward card figures (GBP80-140, offsetting a GBP30 setup fee, already "
            "documented on BT's own provider page) and NOW Broadband's real GBP70-75 "
            "voucher pattern. Added a genuinely concrete counter-example using "
            "Community Fibre's no-reward GBP12.50/mo price to illustrate why a lower "
            "ongoing price can beat a bigger one-off card once the full term is "
            "compared, replacing the old page's entirely hypothetical framing. Word "
            "count 622 -> 826."
        ),
        "priority_score": 25,
        "impact_score": 23,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (broadband-deals-with-cashback), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-broadband-deals-no-setup-fee-guide",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: Broadband Deals With No Setup Fee guide (36th content-priority-analysis target)",
        "description": (
            "Fourth target in the second 13-page batch: 648 words with a real error -- "
            "BT was listed in the 'mainstream, no setup fee' table row alongside Sky "
            "and EE, but BT genuinely charges a GBP30 setup fee (confirmed on its own "
            "already-corrected provider page this session), only offset by a reward "
            "card that must be actively claimed. Corrected the table and matching FAQ, "
            "and added a new section distinguishing a genuine fee (BT's GBP30, Virgin "
            "Media's GBP35 with no offset) from a fee that functions as free (NOW "
            "Broadband's GBP5 advance fee, credited back to the first bill) -- a real "
            "distinction the old page collapsed into a single 'usually fee-charging' "
            "bucket. Word count 648 -> 812."
        ),
        "priority_score": 30,
        "impact_score": 32,
        "effort": "Medium",
        "target": "app/guides/[slug]/page.tsx (broadband-deals-with-no-setup-fee), data/guides.ts",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "content-refresh-4-more-thin-comparison-pages",
        "type": "Content",
        "pillar": "Content",
        "title": "Deep rewrite: 4 more thin comparison pages (37th-40th content-priority-analysis targets)",
        "description": (
            "Continuing the second 13-page batch: 4 more comparison pages, 3 of which "
            "repeated the same 'Vodafone is cheaper' error already found and fixed twice "
            "this session on different pairs -- bt-vs-vodafone and sky-vs-vodafone both "
            "claimed Vodafone was the value pick, when Vodafone (GBP25) is now more "
            "expensive than both BT (GBP23.99) and Sky (GBP23.00). Rewrote all 4 with "
            "real current prices, Ofcom Q1 2026 complaints figures and populated "
            "factSnapshot fields: bt-vs-vodafone (BT now cheaper AND better complaints), "
            "sky-vs-vodafone (Sky now cheaper AND far better complaints), ee-vs-sky "
            "(prices within 1p of each other, both genuinely strong on Ofcom -- a real "
            "penny-and-feature-based decision replacing entirely vague marketing copy), "
            "plusnet-vs-bt (Plusnet has the best Ofcom complaints record of any major UK "
            "provider AND is cheaper than its own parent brand BT, a much stronger "
            "differentiator than the old page's vague 'UK support' framing). Word "
            "counts: 455->600, 504->608, 500->648, 507->604."
        ),
        "priority_score": 35,
        "impact_score": 34,
        "effort": "High",
        "target": "data/provider-comparisons.ts (bt-vs-vodafone, sky-vs-vodafone, ee-vs-sky, plusnet-vs-bt)",
        "dependencies": "ops-content-priority-analysis-tooling",
        "source": "User request 2026-08-25 — researched before implementing per the Strategic Lens process",
    },
    {
        "item_id": "bet-decouple-content-from-code",
        "type": "Bigger bet",
        "pillar": "Content",
        "title": "Decouple guide content from code deploys",
        "description": (
            "Every guide lives as a hand-written object in data/guides.ts plus "
            "matching JSX in app/guides/[slug]/page.tsx. Fine at 30 guides; at 100+ "
            "it becomes the bottleneck. Plan a move to MDX-per-guide or a lightweight "
            "headless CMS before volume forces it."
        ),
        "priority_score": 35,
        "impact_score": 48,
        "effort": "High",
        "target": "data/guides.ts, app/guides/[slug]/page.tsx",
        "dependencies": "None, but easier before the guide count grows further",
        "source": "Growth playbook — Content pillar",
    },
]

PAGE_TYPE_TO_PILLAR = {
    "Guide": "Content",
    "Provider page": "Content",
    "Comparison page": "Content",
    "City hub": "Content",
    "Interactive tool": "Functionality",
    "Research page": "Content",
}


def git(*args: str) -> str:
    try:
        return subprocess.run(
            ["git", *args], cwd=ROOT, capture_output=True, text=True, timeout=15
        ).stdout.strip()
    except Exception:
        return ""


def repo_sync_status() -> str:
    """One global snapshot fact, not a fragile per-row git-blame lookup:
    is the working tree currently dirty in areas that hold page-build content?"""
    watched = (
        "data/guides.ts", "data/provider-comparisons.ts", "data/priority-pages.ts",
        "data/providers.ts", "app/postcode", "app/guides", "app/providers",
    )
    dirty = git("status", "--short", *watched)
    return "Live but not committed to git" if dirty else "Committed"


def load_page_builds(source: Path) -> list[dict[str, Any]]:
    wb = load_workbook(source, data_only=True)
    ws = wb["Content Gap Roadmap"]
    rows = list(ws.iter_rows(values_only=True))
    headers, data_rows = rows[0], rows[1:]
    positions = {name: i for i, name in enumerate(headers)}
    sync = repo_sync_status()

    items: list[dict[str, Any]] = []
    for row in data_rows:
        def get(col: str) -> Any:
            return row[positions[col]] if positions.get(col) is not None else None

        url = get("Recommended URL") or ""
        slug = url.replace("https://broadbandpicker.co.uk/", "")
        status_field = get("Current Status") or ""
        page_type = get("Page Type") or ""
        items.append({
            "item_id": f"page-{slug}",
            "type": "Page",
            "pillar": PAGE_TYPE_TO_PILLAR.get(page_type, "Content"),
            "title": get("Page Title") or slug,
            "description": get("Why This Position") or "",
            "priority_score": get("SEO Build Priority Score") or 0,
            "impact_score": get("Revenue Priority Score") or 0,
            "effort": get("Ranking Difficulty Band") or "",
            "target": url,
            "dependencies": "",
            "source": "docs/broadbandpicker-keyword-mapping.xlsx — Content Gap Roadmap",
            "build_status": status_field,
            "repo_sync": sync if status_field.startswith("Built") else "",
            "cluster": get("Topic Cluster") or "",
            "keywords": get("Primary Keywords") or "",
            "volume": get("Combined Est. Volume") or "",
        })
    return items


def default_status_for(item: dict[str, Any]) -> str:
    build_status = item.get("build_status", "")
    if build_status.startswith("Built"):
        return "Done"
    return item.get("initial_status", DEFAULT_STATUS)


def load_manual_overrides(output: Path) -> dict[str, dict[str, Any]]:
    if not output.exists():
        return {}
    try:
        wb = load_workbook(output, data_only=True)
        ws = wb["Master Tracker"]
    except Exception:
        return {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers, data_rows = rows[0], rows[1:]
    positions = {name: i for i, name in enumerate(headers)}
    if "Item ID" not in positions:
        return {}
    overrides: dict[str, dict[str, Any]] = {}
    for row in data_rows:
        item_id = row[positions["Item ID"]]
        if not item_id:
            continue
        overrides[item_id] = {
            header: row[positions[header]]
            for header in MANUAL_HEADERS
            if header in positions
        }
    return overrides


def load_weekly_seo_actions(output: Path) -> tuple[list[str], list[list[Any]]] | None:
    """Carry the API-generated weekly action tab through tracker rebuilds."""
    if not output.exists():
        return None
    try:
        wb = load_workbook(output, data_only=False, read_only=True)
        if "Weekly SEO Actions" not in wb.sheetnames:
            return None
        rows = list(wb["Weekly SEO Actions"].iter_rows(values_only=True))
    except Exception:
        return None
    if not rows:
        return None
    return list(rows[0]), [list(row) for row in rows[1:] if any(value is not None for value in row)]


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[Any]]) -> Any:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    if rows and headers:
        import re
        ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        table = Table(displayName=re.sub(r"\W+", "", name)[:20] + "Table", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    return ws


def style_sheet(ws: Any) -> None:
    navy, pale = "0F172A", "E2F2F6"
    thin = Side(style="thin", color="D7E1E8")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)
    for column in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, column).value or "") for row in range(1, min(ws.max_row, 80) + 1)]
        width = min(50, max(10, max(len(v) for v in values) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for cell in ws[row]:
                cell.fill = PatternFill("solid", fgColor=pale)


def build_workbook(
    page_items: list[dict[str, Any]],
    overrides: dict[str, dict[str, Any]],
    weekly_actions: tuple[list[str], list[list[Any]]] | None = None,
) -> Workbook:
    all_items = FEATURE_BUILDS + page_items
    for item in all_items:
        manual = overrides.get(item["item_id"], {})
        computed_status = default_status_for(item)
        manual_status = manual.get("Status")
        if item["type"] == "Page":
            # Page status is independently verified by the keyword-mapping
            # pipeline's live crawl (build_status) every run — always trust
            # it fresh. A stale "Not started" written on a page's first
            # appearance in this tracker must never freeze it there once
            # the source data later confirms it shipped.
            item["status"] = computed_status
        elif item.get("verified_status"):
            item["status"] = item["verified_status"]
        else:
            # Feature/audit/bet items have no external verification signal;
            # a hand-set Status ("In progress", "Done", ...) is a genuine
            # manual edit and must survive regeneration.
            item["status"] = manual_status or computed_status
        item["owner"] = manual.get("Owner") or ""
        item["target_date"] = manual.get("Target Date") or ""
        item["notes"] = manual.get("Notes") or item.get("completion_notes", "")
    all_items.sort(key=lambda it: it["priority_score"], reverse=True)

    wb = Workbook()
    wb.remove(wb.active)

    readme = [
        ["Purpose", "One prioritised list of everything still to build for BroadbandPicker.co.uk — new pages and product/UX features together, ranked highest priority first."],
        ["Generated", RUN_DATE],
        ["Sources", "Page builds: docs/broadbandpicker-keyword-mapping.xlsx (Content Gap Roadmap). Feature builds: the growth playbook, scored on the same rough 0-100 scale."],
        ["Scoring", "Page rows use the existing SEO Build Priority Score / Revenue Priority Score from the keyword-mapping pipeline. Feature rows are hand-scored on the same scale — both numbers are directional priority signals, not a single unified formula."],
        ["Status is preserved", "Status, Owner, Target Date and Notes carry forward from the previous docs/master-build-tracker.xlsx on every re-run, matched by Item ID — safe to edit by hand."],
        ["Re-run", "python3 scripts/build_master_tracker.py"],
    ]
    add_sheet(wb, "Read Me", ["Field", "Detail"], readme)

    tracker_headers = [
        "Rank", "Item ID", "Type", "Pillar", "Title", "Priority Score", "Impact Score",
        "Effort", "Status", "Owner", "Target Date", "Notes",
    ]
    tracker_rows = []
    for rank, item in enumerate(all_items, 1):
        tracker_rows.append([
            rank, item["item_id"], item["type"], item["pillar"], item["title"],
            item["priority_score"], item["impact_score"], item["effort"],
            item["status"], item["owner"], item["target_date"], item["notes"],
        ])
    add_sheet(wb, "Master Tracker", tracker_headers, tracker_rows)

    pending_headers = [
        "Pending Rank", "Item ID", "Type", "Pillar", "Title", "Priority Score",
        "Impact Score", "Effort", "Status", "Target", "Dependencies", "Owner",
        "Target Date", "Notes",
    ]
    pending_rows = []
    pending_items = [item for item in all_items if item["status"] != "Done"]
    for rank, item in enumerate(pending_items, 1):
        pending_rows.append([
            rank, item["item_id"], item["type"], item["pillar"], item["title"],
            item["priority_score"], item["impact_score"], item["effort"],
            item["status"], item.get("target", ""), item.get("dependencies", ""),
            item["owner"], item["target_date"], item["notes"],
        ])
    add_sheet(wb, "Pending Build Priority", pending_headers, pending_rows)

    page_headers = [
        "Rank", "Title", "Cluster", "Priority Score", "Revenue Score", "Difficulty Band",
        "Build Status", "Repo Sync", "Status", "Keywords", "Volume/mo", "URL",
    ]
    page_rows = []
    sorted_pages = sorted(page_items, key=lambda it: it["priority_score"], reverse=True)
    for rank, item in enumerate(sorted_pages, 1):
        page_rows.append([
            rank, item["title"], item["cluster"], item["priority_score"], item["impact_score"],
            item["effort"], item["build_status"], item["repo_sync"], item["status"],
            item["keywords"], item["volume"], item["target"],
        ])
    add_sheet(wb, "Page Builds", page_headers, page_rows)

    feature_headers = [
        "Rank", "Type", "Pillar", "Title", "Description", "Priority Score", "Impact Score",
        "Effort", "Target", "Dependencies", "Status", "Source",
    ]
    feature_rows = []
    sorted_features = sorted(FEATURE_BUILDS, key=lambda it: it["priority_score"], reverse=True)
    for rank, item in enumerate(sorted_features, 1):
        feature_rows.append([
            rank, item["type"], item["pillar"], item["title"], item["description"],
            item["priority_score"], item["impact_score"], item["effort"], item["target"],
            item["dependencies"], item.get("status", DEFAULT_STATUS), item["source"],
        ])
    add_sheet(wb, "Feature & Strategy Builds", feature_headers, feature_rows)

    done = sum(1 for it in all_items if it["status"] == "Done")
    methodology = [
        ["Item count", f"{len(all_items)} total ({len(page_items)} page builds, {len(FEATURE_BUILDS)} feature/strategy builds); {done} already marked Done."],
        ["Page priority score", "Existing SEO Build Priority Score from the keyword-mapping pipeline — weighted toward CPC/commercial intent, tuned for affiliate revenue."],
        ["Feature priority score", "Hand-scored 0-100 reflecting the growth playbook's sequencing: infra risk and quick wins first, bigger bets last."],
        ["Repo Sync column", "A single snapshot check of `git status` across the folders that hold page-build content — not a per-row git-blame lookup. 'Live but not committed to git' means production and git have diverged for that batch of pages."],
        ["Pending queue", "Pending Build Priority contains only non-Done items from the combined page, feature, audit and strategy list, ordered by Priority Score."],
        ["Editing this file", "Status/Owner/Target Date/Notes are yours to edit freely — they survive the next `python3 scripts/build_master_tracker.py` run."],
    ]
    add_sheet(wb, "Methodology", ["Aspect", "Detail"], methodology)

    if weekly_actions:
        headers, rows = weekly_actions
        add_sheet(wb, "Weekly SEO Actions", headers, rows)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    page_items = load_page_builds(args.source)
    overrides = load_manual_overrides(args.output)
    weekly_actions = load_weekly_seo_actions(args.output)
    wb = build_workbook(page_items, overrides, weekly_actions)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Saved {args.output.resolve()}")
    print(f"{len(page_items)} page builds + {len(FEATURE_BUILDS)} feature builds = "
          f"{len(page_items) + len(FEATURE_BUILDS)} tracked items")


if __name__ == "__main__":
    main()
