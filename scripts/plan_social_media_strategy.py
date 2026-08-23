#!/usr/bin/env python3
"""Plan BroadbandPicker's social content from the actual workspace, not guesswork.

The script:
1. Reads brand, product, provider, guide, tool, and marketing files in this repo.
2. Builds a brand brief so every joke still points at the real comparison business.
3. Recommends which social channels to open (consumer virality + affiliate / B2B).
4. Writes funny, casual, engagement-first ideas for X, Instagram (Reels, Stories,
   feed posts) and TikTok — British sarcasm, memes, twisted quotes, and puns —
   each with a design brief and a soft path back to broadbandpicker.co.uk.

Usage:
    python3 scripts/plan_social_media_strategy.py
    python3 scripts/plan_social_media_strategy.py --weeks 6
    python3 scripts/plan_social_media_strategy.py --start-date 2026-08-24

Output lives in docs/social media manager/.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "docs" / "social media manager"
IG_DIR = OUT_DIR / "Instagram Content creation"
X_DIR = OUT_DIR / "X"
TT_DIR = OUT_DIR / "tiktok"
SITE = "https://broadbandpicker.co.uk"
HANDLE_X = "@broadbandPicker"
DEFAULT_START = date(2026, 8, 24)

BRAND_PRIMARY = "#0EA5E9"
BRAND_NAVY = "#0F172A"
BRAND_GREEN = "#22C55E"
BRAND_SLATE = "#F8FAFC"
BRAND_WHITE = "#FFFFFF"

SKIP_DIR_NAMES = {
    "node_modules",
    ".git",
    ".next",
    "dist",
    "coverage",
    "__pycache__",
}


# ---------------------------------------------------------------------------
# Workspace analysis
# ---------------------------------------------------------------------------


def read_text(relative: str) -> str:
    path = ROOT / relative
    if not path.exists() or not path.is_file():
        return ""
    try:
        return path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return ""


def unique(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        key = item.strip()
        if key and key not in seen:
            seen.add(key)
            out.append(key)
    return out


@dataclass
class BrandProfile:
    site: str = SITE
    name: str = "BroadbandPicker"
    tagline: str = "Independent UK broadband comparison"
    positioning: str = ""
    revenue_model: str = ""
    colours: dict[str, str] = field(default_factory=dict)
    fonts: list[str] = field(default_factory=list)
    logo_files: list[str] = field(default_factory=list)
    social_handles: dict[str, str] = field(default_factory=dict)
    emails: dict[str, str] = field(default_factory=dict)
    providers: list[str] = field(default_factory=list)
    guides: list[dict[str, str]] = field(default_factory=list)
    tools: list[dict[str, str]] = field(default_factory=list)
    cities: list[str] = field(default_factory=list)
    content_pillars: list[str] = field(default_factory=list)
    existing_x_posts: list[str] = field(default_factory=list)
    audience_notes: list[str] = field(default_factory=list)
    commercial_goals: list[str] = field(default_factory=list)
    trust_pages: list[str] = field(default_factory=list)
    key_urls: dict[str, str] = field(default_factory=dict)
    files_read: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def extract_css_colours(css: str) -> dict[str, str]:
    found = dict(re.findall(r"--brand-([a-z]+):\s*(#[0-9A-Fa-f]{3,8})", css))
    colours = {
        "primary": found.get("primary", BRAND_PRIMARY),
        "navy": found.get("secondary", BRAND_NAVY),
        "accent": found.get("accent", BRAND_GREEN),
        "background": BRAND_SLATE,
        "white": BRAND_WHITE,
    }
    return colours


def extract_providers(source: str) -> list[str]:
    names = re.findall(r"^\s+name:\s+'([^']+)'", source, flags=re.MULTILINE)
    return unique(names)


def extract_guides(source: str) -> list[dict[str, str]]:
    guides: list[dict[str, str]] = []
    for match in re.finditer(
        r"slug:\s*'([^']+)',\s*title:\s*'([^']+)'",
        source,
        flags=re.DOTALL,
    ):
        slug, title = match.group(1), match.group(2)
        if slug in {
            "deals-and-pricing",
            "switching-and-rights",
            "technology-and-speeds",
            "providers-and-comparisons",
            "use-cases-and-lifestyle",
            "affordability",
        }:
            continue
        guides.append(
            {
                "slug": slug,
                "title": title,
                "url": f"{SITE}/guides/{slug}",
            }
        )
    return guides


def extract_emails(text: str) -> list[str]:
    return unique(re.findall(r"[a-zA-Z0-9._%+-]+@broadbandpicker\.co\.uk", text))


def extract_handles(text: str) -> dict[str, str]:
    handles: dict[str, str] = {}
    x = re.search(r"https://x\.com/([A-Za-z0-9_]+)", text)
    if x:
        handles["x"] = f"@{x.group(1)}"
    ig = re.search(r"instagram\.com/([A-Za-z0-9_.]+)", text)
    if ig:
        handles["instagram"] = f"@{ig.group(1)}"
    tt = re.search(r"tiktok\.com/@([A-Za-z0-9_.]+)", text)
    if tt:
        handles["tiktok"] = f"@{tt.group(1)}"
    return handles


def extract_cities() -> list[str]:
    cities: list[str] = []
    postcode_dir = ROOT / "app" / "postcode"
    if postcode_dir.exists():
        for child in sorted(postcode_dir.iterdir()):
            if child.is_dir() and child.name not in {"[area]"} and not child.name.startswith("."):
                cities.append(child.name.replace("-", " ").title())
    source = read_text("data/postcodes.ts")
    cities.extend(re.findall(r"city:\s*'([^']+)'", source))
    return unique(cities)


def list_tools() -> list[dict[str, str]]:
    tools = [
        {"name": "Postcode checker", "url": f"{SITE}", "job": "Show deals available at an address"},
        {"name": "Compare", "url": f"{SITE}/compare", "job": "Side-by-side provider comparison"},
        {"name": "Deals", "url": f"{SITE}/deals", "job": "Live UK broadband deal table"},
        {"name": "Speed test", "url": f"{SITE}/speed-test", "job": "Check whether the line is actually doing what it promised"},
        {"name": "Broadband Match quiz", "url": f"{SITE}/tools/broadband-match", "job": "Match a household to a package type"},
        {"name": "Cost calculator", "url": f"{SITE}/tools/broadband-cost-calculator", "job": "See the real cost of a contract, not just month one"},
        {"name": "Glossary", "url": f"{SITE}/broadband-glossary", "job": "Plain-English broadband terms"},
    ]
    return tools


def collect_logo_files() -> list[str]:
    candidates = [
        "public/logo.svg",
        "public/logo.png",
        "app/icon.svg",
        "docs/branding-and-marketing/broadbandpicker-logo-icon.png",
        "docs/branding-and-marketing/broadbandpicker-logo-icon.svg",
    ]
    return [path for path in candidates if (ROOT / path).exists()]


def analyse_workspace() -> BrandProfile:
    brand = BrandProfile()
    files = {
        "app/globals.css": read_text("app/globals.css"),
        "app/layout.tsx": read_text("app/layout.tsx"),
        "app/page.tsx": read_text("app/page.tsx"),
        "app/about/page.tsx": read_text("app/about/page.tsx"),
        "app/how-we-make-money/page.tsx": read_text("app/how-we-make-money/page.tsx"),
        "app/contact/page.tsx": read_text("app/contact/page.tsx"),
        "data/providers.ts": read_text("data/providers.ts"),
        "data/guides.ts": read_text("data/guides.ts"),
        "docs/content-plan.md": read_text("docs/content-plan.md"),
        "docs/seo-geo-content-strategy-2026-06.md": read_text("docs/seo-geo-content-strategy-2026-06.md"),
        "docs/affiliate-outreach-email.txt": read_text("docs/affiliate-outreach-email.txt"),
        "docs/branding-and-marketing/x-content-calendar.md": read_text(
            "docs/branding-and-marketing/x-content-calendar.md"
        ),
        "docs/branding-and-marketing/README.md": read_text("docs/branding-and-marketing/README.md"),
        "components/Logo.tsx": read_text("components/Logo.tsx"),
    }
    brand.files_read = [path for path, text in files.items() if text]
    css, layout, home, about, money, contact = (
        files["app/globals.css"],
        files["app/layout.tsx"],
        files["app/page.tsx"],
        files["app/about/page.tsx"],
        files["app/how-we-make-money/page.tsx"],
        files["app/contact/page.tsx"],
    )

    brand.colours = extract_css_colours(css)
    if "Inter" in layout:
        brand.fonts = ["Inter", "system-ui"]
    brand.logo_files = collect_logo_files()
    brand.social_handles = extract_handles(about + layout)
    if "x" not in brand.social_handles:
        brand.social_handles["x"] = HANDLE_X
        brand.warnings.append("X handle inferred from existing footer; confirm the live username.")
    brand.social_handles.setdefault("instagram", "@broadbandpicker")
    brand.social_handles.setdefault("tiktok", "@broadbandpicker")
    brand.social_handles.setdefault("linkedin", "BroadbandPicker")

    emails = extract_emails(contact + files["docs/affiliate-outreach-email.txt"] + about)
    brand.emails = {
        "partnerships": next((e for e in emails if e.startswith("partnerships")), "partnerships@broadbandpicker.co.uk"),
        "editorial": next((e for e in emails if e.startswith("editorial")), "editorial@broadbandpicker.co.uk"),
        "hello": next((e for e in emails if e.startswith("hello")), "hello@broadbandpicker.co.uk"),
    }

    brand.providers = extract_providers(files["data/providers.ts"])
    brand.guides = extract_guides(files["data/guides.ts"])
    brand.tools = list_tools()
    brand.cities = extract_cities()
    brand.content_pillars = [
        "Deals and pricing (loyalty tax, out-of-contract, 'up to' speeds)",
        "Switching and rights (One Touch Switch, exit fees, April price rises)",
        "British household chaos (WFH, football, flatshares, rural waits)",
        "Technology without the lecture (full fibre vs 'fibre', routers in cupboards)",
        "Affordability without punching down (social tariffs as a scandal of awareness)",
        "Brand trust (independent comparison, affiliate honesty, postcode truth)",
    ]
    brand.trust_pages = [
        f"{SITE}/about",
        f"{SITE}/how-we-make-money",
        f"{SITE}/how-we-review-broadband",
        f"{SITE}/editorial-policy",
        f"{SITE}/contact",
    ]
    brand.key_urls = {
        "home": SITE,
        "deals": f"{SITE}/deals",
        "compare": f"{SITE}/compare",
        "providers": f"{SITE}/providers",
        "guides": f"{SITE}/guides",
        "speed_test": f"{SITE}/speed-test",
        "match": f"{SITE}/tools/broadband-match",
        "calculator": f"{SITE}/tools/broadband-cost-calculator",
        "social_tariffs": f"{SITE}/guides/broadband-social-tariffs-uk",
        "price_rises": f"{SITE}/guides/broadband-price-rises-2026",
        "switch": f"{SITE}/guides/how-to-switch-broadband-uk",
        "partnerships_mail": f"mailto:{brand.emails['partnerships']}",
    }

    if "independent UK broadband comparison" in about.lower() or "independent UK broadband comparison service" in about:
        brand.positioning = (
            "Independent UK broadband comparison. Compare deals by postcode, understand "
            "speeds and contracts, and switch without spending an evening on hold."
        )
    else:
        brand.positioning = "UK broadband comparison by postcode, provider, speed and price."

    if "affiliate commissions" in money.lower():
        brand.revenue_model = (
            "Free for consumers. Revenue from affiliate commissions when someone clicks "
            "through and signs up. Rankings are not sold. Providers can be listed with or "
            "without an affiliate relationship."
        )
    else:
        brand.revenue_model = "UK broadband affiliate comparison."

    brand.audience_notes = [
        "UK households paying a loyalty tax after the cheap intro period.",
        "Renters, students, and shared houses who treat the router like a housegod.",
        "WFH people whose upload speed is a personality flaw.",
        "Football-and-Netflix households who discover contention at 7pm.",
        "Rural and not-spot users who have been promised full fibre since the Coalition.",
        "Affiliate managers, ISP partnerships teams, and Awin advertisers who need a serious publisher.",
    ]
    brand.commercial_goals = [
        "Make broadbandpicker.co.uk the default place Brits go to compare broadband.",
        "Grow branded search and direct traffic so comparison sessions convert.",
        "Earn affiliate sign-ups without sounding like a deal farm.",
        "Look established enough that providers approach BroadbandPicker to join the affiliate roster.",
        "Use social proof (audience, engagement, content quality) as an Awin and advertiser signal.",
    ]

    calendar = files["docs/branding-and-marketing/x-content-calendar.md"]
    brand.existing_x_posts = [
        line.strip()
        for line in calendar.splitlines()
        if line.startswith("Compare ") or line.startswith("What speed") or line.startswith("Out of contract")
    ]
    if calendar:
        brand.warnings.append(
            "Existing X calendar in docs/branding-and-marketing is clear and useful, but too polite "
            "to travel. This plan keeps the facts and swaps the tone for sarcasm, memes and hooks."
        )

    if not brand.providers:
        brand.warnings.append("Could not parse provider names from data/providers.ts.")
    if not brand.guides:
        brand.warnings.append("Could not parse guides from data/guides.ts.")

    _ = home  # homepage scanned so the brand brief stays tied to live copy
    return brand


# ---------------------------------------------------------------------------
# Channel strategy
# ---------------------------------------------------------------------------


def channel_rows(brand: BrandProfile) -> list[dict[str, str]]:
    return [
        {
            "channel": "X (Twitter)",
            "priority": "P0 — already live, keep and sharpen",
            "open": "Yes now",
            "handle": brand.social_handles.get("x", HANDLE_X),
            "why": (
                "Where UK broadband rage, Ofcom news, outages and journalist chatter already live. "
                "Best channel for quote-tweets, one-liners, and being the funny account that still "
                "knows the rules. Footer already points here."
            ),
            "audience": "Switchers, tech-adjacent adults, journalists, altnet watchers, complainers at 7pm.",
            "content": "Sarcastic one-liners, polls, quote remixes, outage colour commentary, deal 'public health warnings'.",
            "cadence": "5–7 posts/week + real-time when a provider or Ofcom does something stupid.",
            "success": "Profile clicks, quote-tweets, follows from journalists, branded search lift.",
            "b2b": "Industry people lurk here. A sharp account is a walking media kit.",
        },
        {
            "channel": "Instagram",
            "priority": "P0 — open this week",
            "open": "Yes now",
            "handle": brand.social_handles.get("instagram", "@broadbandpicker"),
            "why": (
                "Reels are the UK meme engine. Stories build habit. Carousels teach without feeling "
                "like a guide. This is how BroadbandPicker stops looking like a spreadsheet and starts "
                "looking like a brand people screenshot to their group chat."
            ),
            "audience": "25–40, renters, young families, students' older siblings, 'I pay the bill' people.",
            "content": "Reels, Stories, meme posts, quote cards, 7-slide carousels, sticker polls.",
            "cadence": "4 Reels/week, 1 feed post/weekday, Stories on posting days.",
            "success": "Shares, saves, sticker replies, profile visits, link-in-bio clicks to postcode checker.",
            "b2b": "A living visual identity. Providers judge professionalism in 3 seconds.",
        },
        {
            "channel": "TikTok",
            "priority": "P0 — open this week, same content engine as Reels",
            "open": "Yes now",
            "handle": brand.social_handles.get("tiktok", "@broadbandpicker"),
            "why": (
                "Discovery graph, not friend graph. A 12-second joke about buffering during the football "
                "will find people who have never heard of a comparison site. Students, first-time switchers, "
                "and 'my dad still pays £40 for 30 Mbps' kids live here."
            ),
            "audience": "16–34 UK, students, HMO flatshares, first adult bills.",
            "content": "Native vertical sketches, POV, green-screen rants, fake TfL/weather warnings.",
            "cadence": "1/day for 30 days, then 5/week. Recycle the best as IG Reels 24h later.",
            "success": "Watch-through, comments, stitches, follows, traffic via link in bio / Spark Ads later.",
            "b2b": "Reach numbers you can put in an affiliate deck. 'We made broadband funny' is a pitch.",
        },
        {
            "channel": "LinkedIn",
            "priority": "P0 — open this week for provider / Awin gravity",
            "open": "Yes now",
            "handle": "Company page: BroadbandPicker",
            "why": (
                "This is how broadband providers and affiliate managers find you, not how consumers have a laugh. "
                "A crisp company page plus founder posts about methodology, coverage, and traffic is the difference "
                "between chasing programmes and being invited onto them."
            ),
            "audience": "ISP partnerships, Awin account managers, altnet marketers, comparison-site peers.",
            "content": "Site milestones, honest methodology, 'we list you even without an affiliate link', market notes.",
            "cadence": "2–3 posts/week. Founder voice, not corporate slurry.",
            "success": "Inbound to partnerships@, profile views from provider domains, connection requests.",
            "b2b": "Primary B2B channel. Pin a partnerships post. Put the site and Awin-ready trust pages in the featured section.",
        },
        {
            "channel": "YouTube (Shorts first, long-form later)",
            "priority": "P1 — start Shorts in week 3 by mirroring TikTok",
            "open": "Create the channel now, post from week 3",
            "handle": "BroadbandPicker",
            "why": "Shorts are free distribution. Long-form 'How to switch' and 'price rise 2026' videos become SEO assets.",
            "audience": "Same as TikTok, plus people who Google 'how to switch broadband' and click Videos.",
            "content": "Mirrored Shorts; later 6–8 min explainers with chapters and on-screen CTAs.",
            "cadence": "Shorts: 4/week once TikTok is flowing. Long-form: 1/month from month 2.",
            "success": "Subs, suggested traffic, branded views, later AdSense-adjacent watch time.",
            "b2b": "A real media property. Providers like 'hosted on YouTube', not just 'a tweet'.",
        },
        {
            "channel": "Facebook Page",
            "priority": "P1 — month 2",
            "open": "Create now, post from month 2 (or auto-share IG)",
            "handle": "BroadbandPicker",
            "why": (
                "Still where 35–60 UK shares rants about bills. Price-rise explainers and 'are you out of contract?' "
                "graphics travel in local groups if you are useful, not spammy."
            ),
            "audience": "Bill-payers, parents, local community groups.",
            "content": "Native graphics, Reels auto-share, calm explainers with a joke in the first line.",
            "cadence": "3–4/week, mostly recycled.",
            "success": "Shares, group discussion, 40+ traffic.",
            "b2b": "Low. Do it for reach, not for Awin theatre.",
        },
        {
            "channel": "WhatsApp Channel",
            "priority": "P1 — month 2",
            "open": "After IG/TikTok have a following to point at it",
            "handle": "BroadbandPicker Deals",
            "why": "UK usage is high. A broadcast channel for 'your contract is a trap' alerts and April price-rise week.",
            "audience": "People who will not open another app but will read a once-a-week ping.",
            "content": "1–2 notes/week. No spam. Link to a specific page.",
            "cadence": "Weekly, plus event-led.",
            "success": "Follows, link taps.",
            "b2b": "None. Pure consumer.",
        },
        {
            "channel": "Pinterest",
            "priority": "P2 — month 3",
            "open": "Only after 20 evergreen graphics exist",
            "handle": "BroadbandPicker",
            "why": "Evergreen infographics ('speed you actually need', 'social tariff checklist') keep sending traffic for years.",
            "audience": "Planners, movers, students' parents, 'save this' people.",
            "content": "Tall infographics, checklist pins, quote cards.",
            "cadence": "Batch 15 pins, then 3/week.",
            "success": "Saves and outbound clicks.",
            "b2b": "None.",
        },
        {
            "channel": "Threads",
            "priority": "P2 — only if IG is working",
            "open": "Same login as Instagram, post when a joke is too long for a caption",
            "handle": brand.social_handles.get("instagram", "@broadbandpicker"),
            "why": "Cheap extra surface. Do not build a separate strategy.",
            "audience": "IG overlap.",
            "content": "X-style text, no extra production.",
            "cadence": "Opportunistic.",
            "success": "Negligible until proven.",
            "b2b": "None.",
        },
        {
            "channel": "Reddit (presence, not a brand megaphone)",
            "priority": "P2 — contribute, do not 'launch an account'",
            "open": "Personal/expert account, not a mascot",
            "handle": "Transparent username, disclose affiliation",
            "why": "r/broadband, r/UKPersonalFinance, r/HousingUK. Helpful answers with a link only when asked beat any ad.",
            "audience": "High-intent researchers.",
            "content": "Genuine answers. Never dump deals.",
            "cadence": "2–3 comments/week.",
            "success": "Referral traffic, trust, the odd journalist lurker.",
            "b2b": "Indirect — shows you understand the market.",
        },
        {
            "channel": "Bluesky",
            "priority": "P3 — skip for now",
            "open": "No",
            "handle": "—",
            "why": "Too small for the operating cost. Revisit if UK broadband Twitter migrates.",
            "audience": "Tech early adopters.",
            "content": "—",
            "cadence": "—",
            "success": "—",
            "b2b": "Low.",
        },
        {
            "channel": "Snapchat",
            "priority": "P3 — skip",
            "open": "No",
            "handle": "—",
            "why": "Ads-only unless you have a youth brand budget. TikTok covers this audience.",
            "audience": "Under 25.",
            "content": "—",
            "cadence": "—",
            "success": "—",
            "b2b": "None.",
        },
    ]


# ---------------------------------------------------------------------------
# Content ideas
# ---------------------------------------------------------------------------


@dataclass
class DesignBrief:
    format_name: str
    dimensions: str
    duration: str
    visual_concept: str
    on_screen_text: list[str]
    palette: str
    type: str
    logo: str
    audio: str
    shots: list[str]
    do_not: list[str]


@dataclass
class Idea:
    id: str
    pillar: str
    title: str
    hook: str
    copy: str
    alt_copy: str
    cta: str
    url_key: str
    engagement: str
    hashtags: list[str]
    platforms: list[str]
    formats: list[str]
    why: str
    compliance: str
    design: DesignBrief


def standard_palette() -> str:
    return (
        f"Navy {BRAND_NAVY} background, sky {BRAND_PRIMARY} highlights, white type, "
        f"green {BRAND_GREEN} only for a genuine 'good news' sting. High contrast, no grey-on-grey."
    )


def logo_rule() -> str:
    return (
        "Native joke first, brand last. Wifi-mark + wordmark on a 0.4s end card or bottom-left "
        "safe zone. Never put the logo over the punchline."
    )


def ideas(brand: BrandProfile) -> list[Idea]:
    providers_sample = ", ".join(brand.providers[:6]) or "BT, Sky, Virgin Media, EE"
    home = brand.key_urls["home"]
    deals = brand.key_urls["deals"]
    compare = brand.key_urls["compare"]
    speed = brand.key_urls["speed_test"]
    social = brand.key_urls["social_tariffs"]
    rises = brand.key_urls["price_rises"]
    switch = brand.key_urls["switch"]

    bank: list[Idea] = []

    def add(idea: Idea) -> None:
        bank.append(idea)

    add(
        Idea(
            id="X-001",
            pillar="sarcasm",
            title="Loyalty is a hostage situation",
            hook="Been with them 8 years.",
            copy=(
                "Been with the same broadband lot for 8 years.\n"
                "New customers: £19.\n"
                "Me: £41.\n\n"
                "That's not loyalty. That's a hostage situation with slightly better Wi-Fi.\n\n"
                "If you're out of contract you're not a valued customer. You're the business model."
            ),
            alt_copy="Out of contract on broadband is just a standing order you forgot to argue with.",
            cta=f"Check you're not the business model: {deals}",
            url_key="deals",
            engagement="Quote this with what you actually pay. No judgement. Well, a bit of judgement.",
            hashtags=["#UKBroadband"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "story", "tiktok", "carousel"],
            why="Loyalty-tax rage is the most shareable broadband emotion in Britain. People tag the bill-payer.",
            compliance="Do not invent a specific provider price. Talk about the industry pattern. Deals vary by postcode.",
            design=DesignBrief(
                format_name="Split bill meme / end card",
                dimensions="1080x1350 feed, 1080x1920 Reel/TikTok, 1600x900 for X",
                duration="0:07–0:12 video or static",
                visual_concept=(
                    "Two fake bills side by side. Left labelled NEW CUSTOMER in sky blue. "
                    "Right labelled LOYAL CUSTOMER in angry red-navy. Same tiny print energy as a real ISP email."
                ),
                on_screen_text=["New customer: £19", "You: £41", "That's not loyalty.", "That's a hostage situation."],
                palette=standard_palette(),
                type="Anton or Impact for the joke; Inter Bold for the end card.",
                logo=logo_rule(),
                audio="Dry VO, northern or estuary, unimpressed. Optional: hold-music sting at the end.",
                shots=["0-2s bills slam on", "2-7s punchline types on", "7-12s BroadbandPicker end card + URL"],
                do_not=["Don't use a real customer's bill", "Don't name a fake official tariff"],
            ),
        )
    )
    add(
        Idea(
            id="X-002",
            pillar="meme",
            title="Yellow warning of buffering",
            hook="YELLOW WARNING OF BUFFERING",
            copy=(
                "YELLOW WARNING OF BUFFERING\n"
                "Issued: 19:42, every evening.\n"
                "Affected: anyone who dared to watch the football, boil a kettle, and exist.\n\n"
                "Met Office couldn't be reached. They were buffering."
            ),
            alt_copy="Weather warning, but it's your Wi-Fi every time MOTD kicks off.",
            cta=f"See if your postcode has actually grown up: {home}",
            url_key="home",
            engagement="Reply with your local buffering time. Ours is 7pm like clockwork. Britain is a contended nation.",
            hashtags=["#Buffering", "#UKBroadband"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "story", "tiktok", "feed"],
            why="Met Office warning format is instantly native, screenshottable, and seasonless.",
            compliance="Parody of a public warning style is fine. Do not impersonate the Met Office logo exactly; pastiche it.",
            design=DesignBrief(
                format_name="Met Office pastiche",
                dimensions="1080x1920 and 1080x1080",
                duration="0:08 with siren-to-silence, or static",
                visual_concept=(
                    "Yellow warning triangle, map of the UK with a blob over 'wherever you live', "
                    "ticker: BUFFERING. Bottom: 'Check your postcode, not the clouds.'"
                ),
                on_screen_text=["YELLOW WARNING OF BUFFERING", "From 7pm until you give up", "Likely: Netflix, Zoom, dignity"],
                palette="Warning yellow #F4D35E on navy, white type, sky-blue BroadbandPicker footer.",
                type="A Met-Office-ish sans. Keep it official-looking then undercut it.",
                logo=logo_rule(),
                audio="Actual weather-beep then a sad buffering spinner sound.",
                shots=["0-1s warning sting", "1-6s map + copy", "6-8s logo"],
                do_not=["Don't use the real Met Office trademark lockup"],
            ),
        )
    )
    add(
        Idea(
            id="X-003",
            pillar="quote",
            title="Keep calm and restart the router",
            hook="Keep calm and restart the router",
            copy=(
                "Keep calm and restart the router.\n\n"
                "The official British response to every crisis since 2010:\n"
                "the economy, the trains, and the Wi-Fi in the back bedroom."
            ),
            alt_copy="We replaced the stiff upper lip with a blinking orange light.",
            cta=f"If restarting it is a lifestyle, compare actual full fibre: {compare}",
            url_key="compare",
            engagement="Tag the household IT department. You know the one. They are 14.",
            hashtags=["#KeepCalm", "#BritishProblems"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "feed", "story", "reel", "tiktok"],
            why="The Keep Calm format is visual shorthand for British coping. Perfect quote-card bait.",
            compliance="Keep Calm is a widely used parody. Don't sell merch of the Crown's actual poster without a licence; social parody is the use.",
            design=DesignBrief(
                format_name="Keep Calm poster",
                dimensions="1080x1350",
                duration="static + 0:06 kinetic type for Reels",
                visual_concept="Classic cream poster, crown replaced with the BroadbandPicker wifi mark, red strip KEEP CALM AND RESTART THE ROUTER.",
                on_screen_text=["KEEP CALM", "AND", "RESTART THE ROUTER"],
                palette="Ministry cream, poster red, navy type, sky wifi-mark.",
                type="A Times-like serif or a poster gothic. Not Inter for this one.",
                logo="Wifi-mark as the 'crown'. Wordmark tiny at the bottom.",
                audio="WW2 newsreel crackle, then a router reboot chord.",
                shots=["Poster hold", "slight paper zoom", "end URL"],
                do_not=["Don't add six other jokes on the poster. One line."],
            ),
        )
    )
    add(
        Idea(
            id="X-004",
            pillar="pun",
            title="Up to",
            hook="“Up to 67 Mbps”",
            copy=(
                "“Up to 67 Mbps.”\n\n"
                "Up to a biscuit with your tea.\n"
                "Up to a weekend in Wales if the rain stops.\n"
                "Up to your nan's idea of 'soon'.\n\n"
                "British broadband: a nation raised on small print."
            ),
            alt_copy="Mbps. Maybe Per Second.",
            cta=f"Look at speeds for your actual address, not the advert: {home}",
            url_key="home",
            engagement="Finish the sentence: “Up to ______.” Best reply gets a virtual biscuit.",
            hashtags=["#UpTo", "#UKBroadband"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "story", "tiktok"],
            why="Everyone has been lied to by 'up to'. It's a national in-joke with a regulator behind it.",
            compliance="Ofcom-aware: advertised speeds are typical/peak, availability is address-level. Don't claim a provider is lying; mock the industry phrasing.",
            design=DesignBrief(
                format_name="Small-print zoom",
                dimensions="1080x1920",
                duration="0:10",
                visual_concept="Giant 'UP TO 67 Mbps' in advert-land, camera zooms into microscopic footnotes until it becomes a novel.",
                on_screen_text=["UP TO 67 Mbps", "*up to a biscuit", "*up to your nan's 'soon'"],
                palette=standard_palette(),
                type="Advert grotesk, then Courier for the footnotes.",
                logo=logo_rule(),
                audio="Sales VO that gets quieter as the footnote takes over.",
                shots=["Hero claim", "zoom to footnote", "punchline", "end card"],
                do_not=["Don't fake a real provider's current advertised number"],
            ),
        )
    )
    add(
        Idea(
            id="X-005",
            pillar="british-life",
            title="89th minute",
            hook="Nothing unites Britain like the Wi-Fi dying in the 89th minute.",
            copy=(
                "Nothing unites this country like the Wi-Fi dying in the 89th minute.\n\n"
                "Left-wing. Right-wing. United in screaming at a plastic box on the skirting board.\n\n"
                "The router has never watched a game of football and it shows."
            ),
            alt_copy="Peak British sport: not the match. The buffering circle.",
            cta=f"If your household treats MOTD as a stress test, start here: {home}",
            url_key="home",
            engagement="Reply ⚽ if this has happened to you. Reply 📉 if it happened tonight.",
            hashtags=["#MOTD", "#BritishProblems"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "tiktok", "story"],
            why="Sport + broadband failure is appointment virality. Post Saturday 7–10pm and midweek fixtures.",
            compliance="Don't show pirated streams. A TV-in-living-room shot or a fake score bug is enough.",
            design=DesignBrief(
                format_name="Living room disaster sketch",
                dimensions="1080x1920",
                duration="0:09–0:15",
                visual_concept="Over-the-sofa POV, match on TV, buffering spinner, cut to a man kneeling before a router like it's a shrine.",
                on_screen_text=["89:00", "buffering", "the national sport"],
                palette="Warm living-room, then navy end card.",
                type="Scoreboard type for the time, Inter for the caption.",
                logo=logo_rule(),
                audio="Crowd roar that collapses into a buffering blip. No copyrighted commentary.",
                shots=["Match", "spinner", "router shrine", "end"],
                do_not=["Don't use Premier League footage"],
            ),
        )
    )
    add(
        Idea(
            id="X-006",
            pillar="quote",
            title="Socrates pays the bill",
            hook="The unexamined broadband bill is not worth paying.",
            copy=(
                "“The unexamined broadband bill is not worth paying.”\n"
                "— Socrates, probably, after Sky put him up £4 in April.\n\n"
                "If you have not looked at your contract since Johnson was in Downing Street, this is your sign."
            ),
            alt_copy="Know thyself. Especially the out-of-contract price on page 7.",
            cta=f"Examine it in 30 seconds: {deals}",
            url_key="deals",
            engagement="Quote-tweet the wisest thing a philosopher never said about your ISP.",
            hashtags=["#Philosophy", "#UKBills"],
            platforms=["x", "instagram"],
            formats=["tweet", "feed", "story", "carousel"],
            why="Wise-quote-plus-undercut is screenshot gold. Feels clever, reads in two seconds.",
            compliance="Parody attribution is obvious. Don't pretend it's a real quotation.",
            design=DesignBrief(
                format_name="Quote card",
                dimensions="1080x1080 and 1080x1350",
                duration="static",
                visual_concept="Marble bust with a cheap ISP router photoshopped as a laurel. Quote in a literary serif. Tiny 'probably' in sky blue.",
                on_screen_text=["The unexamined broadband bill is not worth paying.", "— Socrates, probably"],
                palette="Statue grey, navy, sky pull-quote marks.",
                type="Playfair / Source Serif for the quote, Inter for the probably.",
                logo="Small mark bottom-right, plenty of margin.",
                audio="N/A for static; optional lyre that becomes dial-up.",
                shots=["Hold the card. That's the tweet."],
                do_not=["Don't make the bust a recognisable living person"],
            ),
        )
    )
    add(
        Idea(
            id="X-007",
            pillar="british-life",
            title="Router in the cupboard",
            hook="Your router is in the cupboard under the stairs.",
            copy=(
                "Your router is in the cupboard under the stairs.\n"
                "Next to the ironing board, a bag of compost, and the ghost of a Christmas tree.\n\n"
                "And you want 4K in the back bedroom.\n\n"
                "That's not broadband. That's faith healing."
            ),
            alt_copy="Unpopular opinion: you don't need gigabit. You need the router not living with the hoover.",
            cta=f"Still slow after you move it? Then compare the actual line: {speed}",
            url_key="speed_test",
            engagement="Unpopular opinion time. Router in cupboard: yes or are you a maniac who put it in the attic?",
            hashtags=["#WiFi", "#BritishHomes"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "tiktok", "carousel"],
            why="Everyone has a cupboard router. It's a tag-your-dad post.",
            compliance="Helpful-funny: don't pretend a cupboard is always the reason. Speed-test CTA is the honest next step.",
            design=DesignBrief(
                format_name="Estate-agent tour gone wrong",
                dimensions="1080x1920",
                duration="0:12",
                visual_concept="Phone-torch tour of a British under-stairs cupboard. Reveal: router, blinking, under a pile of coats. Caption like Location Location Location.",
                on_screen_text=["The master bedroom of your Wi-Fi", "Period features: damp, beans, 12 Mbps"],
                palette="Documentary torchlight, then clean end card.",
                type="Channel 4 documentary lower-third.",
                logo=logo_rule(),
                audio="Estate agent VO: 'A generous cupboard with excellent blinking.'",
                shots=["Door opens", "coats", "router hero shot", "end"],
                do_not=["Don't film someone else's house without permission"],
            ),
        )
    )
    add(
        Idea(
            id="X-008",
            pillar="sarcasm",
            title="Engineer window",
            hook="We'll be there between 8 and 6.",
            copy=(
                "Broadband engineer: “We'll be there between 8 and 6.”\n\n"
                "That's not an appointment.\n"
                "That's a hostage video with a van.\n\n"
                "Take the day off work. Brew a flask. Stare at the street like a meerkat. "
                "This is the British wellness industry."
            ),
            alt_copy="The only thing in Britain with a wider window is a by-election.",
            cta=f"When they finally come, make the new deal worth it: {switch}",
            url_key="switch",
            engagement="What's the longest engineer window you've been given? Winner gets absolutely nothing, like the engineer.",
            hashtags=["#BritishProblems"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "tiktok", "story"],
            why="Universal UK pain. Comments become a therapy group. High reply rate.",
            compliance="Don't claim a named provider always misses slots. Mock the culture of the window.",
            design=DesignBrief(
                format_name="Calendar sketch",
                dimensions="1080x1920",
                duration="0:10",
                visual_concept="A calendar where Monday–Sunday are all highlighted. Sticky note: ENGINEER. A flask and a folding chair on the drive.",
                on_screen_text=["Appointment", "8am–6pm", "so… Tuesday?"],
                palette=standard_palette(),
                type="Utility sans, like a job sheet.",
                logo=logo_rule(),
                audio="Distant van that never arrives.",
                shots=["Job sheet", "driveway wait", "empty street", "end"],
                do_not=["Don't use a real Openreach uniform in a mocking way that implies a real engineer"],
            ),
        )
    )
    add(
        Idea(
            id="X-009",
            pillar="british-life",
            title="Full fibre is coming",
            hook="Full fibre is coming to your area.",
            copy=(
                "Full fibre is coming to your area.\n\n"
                "That's what they said about HS2.\n"
                "And a bakers in the high street.\n"
                "And a responsible government.\n\n"
                "I'll believe it when the van is on my drive, not in a press release."
            ),
            alt_copy="The Openreach van is the British Yeti. People swear they've seen it.",
            cta=f"Don't trust the poster. Check the postcode: {home}",
            url_key="home",
            engagement="If full fibre actually arrived at yours, reply with the year you were first promised it. Archaeological survey.",
            hashtags=["#FullFibre", "#UKInfrastructure"],
            platforms=["x", "instagram"],
            formats=["tweet", "feed", "story"],
            why="Infrastructure fatalism is a national sport. Rural and suburban users will share this with violence.",
            compliance="Availability really is address-level. CTA to the postcode checker is both funny and true.",
            design=DesignBrief(
                format_name="Roadworks prophecy",
                dimensions="1080x1350",
                duration="static",
                visual_concept="A weathered 'Coming soon: full fibre' council-style sign overgrown with ivy. Date on the sign: 2014.",
                on_screen_text=["COMING SOON", "full fibre", "est. any parliament now"],
                palette="Council green, rust, navy caption bar.",
                type="Highways signage type.",
                logo="Bar at the bottom: Check the postcode, not the sign.",
                audio="N/A",
                shots=["Single poster image"],
                do_not=["Don't fake an official local-authority crest"],
            ),
        )
    )
    add(
        Idea(
            id="X-010",
            pillar="meme",
            title="Drake / out of contract",
            hook="Paying the out-of-contract price vs spending four minutes comparing.",
            copy=(
                "Reject: another year of the 'loyalty' price because switching sounds like admin.\n"
                "Accept: four minutes, a postcode, and the quiet thrill of paying what new people pay.\n\n"
                "One Touch Switch exists now. Britain accidentally made something simple. Don't tell anyone."
            ),
            alt_copy="Switching used to be emigrating. Now it's a Tuesday.",
            cta=f"How switching actually works: {switch}",
            url_key="switch",
            engagement="Duet/stitch: show your actual bill with the price blurred if you like. We'll wait.",
            hashtags=["#OneTouchSwitch"],
            platforms=["instagram", "tiktok", "x"],
            formats=["reel", "tiktok", "feed", "tweet"],
            why="Drake format still prints. Pair it with a genuine consumer win (OTS).",
            compliance="One Touch Switch is real Ofcom policy. Don't promise zero downtime; say it is designed to keep you online.",
            design=DesignBrief(
                format_name="Drake / yes-no meme",
                dimensions="1080x1350",
                duration="static or 0:06",
                visual_concept="Standard two-panel Drake. Top: crumpled bill. Bottom: postcode checker screenshot (use BroadbandPicker UI, not a competitor).",
                on_screen_text=["Paying the loyalty price", "Four minutes on BroadbandPicker"],
                palette="Meme-native, end card in brand colours.",
                type="Impact for panels, Inter on the end card.",
                logo="Only on the accept panel as the UI, plus end card.",
                audio="Original audio or a well-worn meme sound — check commercial use.",
                shots=["Panel 1", "Panel 2", "URL"],
                do_not=["Don't use uncleared celebrity likeness if you can shoot an original 'nah/yeah' with a staffer"],
            ),
        )
    )
    add(
        Idea(
            id="X-011",
            pillar="pun",
            title="Fibre I barely know her",
            hook="They said I was getting fibre.",
            copy=(
                "They said I was getting fibre.\n\n"
                "I was getting fibre to a cabinet three streets away and a copper line that remembers Blair.\n\n"
                "FTTC: Fibre To The Close-enough.\n"
                "FTTP: Finally The Thing's Proper."
            ),
            alt_copy="Part-fibre is the 'seedless' grapes of infrastructure. Technically true. Spiritually a con.",
            cta=f"Full fibre vs the other thing: {brand.key_urls['guides']}/fttp-vs-fttc-explained",
            url_key="guides",
            engagement="If your 'fibre' still has a copper aftertaste, drop a 🧡. If you've got proper FTTP, gloat quietly.",
            hashtags=["#FTTP", "#FTTC"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "carousel", "reel", "tiktok"],
            why="The fibre/liar confusion is the most useful joke in the market. Teaches while it dunks.",
            compliance="Define FTTC vs FTTP accurately. Don't say copper is 'fake fibre' in a way Ofcom would hate; say part-fibre.",
            design=DesignBrief(
                format_name="Two-door cartoon",
                dimensions="1080x1350 carousel (3 slides)",
                duration="static carousel + 0:12 Reel",
                visual_concept="Two doors. One ornate 'FULL FIBRE'. One labelled 'fibre' in Comic Sans, leading to a muddy alley.",
                on_screen_text=["Fibre", "Full fibre", "One of these still involves the 1990s"],
                palette=standard_palette(),
                type="Clean Inter on slides 2–3 where you actually explain it.",
                logo="Slide 3.",
                audio="Game-show ding vs sad trombone.",
                shots=["Door gag", "one-line definition", "CTA"],
                do_not=["Don't leave people thinking all 'fibre' is a scam — some FTTC is fine for light use"],
            ),
        )
    )
    add(
        Idea(
            id="X-012",
            pillar="british-life",
            title="Flatshare FIFA",
            hook="4 people. 1 router. 3 FaceTimes. 1 FIFA.",
            copy=(
                "Four people. One router. Three FaceTimes. One FIFA.\n\n"
                "This is not a household.\n"
                "This is a failed state with a shared fridge.\n\n"
                "Student broadband is just peace talks with extra Ethernet."
            ),
            alt_copy="The real degree is conflict resolution over the last 12 Mbps.",
            cta=f"Student and renter deals exist. Shocking, I know. {brand.key_urls['guides']}/best-broadband-for-students",
            url_key="guides",
            engagement="Tag the housemate who streams in 4K 'because it's there'. Go on. Start the war in the comments.",
            hashtags=["#StudentLife", "#HMO"],
            platforms=["tiktok", "instagram", "x"],
            formats=["tiktok", "reel", "story", "tweet"],
            why="September gold. Tagging is built in. Parents will share it too, unironically.",
            compliance="Don't mock people on social tariffs in the same joke. This is chaos, not poverty.",
            design=DesignBrief(
                format_name="Security-cam kitchen",
                dimensions="1080x1920",
                duration="0:12",
                visual_concept="Top-down cheap CCTV vibe of a student kitchen. Each person labelled with the bandwidth they're stealing. FIFA lad has a halo of lag.",
                on_screen_text=["National Grid, but for arguments", "Who touched the QoS settings"],
                palette="Sickly fluorescent, then navy end card.",
                type="CCTV timestamp font.",
                logo=logo_rule(),
                audio="Overlapping Zoom hold music and a goal horn that never finishes.",
                shots=["Labels pop on each housemate", "router sparks (fake)", "end"],
                do_not=["Don't film real students without releases"],
            ),
        )
    )
    add(
        Idea(
            id="X-013",
            pillar="useful-funny",
            title="Social tariff scandal",
            hook="There's cheaper broadband if you're on certain benefits. Most eligible people have never heard of it.",
            copy=(
                "Not a joke, for once.\n\n"
                "If you're on Universal Credit, Pension Credit and a few other benefits, there are social tariffs "
                "from around the price of a couple of pints.\n\n"
                "Ofcom's own figures say most eligible households still don't know they exist.\n\n"
                "That's not a 'deal'. That's a scandal with a PDF."
            ),
            alt_copy="Funny account, serious post. Share this one even if you don't need it — someone on your street might.",
            cta=f"Plain-English explainer: {social}",
            url_key="social_tariffs",
            engagement="Share this even if it isn't you. Especially if it might be your parents.",
            hashtags=["#SocialTariff", "#UKBills"],
            platforms=["x", "instagram", "facebook"],
            formats=["tweet", "carousel", "story", "feed"],
            why="Trust-builder. Makes the funny account look like a decent citizen. Providers and Awin notice the editorial spine.",
            compliance="Eligibility is specific. Don't say 'everyone on benefits'. Link the guide. No punching down, no 'scrounger' framing.",
            design=DesignBrief(
                format_name="Calm carousel",
                dimensions="1080x1350, 5 slides",
                duration="static",
                visual_concept="Drop the meme face. Clean navy slides, big type, one fact each, final slide is the URL. This is the adult in the room.",
                on_screen_text=[
                    "If you're on certain benefits, cheaper broadband exists",
                    "Most eligible people have never heard of it",
                    "It's called a social tariff",
                    "We explain who qualifies, without the waffle",
                ],
                palette=f"Navy {BRAND_NAVY}, white type, green {BRAND_GREEN} ticks. No yellow jokes.",
                type="Inter ExtraBold headlines, Inter Regular body.",
                logo="On every slide, small, because this is branded public service.",
                audio="N/A",
                shots=["5 slides"],
                do_not=["Don't put a clown meme on this one"],
            ),
        )
    )
    add(
        Idea(
            id="X-014",
            pillar="quote",
            title="Ask not what your ISP can do",
            hook="Ask not what your ISP can do for you.",
            copy=(
                "Ask not what your ISP can do for you.\n"
                "They already told you.\n"
                "“Have you tried turning it off and on again?”\n\n"
                "— JFK, if JFK had a TalkTalk hub and a dream."
            ),
            alt_copy="The only inaugural address your ISP will ever give is a reboot instruction.",
            cta=f"If you've already rebooted it, compare a line that doesn't need a sermon: {compare}",
            url_key="compare",
            engagement="What's the most useless thing an ISP chatbot has said to you? We'll make a shrine.",
            hashtags=["#CustomerService"],
            platforms=["x", "instagram"],
            formats=["tweet", "feed", "story"],
            why="Classic quote + IT Crowd callback. Two generations in one joke.",
            compliance="Parody. Don't use a presidential seal. 'TalkTalk hub' is illustrative, not a product claim.",
            design=DesignBrief(
                format_name="Oval Office / airing cupboard",
                dimensions="1080x1080",
                duration="static",
                visual_concept="Presidential speech desk. Behind: not a flag, a router with a sad orange light.",
                on_screen_text=["HAVE YOU TRIED", "TURNING IT OFF", "AND ON AGAIN"],
                palette="Navy, red stripe as a visual gag not a party colour, sky wifi-mark as a lapel pin.",
                type="Inscribed serif.",
                logo="Lapel-pin wifi mark + bottom URL.",
                audio="N/A",
                shots=["Single image"],
                do_not=["Don't use an official US or UK government lockup"],
            ),
        )
    )
    add(
        Idea(
            id="X-015",
            pillar="pun",
            title="Bandwidth like a British summer",
            hook="My bandwidth is like a British summer.",
            copy=(
                "My bandwidth is like a British summer.\n"
                "Brief.\n"
                "Disappointing.\n"
                "And somehow still £32 a month.\n\n"
                "Pack a cardigan and a speed test."
            ),
            alt_copy="Two weeks of Wimbledon, then packet loss until April.",
            cta=f"Run the speed test. Take a cardigan. {speed}",
            url_key="speed_test",
            engagement="Other British things my broadband is like: a queue, a rail replacement bus, a barbecue in March…",
            hashtags=["#BritishSummer", "#UKBroadband"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "story", "reel", "tiktok"],
            why="Weather is the UK's second language. Easy stitches and duet templates.",
            compliance="Humour, not a measured claim about a named network.",
            design=DesignBrief(
                format_name="Holiday postcard",
                dimensions="1080x1080",
                duration="static + 0:06",
                visual_concept="Kiss Me Quick postcard: deckchairs, grey sky, a speed-test result stamped like a passport: 11 Mbps.",
                on_screen_text=["WISH YOU WEREN'T HERE", "bandwidth as brief as the summer"],
                palette="Faded postcard reds, navy caption.",
                type="Retro script + Inter.",
                logo="Stamp in the corner using the wifi mark.",
                audio="Seagulls and rain.",
                shots=["Postcard zoom"],
                do_not=["Don't use a real person's holiday photo"],
            ),
        )
    )
    add(
        Idea(
            id="X-016",
            pillar="sarcasm",
            title="April, the other tax month",
            hook="April: council tax, broadband, and the urge to move to France.",
            copy=(
                "April in Britain:\n"
                "• council tax up\n"
                "• broadband up\n"
                "• sudden interest in moving to Brittany\n\n"
                "Ofcom made them put the rise in pounds, not 'CPI plus a vibe'.\n"
                "You can still leave. That's the bit they don't put on the fridge magnet."
            ),
            alt_copy="Price-rise season is just Halloween for your standing order.",
            cta=f"What a rise actually means for your contract: {rises}",
            url_key="price_rises",
            engagement="If a provider has emailed you an April rise, drop the £ amount (not your account number, we're not animals).",
            hashtags=["#CostOfLiving", "#Ofcom"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "carousel", "reel", "tiktok", "story"],
            why="Cyclical viral moment. Prep creative in February, ship when the emails land.",
            compliance="Cite the pounds-and-pence rule accurately. Leaving still depends on the contract. Link the rights guide.",
            design=DesignBrief(
                format_name="Fridge-magnet budget",
                dimensions="1080x1350 carousel",
                duration="static + 0:15 explain-with-a-smirk Reel",
                visual_concept="A British fridge: council tax letter, broadband rise letter, a faded photo of France.",
                on_screen_text=["April", "the other tax month", "You can still leave"],
                palette=standard_palette(),
                type="Inter.",
                logo="Last slide.",
                audio="Countdown to the 1st of April, then a kettle.",
                shots=["Fridge", "letters", "rights line", "CTA"],
                do_not=["Don't tell people they can always leave fee-free — that's not always true"],
            ),
        )
    )
    add(
        Idea(
            id="X-017",
            pillar="meme",
            title="Wi-Fi name personality test",
            hook="Your Wi-Fi name is a personality test.",
            copy=(
                "Your Wi-Fi name is a personality test.\n\n"
                "BTHub-19A2 — you have given up.\n"
                "PrettyFlyForAWiFi — you peaked in 2014.\n"
                "FBI Surveillance Van — you have a nephew.\n"
                "The Password Is On The Fridge — you are the backbone of this country.\n\n"
                "Drop yours. We'll diagnose you for free. The broadband, not the personality. That's beyond us."
            ),
            alt_copy="Tell me your SSID without telling me your SSID.",
            cta=f"After the roast, compare something that isn't named after a crime: {home}",
            url_key="home",
            engagement="PRIMARY CTA is the comments. This post exists to farm replies and quote-tweets.",
            hashtags=["#WiFiName"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "story", "reel", "tiktok"],
            why="Highest projected comment rate in the bank. Names are identity. People tag the person who named it.",
            compliance="Don't encourage passwords in comments. Names only. Joke-ban anyone who posts a password.",
            design=DesignBrief(
                format_name="Phone Wi-Fi picker",
                dimensions="1080x1920",
                duration="0:10",
                visual_concept="iOS/Android Wi-Fi list with joke names. A finger hovers, never connects.",
                on_screen_text=["Choose your fighter", "password in the comments = banned"],
                palette="OS-native then brand end card.",
                type="SF/Roboto for the fake OS, Inter at the end.",
                logo=logo_rule(),
                audio="UI clicks.",
                shots=["Scroll the list", "tap a ridiculous name", "end"],
                do_not=["Don't include a real neighbour's network from a captured screenshot"],
            ),
        )
    )
    add(
        Idea(
            id="X-018",
            pillar="useful-funny",
            title="Speed you actually need",
            hook="You do not need a gigabit to email your solicitor.",
            copy=(
                "You do not need a gigabit to email your solicitor.\n\n"
                "One person, Netflix in HD: a few dozen Mbps will do.\n"
                "Family, WFH, 4K, a teenager: now we can talk.\n"
                "Gaming: ping matters more than a vanity download number.\n\n"
                "The upgrade you need might be moving the router. The upgrade you want is bragging at the pub."
            ),
            alt_copy="Gigabit is a personality. 150 Mbps is a household.",
            cta=f"Match the household, not the advert: {brand.key_urls['match']}",
            url_key="match",
            engagement="Reply with: number of humans vs number of devices. We'll guess your monthly pain.",
            hashtags=["#BroadbandSpeed"],
            platforms=["instagram", "x", "tiktok"],
            formats=["carousel", "reel", "tweet", "tiktok"],
            why="Saves people money, which is how you become the trusted comparison brand rather than a hype merchant.",
            compliance="Use ranges, not fake precision. Always 'typical household' language.",
            design=DesignBrief(
                format_name="Pub-brag vs reality carousel",
                dimensions="1080x1350, 6 slides",
                duration="static + 0:20 Reel voiceover",
                visual_concept="Slide 1 joke, slides 2–5 useful table, slide 6 Broadband Match CTA. Let the joke earn the table.",
                on_screen_text=["Gigabit is a personality", "Here's what you actually need"],
                palette=standard_palette(),
                type="Inter throughout. This one should look like the website grew a sense of humour.",
                logo="All slides, small.",
                audio="Pub chatter under the VO, then clean.",
                shots=["Joke", "table", "quiz CTA"],
                do_not=["Don't invent a 'perfect Mbps' for every home"],
            ),
        )
    )
    add(
        Idea(
            id="X-019",
            pillar="sarcasm",
            title="Same copper, different jumper",
            hook="Choosing broadband by TV advert is like choosing a pension by jingle.",
            copy=(
                "Half the 'different' broadband brands are the same Openreach line in a different jumper.\n\n"
                "You're not picking a network.\n"
                "You're picking a mail-merge, a modem, and how badly they hide the price in month 19.\n\n"
                f"We put {providers_sample} and the rest in one place, because life is too short to open twelve tabs."
            ),
            alt_copy="It's the same picture. It's often the same cabinet.",
            cta=f"One table, several jumpers: {compare}",
            url_key="compare",
            engagement="Reply with the brand you thought was 'the network' and we'll gently ruin Christmas.",
            hashtags=["#Openreach", "#Altnets"],
            platforms=["x", "instagram"],
            formats=["tweet", "carousel", "feed"],
            why="Makes BroadbandPicker look like the adult comparison site. Also invites altnet fans to argue, which is free reach.",
            compliance="True for many Openreach resellers; NOT true for Virgin Media, Hyperoptic, Community Fibre, etc. Say 'often', show the exceptions.",
            design=DesignBrief(
                format_name="It's the same picture",
                dimensions="1080x1080",
                duration="static",
                visual_concept="Spider-Man pointing meme energy, but two identical Openreach cabinets wearing different knitted jumpers with brand-ish colours (not logos).",
                on_screen_text=["It's the same cabinet", "Different jumper"],
                palette="Street photography + brand bar.",
                type="Inter.",
                logo="Bar.",
                audio="N/A",
                shots=["Single image"],
                do_not=["Don't put real ISP logos on the jumpers"],
            ),
        )
    )
    add(
        Idea(
            id="X-020",
            pillar="british-life",
            title="Zoom, you're muted, no it's the line",
            hook="You're muted. No — that's the broadband.",
            copy=(
                "“You're muted.”\n"
                "“I'm not muted.”\n"
                "“We can see you talking.”\n"
                "“That's the broadband doing impressionism.”\n\n"
                "WFH Britain: professional on the top half, 2005 on the connection."
            ),
            alt_copy="Upload speed is a personality flaw and HR knows.",
            cta=f"Best broadband for working from home, minus the TED talk: {brand.key_urls['guides']}/best-broadband-for-working-from-home",
            url_key="guides",
            engagement="HR people, drop the funniest thing you've seen a lag do to a grown adult.",
            hashtags=["#WFH", "#Zoom"],
            platforms=["tiktok", "instagram", "x"],
            formats=["tiktok", "reel", "tweet", "story"],
            why="WFH is still a default UK identity. Employers will share it. Dangerous in a good way.",
            compliance="Don't use real Zoom UI in a way that implies endorsement. Generic grid is enough.",
            design=DesignBrief(
                format_name="Video-call grid sketch",
                dimensions="1080x1920",
                duration="0:11",
                visual_concept="2x2 grid. Three crisp colleagues. One pixelated impressionist painting that used to be a man in Stevenage.",
                on_screen_text=["You're muted", "That's the broadband"],
                palette="App greys, then brand end card.",
                type="UI type.",
                logo=logo_rule(),
                audio="Warped VO of 'can you hear me now'.",
                shots=["Grid", "pixel man talks", "others freeze", "end"],
                do_not=["Don't mock a real person's appearance; keep it about the pixels"],
            ),
        )
    )
    add(
        Idea(
            id="X-021",
            pillar="pun",
            title="Packet loss packed lunch",
            hook="It's not you. It's packet loss.",
            copy=(
                "It's not you. It's packet loss.\n\n"
                "Your relationship with this ISP has run its course.\n"
                "The packets left and they're not coming back.\n"
                "Take the router. Leave the contract."
            ),
            alt_copy="Let's split the Ethernet. I'll take the kids' iPads.",
            cta=f"Break up properly: {switch}",
            url_key="switch",
            engagement="Break-up texts for your ISP. We'll post the best ones. Keep it PG, this is broadband not Love Island.",
            hashtags=["#PacketLoss"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "tiktok", "story"],
            why="Romance-breakup format is a native meme. Easy stitches.",
            compliance="Humour. Don't claim a named ISP has unusual packet loss without data.",
            design=DesignBrief(
                format_name="Break-up text",
                dimensions="1080x1920",
                duration="0:08",
                visual_concept="iMessage conversation with 'Your ISP'. Grey bubble: 'it's not you it's packet loss.'",
                on_screen_text=["Delivered", "read at 19:42", "during the football, obviously"],
                palette="iMessage blue/grey, navy end card.",
                type="SF-like.",
                logo=logo_rule(),
                audio="Text-send pings.",
                shots=["Messages appear", "end"],
                do_not=["Don't use Apple's trademarked iMessage lockup in ads — organic pastiche is the usual risk level"],
            ),
        )
    )
    add(
        Idea(
            id="X-022",
            pillar="british-life",
            title="Greggs vs gigabit",
            hook="I'd give up a sausage roll for 900 Mbps.",
            copy=(
                "I'd give up a sausage roll for 900 Mbps.\n\n"
                "I would not give up the steak bake.\n"
                "There is a line, even in this economy.\n\n"
                "Comment your price. We'll tell you if it's a steak bake or a sad vegan roll of a contract."
            ),
            alt_copy="Britain's two religions: pastry and a solid connection in the kitchen.",
            cta=f"Find out if you're overpaying for crumbs: {deals}",
            url_key="deals",
            engagement="Sausage roll or steak bake: what is your broadband actually worth?",
            hashtags=["#Greggs", "#Britain"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "story", "reel", "tiktok"],
            why="Greggs is UK engagement catnip. Harmless, visual, comment-driven.",
            compliance="Don't use Greggs branding as if it's a partnership. Generic pastry is safer than the logo.",
            design=DesignBrief(
                format_name="Fake menu board",
                dimensions="1080x1350",
                duration="static + 0:07",
                visual_concept="Café menu: Sausage roll / 70 Mbps. Steak bake / full fibre. Vegan roll / out of contract price.",
                on_screen_text=["TODAY'S SPECIALS", "dignity: sold out"],
                palette="Warm bakery, navy footer.",
                type="Menu board gothic.",
                logo="Footer.",
                audio="Café noise.",
                shots=["Menu pan"],
                do_not=["Don't imply Greggs endorses BroadbandPicker"],
            ),
        )
    )
    add(
        Idea(
            id="X-023",
            pillar="meme",
            title="TfL delay board",
            hook="Wi-Fi delayed due to signalling problems in the cupboard.",
            copy=(
                "Wi-Fi delayed due to signalling problems in the cupboard under the stairs.\n\n"
                "Replacement bus service: 4G from your phone, if you stand by the window and believe.\n\n"
                "Mind the packet gap."
            ),
            alt_copy="This is your captain speaking. We are currently held at the cabinet.",
            cta=f"{home}",
            url_key="home",
            engagement="Write the delay reason for your line. 'Leaves on the copper' is already taken.",
            hashtags=["#TfL", "#BritishProblems"],
            platforms=["instagram", "x", "tiktok"],
            formats=["feed", "story", "reel", "tweet", "tiktok"],
            why="Every Londoner, and every Brit who has stared at a departure board, gets it immediately.",
            compliance="Pastiche the board. Don't copy TfL's roundel.",
            design=DesignBrief(
                format_name="Station information board",
                dimensions="1080x1920",
                duration="0:08",
                visual_concept="Dot-matrix board: WIFI  delayed  84 min. Reason: CUPBOARD. A tiny wifi mark as the operator.",
                on_screen_text=["WIFI", "Delayed", "Signalling problems in the cupboard"],
                palette="Departure-board orange on black, sky-blue end card.",
                type="Dot matrix / LED.",
                logo="As the 'operator' code BBP.",
                audio="Platform beep.",
                shots=["Board flicker", "end"],
                do_not=["Don't use the TfL roundel or Johnston type commercially if you're running ads — organic pastiche, original type"],
            ),
        )
    )
    add(
        Idea(
            id="X-024",
            pillar="useful-funny",
            title="Things faster than your broadband",
            hook="Things that are currently faster than your broadband.",
            copy=(
                "Things currently faster than your broadband:\n"
                "• a queue at Greggs\n"
                "• a Southern Rail apology\n"
                "• your nan's gossip\n"
                "• the kettle\n"
                "• the government changing its mind\n\n"
                "If this hurt, it was supposed to. Speed-test it. Then compare it."
            ),
            alt_copy="A ranked list of national disappointments, with your ISP in the medals.",
            cta=f"{speed}",
            url_key="speed_test",
            engagement="Add one. Clean, please. This is a family comparison site with a filthy sense of timing.",
            hashtags=["#BritishProblems"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "tiktok", "carousel"],
            why="Listicle + roast = reply magnet. Each reply is a free extra joke.",
            compliance="Hyperbole. Don't present as a measured benchmark.",
            design=DesignBrief(
                format_name="Olympic medal table",
                dimensions="1080x1350",
                duration="0:12",
                visual_concept="Medal table. Gold: Nan's gossip. Silver: kettle. Bronze: your download speed, looking ashamed.",
                on_screen_text=["MEDAL TABLE", "household events"],
                palette=standard_palette(),
                type="Sports graphic.",
                logo=logo_rule(),
                audio="BBC sport ident pastiche (original, not the real ident).",
                shots=["Table fills in", "speed-test CTA"],
                do_not=["Don't use BBC branding"],
            ),
        )
    )
    add(
        Idea(
            id="X-025",
            pillar="b2b",
            title="Providers: we will list you anyway",
            hook="We will list your broadband even if you never pay us a penny.",
            copy=(
                "BroadbandPicker lists UK providers whether or not we have an affiliate deal.\n\n"
                "If we earn a commission when someone switches, we say so.\n"
                "If we don't, we still show the package, because hiding the market is how comparison sites get a reputation.\n\n"
                f"Partnerships: {brand.emails['partnerships']}\n"
                f"How we make money: {brand.key_urls['home']}/how-we-make-money"
            ),
            alt_copy="Editorial independence is not a vibe. It is a page on the website.",
            cta=f"Affiliate and provider enquiries: {brand.emails['partnerships']}",
            url_key="partnerships_mail",
            engagement="If you run an ISP or an affiliate programme, the DMs are open. If you run a comparison site that hides this, sit down.",
            hashtags=["#AffiliateMarketing", "#UKBroadband"],
            platforms=["linkedin", "x"],
            formats=["tweet", "linkedin"],
            why="This is the post that gets BroadbandPicker onto provider radars. Serious, still a bit dry.",
            compliance="Must match /how-we-make-money. No 'guaranteed placements'. No fake traffic numbers.",
            design=DesignBrief(
                format_name="LinkedIn document post",
                dimensions="1920x1080 document, 3 pages",
                duration="static",
                visual_concept="Clean navy pages: 1) we list you anyway 2) methodology 3) contact. Looks like a mini media kit.",
                on_screen_text=["We list providers with or without an affiliate deal", brand.emails["partnerships"]],
                palette=f"Navy {BRAND_NAVY}, sky {BRAND_PRIMARY}, white.",
                type="Inter only. B2B, not meme.",
                logo="Header on every page.",
                audio="N/A",
                shots=["3-page carousel"],
                do_not=["Don't use sarcasm so hard that a partnerships manager thinks you're unserious"],
            ),
        )
    )
    add(
        Idea(
            id="X-026",
            pillar="meme",
            title="2am vs 7pm speed test",
            hook="Speed test at 2am vs speed test at 7pm.",
            copy=(
                "Speed test at 2am: I am a god.\n"
                "Speed test at 7pm: I am in a group project with the entire cul-de-sac.\n\n"
                "Contention is just British for 'everyone else had the same idea'."
            ),
            alt_copy="Your neighbourhood, buffering in unison. Community, if you squint.",
            cta=f"Test it twice. Then come back: {speed}",
            url_key="speed_test",
            engagement="Post both numbers. 2am and 7pm. We'll nominate a street for a blue plaque.",
            hashtags=["#SpeedTest"],
            platforms=["tiktok", "instagram", "x"],
            formats=["tiktok", "reel", "tweet", "story"],
            why="Participatory. People love posting their numbers. User-generated content without asking for it.",
            compliance="Remind people results vary. Don't diagnose a named network from one screenshot.",
            design=DesignBrief(
                format_name="Before / after speed-test UI",
                dimensions="1080x1920",
                duration="0:08",
                visual_concept="Two phones. Nightstand 2am: huge number. Sofa 7pm: a number that wants to go home.",
                on_screen_text=["02:00", "19:00", "same house, different country"],
                palette=standard_palette(),
                type="App UI + Inter captions.",
                logo=logo_rule(),
                audio="Whoosh from night to evening TV.",
                shots=["2am", "smash cut 7pm", "CTA to our speed test"],
                do_not=["Don't fake a provider's official speed-test result page"],
            ),
        )
    )
    add(
        Idea(
            id="X-027",
            pillar="quote",
            title="Hell is other people's Netflix",
            hook="Hell is other people's Netflix on your connection.",
            copy=(
                "“Hell is other people's Netflix.”\n"
                "— Sartre, on a TalkTalk 40/10, 2013, probably.\n\n"
                "If you can hear your partner's streaming from the ping of your spreadsheet, it is time."
            ),
            alt_copy="L'enfer, c'est les autres onglets.",
            cta=f"{brand.key_urls['guides']}/best-broadband-for-streaming",
            url_key="guides",
            engagement="Who is the Netflix in your relationship? Tag them. Then hide.",
            hashtags=["#Streaming"],
            platforms=["x", "instagram"],
            formats=["tweet", "feed", "story"],
            why="Couple-tag bait. Quote cards get saved.",
            compliance="Parody attribution. No real provider performance claim.",
            design=DesignBrief(
                format_name="Penguin Classics pastiche",
                dimensions="1080x1350",
                duration="static",
                visual_concept="Orange-spine book cover: HELL IS OTHER PEOPLE'S NETFLIX. Author: your ping.",
                on_screen_text=["HELL IS OTHER PEOPLE'S NETFLIX"],
                palette="Penguin orange, navy, cream.",
                type="A literary grotesque. Not the real Penguin lockup.",
                logo="Tiny.",
                audio="N/A",
                shots=["Cover"],
                do_not=["Don't copy Penguin's colophon"],
            ),
        )
    )
    add(
        Idea(
            id="X-028",
            pillar="british-life",
            title="Moving house broadband",
            hook="Moving house is two crises in a trench coat. One of them is broadband.",
            copy=(
                "Moving house checklist:\n"
                "• boxes\n"
                "• kettle first\n"
                "• argue about the sofa\n"
                "• discover the new place 'has fibre' in the same way it 'has storage'\n\n"
                "Start the broadband before you start the argument. One of these is reversible."
            ),
            alt_copy="You can live without a sofa. You cannot live without moaning about the Wi-Fi.",
            cta=f"Moving-house broadband, in order: {brand.key_urls['guides']}/broadband-moving-house",
            url_key="guides",
            engagement="What's the first thing you set up in a new place: kettle, Wi-Fi, or the telly? There is a wrong answer.",
            hashtags=["#MovingHouse"],
            platforms=["instagram", "x", "tiktok"],
            formats=["carousel", "reel", "tweet", "tiktok", "story"],
            why="High commercial intent dressed as a sitcom. Peak May–September, still works year-round.",
            compliance="Setup times vary. Don't promise 'online the day you move' as a universal fact.",
            design=DesignBrief(
                format_name="Checklist on a cardboard box",
                dimensions="1080x1350",
                duration="static + 0:12",
                visual_concept="Sharpie on a moving box. Last item, circled: ACTUALLY CHECK THE POSTCODE.",
                on_screen_text=["Kettle", "Wi-Fi", "Sofa (optional)"],
                palette="Cardboard, navy sharpie, sky circle.",
                type="Handwritten then Inter CTA.",
                logo="Tape-style end card.",
                audio="Packing tape.",
                shots=["Box", "circle the postcode line", "CTA"],
                do_not=["Don't show a real new address"],
            ),
        )
    )
    add(
        Idea(
            id="X-029",
            pillar="sarcasm",
            title="Chatbot of Theseus",
            hook="I asked the ISP chatbot if I could leave. It asked if I'd tried restarting myself.",
            copy=(
                "Me: I'd like to leave, the contract ended in March.\n"
                "Chatbot: I understand this must be frustrating. Have you tried restarting the router?\n"
                "Me: The router isn't the thing keeping me here. You are.\n"
                "Chatbot: I'm sorry you feel that way. Here's a link to our blog about cookies.\n\n"
                "One Touch Switch exists so you never have to finish this play."
            ),
            alt_copy="Customer service as interpretive dance.",
            cta=f"{switch}",
            url_key="switch",
            engagement="Paste the worst chatbot line you've been sent. We'll workshop a tragedy.",
            hashtags=["#CustomerService"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "tiktok", "carousel"],
            why="Everyone has the screenshot. User-generated follow-up content is almost guaranteed.",
            compliance="Fictional composite. Don't reproduce a real private chat with account data.",
            design=DesignBrief(
                format_name="Chat UI sketch",
                dimensions="1080x1920",
                duration="0:14",
                visual_concept="Generic chat bubbles. The bot remains cheerfully useless. A 'human' button that is greyed out until 2029.",
                on_screen_text=["Have you tried restarting yourself?"],
                palette="Chat greys, navy end.",
                type="UI type.",
                logo=logo_rule(),
                audio="Hold music under the captions.",
                shots=["Messages ping in", "end card One Touch Switch"],
                do_not=["Don't use a real ISP's chatbot skin"],
            ),
        )
    )
    add(
        Idea(
            id="X-030",
            pillar="pun",
            title="Broad band",
            hook="It's called broadband because the bills are.",
            copy=(
                "It's called broadband because the bills are.\n\n"
                "Narrow money. Broad cheek.\n"
                "If your tariff no longer fits, that's not a glow-up. That's month 19."
            ),
            alt_copy="Broad band, tight budget, classic Britain.",
            cta=f"{brand.key_urls['calculator']}",
            url_key="calculator",
            engagement="Rate this pun 1–10. 1 is 'leave the country'. 10 is 'put it on a tea towel'.",
            hashtags=["#DadJokes", "#UKBills"],
            platforms=["x", "instagram"],
            formats=["tweet", "story", "feed"],
            why="Dad-joke posts get tagged at dads. Cheap, cheerful, on-brand if you wink at the camera.",
            compliance="None special.",
            design=DesignBrief(
                format_name="Tea towel",
                dimensions="1080x1350",
                duration="static",
                visual_concept="Souvenir tea towel: It's called broadband because the bills are. Red, white, sky blue. Deeply sincere kitsch.",
                on_screen_text=["IT'S CALLED BROADBAND BECAUSE THE BILLS ARE"],
                palette="Kitsch Union-adjacent colours without using the flag as a prop if it feels shouty — navy/sky/white is enough.",
                type="Tea-towel serif.",
                logo="Woven in the corner.",
                audio="N/A",
                shots=["Product shot of a fake tea towel"],
                do_not=["Don't actually print it until trademarks and jokes are cleared"],
            ),
        )
    )
    add(
        Idea(
            id="X-031",
            pillar="useful-funny",
            title="Poll: who is the IT department",
            hook="Who is the IT department in your house?",
            copy=(
                "Who is the IT department in your house?\n\n"
                "A) The eldest child\n"
                "B) The person who pays the bill, bitterly\n"
                "C) A bloke called Steve who moved out in 2019\n"
                "D) The router. We do not question the router.\n\n"
                "Vote. Then tag the victim."
            ),
            alt_copy="Every British home has an unpaid sysadmin.",
            cta=f"{home}",
            url_key="home",
            engagement="This IS the engagement. Use native polls on X, IG Stories, TikTok stickers.",
            hashtags=["#BritishHomes"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "story", "tiktok", "reel"],
            why="Polls are the cheapest viral mechanic. Tags do the distribution.",
            compliance="Fine.",
            design=DesignBrief(
                format_name="Native poll + still",
                dimensions="1080x1920 story, X native poll",
                duration="24h story, tweet lives",
                visual_concept="Four illustrated household saints. Steve has a halo of Ethernet.",
                on_screen_text=["WHO IS IT SUPPORT", "tag them"],
                palette=standard_palette(),
                type="Inter.",
                logo="Corner.",
                audio="Optional VO reading the options.",
                shots=["Question", "options", "sticker"],
                do_not=["Don't hide the poll under a 400-word caption"],
            ),
        )
    )
    add(
        Idea(
            id="X-032",
            pillar="british-life",
            title="Rural Yeti van",
            hook="The fibre van is the British Yeti.",
            copy=(
                "People in the village have seen the fibre van.\n"
                "A man in the pub drew a picture of it on a beer mat.\n"
                "The council newsletter said 'imminently' in 2018.\n\n"
                "If you're rural, 5G home and the honest conversation about not-spots is not a cop-out. It's a plan B with a sim."
            ),
            alt_copy="Not-spot Britain deserves better than another 'coming soon'.",
            cta=f"{brand.key_urls['guides']}/best-broadband-for-rural-areas-uk",
            url_key="guides",
            engagement="Rural lot: year you were first promised fibre, in the replies. We'll plot the grief.",
            hashtags=["#RuralBritain", "#NotSpot"],
            platforms=["x", "instagram"],
            formats=["tweet", "feed", "carousel"],
            why="Underserved audience, high gratitude, high share-to-parish-Facebook potential.",
            compliance="Don't pretend every rural postcode is hopeless. Point to 5G home / satellite / upcoming FTTP where true.",
            design=DesignBrief(
                format_name="Beer-mat cryptozoology",
                dimensions="1080x1080",
                duration="static",
                visual_concept="Pub table, beer mat sketch of a van labelled FIBRE. Like a cryptid sighting.",
                on_screen_text=["SIGHTING", "unconfirmed", "third pint"],
                palette="Pub wood, navy caption.",
                type="BiRo handwriting.",
                logo="Beer-mat corner.",
                audio="N/A",
                shots=["Table still"],
                do_not=["Don't mock rural users as backward — mock the promises"],
            ),
        )
    )
    add(
        Idea(
            id="X-033",
            pillar="meme",
            title="Expanding brain",
            hook="Expanding brain of British broadband.",
            copy=(
                "Restart the router.\n"
                "Shout at the ceiling.\n"
                "Stand on one leg by the window for 4G.\n"
                "Check the postcode and switch, like an adult.\n\n"
                "The last one feels illegal. It isn't. One Touch Switch is sitting there, unpaid, like a gym membership for your dignity."
            ),
            alt_copy="Personal growth, but for standing orders.",
            cta=f"{switch}",
            url_key="switch",
            engagement="Which level are you on. Be honest. The ceiling-shouters always lie.",
            hashtags=["#Memes"],
            platforms=["instagram", "tiktok", "x"],
            formats=["feed", "reel", "tiktok", "tweet"],
            why="Format does the teaching. Last panel is the product.",
            compliance="Switching is usually straightforward now; still mention early-termination fees if in contract.",
            design=DesignBrief(
                format_name="Expanding brain",
                dimensions="1080x1350",
                duration="static",
                visual_concept="Four-panel brain. Final panel is the BroadbandPicker postcode box glowing like the holy grail.",
                on_screen_text=["restart", "shout", "4G window", "switch like an adult"],
                palette="Meme-native, brand on last panel only.",
                type="Impact.",
                logo="Last panel UI.",
                audio="Galaxy brain sting for Reels.",
                shots=["Panels stack", "end"],
                do_not=["Don't make the first three panels so funny that nobody reads the fourth"],
            ),
        )
    )
    add(
        Idea(
            id="X-034",
            pillar="b2b",
            title="Awin-ready publisher energy",
            hook="Comparison sites that only shout 'CHEAPEST DEAL' are why advertisers invented brand guidelines.",
            copy=(
                "BroadbandPicker is being built as a proper UK publisher:\n"
                "independent rankings, public methodology, honest 'how we make money',\n"
                f"and {len(brand.providers) or 12}+ providers on the record.\n\n"
                "If you run a broadband affiliate programme and you want a partner that will still mention the ugly out-of-contract price, you know where we are."
            ),
            alt_copy="We will not give you 5 stars because the cookie is tasty.",
            cta=f"{brand.emails['partnerships']} · {SITE}/how-we-review-broadband",
            url_key="partnerships_mail",
            engagement="Partnerships teams: what does a good publisher look like to you? We'll do that, then the jokes.",
            hashtags=["#Awin", "#Affiliate"],
            platforms=["linkedin", "x"],
            formats=["linkedin", "tweet"],
            why="Directly serves the 'get providers to approach us' goal.",
            compliance="No fake traffic. No 'official partner of' until you are.",
            design=DesignBrief(
                format_name="Media-kit still",
                dimensions="1200x627 LinkedIn",
                duration="static",
                visual_concept="Simple stat tiles from the workspace: provider count, guide count, trust pages live.",
                on_screen_text=["Independent comparison", "Affiliate-honest", "Built for UK broadband"],
                palette=f"Navy {BRAND_NAVY} + sky {BRAND_PRIMARY}.",
                type="Inter ExtraBold.",
                logo="Top-left.",
                audio="N/A",
                shots=["One still"],
                do_not=["Don't put meme Impact font on LinkedIn"],
            ),
        )
    )
    add(
        Idea(
            id="X-035",
            pillar="useful-funny",
            title="Black Friday is not a personality",
            hook="Black Friday broadband is just a sale on the same copper with a louder email.",
            copy=(
                "Black Friday broadband is just a sale on the same copper with a louder email.\n\n"
                "If the postcode doesn't get the network, 40% off nothing is still nothing.\n"
                "We'll track the deals that are actually deals. The rest can fight in your spam folder."
            ),
            alt_copy="Cyber Monday, but make it an 18-month contract. Romantic.",
            cta=f"{brand.key_urls['guides']}/black-friday-broadband-deals-uk",
            url_key="guides",
            engagement="What was last year's 'deal' that turned into a pumpkin in month 13?",
            hashtags=["#BlackFriday"],
            platforms=["x", "instagram", "tiktok"],
            formats=["tweet", "reel", "tiktok", "carousel", "story"],
            why="Seasonal search spike. Funny scepticism builds more trust than screaming SALE.",
            compliance="Date the prices. Don't advertise a specific saving unless verified that day.",
            design=DesignBrief(
                format_name="Red-slash sale parody",
                dimensions="1080x1920",
                duration="0:10",
                visual_concept="Screaming red SALE graphics that peel off to reveal a boring navy fact about postcodes.",
                on_screen_text=["40% OFF", "the same cabinet", "check the postcode"],
                palette="Sale red then brand navy — the point is the peel.",
                type="Shouty then Inter.",
                logo=logo_rule(),
                audio="Shop tannoy that becomes a kettle.",
                shots=["SALE", "peel", "postcode"],
                do_not=["Don't run unverified 'from £x' in the caption"],
            ),
        )
    )
    add(
        Idea(
            id="X-036",
            pillar="quote",
            title="To switch or not to switch",
            hook="To switch, or not to switch.",
            copy=(
                "To switch, or not to switch — that is the question.\n"
                "(It's to switch, Derek. You're out of contract and they're charging you like you personally laid the fibre.)\n\n"
                "Whether 'tis nobler on the sofa to suffer the slings and arrows of outrageous ping…"
            ),
            alt_copy="Shakespeare, if Shakespeare had Virgin Media and a deadline.",
            cta=f"{deals}",
            url_key="deals",
            engagement="Finish the soliloquy. Best iambic pentameter about a hub gets immortalised in a Story.",
            hashtags=["#Shakespeare", "#UKBroadband"],
            platforms=["x", "instagram"],
            formats=["tweet", "feed", "story"],
            why="Quote-remix pillar in its purest form. Attracts the 'actually' crowd, who share loudly.",
            compliance="Obvious parody.",
            design=DesignBrief(
                format_name="Folio page",
                dimensions="1080x1350",
                duration="static",
                visual_concept="Fake First Folio page. Stage direction: Enter ROUTER, blinking. Derek, on the sofa, pays £41.",
                on_screen_text=["To switch, or not to switch", "Enter ROUTER"],
                palette="Aged paper, blackletter-ish (readable), sky stage directions.",
                type="A readable old-style serif, not unreadable blackletter.",
                logo="Printer's mark wifi icon.",
                audio="Optional lute that becomes buffering.",
                shots=["Page"],
                do_not=["Don't make it so 'designy' that nobody reads Derek"],
            ),
        )
    )
    return bank


# ---------------------------------------------------------------------------
# Calendars
# ---------------------------------------------------------------------------

WEEKDAY_SLOTS = {
    0: ("08:30", "18:30"),  # Mon
    1: ("12:30", "19:00"),  # Tue
    2: ("08:30", "18:30"),  # Wed
    3: ("12:30", "19:00"),  # Thu
    4: ("12:00", "19:30"),  # Fri
    5: ("10:00", "19:00"),  # Sat
    6: ("10:00", "18:00"),  # Sun
}


def monday_on_or_after(start: date) -> date:
    return start if start.weekday() == 0 else start + timedelta(days=(7 - start.weekday()))


def build_calendar(start: date, weeks: int, bank: list[Idea]) -> list[dict[str, str]]:
    """Four-plus weeks of mixed-platform slots, rotating the idea bank."""
    rows: list[dict[str, str]] = []
    usable = [idea for idea in bank if "b2b" not in idea.pillar]
    b2b = [idea for idea in bank if idea.pillar == "b2b"]
    idx = 0
    for week in range(weeks):
        week_start = start + timedelta(days=7 * week)
        for day in range(7):
            day_date = week_start + timedelta(days=day)
            times = WEEKDAY_SLOTS[day_date.weekday()]
            # Consumer beat
            idea = usable[idx % len(usable)]
            idx += 1
            platform_today = ["x", "instagram", "tiktok"][day % 3]
            if platform_today not in idea.platforms:
                platform_today = idea.platforms[0]
            fmt = idea.formats[0]
            if platform_today == "instagram":
                fmt = "reel" if day % 2 == 0 else "feed"
            elif platform_today == "tiktok":
                fmt = "tiktok"
            elif platform_today == "x":
                fmt = "tweet" if "tweet" in idea.formats else idea.formats[0]
            rows.append(
                {
                    "date": day_date.isoformat(),
                    "day": day_date.strftime("%A"),
                    "time_uk": times[0],
                    "week": str(week + 1),
                    "platform": platform_today,
                    "format": fmt,
                    "idea_id": idea.id,
                    "pillar": idea.pillar,
                    "title": idea.title,
                    "hook": idea.hook,
                    "status": "Draft",
                }
            )
            # Second daily beat on weekdays for X / Stories
            if day_date.weekday() < 5:
                story_idea = usable[(idx + 3) % len(usable)]
                rows.append(
                    {
                        "date": day_date.isoformat(),
                        "day": day_date.strftime("%A"),
                        "time_uk": times[1],
                        "week": str(week + 1),
                        "platform": "instagram",
                        "format": "story",
                        "idea_id": story_idea.id,
                        "pillar": story_idea.pillar,
                        "title": f"Story: {story_idea.title}",
                        "hook": story_idea.engagement,
                        "status": "Draft",
                    }
                )
        # LinkedIn twice a week
        for offset, idea in zip((1, 4), b2b or usable[:2]):
            day_date = week_start + timedelta(days=offset)
            rows.append(
                {
                    "date": day_date.isoformat(),
                    "day": day_date.strftime("%A"),
                    "time_uk": "09:30",
                    "week": str(week + 1),
                    "platform": "linkedin",
                    "format": "linkedin",
                    "idea_id": idea.id,
                    "pillar": idea.pillar,
                    "title": idea.title,
                    "hook": idea.hook,
                    "status": "Draft",
                }
            )
    return rows


# ---------------------------------------------------------------------------
# Markdown writers
# ---------------------------------------------------------------------------


def md_list(items: list[str]) -> str:
    return "\n".join(f"- {item}" for item in items)


def write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content.strip() + "\n", encoding="utf-8")


def render_design(design: DesignBrief) -> str:
    shots = "\n".join(f"  - {s}" for s in design.shots)
    dont = "\n".join(f"  - {s}" for s in design.do_not)
    text = "\n".join(f"  - {s}" for s in design.on_screen_text)
    return f"""**Format:** {design.format_name}
**Dimensions:** {design.dimensions}
**Duration:** {design.duration}
**Visual concept:** {design.visual_concept}
**On-screen text:**
{text}
**Palette:** {design.palette}
**Type:** {design.type}
**Logo:** {design.logo}
**Audio:** {design.audio}
**Shot list:**
{shots}
**Do not:**
{dont}"""


def render_idea(idea: Idea, brand: BrandProfile, *, platform: str | None = None) -> str:
    url = brand.key_urls.get(idea.url_key, SITE)
    tags = " ".join(idea.hashtags)
    return f"""### {idea.id} — {idea.title}

- **Pillar:** {idea.pillar}
- **Hook:** {idea.hook}
- **Best on:** {", ".join(idea.platforms)} · {", ".join(idea.formats)}
- **Why it should travel:** {idea.why}
- **Engagement mechanic:** {idea.engagement}
- **CTA / URL:** {idea.cta}
- **Hashtags (use sparingly):** {tags or "none"}
- **Compliance:** {idea.compliance}

**Copy**

{idea.copy}

**Alt / shorter**

{idea.alt_copy}

**Design brief**

{render_design(idea.design)}
"""


def render_channel_markdown(brand: BrandProfile) -> str:
    rows = channel_rows(brand)
    blocks = []
    for row in rows:
        blocks.append(
            f"""### {row['channel']}

| | |
|---|---|
| **Priority** | {row['priority']} |
| **Open an account?** | {row['open']} |
| **Handle** | `{row['handle']}` |
| **Why this channel** | {row['why']} |
| **Audience** | {row['audience']} |
| **Content** | {row['content']} |
| **Cadence** | {row['cadence']} |
| **Success looks like** | {row['success']} |
| **Affiliate / provider value** | {row['b2b']} |
"""
        )
    return "\n".join(blocks)


def master_strategy_md(brand: BrandProfile, bank: list[Idea], weeks: int) -> str:
    provider_line = ", ".join(brand.providers) if brand.providers else "major UK ISPs"
    guide_count = len(brand.guides)
    return f"""# BroadbandPicker social strategy

Generated from the workspace on {date.today().isoformat()}.
Site: {brand.site}

## What we are actually building

{brand.positioning}

**Revenue:** {brand.revenue_model}

The job of social is not "post deals". The job is:

1. Make BroadbandPicker the funny, trusted British broadband brand people screenshot to their group chat.
2. Send high-intent traffic to the postcode checker, deals, guides and tools.
3. Look like a real publisher so Awin advertisers and ISP partnership teams approach **us**.

Consumer joke → brand memory → comparison session → affiliate sign-up.
B2B proof → inbound affiliate applications → more complete deal tables → more consumer trust.

## Brand facts this plan is based on

- **Providers in the dataset:** {provider_line}
- **Guides on the site:** {guide_count}
- **Tools:** {", ".join(t["name"] for t in brand.tools)}
- **City / postcode hubs:** {", ".join(brand.cities[:12])}
- **Existing social:** X {brand.social_handles.get("x")} (linked from the site footer and About)
- **Partnerships:** {brand.emails["partnerships"]}
- **Colours:** primary {brand.colours.get("primary")}, navy {brand.colours.get("navy")}, accent {brand.colours.get("accent")}
- **Type on site:** {", ".join(brand.fonts) or "Inter"}
- **Logo files:** {", ".join(brand.logo_files) or "see public/logo.svg"}

### Files read

{md_list(brand.files_read)}

### Warnings from the audit

{md_list(brand.warnings) if brand.warnings else "- None."}

## Positioning on social (one sentence)

BroadbandPicker is the UK broadband comparison site that talks like a mate in the pub, ranks like a grown-up publisher, and will not pretend your 'fibre' is full fibre if it still involves 1998.

Tone: British sarcasm, memes, twisted wise quotes, puns. Never cringe-corporate. Never punch down at people who cannot afford the bill. Punch **up** at loyalty tax, "up to" speeds, 8am–6pm engineer windows, and April rises.

## The 80 / 15 / 5 mix

| Share | Job | Examples |
|---|---|---|
| 80% | Travel. Jokes, polls, quote cards, sketches. Barely a URL. | Buffering weather warning, Keep Calm router, Wi-Fi name roast |
| 15% | Useful-funny. A joke that earns a table. | Speed you actually need, social tariffs, One Touch Switch |
| 5% | Soft sell and B2B. | Out-of-contract public-health warning, LinkedIn methodology |

If a post could have been written by Uswitch's legal team, rewrite it.

## Content pillars

{md_list(brand.content_pillars)}

## Funnel

```
Awareness     TikTok / Reels / X jokes     "this is so Britain"
Consideration Carousels, Stories, threads  "wait, I am out of contract"
Conversion    Bio link, sticker link, X    postcode checker / deals / match quiz
Trust         Social tariff posts, how we  Awin, journalists, ISP partners
              make money, LinkedIn
```

Every joke still has a door home: {SITE}

Do **not** affiliate-dump in captions. One relevant URL. ASA-friendly: prices dated, speeds address-level, commissions disclosed on the site (and on ads, with #ad, if you ever boost a specific deal).

## 90-day operating plan

### Days 1–7 — Claim the name

- Keep and pin a better X bio. Handle is already on the website.
- Open Instagram, TikTok, LinkedIn company page, YouTube (dormant until week 3).
- Usernames: `broadbandpicker` everywhere. If taken, `broadbandpickeruk`.
- Avatars: `docs/branding-and-marketing/broadbandpicker-logo-icon.png`.
- Bios: one joke + one job. Example: "UK broadband comparison. We mock the loyalty tax. Then we help you leave it. {SITE}"
- Link in bio: {SITE} (later a Linktree-style page with Deals / Speed test / Match / Partnerships).
- Cross-link from the site footer once IG and TikTok exist (X is already there).

### Days 8–30 — Daily native comedy

- TikTok daily, Reels 4x/week (shift the best TikTok 24h later), X 1–2x/day, IG Stories on weekdays, LinkedIn 2x/week.
- Do not wait for a studio. Phone, navy end card, captions burned in.
- Reply to **every** comment for 30 days. The algorithm is a conversation.

### Days 31–60 — Turn jokes into a machine

- Repeat winners. Kill flops without sentiment.
- Start YouTube Shorts mirroring.
- One 'serious' carousel a week (social tariffs, rights, methodology).
- Founder LinkedIn: traffic milestones only if true; otherwise methodology and market notes.

### Days 61–90 — Make it commercially loud

- WhatsApp Channel + Facebook page.
- First modest TikTok Spark / IG boost **only** on posts that already organic-work.
- One-pager media kit PDF from the LinkedIn posts, sent with `docs/affiliate-outreach-email.txt`.
- Invite 5 providers that are on the site but not in affiliate yet.

## KPI that match the actual business

| Stage | Metric | 90-day aim (honest, small site) |
|---|---|---|
| Attention | Follows + average watch-through | 3–5k combined IG+TT if the jokes land |
| Memory | Branded Google searches, direct traffic | Up and to the right, even if small |
| Action | Clicks to postcode / deals / match | UTM `?utm_source=tiktok&utm_medium=social&utm_campaign=launch` |
| Money | Affiliate conversions assisted by social | Track in Awin + GA4 |
| B2B | Inbound to {brand.emails['partnerships']} | 1 serious conversation is a win |

## Repurposing engine (one idea, four assets)

1. Write the joke in a Notes app. If it doesn't make you smirk, it will not make a stranger share it.
2. Film a 10-second vertical (TikTok).
3. Post to Reels next day with a different first frame.
4. Cut a quote card for IG feed + X.
5. Put a poll sticker in Stories.
6. If it teaches something, LinkedIn it without the meme font.

## Compliance (non-negotiable)

- British English. Fibre, not fiber. £, not $.
- No fake "cheapest in the UK" without a dated, qualified claim.
- Speeds are typical / address-level. Always a postcode caveat on deal posts.
- Affiliate: the website already discloses. Paid ads and specific "get this deal" posts need clear #ad / #affiliate.
- Satire of **industry tropes**, not defamation of a named provider's measured performance.
- Social tariffs: public-service tone. No punching down.
- Don't encourage people to post account numbers, passwords, or bills with personal data.

## Idea bank size in this run

{len(bank)} named concepts, scheduled across {weeks} weeks, written out per channel in the folders beside this file.
"""


def voice_playbook_md() -> str:
    return f"""# Voice, jokes, and how not to be LinkedIn

## The voice in one paragraph

Talk like the funniest person in a British kitchen when the Wi-Fi dies during the football. Dry. Specific. A bit angry, never cruel. You are on the household's side against the loyalty tax, the engineer window, and the words "up to". You are not a deal spruiker in a headset.

## Do

- Understatement, then the knife. "That's not an appointment. That's a hostage video with a van."
- Concrete British props: kettle, cupboard under the stairs, Greggs, TfL boards, MOTD, April bills, HS2, the pub.
- Teach in the second beat, not the first. Joke, then the useful bit.
- Invite a tag, a poll, a reply, a roast of a Wi-Fi name.
- Let some posts have no URL. Brand memory is a conversion channel.

## Do not

- "We're so excited to announce." Nobody in Britain is excited to announce broadband.
- Americanisms. No "cell phone", "fiber", "awesome deals!!!".
- Emoji salads. One, maybe two, if the platform is Instagram. None on X unless a poll.
- Punching down: poverty, accents, "your nan can't work the remote" as the whole joke. Nan's gossip can be *fast*. Nan is not the punchline.
- Fake intimacy. Don't call the audience "guys" every third line.

## Four joke formulas (steal these)

1. **Industry phrase, taken literally.** "Up to 67 Mbps" → "up to a biscuit."
2. **Official format, unofficial crisis.** Met Office warning, TfL board, Keep Calm poster, fridge letter.
3. **Wise quote, unwise bill.** Socrates, JFK, Sartre, Hamlet, all forced to look at a hub.
4. **Two-column injustice.** New customer £19 / you £41.

## Pun drawer (use one per week, not one per post)

- Mbps — Maybe Per Second
- FTTP — Finally The Thing's Proper
- FTTC — Fibre To The Close-enough
- Packet loss — it's not you
- Broadband — because the bills are
- Router / "rooter" arguments as a Story poll
- ISP — I Spend Plenty
- Latency — late-ncy, the national rail of data
- Openreach — open *to* a van, eventually
- Contention — everyone else had the same idea

## British life × broadband map

| Life | Broadband version |
|---|---|
| Loyalty to a bank | Out-of-contract price |
| Engineer / courier windows | 8 until 6, bring a flask |
| "Coming soon" infrastructure | Full fibre since 2014 |
| Football | 89th-minute buffering |
| Cupboard under the stairs | Where the router goes to die |
| Flatshares | FIFA vs FaceTime peace talks |
| WFH | Impressionist Zoom |
| Rural promises | Yeti van |
| April | Council tax + broadband |
| Small print | "Up to" |

## Engagement mechanics that actually get tags

- "Quote with what you pay."
- "Drop your Wi-Fi name."
- "Tag the unpaid IT department."
- Native polls with a stupid fourth option.
- "2am speed vs 7pm speed."
- "Finish this soliloquy."
- "Add one thing faster than your broadband."

If a post does not ask for a human behaviour, it is a poster. Posters rarely travel.

## Design rules of thumb

- Joke is native. Brand is the end card.
- Burned-in captions always (Reels/TikTok are watched muted).
- Safe zones: keep type out of the bottom 250px (UI chrome).
- Wifi-mark from the logo SVG, not a random stock wifi icon.
- End card: navy field, sky mark, one URL, no QR codes unless print.
- Meme fonts for memes. Inter for anything that looks like the company.

## When to be serious

Social tariffs, complaints, outages with real harm, anything involving money a vulnerable person might misread. Switch the clown off. Keep the clarity.
"""


def channel_md(brand: BrandProfile) -> str:
    p0 = [r for r in channel_rows(brand) if r["priority"].startswith("P0")]
    return f"""# Which social accounts BroadbandPicker should actually open

## Short answer

Open **X (keep), Instagram, TikTok, LinkedIn, and a YouTube channel** now.

Add **Facebook + WhatsApp Channel** in month 2.

Treat **Pinterest** as a later evergreen dump. **Reddit** as a human, not a mascot. **Skip Snapchat and Bluesky** until the P0 channels are a habit.

X already exists as {brand.social_handles.get("x")} on the site. That is the only live consumer channel today. The current X calendar in `docs/branding-and-marketing/` is too polite to get you retweets. Keep the account; change the voice.

## Why these, given the business

BroadbandPicker makes money when a human compares a deal and signs up. The site is free; providers pay affiliate commission. Rankings are not for sale.

That needs:

- **Reach among bill-payers and switchers** → TikTok, Instagram, X, Facebook
- **Brand that doesn't feel like a voucher site** → jokes people want to be seen sharing
- **Proof for Awin and ISP partnership teams** → LinkedIn + a site that already has `/how-we-make-money` and methodology
- **Somewhere journalists already argue about Ofcom** → X

If you only have hours, the order is:

1. X voice-lift (same day)
2. Instagram + TikTok (same content engine)
3. LinkedIn company page (B2B, this is how providers approach you)
4. YouTube Shorts (free extra)
5. Everything else

## Full channel grid

{render_channel_markdown(brand)}

## Handles and bios (copy/paste)

**X** — {brand.social_handles.get("x")}
UK broadband comparison. We mock the loyalty tax. Then we help you leave it.
{SITE}

**Instagram / TikTok** — {brand.social_handles.get("instagram")}
British broadband, minus the small print energy.
Jokes about routers. Help with switching.
Link in bio: {SITE}

**LinkedIn**
Independent UK broadband comparison. Postcode checker, provider reviews, guides.
Partnerships: {brand.emails["partnerships"]}

**YouTube**
BroadbandPicker — UK broadband explained without the coma.

## Profile kit

- Square avatar: `docs/branding-and-marketing/broadbandpicker-logo-icon.png`
- Banner: navy {brand.colours.get("navy")} with the wifi-mark and "Compare broadband. Mock the loyalty tax."
- Website field: {SITE}
- Contact: {brand.emails["hello"]} / {brand.emails["partnerships"]}

## How this gets providers to call you

Consumer memes make the brand look alive. LinkedIn + the trust pages make it look safe. The outreach email in `docs/affiliate-outreach-email.txt` should go out **with** a LinkedIn URL and "here is what we post" — not instead of social.

Pin on LinkedIn: we list providers with or without affiliate deals, editorial is independent, here is `/how-we-make-money`.

That is the opposite of a link farm. That is the point.
"""


def x_posts_md(brand: BrandProfile, bank: list[Idea]) -> str:
    x_ideas = [i for i in bank if "x" in i.platforms]
    body = "\n\n".join(render_idea(i, brand) for i in x_ideas)
    return f"""# X posts, threads and polls

Ready-to-ship copy. Shorten on the day if the news has moved. One URL, not three.

{body}
"""


def x_strategy_md(brand: BrandProfile, bank: list[Idea]) -> str:
    x_count = sum(1 for i in bank if "x" in i.platforms)
    return f"""# X / Twitter strategy — {brand.social_handles.get("x")}

## Why X first

The website already points here. UK broadband discourse already lives here. Journalists, altnet people, and the 7pm complainers are here. You do not need to find a room. You need to be the funniest competent person in it.

## Bio, pin, profile

- **Bio:** UK broadband comparison. We mock the loyalty tax. Then we help you leave it. {SITE}
- **Location:** United Kingdom
- **Pin:** a Keep Calm / hostage-loyalty post, not a corporate intro.
- **Avatar:** logo icon PNG.

## Cadence

- 1–2 original posts a day.
- Reply to Ofcom, ISPreview-adjacent chatter, and provider outages in-character (never dump a link under a real outage until you've been useful).
- 1 poll per week.
- 1 thread per week (useful-funny).
- Quote-tweet with a one-liner more than you like-and-leave.

Best windows: 08:30, 12:30, 18:00–20:00 UK, plus live at 89 minutes of a big match.

## What changes vs the old calendar

The old calendar taught speeds and switching like a leaflet. Fine for SEO. Death on X.

New mix: 70% jokes and hooks, 20% useful-funny, 10% "here is the actual page".

## Thread skeleton (weekly)

1. Unreasonable opening line.
2. Three facts that hurt because they're true.
3. The one thing to do (URL).
4. A question so people reply.

## Hashtags

Usually none. Occasionally #UKBroadband. Never a pile.

## Posts

{x_count} ideas are written out in `posts.md` with full copy and design briefs.
"""


def ig_strategy_md(brand: BrandProfile) -> str:
    return f"""# Instagram strategy — {brand.social_handles.get("instagram")}

Instagram is the brand gallery and the second home of the TikTok engine. If X is the pub argument, Instagram is the postcard you stick on the fridge.

## Three surfaces, three jobs

| Surface | Job | Cadence |
|---|---|---|
| **Reels** | Reach strangers. Same sketches as TikTok, different first frame. | 4/week |
| **Stories** | Habit and replies. Polls, sliders, 'this or that', behind-the-bit. | Weekdays |
| **Feed** | Memory. Quote cards, carousels, memes that still look like you in 12 months. | 5/week |

## Grid rules

- First 9 posts should already look like a brand: 3 quote cards, 3 memes, 3 useful carousels.
- Navy and sky should be recognisable even when the joke is a TfL board.
- Link in bio only. No "link in comments" spam.

## Highlights to create on day one

1. **Deals** — latest useful-funny deal posts
2. **Switch** — One Touch Switch carousel
3. **Tariffs** — social tariff public service
4. **Jokes** — the greatest hits
5. **About** — 3 Story frames: who we are, how we make money, the site

## Audio

Prefer original VO. If you use a trending sound, the **on-screen joke must work muted**.

## Hashtags (3–6)

#UKBroadband #BritishProblems #Broadband #WiFiProblems plus one topical.

## Sharing

Every Reel also goes to Facebook later. Cross-post to Stories with a poll sticker ("too real?").
"""


def ig_reels_md(brand: BrandProfile, bank: list[Idea]) -> str:
    reels = [i for i in bank if "reel" in i.formats]
    body = "\n\n".join(render_idea(i, brand) for i in reels)
    return f"""# Instagram Reels

Shoot vertical 1080x1920, 7–15 seconds unless the useful-funny needs 20. Captions burned in. First frame must read as a poster because that is what the feed shows.

Hook in 0.8 seconds. If the first line needs a preface, it is a blog.

## Ideas

{body}
"""


def ig_stories_md(brand: BrandProfile, bank: list[Idea]) -> str:
    stories = [i for i in bank if "story" in i.formats]
    frames = []
    for idea in stories:
        frames.append(
            f"""### Story set — {idea.id} {idea.title}

Sequence of 3–5 frames, 1080x1920, 5 seconds each, navy or native meme.

1. **Hook frame.** {idea.hook}
2. **Punchline frame.** Pull the best line from the copy.
3. **Sticker.** {idea.engagement}
4. **Optional useful frame.** One fact, huge type.
5. **Link sticker / "see more".** {idea.cta}

Design: same brief as the parent idea, simplified. Type must sit above the sticker tray.

{render_design(idea.design)}
"""
        )
    extras = """
## Recurring Story series (put these on a weekday rotation)

| Series | Day | Mechanic |
|---|---|---|
| **Bill or it didn't happen** | Monday | Slider: how much do you pay? |
| **Cupboard or attic** | Tuesday | This-or-that sticker |
| **Speed-test o'clock** | Wednesday | Add-yours sticker of 2am vs 7pm |
| **Wi-Fi names** | Thursday | Question sticker |
| **Out of contract?** | Friday | Poll, then a link to /deals |
| **Matchday buffering** | Saturday | Live react if a big fixture |
| **Hangover admin** | Sunday | "If you do one useful thing…" switch reminder |

## Story ads / highlights hygiene

- No more than one link sticker per set.
- Never screenshot a customer's bill with visible address.
- Save the best reply stickers into a Highlight called "You lot".
"""
    return "# Instagram Stories\n\n" + "\n".join(frames) + extras


def ig_posts_md(brand: BrandProfile, bank: list[Idea]) -> str:
    posts = [i for i in bank if any(f in i.formats for f in ("feed", "carousel"))]
    body = "\n\n".join(render_idea(i, brand) for i in posts)
    return f"""# Instagram feed posts (single image, carousel, quote card)

Feed is the museum. If it won't still be funny in six months, make it a Story instead.

Carousels: slide 1 is the joke, slides 2–4 earn the save, last slide is the URL.

Caption formula: line 1 hook, line 2–5 the bit, last line the ask (tag / comment). URL at the end or not at all — bio does the rest.

## Ideas

{body}
"""


def ig_design_md(brand: BrandProfile) -> str:
    return f"""# Instagram design system

## Colour

| Token | Hex | Use |
|---|---|---|
| Sky | {brand.colours.get("primary", BRAND_PRIMARY)} | Highlights, wifi-mark, links |
| Navy | {brand.colours.get("navy", BRAND_NAVY)} | End cards, carousels, Stories |
| Green | {brand.colours.get("accent", BRAND_GREEN)} | Genuine good-news ticks only |
| Paper | {BRAND_SLATE} | Quote-card background alt |
| White | {BRAND_WHITE} | Type on navy |

## Type

- Site-faithful: **Inter ExtraBold** headlines, Inter Regular body.
- Meme-native: Anton / Impact / condensed gothic.
- Quote remix: a readable serif (Playfair / Source Serif).
- Official-pastiche: whatever the format needs (dot matrix, highways, Met Office), always slightly wrong so it reads as parody.

## Logo

Source: `docs/branding-and-marketing/broadbandpicker-logo-icon.svg` and `public/logo.svg`.

- End card: mark + wordmark, centred, URL under.
- Feed: mark bottom-left, 64px at 1080, 5% margin.
- Never stretch. Never recolour the mark to green.

## Sizes

| Placement | Size | Safe |
|---|---|---|
| Reel / Story | 1080x1920 | Keep type out of bottom 250px and top 150px |
| Feed 4:5 | 1080x1350 | Preferred |
| Feed 1:1 | 1080x1080 | Quote cards |
| Carousel | 1080x1350 | Same type size every slide |

## End card (standard 0.4–1.0s or last carousel slide)

Navy field. Wifi-mark. **BroadbandPicker**. One line joke leftover or "Compare your postcode". {SITE}

## Production default when there is no designer

iPhone, a navy full-screen title in CapCut / VN, burned-in captions, staff or founder VO. Ship it. Polish winners later.
"""


def tiktok_strategy_md(brand: BrandProfile) -> str:
    return f"""# TikTok strategy — {brand.social_handles.get("tiktok")}

TikTok is how strangers meet BroadbandPicker. Play native or do not play.

## Account

- Name: BroadbandPicker
- Bio: British broadband, minus the small print energy. {SITE}
- Photo: logo icon
- Post vertical only. No watermarks from other apps (film IG *after* TikTok, not before).

## Cadence

Daily for 30 days. Then 5 a week. Evening slots (18:30–21:00) and lunch.

## Native patterns that fit this brand

- POV: "you're the router in a British cupboard"
- Green screen: a fake bill, a fake warning, a fake TfL board
- Storytime: 20s, one specific incident
- Street-interview energy (even if it's just you on a sofa): "what do you pay for broadband"
- Duet / stitch a provider ad (fair-dealing commentary, don't just dunk with music)

## Captions

One or two lines, the punchline, a question. Hashtags: 3. #broadband #britishhumour #ukinternet is plenty.

## Comments

Pin a question. Reply in-character. Pull the best comment into tomorrow's video.

## Advertising later

Only Spark Ads on videos that already have strong watch-through. Never boost a dud. Never run a hard "sign up with X provider" as the first paid thing — that's how you look like a voucher account. Boost the joke, land on the postcode checker.
"""


def tiktok_ideas_md(brand: BrandProfile, bank: list[Idea]) -> str:
    clips = [i for i in bank if "tiktok" in i.formats or "tiktok" in i.platforms]
    body = "\n\n".join(render_idea(i, brand) for i in clips)
    return f"""# TikTok video ideas

All 7–15s unless noted. Hook is spoken **and** burned in. End on the face or the joke, then a 0.4s brand sting.

## Ideas

{body}
"""


def tiktok_design_md(brand: BrandProfile) -> str:
    return f"""# TikTok design briefs — house rules

- 1080x1920, 30fps, captions always.
- First frame = poster. Assume the algorithm shows a still.
- VO: dry UK, close-miked, unimpressed. No radio-ad warmth.
- Text: max 6 words on screen at a time.
- End sting: navy, wifi-mark, `{SITE.replace("https://", "")}`.
- Do not use other platforms' watermarks.
- Do not use uncleared Premier League, BBC, TfL or Greggs marks on paid ads.
- Film once, cut two hooks, post the better one, save the other for IG tomorrow.

Palette: {brand.colours}
"""


def x_design_md(brand: BrandProfile) -> str:
    return f"""# X design briefs — house rules

- Prefer native text. Images when the joke **is** a poster (Keep Calm, Met Office, TfL, quote card).
- Image size: 1600x900 (16:9) or 1080x1080. Avoid 9:16 as the main tweet image — it collapses.
- Alt text: write the joke again, for once.
- No more than one image unless it is a before/after bill.
- Brand mark small. If they cannot read the tweet without the image, the tweet is too weak.

Palette: {brand.colours}
"""


def readme_md(brand: BrandProfile, weeks: int) -> str:
    return f"""# Social media manager

This folder is BroadbandPicker's social operating kit. It is generated by:

```bash
python3 scripts/plan_social_media_strategy.py
```

The script **reads this workspace** (about page, how-we-make-money, providers, guides, tools, colours, existing X calendar, content plan) and writes a brand-aware strategy plus jokes you can actually post.

## Folders

| Folder | What is in it |
|---|---|
| `Instagram Content creation/` | Reels, Stories, feed/carousel ideas, IG design system, calendar |
| `X/` | Voice-lifted X strategy, posts, threads, design rules, calendar |
| `tiktok/` | Native video ideas, design rules, calendar |

Also in this root: master strategy, channel recommendations, voice playbook, brand intelligence JSON.

## What to open

**Now:** X (already live as {brand.social_handles.get("x")}), Instagram, TikTok, LinkedIn, YouTube.
**Month 2:** Facebook page, WhatsApp Channel.
**Later / maybe:** Pinterest, Reddit-as-a-human.
**Skip:** Snapchat, Bluesky.

Details: `01-channel-recommendations.md`.

## Brand colours

- Sky `{brand.colours.get("primary", BRAND_PRIMARY)}`
- Navy `{brand.colours.get("navy", BRAND_NAVY)}`
- Accent `{brand.colours.get("accent", BRAND_GREEN)}`

Avatar file: `docs/branding-and-marketing/broadbandpicker-logo-icon.png`

## Commercial point of all this

Jokes get reach. Reach gets branded traffic. Traffic gets compared deals. Compared deals get affiliate commissions. A loud, clearly independent brand gets **providers applying to you** instead of the other way round.

Partnerships: {brand.emails["partnerships"]}
"""


def brand_json(brand: BrandProfile) -> dict[str, Any]:
    payload = asdict(brand)
    payload["generated_at"] = datetime.now().isoformat(timespec="seconds")
    payload["idea_pillars"] = [
        "british sarcasm",
        "memes",
        "wise quotes made stupid and British",
        "puns",
        "useful-funny",
        "B2B affiliate gravity",
    ]
    return payload


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, brand: BrandProfile, bank: list[Idea], calendar: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    wb = Workbook()

    def style_header(ws) -> None:
        fill = PatternFill("solid", fgColor="0F172A")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    def autosize(ws) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = 12
            for cell in col[:40]:
                width = max(width, min(48, len(str(cell.value or "")) + 2))
            ws.column_dimensions[letter].width = width

    ws = wb.active
    ws.title = "Brand"
    ws.append(["Field", "Value"])
    for key, value in asdict(brand).items():
        if isinstance(value, (list, dict)):
            value = json.dumps(value, ensure_ascii=False)[:5000]
        ws.append([key, value])
    style_header(ws)
    autosize(ws)

    ws = wb.create_sheet("Channels")
    rows = channel_rows(brand)
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    style_header(ws)
    autosize(ws)

    ws = wb.create_sheet("Idea bank")
    ws.append(["id", "pillar", "title", "hook", "platforms", "formats", "url_key", "engagement", "why"])
    for idea in bank:
        ws.append(
            [
                idea.id,
                idea.pillar,
                idea.title,
                idea.hook,
                ", ".join(idea.platforms),
                ", ".join(idea.formats),
                idea.url_key,
                idea.engagement,
                idea.why,
            ]
        )
    style_header(ws)
    autosize(ws)

    ws = wb.create_sheet("Calendar")
    headers = list(calendar[0].keys()) if calendar else ["date"]
    ws.append(headers)
    for row in calendar:
        ws.append([row[h] for h in headers])
    style_header(ws)
    autosize(ws)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def filter_cal(calendar: list[dict[str, str]], platform: str) -> list[dict[str, str]]:
    return [row for row in calendar if row["platform"] == platform]


def main() -> int:
    parser = argparse.ArgumentParser(description="Plan BroadbandPicker social content from this repo.")
    parser.add_argument("--weeks", type=int, default=4, help="How many weeks of calendar to schedule.")
    parser.add_argument("--start-date", default=DEFAULT_START.isoformat(), help="YYYY-MM-DD, snapped to the next Monday.")
    parser.add_argument("--output", default=str(OUT_DIR), help="Output folder.")
    args = parser.parse_args()

    start = monday_on_or_after(date.fromisoformat(args.start_date))
    weeks = max(2, args.weeks)
    out = Path(args.output)
    ig = out / "Instagram Content creation"
    xdir = out / "X"
    tdir = out / "tiktok"
    for folder in (out, ig, xdir, tdir):
        folder.mkdir(parents=True, exist_ok=True)

    brand = analyse_workspace()
    bank = ideas(brand)
    calendar = build_calendar(start, weeks, bank)

    write(out / "README.md", readme_md(brand, weeks))
    write(out / "00-brand-intelligence.json", json.dumps(brand_json(brand), indent=2, ensure_ascii=False))
    write(out / "01-channel-recommendations.md", channel_md(brand))
    write(out / "02-voice-and-humour-playbook.md", voice_playbook_md())
    write(out / "03-master-strategy.md", master_strategy_md(brand, bank, weeks))
    write_csv(out / "content-calendar-all.csv", calendar)
    write_xlsx(out / "social-content-planner.xlsx", brand, bank, calendar)

    write(xdir / "strategy.md", x_strategy_md(brand, bank))
    write(xdir / "posts.md", x_posts_md(brand, bank))
    write(xdir / "design-briefs.md", x_design_md(brand))
    write_csv(xdir / "content-calendar.csv", filter_cal(calendar, "x"))

    write(ig / "strategy.md", ig_strategy_md(brand))
    write(ig / "reels.md", ig_reels_md(brand, bank))
    write(ig / "stories.md", ig_stories_md(brand, bank))
    write(ig / "feed-posts.md", ig_posts_md(brand, bank))
    write(ig / "design-system.md", ig_design_md(brand))
    write_csv(
        ig / "content-calendar.csv",
        [row for row in calendar if row["platform"] == "instagram"],
    )

    write(tdir / "strategy.md", tiktok_strategy_md(brand))
    write(tdir / "video-ideas.md", tiktok_ideas_md(brand, bank))
    write(tdir / "design-briefs.md", tiktok_design_md(brand))
    write_csv(tdir / "content-calendar.csv", filter_cal(calendar, "tiktok"))

    print(f"Brand files read: {len(brand.files_read)}")
    print(f"Providers: {len(brand.providers)} · Guides: {len(brand.guides)}")
    print(f"Ideas: {len(bank)} · Calendar rows: {len(calendar)}")
    print(f"Wrote {out.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
