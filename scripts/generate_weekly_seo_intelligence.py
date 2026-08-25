#!/usr/bin/env python3
"""Generate a joined weekly GSC + GA4 SEO/GEO action report.

The join is deliberately page-level. Search Console and Analytics data must
never be joined at user level. The workbook update replaces only the dedicated
"Weekly SEO Actions" tab and preserves every existing tracker row/status.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from statistics import median

from google.auth.transport.requests import Request
from google.oauth2 import service_account
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CREDENTIALS = Path("/Users/sahilrafiqsayed/broadbandpicker-ga4-credentials.json")
DEFAULT_JSON = ROOT / "docs" / "weekly-seo-intelligence.json"
DEFAULT_MARKDOWN = ROOT / "docs" / "weekly-seo-intelligence.md"
DEFAULT_TRACKER = ROOT / "docs" / "master-build-tracker.xlsx"
SITE = "sc-domain:broadbandpicker.co.uk"
SITE_ORIGIN = "https://broadbandpicker.co.uk"
GA4_PROPERTY = "551202232"
GSC_BASE = "https://www.googleapis.com/webmasters/v3"
GA_BASE = "https://analyticsdata.googleapis.com/v1beta"
SCOPES = [
    "https://www.googleapis.com/auth/webmasters.readonly",
    "https://www.googleapis.com/auth/analytics.readonly",
]


@dataclass
class PageOpportunity:
    page: str
    clicks: float
    impressions: float
    ctr: float
    position: float
    previous_clicks: float
    previous_impressions: float
    impression_change_pct: float | None
    sessions: float
    engaged_sessions: float
    engagement_rate: float | None
    key_events: float
    affiliate_clicks: float
    ai_referral_visits: float
    score: float
    category: str
    recommended_action: str
    top_queries: list[str]


class GoogleClient:
    def __init__(self, credentials_path: Path) -> None:
        self.credentials = service_account.Credentials.from_service_account_file(
            str(credentials_path), scopes=SCOPES
        )

    def request(self, method: str, url: str, payload: dict | None = None) -> dict:
        if not self.credentials.valid:
            self.credentials.refresh(Request())
        body = None if payload is None else json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            url,
            data=body,
            method=method,
            headers={
                "Authorization": f"Bearer {self.credentials.token}",
                "Content-Type": "application/json",
            },
        )
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                raw = response.read().decode("utf-8")
                return json.loads(raw) if raw else {}
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Google API returned HTTP {exc.code}: {detail}") from exc


def canonical_path(value: str) -> str:
    if not value or value == "(not set)":
        return "/"
    try:
        parsed = urllib.parse.urlsplit(value if "://" in value else f"{SITE_ORIGIN}{value}")
        path = parsed.path or "/"
    except ValueError:
        path = value.split("?", 1)[0] or "/"
    return path.rstrip("/") or "/"


def gsc_query(client: GoogleClient, start: date, end: date, dimensions: list[str]) -> list[dict]:
    site = urllib.parse.quote(SITE, safe="")
    response = client.request(
        "POST",
        f"{GSC_BASE}/sites/{site}/searchAnalytics/query",
        {
            "startDate": start.isoformat(),
            "endDate": end.isoformat(),
            "type": "web",
            "dimensions": dimensions,
            "dataState": "final",
            "rowLimit": 25000,
            "aggregationType": "auto",
        },
    )
    return response.get("rows", [])


def ga4_report(
    client: GoogleClient,
    start: date,
    end: date,
    dimensions: list[str],
    metrics: list[str],
    dimension_filter: dict | None = None,
) -> list[dict]:
    payload: dict = {
        "dateRanges": [{"startDate": start.isoformat(), "endDate": end.isoformat()}],
        "dimensions": [{"name": name} for name in dimensions],
        "metrics": [{"name": name} for name in metrics],
        "limit": "100000",
    }
    if dimension_filter:
        payload["dimensionFilter"] = dimension_filter
    response = client.request(
        "POST", f"{GA_BASE}/properties/{GA4_PROPERTY}:runReport", payload
    )
    rows = []
    for row in response.get("rows", []):
        rows.append(
            {
                "dimensions": [value.get("value", "") for value in row.get("dimensionValues", [])],
                "metrics": [float(value.get("value", 0) or 0) for value in row.get("metricValues", [])],
            }
        )
    return rows


def page_map(rows: list[dict]) -> dict[str, dict[str, float]]:
    result: dict[str, dict[str, float]] = {}
    for row in rows:
        page = canonical_path(row["keys"][0])
        result[page] = {
            "clicks": float(row.get("clicks", 0)),
            "impressions": float(row.get("impressions", 0)),
            "ctr": float(row.get("ctr", 0)),
            "position": float(row.get("position", 0)),
        }
    return result


def suspicious_query(value: str) -> bool:
    """Keep anomalous non-English query rows from manufacturing SEO wins."""
    if not value:
        return False
    non_ascii = sum(1 for character in value if ord(character) > 127)
    return non_ascii / len(value) >= 0.25


def remove_query_anomalies(
    pages: dict[str, dict[str, float]], query_rows: list[dict]
) -> tuple[dict[str, dict[str, float]], list[str]]:
    excluded: dict[str, dict[str, float]] = defaultdict(
        lambda: {"clicks": 0.0, "impressions": 0.0, "position_weight": 0.0}
    )
    labels: list[str] = []
    for row in query_rows:
        query = row["keys"][1]
        if not suspicious_query(query):
            continue
        page = canonical_path(row["keys"][0])
        impressions = float(row.get("impressions", 0))
        excluded[page]["clicks"] += float(row.get("clicks", 0))
        excluded[page]["impressions"] += impressions
        excluded[page]["position_weight"] += impressions * float(row.get("position", 0))
        labels.append(f"{page}: {query}")

    adjusted = {page: metrics.copy() for page, metrics in pages.items()}
    for page, removed in excluded.items():
        if page not in adjusted:
            continue
        original = adjusted[page]
        impressions = max(0.0, original["impressions"] - removed["impressions"])
        clicks = max(0.0, original["clicks"] - removed["clicks"])
        position_weight = original["position"] * original["impressions"] - removed["position_weight"]
        original.update({
            "clicks": clicks,
            "impressions": impressions,
            "ctr": clicks / impressions if impressions else 0.0,
            "position": max(0.0, position_weight / impressions) if impressions else 0.0,
        })
    return adjusted, labels


def pct_change(current: float, previous: float) -> float | None:
    if previous == 0:
        return None if current == 0 else 100.0
    return round((current - previous) / previous * 100, 1)


def build_opportunities(
    current_pages: dict[str, dict[str, float]],
    previous_pages: dict[str, dict[str, float]],
    query_rows: list[dict],
    ga_pages: dict[str, dict[str, float]],
    affiliate_by_page: dict[str, float],
    ai_by_page: dict[str, float],
) -> list[PageOpportunity]:
    query_groups: dict[str, list[tuple[float, str]]] = defaultdict(list)
    for row in query_rows:
        page = canonical_path(row["keys"][0])
        if suspicious_query(row["keys"][1]):
            continue
        query_groups[page].append((float(row.get("impressions", 0)), row["keys"][1]))

    ctr_values = [metrics["ctr"] for metrics in current_pages.values() if metrics["impressions"] >= 10]
    site_median_ctr = median(ctr_values) if ctr_values else 0.0
    max_impressions = max((metrics["impressions"] for metrics in current_pages.values()), default=1)
    opportunities: list[PageOpportunity] = []

    for page, current in current_pages.items():
        previous = previous_pages.get(page, {"clicks": 0, "impressions": 0})
        ga = ga_pages.get(page, {})
        impressions = current["impressions"]
        position = current["position"]
        ctr = current["ctr"]
        sessions = ga.get("sessions", 0)
        engaged = ga.get("engaged_sessions", 0)
        engagement_rate = engaged / sessions if sessions else None
        trend = pct_change(impressions, previous["impressions"])

        demand_score = 35 * (math.log1p(impressions) / math.log1p(max_impressions)) if impressions else 0
        if impressions >= 40 and 4 <= position <= 15:
            rank_score = 30
            category = "Quick ranking win"
        elif impressions >= 40 and 15 < position <= 30:
            rank_score = 22
            category = "Page-two opportunity"
        elif position < 4 and impressions >= 25:
            rank_score = 15
            category = "Defend winner"
        else:
            rank_score = 8
            category = "Monitor"

        ctr_score = 20 if impressions >= 20 and ctr < site_median_ctr else 5
        trend_score = 10 if trend is not None and trend < -10 else 3
        behaviour_score = 0
        if engagement_rate is not None:
            behaviour_score = 5 if engagement_rate >= 0.6 else 10
            if engagement_rate < 0.4:
                category = "Content improvement"
        if sessions >= 5 and affiliate_by_page.get(page, 0) == 0 and (engagement_rate or 0) >= 0.5:
            category = "Conversion improvement"

        score = round(min(100, demand_score + rank_score + ctr_score + trend_score + behaviour_score), 1)
        if category == "Quick ranking win":
            action = "Strengthen the title/snippet, answer the leading queries directly, add internal links and refresh supporting evidence."
        elif category == "Page-two opportunity":
            action = "Expand missing subtopics for the leading queries, improve topical internal links and validate intent alignment."
        elif category == "Content improvement":
            action = "Improve answer-first copy, section hierarchy, evidence, comparison usefulness and mobile reading flow."
        elif category == "Conversion improvement":
            action = "Test clearer comparison and provider CTAs while preserving the informational answer and editorial independence."
        elif category == "Defend winner":
            action = "Protect rankings with freshness checks, authoritative evidence, stable URLs and supporting internal links."
        else:
            action = "Monitor until the page accumulates enough demand or behavioural evidence for a confident intervention."

        top_queries = [query for _, query in sorted(query_groups[page], reverse=True)[:5]]
        opportunities.append(
            PageOpportunity(
                page=page,
                clicks=current["clicks"],
                impressions=impressions,
                ctr=ctr,
                position=position,
                previous_clicks=previous["clicks"],
                previous_impressions=previous["impressions"],
                impression_change_pct=trend,
                sessions=sessions,
                engaged_sessions=engaged,
                engagement_rate=engagement_rate,
                key_events=ga.get("key_events", 0),
                affiliate_clicks=affiliate_by_page.get(page, 0),
                ai_referral_visits=ai_by_page.get(page, 0),
                score=score,
                category=category,
                recommended_action=action,
                top_queries=top_queries,
            )
        )
    return sorted(opportunities, key=lambda item: (-item.score, -item.impressions, item.page))


def render_markdown(payload: dict) -> str:
    lines = [
        "# BroadbandPicker weekly SEO + GEO intelligence",
        "",
        f"Generated: {payload['generated_at']}",
        f"Finalised comparison: {payload['current_window']['start']} to {payload['current_window']['end']} vs {payload['previous_window']['start']} to {payload['previous_window']['end']}",
        "",
        "## Executive summary",
        "",
        f"- GSC clicks: **{payload['summary']['gsc_clicks']:g}**",
        f"- GSC impressions: **{payload['summary']['gsc_impressions']:g}**",
        f"- GA4 organic sessions available: **{payload['summary']['ga4_organic_sessions']:g}**",
        f"- Identifiable AI referral visits: **{payload['summary']['ai_referral_visits']:g}**",
        "",
        "## Next five actions",
        "",
        "| Rank | Page | Opportunity | Score | Impressions | Position | CTR | Recommended action |",
        "| --- | --- | --- | ---: | ---: | ---: | ---: | --- |",
    ]
    for rank, row in enumerate(payload["next_five"], 1):
        lines.append(
            f"| {rank} | `{row['page']}` | {row['category']} | {row['score']:.1f} | "
            f"{row['impressions']:.0f} | {row['position']:.1f} | {row['ctr']:.1%} | {row['recommended_action']} |"
        )
    lines += [
        "",
        "## Measurement interpretation",
        "",
        "- Search Console and GA4 are joined only by canonical landing page, never by user.",
        "- A high score prioritises evidence-backed opportunity; it is not a ranking guarantee.",
        "- AI-referral counts are a minimum because some assistants suppress referrer information.",
        "- Google AI Overview clicks usually remain part of Google organic traffic. Use Search Console's generative-AI report when available for visibility-specific analysis.",
        "- GA4 custom dimensions can take 24-48 hours to populate after registration.",
        "",
    ]
    return "\n".join(lines)


def update_tracker(path: Path, actions: list[dict], generated_at: str) -> None:
    workbook = load_workbook(path)
    sheet_name = "Weekly SEO Actions"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name, 1)
    headers = [
        "Rank", "Page", "Opportunity", "Score", "Clicks", "Impressions", "CTR", "Position",
        "Impression Change", "GA4 Sessions", "Engagement Rate", "Affiliate Clicks",
        "AI Referral Visits", "Top Queries", "Recommended Action", "Generated",
    ]
    ws.append(headers)
    for rank, row in enumerate(actions, 1):
        ws.append([
            rank, row["page"], row["category"], row["score"], row["clicks"], row["impressions"],
            row["ctr"], row["position"], row["impression_change_pct"], row["sessions"],
            row["engagement_rate"], row["affiliate_clicks"], row["ai_referral_visits"],
            "; ".join(row["top_queries"]), row["recommended_action"], generated_at,
        ])
    header_fill = PatternFill("solid", fgColor="0F172A")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [8, 44, 24, 10, 10, 14, 10, 10, 18, 14, 16, 16, 18, 54, 76, 24]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws["G"][1:]:
        cell.number_format = "0.00%"
    for cell in ws["K"][1:]:
        cell.number_format = "0.00%"
    for cell in ws["I"][1:]:
        cell.number_format = '0.0"%"'

    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=path.parent, delete=False) as tmp:
        temp_path = Path(tmp.name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        temp_path.unlink(missing_ok=True)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--credentials", type=Path, default=DEFAULT_CREDENTIALS)
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--markdown", type=Path, default=DEFAULT_MARKDOWN)
    parser.add_argument("--tracker", type=Path, default=DEFAULT_TRACKER)
    parser.add_argument("--top", type=int, default=30)
    parser.add_argument("--skip-tracker", action="store_true")
    args = parser.parse_args()
    if not args.credentials.is_file():
        print("Credential file was not found.", file=sys.stderr)
        return 2

    client = GoogleClient(args.credentials)
    current_end = date.today() - timedelta(days=3)
    current_start = current_end - timedelta(days=27)
    previous_end = current_start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=27)

    current_page_rows = gsc_query(client, current_start, current_end, ["page"])
    previous_page_rows = gsc_query(client, previous_start, previous_end, ["page"])
    query_rows = gsc_query(client, current_start, current_end, ["page", "query"])
    previous_query_rows = gsc_query(client, previous_start, previous_end, ["page", "query"])

    organic_filter = {
        "filter": {
            "fieldName": "sessionDefaultChannelGroup",
            "stringFilter": {"matchType": "EXACT", "value": "Organic Search"},
        }
    }
    ga_rows = ga4_report(
        client, current_start, current_end,
        ["landingPagePlusQueryString"],
        ["sessions", "engagedSessions", "keyEvents"],
        organic_filter,
    )
    ga_pages: dict[str, dict[str, float]] = {}
    for row in ga_rows:
        page = canonical_path(row["dimensions"][0])
        sessions, engaged, key_events = row["metrics"]
        ga_pages[page] = {"sessions": sessions, "engaged_sessions": engaged, "key_events": key_events}

    event_rows = ga4_report(
        client, current_start, current_end,
        ["landingPagePlusQueryString", "eventName"], ["eventCount"],
    )
    affiliate_by_page: dict[str, float] = defaultdict(float)
    ai_by_page: dict[str, float] = defaultdict(float)
    for row in event_rows:
        page = canonical_path(row["dimensions"][0])
        event_name = row["dimensions"][1]
        if event_name == "outbound_provider_click":
            affiliate_by_page[page] += row["metrics"][0]
        if event_name == "ai_referral_visit":
            ai_by_page[page] += row["metrics"][0]

    current_pages, excluded_current = remove_query_anomalies(page_map(current_page_rows), query_rows)
    previous_pages, excluded_previous = remove_query_anomalies(page_map(previous_page_rows), previous_query_rows)
    opportunities = build_opportunities(
        current_pages, previous_pages, query_rows, ga_pages, affiliate_by_page, ai_by_page
    )
    generated_at = datetime.now(timezone.utc).isoformat()
    payload = {
        "generated_at": generated_at,
        "current_window": {"start": current_start.isoformat(), "end": current_end.isoformat()},
        "previous_window": {"start": previous_start.isoformat(), "end": previous_end.isoformat()},
        "summary": {
            "gsc_clicks": sum(item.clicks for item in opportunities),
            "gsc_impressions": sum(item.impressions for item in opportunities),
            "ga4_organic_sessions": sum(item.sessions for item in opportunities),
            "ai_referral_visits": sum(item.ai_referral_visits for item in opportunities),
            "pages_scored": len(opportunities),
            "anomalous_query_rows_excluded": len(excluded_current) + len(excluded_previous),
        },
        "next_five": [asdict(item) for item in opportunities[:5]],
        "opportunities": [asdict(item) for item in opportunities[: args.top]],
    }
    args.json.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    args.markdown.write_text(render_markdown(payload), encoding="utf-8")
    if not args.skip_tracker:
        update_tracker(args.tracker, payload["opportunities"], generated_at)

    print(json.dumps(payload["summary"], indent=2))
    print("Next five actions:")
    for rank, item in enumerate(payload["next_five"], 1):
        print(f"  {rank}. {item['page']} — {item['category']} ({item['score']})")
    print(f"Saved {args.json.resolve()}")
    print(f"Saved {args.markdown.resolve()}")
    if not args.skip_tracker:
        print(f"Updated {args.tracker.resolve()} [{len(payload['opportunities'])} action rows]")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
