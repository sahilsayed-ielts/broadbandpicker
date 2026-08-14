#!/usr/bin/env python3
"""Build an evidence-led UK broadband SEO/GEO plan as an Excel workbook.

The script:
1. Crawls BroadbandPicker's XML sitemap and audits basic on-page signals.
2. Collects Google UK autocomplete suggestions for explicitly configured seeds.
3. Combines those findings with a dated snapshot of Ubersuggest UK metrics.
4. Produces a prioritised landing-page, content, GEO and Awin-readiness plan.

Usage:
    python3 scripts/build_uk_broadband_seo_plan.py
    python3 scripts/build_uk_broadband_seo_plan.py --offline
    python3 scripts/build_uk_broadband_seo_plan.py --output docs/my-plan.xlsx
    python3 scripts/build_uk_broadband_seo_plan.py --sync-google-sheet

Search volumes are directional third-party estimates, not Search Console data.
Re-run the workbook monthly and replace estimates with first-party GSC conversion
and query data once sufficient data is available.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import urllib.parse
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from lxml import html
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "docs" / "uk-broadband-seo-geo-plan.xlsx"
DEFAULT_GOOGLE_SHEET_ID = "19zz-7Zy1k1tED4MqpMNv2msx0JvNpbE_hyfTymu5GZc"
SITE = "https://broadbandpicker.co.uk"
RUN_DATE = datetime.now(timezone.utc).date().isoformat()
USER_AGENT = (
    "BroadbandPickerResearchBot/1.0 "
    "(SEO research; contact: https://broadbandpicker.co.uk/contact)"
)

# Dated live Ubersuggest snapshot collected 2026-07-29, UK/en.
# Only non-zero rows are included because zero can mean "no reported data",
# not necessarily no searches.
KEYWORDS = [
    ("broadband deals", 165000, 71, 8.32, "Commercial", "Deals"),
    ("deals for broadband", 201000, 70, 6.815, "Commercial", "Deals"),
    ("great broadband deals", 49500, 54, 5.875, "Commercial", "Deals"),
    ("best broadband deals", 40500, 70, 7.02, "Commercial", "Deals"),
    ("broadband deals cheapest", 18100, 45, 4.58, "Commercial", "Deals"),
    ("uk broadband deals", 14800, 71, 6.93, "Commercial", "Deals"),
    ("broadband test speed test", 246000, 39, 1.14, "Tool", "Speed"),
    ("speed check of broadband", 201000, 61, 1.09, "Tool", "Speed"),
    ("broadband speed check up", 40500, 24, 1.145, "Tool", "Speed"),
    ("broadband", 49500, 66, 11.38, "Mixed", "Core"),
    ("broadband providers", 6600, 69, 11.05, "Commercial", "Providers"),
    ("broadband providers best", 8100, 41, 5.395, "Commercial", "Providers"),
    ("best broadband providers", 5400, 69, 6.07, "Commercial", "Providers"),
    ("broadband providers for my area", 5400, 61, 7.31, "Local commercial", "Availability"),
    ("broadband providers my area", 4400, 62, 6.78, "Local commercial", "Availability"),
    ("comparison of broadband providers", 4400, 28, 6.605, "Commercial", "Comparison"),
    ("uk broadband providers comparison", 4400, 72, 6.95, "Commercial", "Comparison"),
    ("broadband providers comparison uk", 3600, 66, 7.03, "Commercial", "Comparison"),
    ("switch broadband providers", 3600, 38, 13.07, "Commercial guide", "Switching"),
    ("best broadband providers in uk", 2400, 59, 5.45, "Commercial", "Providers"),
    ("broadband providers london uk", 2400, 28, 6.935, "Local commercial", "Availability"),
    ("broadband providers for business", 1900, 23, 29.35, "B2B commercial", "Business"),
    ("business broadband providers", 1000, 49, 53.80, "B2B commercial", "Business"),
    ("uk broadband providers customer satisfaction", 1600, 25, 3.41, "Research", "Trust"),
    ("broadband providers in the uk", 1600, 24, 4.325, "Commercial", "Providers"),
    ("cheapest broadband providers uk", 1300, 67, 6.89, "Commercial", "Deals"),
    ("satellite broadband providers", 720, 43, 2.41, "Commercial guide", "Alternatives"),
    ("bt broadband deals", 33100, 43, 2.12, "Brand commercial", "Provider deals"),
    ("sky broadband deals", 33100, 45, 1.96, "Brand commercial", "Provider deals"),
    ("ee broadband deals", 33100, 29, 1.77, "Brand commercial", "Provider deals"),
    ("broadband deals tv", 27100, 46, 6.35, "Commercial", "Bundles"),
    ("broadband deals and phone", 22200, 67, 4.76, "Commercial", "Bundles"),
    ("broadband deals and tv", 22200, 58, 8.40, "Commercial", "Bundles"),
    ("telephone & broadband deals", 18100, 55, 4.33, "Commercial", "Bundles"),
    ("phone and broadband deals", 14800, 57, 4.07, "Commercial", "Bundles"),
    ("landline broadband deals", 14800, 50, 4.23, "Commercial", "Bundles"),
    ("best landline and broadband deals", 12100, 50, 3.90, "Commercial", "Bundles"),
    ("virgin broadband deals", 12100, 67, 2.67, "Brand commercial", "Provider deals"),
    ("best broadband deals and tv", 12100, 32, 4.49, "Commercial", "Bundles"),
    ("broadband ee", 165000, 64, 1.44, "Brand navigation", "Provider"),
    ("sky broadband", 135000, 70, 1.75, "Brand navigation", "Provider"),
    ("bt broadband", 110000, 69, 2.14, "Brand navigation", "Provider"),
    ("broadband with vodafone", 110000, 47, 2.26, "Brand navigation", "Provider"),
    ("vodafone broadband", 90500, 37, 1.94, "Brand navigation", "Provider"),
    ("broadband virgin", 49500, 65, 2.51, "Brand navigation", "Provider"),
    ("virgin broadband", 40500, 39, 2.39, "Brand navigation", "Provider"),
    ("broadband talktalk", 40500, 40, 2.41, "Brand navigation", "Provider"),
    ("deals with bt broadband", 40500, 54, 1.76, "Brand commercial", "Provider deals"),
]

SEEDS = [
    "broadband", "broadband deals", "broadband providers",
    "broadband speed", "switch broadband", "full fibre broadband",
    "business broadband", "broadband social tariff",
]

# Add every newly built content page here. The workbook will retain a dated,
# auditable snapshot of its primary UK sources and show failed/stale retrievals.
# Source pages should be official providers, regulators or original datasets;
# competitor copy is not a factual source.
PAGE_RESEARCH_SOURCES = {
    "/guides/best-broadband-and-tv-deals": {
        "topic": "Best broadband and TV deals UK",
        "target_terms": [
            "broadband", "TV", "contract", "price", "Sky Sports",
            "TNT Sports", "Netflix", "switch",
        ],
        "sources": [
            ("Sky TV and broadband deals", "Provider",
             "https://www.sky.com/deals?section=tvandbroadband"),
            ("Virgin Media broadband and TV", "Provider",
             "https://www.virginmedia.com/broadband/broadband-and-tv"),
            ("BT and EE TV packages", "Provider",
             "https://www.bt.com/tv/packages"),
            ("Ofcom money-saving tips", "Regulator",
             "https://www.ofcom.org.uk/phones-and-broadband/saving-money/money-saving-tips-for-phone-broadband-and-pay-tv"),
            ("Ofcom switching broadband provider", "Regulator",
             "https://www.ofcom.org.uk/phones-and-broadband/switching-provider/switching-broadband-provider"),
        ],
    },
    "/guides/best-business-broadband-providers-uk": {
        "topic": "Best business broadband providers UK",
        "target_terms": ["business broadband", "static IP", "SLA", "backup", "full fibre"],
        "sources": [
            ("Ofcom Business Broadband Code", "Regulator", "https://www.ofcom.org.uk/phones-and-broadband/coverage-and-speeds/business-broadband-cop"),
            ("BT Business broadband", "Provider", "https://business.bt.com/sme/business-broadband/"),
        ],
    },
    "/postcode": {
        "topic": "Broadband providers in my area",
        "target_terms": ["coverage", "postcode", "full fibre", "availability", "network"],
        "sources": [("Ofcom coverage checker", "Regulator", "https://checker.ofcom.org.uk/en-gb/broadband-coverage")],
    },
    "/guides/best-phone-and-broadband-deals": {
        "topic": "Phone and broadband deals UK",
        "target_terms": ["digital voice", "home phone", "calls", "broadband", "switch"],
        "sources": [("Ofcom future of landline calls", "Regulator", "https://www.ofcom.org.uk/phones-and-broadband/landline-phones/future-of-landline-calls")],
    },
    "/providers/bt/deals": {"topic": "BT broadband deals", "target_terms": ["broadband", "contract", "speed", "price"], "sources": [("BT broadband deals", "Provider", "https://www.bt.com/broadband/deals")]},
    "/providers/sky/deals": {"topic": "Sky broadband deals", "target_terms": ["broadband", "contract", "speed", "price"], "sources": [("Sky broadband deals", "Provider", "https://www.sky.com/deals/broadband")]},
    "/providers/ee/deals": {"topic": "EE broadband deals", "target_terms": ["broadband", "contract", "speed", "price"], "sources": [("EE broadband", "Provider", "https://ee.co.uk/broadband")]},
    "/providers/virgin-media/deals": {"topic": "Virgin Media broadband deals", "target_terms": ["broadband", "contract", "speed", "price"], "sources": [("Virgin Media broadband deals", "Provider", "https://www.virginmedia.com/broadband/broadband-deals")]},
    "/research/uk-broadband-customer-satisfaction": {
        "topic": "UK broadband customer satisfaction",
        "target_terms": ["complaints", "customer service", "satisfaction", "broadband"],
        "sources": [
            ("Ofcom telecoms complaints", "Regulator", "https://www.ofcom.org.uk/phones-and-broadband/service-quality/telecoms-and-pay-tv-complaints"),
            ("Ofcom comparing customer service", "Regulator", "https://www.ofcom.org.uk/phones-and-broadband/service-quality/comparing-customer-service"),
        ],
    },
    "/guides/satellite-broadband-uk": {
        "topic": "Satellite broadband UK",
        "target_terms": ["satellite", "rural", "latency", "equipment", "coverage"],
        "sources": [
            ("Starlink UK residential", "Provider", "https://www.starlink.com/gb/residential"),
            ("Ofcom Connected Nations", "Regulator", "https://www.ofcom.org.uk/phones-and-broadband/coverage-and-speeds/connected-nations"),
        ],
    },
    "/postcode/london": {
        "topic": "Broadband providers London",
        "target_terms": ["London", "coverage", "full fibre", "postcode", "providers"],
        "sources": [("Ofcom coverage checker", "Regulator", "https://checker.ofcom.org.uk/en-gb/broadband-coverage")],
    },
}

SOURCES = [
    ("Ubersuggest keyword ideas", "UK monthly volume, SEO difficulty and CPC snapshot", "2026-07-29",
     "https://app.neilpatel.com/en/ubersuggest/keyword_ideas/", "Third-party estimates; directional, not guaranteed."),
    ("Google autocomplete", "Live UK query suggestions collected by this script", RUN_DATE,
     "https://suggestqueries.google.com/complete/search", "Demand signal only; no volume."),
    ("BroadbandPicker sitemap", "Existing live URL inventory and page audit", RUN_DATE,
     f"{SITE}/sitemap.xml", "First-party crawl."),
    ("Ofcom: switching provider", "One Touch Switch consumer guidance", "Updated 2026-07-01",
     "https://www.ofcom.org.uk/phones-and-broadband/switching-provider/switching", "Primary regulator source."),
    ("Ofcom: pricing and consumer engagement", "Market, pricing and switching evidence", "2026-02-26",
     "https://www.ofcom.org.uk/siteassets/resources/documents/research-and-data/multi-sector/pricing/2025/pricing-and-consumer-engagement-report.pdf?v=412887",
     "Primary regulator report."),
    ("ASA/CAP: affiliate marketing", "Affiliate disclosure requirements", "Accessed " + RUN_DATE,
     "https://www.asa.org.uk/advice-online/affiliate-marketing.html", "Primary compliance source."),
    ("Awin FAQs", "Publisher profile and advertiser approval guidance", "Accessed " + RUN_DATE,
     "https://www.awin.com/gb/faqs", "A complete publisher profile improves approval chances."),
    ("Google Search: AI optimisation", "Official generative-search guidance", "Updated 2026",
     "https://developers.google.com/search/docs/fundamentals/ai-optimization-guide",
     "No special GEO schema; unique, useful, crawlable content wins."),
    ("Google Search: helpful content", "People-first content and quality self-assessment", "Updated 2025",
     "https://developers.google.com/search/docs/fundamentals/creating-helpful-content", "Primary search guidance."),
]

LANDING_PAGES = [
    ("P0", "/speed-test", "Broadband speed test", "broadband test speed test", 246000, "Tool",
     "Improve existing", "Fast test, postcode/provider context, result interpretation, next-step deals",
     "WebApplication + FAQ where visible", "Speed cluster; link results to /compare"),
    ("P0", "/deals", "Best broadband deals UK", "broadband deals", 165000, "Commercial",
     "Improve existing", "Postcode first, total contract cost, setup fees, rises, speed and verified timestamp",
     "ItemList", "Canonical owner for generic deal terms"),
    ("P0", "/guides/best-broadband-providers-uk", "Best broadband providers UK", "broadband providers best", 8100,
     "Commercial research", "Improve existing", "Transparent scoring, negatives, use cases, Ofcom evidence, author/test method",
     "Article + ItemList", "Do not make /providers compete for 'best'"),
    ("P0", "/compare", "Compare broadband providers", "comparison of broadband providers", 4400,
     "Commercial", "Improve existing", "Postcode availability, sortable true-cost table, clear commission disclosure",
     "WebApplication/ItemList as eligible", "Comparison head term owner"),
    ("P0", "/guides/how-to-switch-broadband-uk", "Switch broadband provider", "switch broadband providers", 3600,
     "Commercial guide", "Improve existing", "One Touch Switch steps, charges, downtime, checklist, last verified date",
     "HowTo only if all visible + Article", "Link to deals, contract-end and moving pages"),
    ("P0", "/guides/best-business-broadband-providers-uk", "Best business broadband providers UK", "broadband providers for business", 1900,
     "B2B commercial", "Create", "SLA, static IP, backup, support hours, leased line vs FTTP, quote CTA",
     "Article + ItemList", "Keep inside the established /guides hierarchy; feature from Guides and footer"),
    ("P0", "/postcode", "Broadband providers in my area", "broadband providers for my area", 5400,
     "Local commercial", "Create parent hub", "Address/postcode checker, networks vs retailers, area links, methodology",
     "WebApplication + FAQ", "Parent for existing /postcode/[area] pages; prevents postcode pages being orphaned"),
    ("P1", "/guides/best-broadband-and-tv-deals", "Best broadband and TV deals", "broadband deals and tv", 22200,
     "Commercial", "Improve existing", "Sports/entertainment segmentation, total bundle cost, exit implications",
     "ItemList + Article", "Own TV bundle cluster"),
    ("P1", "/guides/best-phone-and-broadband-deals", "Phone and broadband deals", "broadband deals and phone", 22200,
     "Commercial", "Create/consolidate", "Call packages, line rental, digital voice, vulnerable-user considerations",
     "Article + ItemList", "Consolidate telephone/phone/landline variants"),
    ("P1", "/guides/cheapest-broadband-uk", "Cheapest broadband UK", "broadband deals cheapest", 18100,
     "Commercial", "Improve existing", "True monthly cost, setup, rises, reward valuation, minimum speed",
     "ItemList + Article", "Distinct value angle from /deals"),
    ("P1", "/providers/bt/deals", "BT broadband deals", "bt broadband deals", 33100,
     "Brand commercial", "Create only with live feed", "Current offers, new vs existing customer, verified prices, neutral alternatives",
     "ItemList", "Template for approved providers; avoid thin doorway pages"),
    ("P1", "/providers/sky/deals", "Sky broadband deals", "sky broadband deals", 33100,
     "Brand commercial", "Create only with live feed", "Current offers, TV bundle cross-over, price verification, alternatives",
     "ItemList", "Template for approved providers; avoid thin doorway pages"),
    ("P1", "/providers/ee/deals", "EE broadband deals", "ee broadband deals", 33100,
     "Brand commercial", "Create only with live feed", "Current offers, mobile bundle value, Smart Hub facts, alternatives",
     "ItemList", "Template for approved providers; avoid thin doorway pages"),
    ("P1", "/providers/virgin-media/deals", "Virgin Media broadband deals", "virgin broadband deals", 12100,
     "Brand commercial", "Create only with live feed", "Cable/full-fibre availability, bundles, price verification, alternatives",
     "ItemList", "Template for approved providers; avoid thin doorway pages"),
    ("P1", "/research/uk-broadband-customer-satisfaction", "UK broadband customer satisfaction", "uk broadband providers customer satisfaction", 1600,
     "Research/link earning", "Create", "Ofcom complaints, satisfaction, method, downloadable CSV, limitations",
     "Dataset + Article", "Original synthesis; refresh on regulator release"),
    ("P1", "/guides/broadband-social-tariffs-uk", "Broadband social tariffs", "broadband social tariff", 0,
     "Informational/public service", "Improve existing", "Eligibility, prices, evidence date, how to apply, provider table",
     "Article + ItemList", "Trust asset; volume to validate in GSC/Ads"),
    ("P2", "/guides/satellite-broadband-uk", "Satellite broadband providers UK", "satellite broadband providers", 720,
     "Commercial guide", "Create", "Coverage, latency, data, installation, Starlink and alternatives",
     "Article + ItemList", "Rural/alternative access cluster"),
    ("P2", "/postcode/london", "Broadband providers London", "broadband providers london uk", 2400,
     "Local commercial", "Create after data", "Borough/network availability, median speeds, provider footprints, postcode CTA",
     "Article + Dataset", "Place below /postcode and above London postcode-prefix pages; launch only with unique evidence"),
]

BUILD_STATUS = [
    ["Platform", "Production Next.js site", "Built", "/", "Keep dependencies and deployment checks current"],
    ["Commercial", "Deals and comparison journeys", "Built", "/deals; /compare", "Refresh verified prices and conversion data"],
    ["Tools", "Broadband speed test", "Built", "/speed-test", "Monitor completion and outbound-click events"],
    ["Content", "Priority guide pages", "Built", "/guides", "Refresh facts, sources and internal links monthly"],
    ["Providers", "Provider reviews and comparisons", "Built", "/providers; /providers/compare", "Expand only where evidence and demand justify it"],
    ["Providers", "Live provider deal pilots", "Built", "/providers/bt/deals; /providers/sky/deals; /providers/ee/deals; /providers/virgin-media/deals", "Monitor feed freshness and programme terms"],
    ["Local", "Postcode hub, London and area pages", "Built", "/postcode; /postcode/london; /postcode/[area]", "Add locations only with unique coverage evidence"],
    ["Research", "UK broadband customer satisfaction report", "Built", "/research/uk-broadband-customer-satisfaction", "Refresh on each relevant Ofcom release"],
    ["Trust", "About, editorial, monetisation and legal policies", "Built", "/about; /editorial-policy; /how-we-make-money; /privacy-policy; /cookie-policy; /terms", "Review when commercial or data practices change"],
    ["Monetisation", "Awin publisher setup and Highland Broadband programme", "Built", "Awin publisher 2942019", "Continue targeted advertiser outreach"],
    ["Monetisation", "AdSense ownership, consent message and ads.txt", "In progress", "/ads.txt", "Wait for site review and ads.txt crawler authorisation"],
    ["Measurement", "GSC, analytics and conversion reporting loop", "In progress", "Sitewide", "Use query, engagement and assisted-conversion data to reprioritise"],
    ["Growth", "Original research outreach and digital PR", "Pending", "/research", "Create outreach list and earn citations/links"],
    ["Growth", "Additional evidence-led local pages", "Pending", "/postcode", "Pilot only after coverage data and Search Console demand validate them"],
]


def fetch(url: str, timeout: int = 20) -> requests.Response:
    response = requests.get(url, timeout=timeout, headers={"User-Agent": USER_AGENT})
    response.raise_for_status()
    return response


def get_sitemap_urls(offline: bool) -> list[str]:
    if offline:
        return []
    try:
        root = ET.fromstring(fetch(f"{SITE}/sitemap.xml").content)
        return [node.text.strip() for node in root.findall(".//{*}loc") if node.text]
    except Exception as exc:
        print(f"Warning: sitemap fetch failed: {exc}")
        return []


def audit_page(url: str) -> dict[str, Any]:
    row = {"url": url, "status": "", "title": "", "title_len": 0, "description": "",
           "description_len": 0, "h1_count": 0, "canonical": "", "json_ld": 0, "words": 0,
           "affiliate_disclosure": False, "last_verified": False, "notes": "",
           "indexable": False, "robots": "", "canonical_match": False,
           "h2_count": 0, "schema_types": "", "internal_links": 0,
           "external_links": 0, "images": 0, "images_missing_alt": 0,
           "html_lang": "", "viewport": False, "og_complete": False,
           "author_signal": False, "visible_sources": False,
           "question_headings": 0, "concise_answers": 0, "seo_score": 0,
           "geo_score": 0, "diagnosis": "", "priority_actions": ""}
    try:
        response = fetch(url)
        tree = html.fromstring(response.content)
        title = " ".join(tree.xpath("//title/text()")).strip()
        descriptions = tree.xpath("//meta[translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='description']/@content")
        description = descriptions[0].strip() if descriptions else ""
        canonicals = tree.xpath("//link[translate(@rel,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='canonical']/@href")
        text_content = " ".join(tree.xpath("//body//text()[normalize-space()]"))
        clean_text = re.sub(r"\s+", " ", text_content).strip()
        lower = clean_text.lower()
        freshness_signal = any(
            label in lower for label in ("last verified", "last reviewed", "last tested", "last updated", "last researched", "updated")
        )
        robots_values = tree.xpath(
            "//meta[translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='robots']/@content"
        )
        robots = ", ".join(robots_values).lower()
        x_robots = response.headers.get("X-Robots-Tag", "").lower()
        canonical = canonicals[0].strip() if canonicals else ""
        canonical_match = canonical.rstrip("/") == url.rstrip("/")
        links = tree.xpath("//a[@href]/@href")
        internal_links = sum(
            1 for link in links
            if link.startswith("/") or urllib.parse.urlparse(link).netloc in {"", "broadbandpicker.co.uk", "www.broadbandpicker.co.uk"}
        )
        external_links = sum(1 for link in links if link.startswith("http") and "broadbandpicker.co.uk" not in link)
        images = tree.xpath("//img")
        images_missing_alt = sum(1 for image in images if not (image.get("alt") or "").strip())
        schema_types: set[str] = set()
        for raw_schema in tree.xpath("//script[@type='application/ld+json']/text()"):
            try:
                payload = json.loads(raw_schema)
                nodes = payload if isinstance(payload, list) else [payload]
                for node in nodes:
                    if isinstance(node, dict) and "@graph" in node:
                        nodes.extend(item for item in node["@graph"] if isinstance(item, dict))
                    if isinstance(node, dict) and node.get("@type"):
                        value = node["@type"]
                        schema_types.update(value if isinstance(value, list) else [str(value)])
            except (json.JSONDecodeError, TypeError):
                schema_types.add("Invalid JSON-LD")
        headings = [re.sub(r"\s+", " ", value).strip() for value in tree.xpath("//h2//text() | //h3//text() | //summary//text()")]
        question_headings = sum(1 for value in headings if "?" in value or re.match(r"(?i)^(what|how|why|when|which|is|can|does|do)\b", value))
        paragraphs = [
            re.sub(r"\s+", " ", element.text_content()).strip()
            for element in tree.xpath("//main//p | //article//p")
        ]
        concise_answers = sum(1 for value in paragraphs if 35 <= len(value.split()) <= 90)
        source_domains = ("ofcom.org.uk", "gov.uk", "asa.org.uk", "openreach.com")
        visible_sources = any(any(domain in link for domain in source_domains) for link in links)
        author_signal = bool(tree.xpath("//*[@rel='author'] | //*[contains(translate(@class,'AUTHOR','author'),'author')]")) or "reviewed by" in lower
        og_title = tree.xpath("//meta[@property='og:title']/@content")
        og_description = tree.xpath("//meta[@property='og:description']/@content")
        html_lang = " ".join(tree.xpath("/html/@lang")).strip()
        viewport = bool(tree.xpath("//meta[translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='viewport']"))
        h1_count = len(tree.xpath("//h1"))
        h2_count = len(tree.xpath("//h2"))
        word_count = len(re.findall(r"\b[\w'-]+\b", clean_text))
        indexable = response.status_code == 200 and "noindex" not in robots and "noindex" not in x_robots
        title_ok = 30 <= len(title) <= 65
        description_ok = 70 <= len(description) <= 165
        expected_schema = "WebApplication" if urllib.parse.urlparse(url).path == "/speed-test" else ""
        schema_ok = bool(schema_types) and (not expected_schema or expected_schema in schema_types)
        seo_checks = [indexable, title_ok, description_ok, h1_count == 1, h2_count >= 2,
                      canonical_match, viewport, bool(html_lang), len(images) == 0 or images_missing_alt == 0,
                      internal_links >= 3, schema_ok]
        seo_score = round(100 * sum(seo_checks) / len(seo_checks))
        geo_checks = [word_count >= 500, question_headings >= 2, concise_answers >= 2,
                      visible_sources, author_signal, bool(schema_types), external_links >= 1,
                      freshness_signal]
        geo_score = round(100 * sum(geo_checks) / len(geo_checks))
        actions: list[str] = []
        if not indexable: actions.append("Remove noindex/blocking and confirm a 200 response")
        if not title_ok: actions.append("Write a unique, concise title that states the page topic and UK intent")
        if not description_ok: actions.append("Add a unique benefit-led meta description")
        if h1_count != 1: actions.append("Use one descriptive visible H1")
        if h2_count < 2: actions.append("Add scannable H2 sections matching user follow-up questions")
        if not canonical_match: actions.append("Set a self-referencing canonical")
        if not schema_types: actions.append("Add valid structured data that matches the visible page")
        elif expected_schema and expected_schema not in schema_types:
            actions.append(f"Add valid {expected_schema} JSON-LD for the visible interactive tool")
        if question_headings < 2 or concise_answers < 2: actions.append("Add answer-first question sections with concise factual summaries")
        if not visible_sources: actions.append("Cite visible primary UK sources such as Ofcom")
        if not author_signal: actions.append("Show author/reviewer, methodology and editorial responsibility")
        if not freshness_signal: actions.append("Show a genuine last-tested/updated date")
        if internal_links < 3: actions.append("Add contextual links to comparison, deals and speed guidance")
        diagnosis = "Strong" if min(seo_score, geo_score) >= 80 else "Needs improvement" if min(seo_score, geo_score) >= 55 else "High-priority rebuild"
        row.update({
            "status": response.status_code, "title": title, "title_len": len(title),
            "description": description, "description_len": len(description),
            "h1_count": h1_count, "canonical": canonical,
            "json_ld": len(tree.xpath("//script[@type='application/ld+json']")),
            "words": word_count,
            "affiliate_disclosure": ("commission" in lower or "affiliate" in lower),
            "last_verified": freshness_signal,
            "indexable": indexable, "robots": robots or x_robots or "index/follow implied",
            "canonical_match": canonical_match, "h2_count": h2_count,
            "schema_types": ", ".join(sorted(schema_types)), "internal_links": internal_links,
            "external_links": external_links, "images": len(images),
            "images_missing_alt": images_missing_alt, "html_lang": html_lang,
            "viewport": viewport, "og_complete": bool(og_title and og_description),
            "author_signal": author_signal, "visible_sources": visible_sources,
            "question_headings": question_headings, "concise_answers": concise_answers,
            "seo_score": seo_score, "geo_score": geo_score, "diagnosis": diagnosis,
            "priority_actions": "; ".join(actions[:8]) or "Monitor in Search Console and refresh evidence",
        })
    except Exception as exc:
        row["notes"] = str(exc)[:250]
    return row


def collect_autocomplete(offline: bool) -> list[tuple[str, str]]:
    if offline:
        return []
    collected: set[tuple[str, str]] = set()
    endpoint = "https://suggestqueries.google.com/complete/search"
    for seed in SEEDS:
        try:
            response = fetch(endpoint + "?" + urllib.parse.urlencode({
                "client": "firefox", "q": seed, "gl": "uk", "hl": "en",
            }))
            payload = response.json()
            for suggestion in payload[1]:
                collected.add((seed, suggestion))
            time.sleep(0.2)
        except Exception as exc:
            print(f"Warning: autocomplete failed for {seed!r}: {exc}")
    return sorted(collected)


def collect_page_research(offline: bool) -> list[list[Any]]:
    """Scrape registered primary sources into an evidence log for content refreshes."""
    rows: list[list[Any]] = []
    for page_path, config in PAGE_RESEARCH_SOURCES.items():
        terms = config["target_terms"]
        for label, source_type, url in config["sources"]:
            result = {
                "status": "Offline", "http": "", "title": "", "h1": "",
                "word_count": 0, "matched": "", "extract": "",
            }
            if not offline:
                try:
                    response = fetch(url, timeout=30)
                    tree = html.fromstring(response.content)
                    title = " ".join(tree.xpath("//title/text()")).strip()
                    h1 = " ".join(tree.xpath("//h1//text()")).strip()
                    for noisy_node in tree.xpath("//script|//style|//noscript|//svg"):
                        noisy_node.drop_tree()
                    raw_text = " ".join(tree.xpath("//body//text()[normalize-space()]"))
                    clean_text = re.sub(r"\s+", " ", raw_text).strip()
                    matched = [term for term in terms if term.lower() in clean_text.lower()]
                    result.update({
                        "status": "Retrieved",
                        "http": response.status_code,
                        "title": title,
                        "h1": h1,
                        "word_count": len(re.findall(r"\b[\w'-]+\b", clean_text)),
                        "matched": ", ".join(matched),
                        "extract": clean_text[:700],
                    })
                except Exception as exc:
                    result["status"] = "Failed"
                    result["extract"] = str(exc)[:700]
            rows.append([
                page_path, config["topic"], label, source_type, url, RUN_DATE,
                result["status"], result["http"], result["title"], result["h1"],
                result["word_count"], result["matched"], result["extract"],
                "Human verification required before publishing prices, contract terms or rankings",
            ])
    return rows


def route_family(path: str) -> str:
    parts = [part for part in path.strip("/").split("/") if part]
    if not parts:
        return "Homepage"
    if parts[0] == "providers" and len(parts) > 1 and parts[1] == "compare":
        return "Provider comparisons"
    return {
        "guides": "Guides",
        "providers": "Providers",
        "postcode": "Postcode/area",
        "compare": "Compare tool",
        "deals": "Deals",
        "speed-test": "Speed test",
    }.get(parts[0], "Trust/static")


def parent_path(path: str) -> str:
    parts = [part for part in path.strip("/").split("/") if part]
    return "/" if len(parts) <= 1 else "/" + "/".join(parts[:-1])


def breadcrumb_for(path: str, title: str) -> str:
    labels = {
        "guides": "Guides", "providers": "Providers", "compare": "Compare",
        "postcode": "Broadband in your area", "deals": "Deals",
        "speed-test": "Speed Test", "research": "Research",
    }
    parts = [part for part in path.strip("/").split("/") if part]
    crumbs = ["Home"]
    built: list[str] = []
    for index, part in enumerate(parts):
        built.append(part)
        if index == len(parts) - 1:
            crumbs.append(title)
        else:
            crumbs.append(labels.get(part, part.replace("-", " ").title()))
    return " > ".join(crumbs)


def navigation_placement(path: str, priority: str) -> tuple[str, str]:
    family = route_family(path)
    if path in {"/deals", "/compare", "/providers", "/guides", "/speed-test"}:
        return "Primary header", "Existing top-level destination"
    if path == "/postcode":
        return "Primary header or prominent postcode control", "Add 'In your area' destination; link every /postcode/[area] page back"
    if family == "Provider comparisons":
        return "Providers hub + provider pages", "Link contextually from both compared provider pages"
    if family == "Providers":
        return "Providers hub + relevant provider page", "Do not add every provider/deal URL to the header"
    if family == "Guides":
        return "Guides hub + relevant category hub", "Feature P0/P1 pages in category cards; footer only for durable high-value guides"
    if family == "Postcode/area":
        return "/postcode hub + regional parent", "Use hierarchical area links; avoid a flat list of all prefixes in global navigation"
    if path.startswith("/research"):
        return "Guides hub initially; Research hub when 3+ assets exist", "Do not create an empty global-nav section"
    return ("Contextual links", "Add from the closest hub and at least two relevant pages")


def priority_score(volume: int, difficulty: int, cpc: float, intent: str) -> float:
    intent_weight = 1.25 if "Commercial" in intent or "commercial" in intent else 1.0
    demand = min(100, (volume ** 0.5) / 4) if volume else 5
    opportunity = max(5, 100 - difficulty)
    value = min(100, cpc * 4)
    return round((demand * 0.45 + opportunity * 0.35 + value * 0.20) * intent_weight, 1)


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[Any]]) -> Any:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    if rows and headers:
        ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        table = Table(displayName=re.sub(r"\W+", "", name)[:20] + "Table", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    return ws


def style_sheet(ws: Any) -> None:
    navy, aqua, pale = "12304A", "11A8A0", "EAF6F5"
    thin = Side(style="thin", color="D7E1E8")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)
    for column in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, column).value or "") for row in range(1, min(ws.max_row, 80) + 1)]
        width = min(48, max(10, max(len(v) for v in values) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for cell in ws[row]:
                cell.fill = PatternFill("solid", fgColor=pale)


def build_workbook(output: Path, urls: list[str], audit: list[dict[str, Any]],
                   autocomplete: list[tuple[str, str]], page_research: list[list[Any]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    readme = [
        ["Purpose", "A practical, evidence-led SEO and GEO roadmap for BroadbandPicker.co.uk, including Awin advertiser readiness."],
        ["Generated", RUN_DATE],
        ["Scope", "UK broadband; SEO, generative-search visibility, landing pages, content governance, internal linking and affiliate compliance."],
        ["Evidence hierarchy", "First-party site crawl + primary Ofcom/ASA/Awin/Google guidance + directional Ubersuggest estimates + Google autocomplete."],
        ["Volume caveat", "Monthly volume is a third-party estimate. It is not additive across variants and must not be presented as a traffic forecast."],
        ["Zero-volume caveat", "A zero means not measured/not supplied, not no demand. Validate with Google Ads Keyword Planner and Search Console."],
        ["GEO principle", "There is no special GEO schema. Build indexable, source-backed, original, answer-first pages that are easy to quote and keep current."],
        ["Awin principle", "Approval cannot be guaranteed. A complete profile, real audience evidence, transparent promotional methods and programme-specific compliance improve the case."],
        ["How to use", "Start with P0 pages and Awin controls. Assign one canonical URL per cluster. Review metrics monthly and refresh commercial facts at least monthly."],
        ["Re-run", "python3 scripts/build_uk_broadband_seo_plan.py --sync-google-sheet"],
    ]
    add_sheet(wb, "Read Me", ["Field", "Detail"], readme)

    add_sheet(
        wb,
        "Build Status",
        ["Workstream", "Deliverable", "Status", "Evidence", "Next action"],
        BUILD_STATUS,
    )

    executive = [
        ["1", "Fix trust/data risks before outreach", "P0", "Owner: founder/editor",
         "Named accountable publisher; company/contact details; visible disclosures; sourced ratings; verified deal timestamps", "Weeks 1-2"],
        ["2", "Strengthen money-page utility", "P0", "Owner: product/content",
         "Postcode-first comparison, true contract cost, price-rise notes, availability and impartial ranking method", "Weeks 1-4"],
        ["3", "Win high-demand tools and commercial clusters", "P0", "Owner: SEO/product",
         "Speed test, deals, providers, comparison, switching, area checker, business broadband", "Weeks 2-8"],
        ["4", "Create citation-worthy evidence", "P1", "Owner: research/editor",
         "Ofcom-backed customer satisfaction dataset, source notes, limitations and downloadable tables", "Weeks 5-10"],
        ["5", "Apply to Awin advertisers selectively", "P0", "Owner: partnerships",
         "Submit only after QA; tailor pitch with audience, traffic, content examples, promotional method and compliance links", "Weeks 4-8"],
        ["6", "Scale only proven templates", "P1", "Owner: SEO/editor",
         "Provider deal pages require live feed; local pages require unique coverage/speed data; no mass thin variants", "Weeks 8-13"],
    ]
    add_sheet(wb, "Executive Plan", ["#", "Workstream", "Priority", "Owner", "Definition of done", "Timing"], executive)

    kw_rows = []
    for keyword, volume, sd, cpc, intent, cluster in KEYWORDS:
        kw_rows.append([keyword, volume, sd, cpc, intent, cluster,
                        priority_score(volume, sd, cpc, intent), "Ubersuggest UK/en", "2026-07-29"])
    kw_rows.sort(key=lambda row: row[6], reverse=True)
    ws_kw = add_sheet(wb, "Keyword Research",
                      ["Keyword", "UK monthly volume", "SEO difficulty", "CPC GBP", "Intent", "Cluster",
                       "Opportunity score", "Source", "Snapshot date"], kw_rows)
    ws_kw.conditional_formatting.add(f"G2:G{ws_kw.max_row}",
                                     ColorScaleRule(start_type="min", start_color="F8696B",
                                                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                                    end_type="max", end_color="63BE7B"))
    for cell in ws_kw["D"][1:]:
        cell.number_format = '£0.00'
    for cell in ws_kw["B"][1:]:
        cell.number_format = '#,##0'

    add_sheet(wb, "Autocomplete",
              ["Seed", "Live Google UK suggestion", "Use"],
              [[seed, suggestion, "Map to existing cluster before creating any URL"] for seed, suggestion in autocomplete]
              or [["Offline/no data", "Re-run without --offline to collect live suggestions", "No action"]])

    add_sheet(wb, "Page Research Evidence",
              ["Target page", "Topic", "Source", "Source type", "URL", "Retrieved date",
               "Retrieval status", "HTTP", "Source title", "Source H1", "Approx. words",
               "Target terms found", "Evidence extract", "Publication control"],
              page_research or [["No registered sources", "", "", "", "", RUN_DATE, "", "", "", "",
                                 "", "", "", "Add the page to PAGE_RESEARCH_SOURCES"]])

    lp_rows = []
    for row in LANDING_PAGES:
        priority, path, title, keyword, volume, intent, action, modules, schema, notes = row
        existing = "Yes" if any(u.rstrip("/").endswith(path.rstrip("/")) for u in urls) else "No"
        current_action = "Maintain and improve" if existing == "Yes" else action
        nav_location, nav_reason = navigation_placement(path, priority)
        lp_rows.append([priority, path, parent_path(path), route_family(path), title, keyword, volume, intent,
                        existing, current_action, nav_location, breadcrumb_for(path, title), modules, schema,
                        nav_reason + "; " + notes])
    add_sheet(wb, "Landing Pages",
              ["Priority", "Recommended URL", "Parent URL", "Current route family", "Page/H1", "Primary keyword",
               "Est. volume", "Intent", "Exists live", "Action", "Navigation placement", "Breadcrumb trail",
               "Required differentiators", "Schema guidance", "Navigation/cannibalisation note"],
              lp_rows)

    paths = [urllib.parse.urlparse(url).path or "/" for url in urls]
    family_counts = Counter(route_family(path) for path in paths)
    hierarchy_rows = []
    for family, count in sorted(family_counts.items(), key=lambda item: (-item[1], item[0])):
        examples = ", ".join(path for path in paths if route_family(path) == family)[:350]
        hierarchy_rows.append([family, count, examples])
    add_sheet(wb, "Current URL Hierarchy",
              ["Live route family", "URL count", "Example live paths"],
              hierarchy_rows or [["No live hierarchy", 0, "Run without --offline"]])

    nav_rows = []
    for priority, path, title, keyword, volume, intent, action, modules, schema, notes in LANDING_PAGES:
        nav_location, nav_reason = navigation_placement(path, priority)
        nav_rows.append([
            priority, path, parent_path(path), route_family(path), nav_location,
            breadcrumb_for(path, title), "Parent hub + 2 contextual links minimum",
            "Yes" if priority == "P0" else "No",
            nav_reason, "Index only when unique, complete and linked from its parent",
        ])
    add_sheet(wb, "Navigation Architecture",
              ["Priority", "Page URL", "Parent/hub URL", "Route family", "Discovery placement",
               "Recommended breadcrumb", "Minimum inbound links", "Feature prominently",
               "Navigation rationale", "Indexation gate"], nav_rows)

    clusters = [
        ["Deals & value", "/deals", "Best, cheapest, no price rise, cashback, setup fee, short contracts",
         "Provider deal pages; bundles; social tariffs", "Compare/deals CTA", "Monthly + on provider change"],
        ["Providers & reviews", "/providers", "Provider entity pages and best-provider methodology",
         "Provider comparisons; deals; complaints/satisfaction", "Provider/availability CTA", "Quarterly + Ofcom releases"],
        ["Availability & local", "/postcode", "Networks, postcode/address eligibility, selective cities",
         "Full fibre; rural; provider footprints", "Postcode checker", "Coverage refresh monthly"],
        ["Speed & technology", "/speed-test", "Speed test, speed needs, FTTP/FTTC, latency, Wi-Fi",
         "Gaming, streaming, WFH, routers", "Test then compare", "Technical pages 6-monthly"],
        ["Switching & rights", "/guides/how-to-switch-broadband-uk", "OTS, contract end, price rises, moving, cancellation",
         "Complaints, compensation, vulnerable users", "Switch/compare CTA", "On Ofcom rule change"],
        ["Bundles", "/guides/best-broadband-and-tv-deals", "TV, phone, landline, mobile bundles",
         "Sky/Virgin/BT bundle comparisons", "Bundle comparison", "Monthly"],
        ["Business", "/guides/best-business-broadband-providers-uk", "SME providers, SLA, leased lines, static IP, backup",
         "Business by need/size; buyer guides", "Quote/lead CTA", "Quarterly"],
        ["Research", "/research", "Customer satisfaction, complaints, price trackers, coverage datasets",
         "Methodology and downloadable data", "Newsletter/citation", "On source release"],
    ]
    add_sheet(wb, "Content Clusters",
              ["Cluster", "Pillar URL", "Core coverage", "Supporting content", "Primary conversion", "Freshness SLA"], clusters)

    roadmap = [
        ["1-15", "Trust + measurement", "Publish named ownership/entity details; verify contact; affiliate disclosure above first commercial CTA; remove unsupported rating counts; configure GSC/Bing/GA4 and conversion events; fix crawl/index issues", "P0", "All controls pass; baseline captured", "In progress", "Trust pages and disclosures are live; measurement review remains ongoing"],
        ["1-15", "Awin pack", "Complete publisher profile; document audience, traffic sources and promotional methods; prepare 3 tailored application examples; add legal entity and domain email", "P0", "Application pack QA complete", "Built", "Publisher 2942019 is active and Highland Broadband invitation was accepted"],
        ["16-30", "Core money pages", "Upgrade /deals, /compare, best providers and switching pages with true-cost, dates, methodology, sources and unique decision tables", "P0", "4 pages published; commercial facts verified", "Built", "/deals, /compare and priority guides are live"],
        ["16-30", "Speed test", "Improve result explanation, contextual next steps, shareable result, FAQ and internal links; measure completed tests", "P0", "Test completion and CTA events tracked", "Built", "/speed-test is live; continue event-quality monitoring"],
        ["31-45", "New demand pages", "Launch /postcode parent hub and /guides/best-business-broadband-providers-uk with real data and expert review", "P0", "2 indexable differentiated pages", "Built", "Both planned URLs are live"],
        ["31-45", "Awin outreach", "Apply first to best-fit/auto-approval programmes, then priority providers with tailored pitch and relevant URLs", "P0", "Applications logged; feedback recorded", "In progress", "Highland Broadband joined; continue selective outreach"],
        ["46-60", "Bundles", "Consolidate phone/telephone/landline intent; strengthen TV bundles; create hub-to-supporting internal links", "P1", "No keyword cannibalisation; 2 clusters live", "Built", "TV and phone/broadband guide pages are live"],
        ["46-60", "Provider deal pilots", "Create 2 provider-deal templates only where feeds are reliable and programme terms allow use", "P1", "Prices stamped and monitored", "Built", "BT, Sky, EE and Virgin Media deal pages are live with feed monitoring"],
        ["61-75", "Original research", "Publish customer satisfaction/complaints dataset with method, source downloads, caveats, author and update log", "P1", "Dataset + outreach list live", "In progress", "Research page is live; outreach and citation work remains"],
        ["76-90", "Selective local scale", "Pilot London plus 2 areas only where unique network/coverage data supports a useful page", "P2", "No thin/location-swapped copy", "Built", "London hub and evidence-led postcode area pages are live"],
        ["76-90", "Review and iterate", "Use GSC query/page data, assisted conversions, Awin decisions and engagement to reprioritise quarter two", "P0", "Decision log and next-quarter plan", "Pending", "Complete after sufficient Search Console, AdSense and affiliate data accrues"],
    ]
    add_sheet(wb, "90 Day Roadmap",
              ["Days", "Workstream", "Actions", "Priority", "Exit criterion", "Status", "Evidence / next step"], roadmap)

    geo = [
        ["Answer-first", "Put a direct 40-80 word answer below H1; then decision table, evidence and nuance", "All guide/comparison pages", "Snippet and cited-answer eligibility"],
        ["Entity clarity", "Use consistent provider names, network ownership, service area, authors and publisher details", "Templates + About", "Reduced ambiguity"],
        ["Evidence blocks", "Cite Ofcom/provider primary sources beside claims; show last verified and limitations", "Commercial and rights pages", "Citation trust"],
        ["Original information", "Publish tests, scoring inputs, complaints/satisfaction synthesis, price histories and downloadable data", "Research/provider pages", "Link and citation earning"],
        ["Atomic sections", "One descriptive H2 per question; short self-contained answer followed by supporting detail", "Guides", "Easy extraction without fragment spam"],
        ["Tables", "Use accessible HTML tables with units, dates, source notes and row-level context", "Deals/comparisons/research", "Machine and human clarity"],
        ["Freshness", "Separate published, updated and price-verified dates; keep an edit log for high-risk pages", "All money pages", "Reliable current answers"],
        ["Technical", "Indexable SSR content, canonical URLs, clean internal links, XML sitemap, good CWV and crawl access", "Sitewide", "Search foundation"],
        ["Schema restraint", "Markup only visible content with supported types; no special GEO schema or fabricated ratings", "Sitewide", "Eligibility without policy risk"],
        ["Multimedia", "Add original charts/screenshots/video where they materially explain tests or comparisons", "Research/tools", "More distinct evidence"],
        ["Do not do", "No mass AI city pages, FAQ variations, copied provider copy, hidden text or schema-only content", "Governance", "Avoid scaled-content abuse"],
    ]
    add_sheet(wb, "GEO Playbook", ["Pattern", "Implementation", "Where", "Why"], geo)

    awin = [
        ["Publisher identity", "Named owner/editor, legal/trading identity where applicable, domain email and working contact route", "Must fix", "About + footer + contact", "Trust"],
        ["Complete Awin profile", "Audience, UK focus, promotional type, website URL, traffic sources and content examples", "Must fix", "Awin profile", "Awin says completeness improves approval odds"],
        ["Affiliate disclosure", "Commercial nature clear before engagement and before/near affiliate CTAs", "Must fix", "Money pages + sitewide disclosure", "ASA/CAP 2.1/2.3"],
        ["Editorial independence", "Explain funding, ranking, inclusion, correction and update processes", "Pass/verify", "/how-we-make-money + editorial policy", "Advertiser brand safety"],
        ["Review methodology", "Sourced scoring inputs, weighting, negatives, who reviewed, last verified and correction route", "Must improve", "/how-we-review-broadband", "Credibility"],
        ["Accurate provider claims", "Use approved creative/copy, current prices, terms, availability caveats and no unsupported superlatives", "Must fix", "All commercial pages", "Programme compliance"],
        ["Ratings/schema", "Remove fabricated or unsupported aggregate review counts and ratings", "Urgent audit", "Provider templates", "Search and advertiser risk"],
        ["User value", "Working comparison/postcode/speed tools plus substantial independent content", "Pass/improve", "Core product", "Shows genuine publisher value"],
        ["Legal/privacy", "GDPR/cookie/privacy/terms match actual analytics, email and affiliate processing", "Verify", "Legal pages + CMP", "Compliance"],
        ["Site quality", "Mobile QA, no broken links/images, fast pages, no placeholders, current content", "Verify", "Sitewide", "Reviewer first impression"],
        ["Audience evidence", "Provide GSC/GA4 trend, engaged sessions, UK share, newsletter/social evidence without exaggeration", "Build", "Application pitch", "Commercial credibility"],
        ["Tailored application", "Name relevant pages, audience fit, promotion method, compliance controls and planned placements", "Required", "Per provider", "Avoid generic applications"],
        ["No guarantee", "Programme owners decide; respect programme-specific rules and document rejection feedback", "Ongoing", "Partner log", "Realistic governance"],
    ]
    add_sheet(wb, "Awin Readiness", ["Control", "Acceptance evidence", "Status", "Location", "Reason"], awin)

    technical = [
        ["Indexation", "GSC coverage, sitemap submitted, canonical self-reference, no accidental noindex", "Weekly", "SEO"],
        ["Rendering", "Key answers, deal facts and links present in server-rendered HTML", "Per release", "Engineering"],
        ["Robots", "Allow search crawlers needed for discovery; verify OAI-SearchBot/Bing/Google policies intentionally", "Quarterly", "Engineering/SEO"],
        ["Structured data", "Validate Organization, Breadcrumb, Article, Dataset, ItemList as actually eligible; no unsupported rating data", "Per template", "Engineering"],
        ["CWV", "Track LCP, INP, CLS by template; reduce JS and image weight", "Monthly", "Engineering"],
        ["Facets", "No indexable duplicate filter/sort URLs; canonical and parameter rules tested", "Per release", "Engineering/SEO"],
        ["Freshness", "Deal feed failures alert; stale commercial pages fall back safely or show verification status", "Daily", "Product"],
        ["404/redirects", "Monitor broken inbound/internal URLs and use relevant one-hop redirects", "Monthly", "SEO"],
        ["Internal links", "Pillars link to every support page; support pages return to pillar and relevant money page", "Per publish", "Content"],
        ["Measurement", "Track postcode submit, speed test complete, filter use, outbound provider click and confirmed sale", "Per release", "Analytics"],
    ]
    add_sheet(wb, "Technical SEO", ["Area", "Requirement", "Cadence", "Owner"], technical)

    links = [
        ["/speed-test", "/guides/broadband-speeds-explained", "Interpret your result", "Result summary"],
        ["/speed-test", "/compare", "Compare faster deals available to you", "After low/average result"],
        ["/deals", "/guides/how-to-switch-broadband-uk", "How switching works", "Pre-table explainer"],
        ["/deals", "/guides/cheapest-broadband-uk", "See the cheapest true-cost options", "Value segment"],
        ["/guides/best-broadband-providers-uk", "/how-we-review-broadband", "How we score providers", "Above ranking"],
        ["/guides/best-broadband-providers-uk", "/compare", "Compare live deals at your address", "Decision CTA"],
        ["/guides/how-to-switch-broadband-uk", "/deals", "Compare deals before you switch", "After checklist"],
        ["/postcode", "/compare", "Check deals by postcode", "Primary CTA"],
        ["/postcode/[area]", "/postcode", "Broadband in your area", "Breadcrumb and related-area module"],
        ["/guides/best-business-broadband-providers-uk", "/guides/broadband-speeds-explained", "Choose the right speed and upload capacity", "Requirements section"],
        ["/research/uk-broadband-customer-satisfaction", "/guides/best-broadband-providers-uk", "See our provider recommendations", "After results"],
        ["/providers/[provider]", "/providers/[provider]/deals", "See current verified deals", "Commercial module"],
        ["/providers/[provider]/deals", "/how-we-make-money", "How we make money", "Disclosure"],
    ]
    add_sheet(wb, "Internal Links", ["From", "To", "Suggested anchor", "Placement"], links)

    audit_rows = [[item.get(key, "") for key in [
        "url", "status", "title", "title_len", "description", "description_len", "h1_count",
        "canonical", "json_ld", "words", "affiliate_disclosure", "last_verified", "notes"
    ]] for item in audit]
    add_sheet(wb, "Live Site Audit",
              ["URL", "Status", "Title", "Title length", "Meta description", "Description length", "H1 count",
               "Canonical", "JSON-LD blocks", "Approx. words", "Disclosure text", "Freshness text", "Notes"],
              audit_rows or [["No live crawl data", "", "", "", "", "", "", "", "", "", "", "", "Run without --offline"]])

    diagnosis_headers = [
        "URL", "Indexable", "Robots", "Canonical matches", "SEO score /100",
        "GEO readiness /100", "Diagnosis", "Title length", "Meta length", "H1s", "H2s",
        "Schema types", "Internal links", "External links", "Images", "Missing alt",
        "HTML lang", "Viewport", "Open Graph complete", "Author/reviewer", "Primary UK sources",
        "Question headings", "Concise answer passages", "Words", "Freshness visible", "Priority actions",
    ]
    diagnosis_rows = [[item.get(key, "") for key in [
        "url", "indexable", "robots", "canonical_match", "seo_score", "geo_score",
        "diagnosis", "title_len", "description_len", "h1_count", "h2_count", "schema_types",
        "internal_links", "external_links", "images", "images_missing_alt", "html_lang",
        "viewport", "og_complete", "author_signal", "visible_sources", "question_headings",
        "concise_answers", "words", "last_verified", "priority_actions",
    ]] for item in audit]
    add_sheet(
        wb, "Page SEO GEO Diagnosis", diagnosis_headers,
        diagnosis_rows or [["No live crawl data"] + [""] * (len(diagnosis_headers) - 2) + ["Run without --offline"]],
    )

    created_page_rows = []
    registered_paths = set(PAGE_RESEARCH_SOURCES)
    for item in audit:
        path = urllib.parse.urlparse(item["url"]).path or "/"
        created_page_rows.append([
            item["url"], path, route_family(path), item["status"], item["title"],
            item["canonical"], item["words"], "Yes" if item["last_verified"] else "No",
            "Yes" if path in registered_paths else "No", RUN_DATE,
            "Refresh registered sources" if path in registered_paths else "Register sources if claims or prices can change",
        ])
    add_sheet(wb, "Created Pages",
              ["URL", "Path", "Route family", "HTTP status", "Title", "Canonical",
               "Approx. words", "Freshness visible", "Research sources registered",
               "Last crawled", "Next maintenance action"],
              created_page_rows or [["No live crawl data", "", "", "", "", "", "", "", "", RUN_DATE,
                                     "Run without --offline"]])

    source_rows = [list(row) for row in SOURCES]
    add_sheet(wb, "Sources", ["Source", "Used for", "Date", "URL", "Caveat"], source_rows)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


SYNC_SHEET_KEYS = {
    "Read Me": ("Field",),
    "Build Status": ("Workstream", "Deliverable"),
    "Page Research Evidence": ("Target page", "URL"),
    "Landing Pages": ("Recommended URL",),
    "Current URL Hierarchy": ("Live route family",),
    "Navigation Architecture": ("Page URL",),
    "Internal Links": ("From", "To"),
    "Live Site Audit": ("URL",),
    "Page SEO GEO Diagnosis": ("URL",),
    "Created Pages": ("URL",),
    "Sources": ("URL",),
}


def google_access_token() -> str:
    """Create a Sheets API token from a service-account secret or ADC file."""
    try:
        from google.oauth2 import service_account
        from google.auth.transport.requests import Request
    except ImportError as exc:
        raise RuntimeError(
            "Google sync requires: pip install -r requirements-seo.txt"
        ) from exc

    raw_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    credentials_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    if raw_json:
        credentials = service_account.Credentials.from_service_account_info(
            json.loads(raw_json), scopes=scopes
        )
    elif credentials_path:
        credentials = service_account.Credentials.from_service_account_file(
            credentials_path, scopes=scopes
        )
    else:
        raise RuntimeError(
            "Set GOOGLE_SERVICE_ACCOUNT_JSON or GOOGLE_APPLICATION_CREDENTIALS, "
            "then share the Google Sheet with that service-account email as Editor."
        )
    credentials.refresh(Request())
    return credentials.token


def sheets_request(token: str, method: str, url: str, **kwargs: Any) -> dict[str, Any]:
    headers = kwargs.pop("headers", {})
    headers.update({"Authorization": f"Bearer {token}", "Content-Type": "application/json"})
    response = requests.request(method, url, headers=headers, timeout=60, **kwargs)
    if not response.ok:
        raise RuntimeError(f"Google Sheets API {response.status_code}: {response.text[:800]}")
    return response.json() if response.content else {}


def a1_column(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def normalise_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def row_key(headers: list[Any], row: list[Any], key_headers: tuple[str, ...]) -> str:
    positions = {str(header): index for index, header in enumerate(headers)}
    return "\u241f".join(
        str(row[positions[name]] if positions[name] < len(row) else "").strip().lower()
        for name in key_headers
    )


def sync_workbook_to_google_sheet(output: Path, spreadsheet_id: str) -> dict[str, Any]:
    """Upsert generated page data while preserving manual columns and formatting."""
    token = google_access_token()
    base = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}"
    workbook = load_workbook(output, read_only=True, data_only=True)
    metadata = sheets_request(
        token,
        "GET",
        base + "?fields=sheets(properties(sheetId,title),tables(tableId,range))",
    )
    sheet_meta = {
        sheet["properties"]["title"]: sheet for sheet in metadata.get("sheets", [])
    }
    created_sheets = {
        name for name in SYNC_SHEET_KEYS
        if name in workbook.sheetnames and name not in sheet_meta
    }
    if created_sheets:
        sheets_request(
            token,
            "POST",
            base + ":batchUpdate",
            json={"requests": [
                {"addSheet": {"properties": {
                    "title": name,
                    "gridProperties": {"frozenRowCount": 1},
                }}}
                for name in sorted(created_sheets)
            ]},
        )
        metadata = sheets_request(
            token,
            "GET",
            base + "?fields=sheets(properties(sheetId,title),tables(tableId,range))",
        )
        sheet_meta = {
            sheet["properties"]["title"]: sheet for sheet in metadata.get("sheets", [])
        }
    value_updates: list[dict[str, Any]] = []
    appended: dict[str, tuple[int, int, int]] = {}
    results = {"updated": 0, "appended": 0, "sheets": 0}

    for sheet_name, key_headers in SYNC_SHEET_KEYS.items():
        if sheet_name not in workbook.sheetnames or sheet_name not in sheet_meta:
            continue
        worksheet = workbook[sheet_name]
        local_rows = [
            [normalise_cell(value) for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        if not local_rows:
            continue
        headers, data_rows = local_rows[0], local_rows[1:]
        quoted = urllib.parse.quote(f"'{sheet_name}'!A1:ZZ", safe="")
        remote = sheets_request(token, "GET", f"{base}/values/{quoted}")
        remote_rows = remote.get("values", [])
        remote_headers = remote_rows[0] if remote_rows else headers
        if any(key not in remote_headers for key in key_headers):
            raise RuntimeError(f"{sheet_name}: missing key columns {key_headers}")
        remote_map = {
            row_key(remote_headers, row, key_headers): index
            for index, row in enumerate(remote_rows[1:], start=2)
            if row_key(remote_headers, row, key_headers).strip("\u241f")
        }
        width = len(headers)
        last_column = a1_column(width)
        value_updates.append({"range": f"'{sheet_name}'!A1:{last_column}1", "values": [headers]})
        new_rows: list[list[Any]] = []
        for row in data_rows:
            key = row_key(headers, row, key_headers)
            if key in remote_map:
                row_number = remote_map[key]
                value_updates.append({
                    "range": f"'{sheet_name}'!A{row_number}:{last_column}{row_number}",
                    "values": [row[:width]],
                })
                results["updated"] += 1
            else:
                new_rows.append(row[:width])
        if new_rows:
            start_row = max(2, len(remote_rows) + 1)
            append_range = urllib.parse.quote(f"'{sheet_name}'!A:{last_column}", safe="")
            endpoint = (
                f"{base}/values/{append_range}:append"
                "?valueInputOption=RAW&insertDataOption=INSERT_ROWS"
            )
            sheets_request(token, "POST", endpoint, json={"values": new_rows})
            appended[sheet_name] = (start_row, start_row + len(new_rows), width)
            results["appended"] += len(new_rows)
        results["sheets"] += 1

    if value_updates:
        sheets_request(
            token,
            "POST",
            base + "/values:batchUpdate",
            json={"valueInputOption": "RAW", "data": value_updates},
        )

    format_requests: list[dict[str, Any]] = []
    for sheet_name, (start_row, end_row, width) in appended.items():
        meta = sheet_meta[sheet_name]
        sheet_id = meta["properties"]["sheetId"]
        if sheet_name in created_sheets:
            format_requests.extend([
                {"repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1,
                              "startColumnIndex": 0, "endColumnIndex": width},
                    "cell": {"userEnteredFormat": {
                        "backgroundColor": {"red": 0.0706, "green": 0.1882, "blue": 0.2902},
                        "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                       "bold": True},
                        "verticalAlignment": "MIDDLE", "wrapStrategy": "WRAP",
                    }},
                    "fields": "userEnteredFormat",
                }},
                {"repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": end_row - 1,
                              "startColumnIndex": 0, "endColumnIndex": width},
                    "cell": {"userEnteredFormat": {
                        "verticalAlignment": "TOP", "wrapStrategy": "WRAP",
                    }},
                    "fields": "userEnteredFormat.verticalAlignment,userEnteredFormat.wrapStrategy",
                }},
                {"autoResizeDimensions": {"dimensions": {
                    "sheetId": sheet_id, "dimension": "COLUMNS",
                    "startIndex": 0, "endIndex": width,
                }}},
            ])
        else:
            format_requests.append({
                "copyPaste": {
                    "source": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2,
                               "startColumnIndex": 0, "endColumnIndex": width},
                    "destination": {"sheetId": sheet_id, "startRowIndex": start_row - 1,
                                    "endRowIndex": end_row - 1, "startColumnIndex": 0,
                                    "endColumnIndex": width},
                    "pasteType": "PASTE_FORMAT",
                }
            })
        tables = meta.get("tables", [])
        if tables:
            table = tables[0]
            table_range = dict(table["range"])
            table_range["endRowIndex"] = max(table_range.get("endRowIndex", 1), end_row - 1)
            table_range["endColumnIndex"] = max(table_range.get("endColumnIndex", width), width)
            format_requests.append({
                "updateTable": {
                    "table": {"tableId": table["tableId"], "range": table_range},
                    "fields": "range",
                }
            })
    if "Automation Log" not in sheet_meta:
        format_requests.append({
            "addSheet": {
                "properties": {
                    "title": "Automation Log",
                    "gridProperties": {"frozenRowCount": 1},
                }
            }
        })
    if format_requests:
        sheets_request(token, "POST", base + ":batchUpdate", json={"requests": format_requests})

    log_values = [["Run timestamp", "Status", "Updated rows", "Appended rows", "Source"],
                  [datetime.now(timezone.utc).isoformat(), "Success", results["updated"],
                   results["appended"], "build_uk_broadband_seo_plan.py"]]
    log_range = urllib.parse.quote("'Automation Log'!A:E", safe="")
    existing_log = sheets_request(token, "GET", f"{base}/values/{log_range}").get("values", [])
    payload = log_values if not existing_log else [log_values[1]]
    sheets_request(
        token,
        "POST",
        f"{base}/values/{log_range}:append?valueInputOption=RAW&insertDataOption=INSERT_ROWS",
        json={"values": payload},
    )
    workbook.close()
    return results


def validate(output: Path) -> None:
    wb = load_workbook(output, read_only=True, data_only=False)
    required = {"Read Me", "Build Status", "Executive Plan", "Keyword Research", "Landing Pages",
                "Page Research Evidence", "Page SEO GEO Diagnosis", "Created Pages", "Current URL Hierarchy", "Navigation Architecture", "90 Day Roadmap",
                "GEO Playbook", "Awin Readiness", "Sources"}
    missing = required.difference(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Workbook missing sheets: {sorted(missing)}")
    if wb["Keyword Research"].max_row < 20 or wb["Landing Pages"].max_row < 10:
        raise RuntimeError("Workbook contains too few research or landing-page rows")
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--offline", action="store_true", help="Skip live sitemap/page/autocomplete requests")
    parser.add_argument("--max-pages", type=int, default=140, help="Maximum sitemap URLs to audit")
    parser.add_argument(
        "--sync-google-sheet",
        action="store_true",
        help="Upsert generated page data into the configured native Google Sheet",
    )
    parser.add_argument(
        "--google-sheet-id",
        default=os.environ.get("BROADBANDPICKER_GOOGLE_SHEET_ID", DEFAULT_GOOGLE_SHEET_ID),
        help="Native Google spreadsheet ID used with --sync-google-sheet",
    )
    args = parser.parse_args()

    urls = get_sitemap_urls(args.offline)
    print(f"Found {len(urls)} live sitemap URLs")
    audit = []
    for index, url in enumerate(urls[:args.max_pages], 1):
        print(f"Auditing {index}/{min(len(urls), args.max_pages)}: {url}")
        audit.append(audit_page(url))
        time.sleep(0.08)
    autocomplete = collect_autocomplete(args.offline)
    print(f"Collected {len(autocomplete)} unique autocomplete suggestions")
    page_research = collect_page_research(args.offline)
    print(f"Collected {len(page_research)} page-specific source snapshots")
    build_workbook(args.output.resolve(), urls, audit, autocomplete, page_research)
    validate(args.output.resolve())
    sync_result = None
    if args.sync_google_sheet:
        sync_result = sync_workbook_to_google_sheet(args.output.resolve(), args.google_sheet_id)
        print(f"Google Sheet sync complete: {sync_result}")
    print(json.dumps({
        "output": str(args.output.resolve()),
        "sheets": load_workbook(args.output.resolve(), read_only=True).sheetnames,
        "site_urls": len(urls), "audited_pages": len(audit),
        "autocomplete_suggestions": len(autocomplete), "keyword_rows": len(KEYWORDS),
        "landing_page_rows": len(LANDING_PAGES), "page_research_sources": len(page_research),
        "google_sheet_sync": sync_result,
    }, indent=2))


if __name__ == "__main__":
    main()
