#!/usr/bin/env python3
"""Detailed SEO + GEO (generative-engine) analysis of every live page, used to
prioritise what to build next (new Content Gap Roadmap pages, already scored
by build_keyword_mapping.py) against what to refresh (existing pages that are
thin, stale, or missing the structural signals that correlate with AI
Overview / chat-answer citation), on one combined, ranked list.

This does not guess at rankings or AI citation. It measures what is
independently checkable: live word count, FAQPage schema presence, a
detectable answer-first paragraph, a checked/reviewed date parsed from
visible text, and the real mapped keyword volume already researched in
docs/broadbandpicker-keyword-mapping.xlsx. Those signals are combined into a
transparent, reproducible score — not a promise of ranking or citation.

Usage:
    python3 scripts/analyze_content_priority.py
    python3 scripts/analyze_content_priority.py --top 15
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests
from lxml import html as lxml_html

ROOT = Path(__file__).resolve().parents[1]
KEYWORD_WORKBOOK = ROOT / "docs" / "broadbandpicker-keyword-mapping.xlsx"
GSC_DIR = ROOT / "data" / "GSC"
OUT = ROOT / "docs" / "content-priority-analysis.json"
BASE = "https://broadbandpicker.co.uk"

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; BroadbandPickerContentAudit/1.0)"}

# Word-count floor below which a page is treated as thin for its type.
# These thresholds come from this session's own established practice: the
# deep provider rewrites (BT, TalkTalk, Zzoomm, etc.) all landed at
# roughly 1,800-2,600 rendered words; comparison pages are naturally
# shorter given their fixed slot structure.
THIN_THRESHOLD = {
    "provider": 1400,
    "comparison": 900,
    "guide": 1200,
    "other": 600,
}

STALE_MONTHS = 6

DATE_RE = re.compile(
    r"(?:Reviewed|Updated|verified|checked(?: against official UK sources)?(?: on)?)\s*"
    r"(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}"
    r"|\d{4}-\d{2}-\d{2})",
    re.IGNORECASE,
)
MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}


@dataclass
class PageAudit:
    path: str
    page_type: str
    status: int = 0
    word_count: int = 0
    has_faq_schema: bool = False
    has_answer_first_paragraph: bool = False
    checked_date: str | None = None
    months_since_checked: float | None = None
    mapped_volume: int = 0
    ai_feature_impressions: int = 0
    thin: bool = False
    stale: bool = False
    missing_geo_signals: list[str] = None
    refresh_score: float = 0.0
    notes: str = ""


def classify(path: str) -> str:
    if path.startswith("/providers/compare/"):
        return "comparison"
    if path.startswith("/providers/") and path.count("/") == 2:
        return "provider"
    if path.startswith("/guides/") and path.count("/") == 2:
        return "guide"
    return "other"


def parse_checked_date(text: str) -> datetime | None:
    match = DATE_RE.search(text)
    if not match:
        return None
    raw = match.group(1)
    try:
        if re.match(r"\d{4}-\d{2}-\d{2}", raw):
            return datetime.strptime(raw, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        day, month_name, year = raw.split()
        month = MONTHS[month_name.lower()]
        return datetime(int(year), month, int(day), tzinfo=timezone.utc)
    except Exception:
        return None


def audit_page(path: str, page_type: str) -> PageAudit:
    audit = PageAudit(path=path, page_type=page_type, missing_geo_signals=[])
    try:
        resp = requests.get(f"{BASE}{path}", headers=HEADERS, timeout=20)
    except requests.RequestException as exc:
        audit.notes = f"Request failed: {exc}"
        return audit

    audit.status = resp.status_code
    if resp.status_code != 200:
        audit.notes = f"Non-200 response ({resp.status_code})"
        return audit

    html_text = resp.text
    # The RSC hydration payload (inline <script> blocks) can be over half the
    # raw HTML and duplicates the page's content as serialized JSON text, so
    # every extraction below must walk the parsed DOM and explicitly exclude
    # script/style content rather than regex the raw response body.
    audit.has_faq_schema = '"@type":"FAQPage"' in html_text or '"@type": "FAQPage"' in html_text

    try:
        tree = lxml_html.fromstring(html_text)
    except Exception as exc:
        audit.notes = f"Parse failed: {exc}"
        return audit

    for bad in tree.xpath("//script | //style"):
        bad.getparent().remove(bad)

    # Scope to <main> where possible so header/nav <p> fragments (mega-menu
    # section labels, footer links) never get counted as page content or
    # mistaken for the answer-first paragraph.
    main = tree.xpath("//main")
    scope = main[0] if main else tree

    visible = " ".join(scope.text_content().split())
    audit.word_count = len(re.findall(r"\b[\w'-]+\b", visible))

    for p in scope.xpath(".//p")[:8]:
        p_text = p.text_content()
        p_words = len(re.findall(r"\b[\w'-]+\b", p_text))
        if 30 <= p_words <= 100:
            audit.has_answer_first_paragraph = True
            break

    checked = parse_checked_date(visible)
    if checked:
        audit.checked_date = checked.date().isoformat()
        now = datetime.now(timezone.utc)
        audit.months_since_checked = round((now - checked).days / 30.44, 1)

    threshold = THIN_THRESHOLD.get(page_type, THIN_THRESHOLD["other"])
    audit.thin = audit.word_count < threshold
    audit.stale = audit.months_since_checked is not None and audit.months_since_checked > STALE_MONTHS

    if audit.thin:
        audit.missing_geo_signals.append(f"below {threshold}-word depth floor for a {page_type} page")
    if not audit.has_faq_schema and page_type in ("provider", "guide"):
        audit.missing_geo_signals.append("no FAQPage schema detected")
    if not audit.has_answer_first_paragraph:
        audit.missing_geo_signals.append("no detectable 30-100 word answer-first paragraph near the top")
    if audit.checked_date is None:
        audit.missing_geo_signals.append("no parseable reviewed/checked date in visible text")
    elif audit.stale:
        audit.missing_geo_signals.append(f"last checked {audit.months_since_checked} months ago")

    return audit


def load_mapped_volume() -> dict[str, int]:
    """Sum estimated monthly UK search volume per mapped URL from the
    already-researched keyword workbook, rather than re-deriving it."""
    if not KEYWORD_WORKBOOK.exists():
        return {}
    wb = openpyxl.load_workbook(KEYWORD_WORKBOOK, data_only=True)
    ws = wb["Keyword Mapping"]
    headers = [c.value for c in ws[1]]
    vol_idx = headers.index("Est. Monthly UK Volume")
    url_idx = headers.index("Mapped / Recommended URL")
    volumes: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[url_idx]
        vol = row[vol_idx]
        if not url or not isinstance(vol, (int, float)):
            continue
        path = url.replace(BASE, "") or "/"
        volumes[path] = volumes.get(path, 0) + int(vol)
    return volumes


def load_ai_feature_impressions() -> dict[str, int]:
    """Real ground truth, not a proxy: per-page impressions from Google
    Search Console's 'Performance on Search Generative AI Features' export
    (a Pages.csv dropped into data/GSC/<export-folder>/). This is actual
    evidence a page has already been surfaced inside an AI Overview or
    similar generative search feature, which no heuristic on this site can
    independently verify -- if this export isn't present, GEO scoring falls
    back to the depth/schema/date proxy signals alone."""
    impressions: dict[str, int] = {}
    if not GSC_DIR.exists():
        return impressions
    for pages_csv in GSC_DIR.glob("*/Pages.csv"):
        with pages_csv.open(newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                url = row.get("Top pages") or row.get("Page") or ""
                count = row.get("Impressions")
                if not url or count is None:
                    continue
                path = url.replace(BASE, "") or "/"
                try:
                    impressions[path] = impressions.get(path, 0) + int(count)
                except ValueError:
                    continue
    return impressions


def load_top_gaps(limit: int = 10) -> list[dict]:
    """Surface the top N-not-yet-built pages already ranked in the Content
    Gap Roadmap, so the new-vs-refresh decision sits on one list."""
    if not KEYWORD_WORKBOOK.exists():
        return []
    wb = openpyxl.load_workbook(KEYWORD_WORKBOOK, data_only=True)
    if "Content Gap Roadmap" not in wb.sheetnames:
        return []
    ws = wb["Content Gap Roadmap"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = row[idx["Current Status"]]
        if status and str(status).startswith("Built"):
            continue
        rows.append({
            "title": row[idx["Page Title"]],
            "url": row[idx["Recommended URL"]],
            "cluster": row[idx["Topic Cluster"]],
            "volume": row[idx["Combined Est. Volume"]],
            "seo_priority_score": row[idx["SEO Build Priority Score"]],
            "revenue_priority_score": row[idx["Revenue Priority Score"]],
        })
    rows.sort(key=lambda r: (r["seo_priority_score"] or 0) + (r["revenue_priority_score"] or 0), reverse=True)
    return rows[:limit]


def get_live_paths() -> list[tuple[str, str]]:
    resp = requests.get(f"{BASE}/sitemap.xml", headers=HEADERS, timeout=30)
    resp.raise_for_status()
    locs = re.findall(r"<loc>(.*?)</loc>", resp.text)
    paths = []
    for loc in locs:
        path = loc.replace(BASE, "") or "/"
        page_type = classify(path)
        if page_type in ("provider", "comparison", "guide"):
            paths.append((path, page_type))
    return paths


def compute_refresh_score(audit: PageAudit) -> float:
    if audit.status != 200:
        return 0.0
    volume_weight = min(audit.mapped_volume / 100, 40)  # cap so one huge keyword doesn't dominate
    # Real GSC evidence the page already earns AI-feature visibility outweighs
    # everything else here: a thin page already being surfaced is the single
    # clearest sign that deepening it converts directly into more AI-feature
    # share, not a hoped-for outcome.
    ai_feature_weight = min(audit.ai_feature_impressions / 5, 60)
    thin_weight = 25 if audit.thin else 0
    stale_weight = 15 if audit.stale else 0
    missing_signal_weight = 5 * len([s for s in audit.missing_geo_signals if "depth floor" not in s and "months ago" not in s])
    return round(volume_weight + ai_feature_weight + thin_weight + stale_weight + missing_signal_weight, 1)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--top", type=int, default=15, help="How many refresh candidates and new-build gaps to report")
    args = parser.parse_args()

    print("Loading mapped keyword volume from the keyword-mapping workbook...")
    volumes = load_mapped_volume()

    print("Loading real Search Generative AI Features impressions from Search Console...")
    ai_impressions = load_ai_feature_impressions()
    if ai_impressions:
        print(f"  Found {len(ai_impressions)} pages with real AI-feature impressions in data/GSC/")
    else:
        print("  No data/GSC export found — GEO scoring falls back to depth/schema/date proxy signals only")

    print("Fetching live sitemap...")
    paths = get_live_paths()
    print(f"Auditing {len(paths)} provider/comparison/guide pages...")

    audits: list[PageAudit] = []
    for i, (path, page_type) in enumerate(paths, 1):
        audit = audit_page(path, page_type)
        audit.mapped_volume = volumes.get(path, 0)
        audit.ai_feature_impressions = ai_impressions.get(path, 0)
        audit.refresh_score = compute_refresh_score(audit)
        audits.append(audit)
        if i % 25 == 0:
            print(f"  ...{i}/{len(paths)}")

    audits.sort(key=lambda a: a.refresh_score, reverse=True)
    top_refresh = [a for a in audits if a.refresh_score > 0][: args.top]

    top_new = load_top_gaps(args.top)

    payload = {
        "generated": datetime.now(timezone.utc).date().isoformat(),
        "method": (
            "Live crawl of every provider/comparison/guide URL in the current sitemap, scored on "
            "word count against this session's own established depth floors, FAQPage schema "
            "presence, a detectable answer-first paragraph, and a checked/reviewed date parsed "
            "from visible text, cross-referenced against real mapped search volume already "
            "researched in docs/broadbandpicker-keyword-mapping.xlsx AND real per-page impression "
            "counts from Google Search Console's 'Performance on Search Generative AI Features' "
            "export (data/GSC/), when present. The GSC figures are actual ground truth that a page "
            "has already been surfaced inside an AI Overview or similar generative feature, not a "
            "proxy, and are weighted heavily in the score for exactly that reason. Word count, "
            "schema and date signals remain proxies for pages with no GSC history yet."
        ),
        "ai_feature_data_present": bool(ai_impressions),
        "pages_audited": len(audits),
        "top_refresh_candidates": [asdict(a) for a in top_refresh],
        "top_new_build_gaps": top_new,
    }
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print(f"\nSaved {OUT}\n")
    print(f"=== Top {len(top_refresh)} pages to UPDATE (weighted by real GSC AI-feature impressions + search volume + thinness) ===")
    for a in top_refresh:
        reasons = ", ".join(a.missing_geo_signals) or "none"
        print(f"  [{a.refresh_score:5.1f}] {a.path}  (AI-feature impr={a.ai_feature_impressions}, vol={a.mapped_volume}, words={a.word_count})  -- {reasons}")

    print(f"\n=== Top {len(top_new)} pages to BUILD (highest-priority Content Gap Roadmap rows not yet live) ===")
    for g in top_new:
        print(f"  [{(g['seo_priority_score'] or 0):5.1f}] {g['title']}  (vol={g['volume']})")


if __name__ == "__main__":
    main()
