#!/usr/bin/env python3
"""Build a professional UK broadband keyword-to-page mapping for BroadbandPicker.co.uk.

The script:
1. Crawls the live sitemap and lightly audits each current page (title, H1, word count)
   so the mapping is checked against what is actually published, not assumptions.
2. Applies a curated, editorially-assigned UK broadband keyword dataset (volume,
   difficulty, CPC, intent, funnel stage) — a directional SEO-intelligence snapshot,
   not live Search Console or Ads data.
3. Assigns every keyword to either an existing URL or a recommended future page,
   scores each row for affiliate-revenue priority, and groups the gaps into a
   publish-order content roadmap.
4. Writes a multi-tab Excel workbook, and optionally:
   - reads extra keyword ideas out of a source Google Sheet you point it at
     (e.g. the working sheet a client shares) and merges them in, deduplicated; and
   - creates a brand-new Google Sheet on Drive containing the finished mapping.

Usage:
    python3 scripts/build_keyword_mapping.py
    python3 scripts/build_keyword_mapping.py --offline
    python3 scripts/build_keyword_mapping.py \\
        --source-sheet-id 1_IeyaXe1fSbXmV8Z4MdVOigMIz-cgK2h \\
        --create-google-doc --share-with sahil.ieltstraining@gmail.com

Search volumes and CPC are directional third-party-style estimates dated to the run.
They are not Google Search Console or Google Ads data. Re-run monthly and replace
with first-party data as it becomes available.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
import subprocess
import sys
import time
import urllib.parse
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import requests
from lxml import html
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "docs" / "broadbandpicker-keyword-mapping.xlsx"
SITE = "https://broadbandpicker.co.uk"
SITEMAP_URL = f"{SITE}/sitemap.xml"
RUN_DATE = datetime.now(timezone.utc).date().isoformat()
USER_AGENT = (
    "BroadbandPickerKeywordMappingBot/1.0 "
    "(SEO research; contact: https://broadbandpicker.co.uk/contact)"
)
# Extracted from the working sheet URL supplied for this run:
# https://docs.google.com/spreadsheets/d/1_IeyaXe1fSbXmV8Z4MdVOigMIz-cgK2h/edit?gid=1523609630
DEFAULT_SOURCE_SHEET_ID = "1_IeyaXe1fSbXmV8Z4MdVOigMIz-cgK2h"
DEFAULT_MAPPING_SHEET_ID = "1Ke0YWo5T-45JRpuXpfqL_0vwmcBbS0i06Da47quRRH0"
DEFAULT_PRIORITY_BATCH_SIZE = 5
PIPELINE_BRIEF = ROOT / "docs" / "page-build-pipeline-brief.md"
PIPELINE_DIR = ROOT / "docs" / "page-build-pipeline"
PAGE_RESEARCH_FILE = PIPELINE_DIR / "current-page-research.json"
CANONICAL_PRIORITY_FILE = PIPELINE_DIR / "canonical-next-priority.json"
PAGE_TYPE_UX_SCAN_FILE = ROOT / "docs" / "home page UX" / "page-type-ux-scan.json"
WEEKLY_SEO_INTELLIGENCE_FILE = ROOT / "docs" / "weekly-seo-intelligence.json"
POST_PUBLICATION_REVIEW_FILE = PIPELINE_DIR / "post-publication-review-queue.json"

PAGE_ROUTES = {
    "providers/compare/": ("data/provider-comparisons.ts", "app/providers/compare/[slug]/page.tsx", "ProviderComparison"),
    "providers/": ("data/providers.ts", "app/providers/[slug]/page.tsx", "Provider"),
    "guides/": ("data/guides.ts", "app/guides/[slug]/page.tsx", "GuideMetadata"),
    "postcode/": ("bespoke", "app/postcode/london/page.tsx", "bespoke"),
    "research/": ("bespoke", "app/research/uk-broadband-customer-satisfaction", "bespoke"),
    "tools/": ("bespoke", "app/speed-test/page.tsx", "bespoke"),
}


# ---------------------------------------------------------------------------
# 1. Current site inventory — pages that already exist and can absorb a keyword
# ---------------------------------------------------------------------------

# Fallback inventory used with --offline, or if the live sitemap is unreachable.
# Kept in sync with app/sitemap.ts as of 2026-08-14; the live crawl below is the
# source of truth whenever the network is available.
STATIC_SITEMAP_URLS = [
    "/", "/deals", "/compare", "/providers", "/providers/compare",
    "/guides", "/speed-test", "/broadband-glossary", "/about", "/contact",
    "/how-we-make-money", "/how-we-review-broadband", "/editorial-policy",
    "/privacy-policy", "/cookie-policy", "/terms",
    "/postcode", "/postcode/london", "/postcode/bristol",
    "/research/uk-broadband-customer-satisfaction",
    "/guides/best-business-broadband-providers-uk",
    "/guides/best-phone-and-broadband-deals",
    "/guides/satellite-broadband-uk",
] + [
    f"/providers/{slug}" for slug in (
        "bt", "sky", "virgin-media", "ee", "talktalk", "plusnet", "vodafone",
        "now-broadband", "hyperoptic", "community-fibre", "zen-internet", "toob",
    )
] + [
    f"/providers/compare/{slug}" for slug in (
        "bt-vs-sky", "bt-vs-virgin-media", "sky-vs-vodafone", "ee-vs-bt",
        "talktalk-vs-now-broadband", "virgin-media-vs-vodafone", "sky-vs-virgin-media",
        "hyperoptic-vs-community-fibre", "bt-vs-vodafone", "ee-vs-sky", "bt-vs-talktalk",
        "ee-vs-vodafone", "sky-vs-talktalk", "plusnet-vs-bt", "plusnet-vs-sky",
        "vodafone-vs-talktalk",
    )
] + [
    f"/guides/{slug}" for slug in (
        "how-to-switch-broadband-uk", "best-broadband-deals-uk",
        "broadband-deals-with-no-mid-contract-price-rise", "best-broadband-and-tv-deals",
        "broadband-deals-under-20", "broadband-deals-with-cashback",
        "broadband-deals-with-no-setup-fee", "best-rolling-monthly-broadband-deals",
        "full-fibre-broadband-explained", "broadband-speeds-explained",
        "cheapest-broadband-uk", "best-broadband-for-working-from-home",
        "best-broadband-for-students", "best-broadband-for-streaming",
        "best-broadband-providers-uk", "broadband-price-rises-2026",
        "can-i-leave-broadband-early-after-price-rise", "broadband-without-phone-line",
        "best-5g-home-broadband-uk", "best-full-fibre-broadband-uk",
        "best-broadband-for-gaming-uk", "broadband-social-tariffs-uk",
        "broadband-moving-house", "best-broadband-for-rural-areas-uk",
        "broadband-for-existing-customers", "one-touch-switching-explained",
        "broadband-contract-end-rights",
    )
]

# URL -> (topic cluster, page type) for pages the live crawl/static list can't infer.
CLUSTER_BY_URL: dict[str, tuple[str, str]] = {
    "/": ("Core commercial", "Home"),
    "/deals": ("Core commercial", "Deals hub"),
    "/compare": ("Core commercial", "Comparison tool"),
    "/providers": ("Provider brand", "Provider directory"),
    "/providers/compare": ("Provider vs comparison", "Comparison directory"),
    "/guides": ("Editorial hub", "Guides directory"),
    "/speed-test": ("Tools", "Interactive tool"),
    "/broadband-glossary": ("Informational", "Glossary"),
    "/postcode": ("Postcode & location", "Availability hub"),
    "/postcode/london": ("Postcode & location", "City hub"),
    "/postcode/bristol": ("Postcode & location", "City hub"),
    "/research/uk-broadband-customer-satisfaction": ("Trust & research", "Research dashboard"),
    "/guides/best-business-broadband-providers-uk": ("Business broadband", "Guide"),
    "/guides/best-phone-and-broadband-deals": ("Phone & bundles", "Guide"),
    "/guides/satellite-broadband-uk": ("Alternative access", "Guide"),
    "/how-we-make-money": ("Trust & compliance", "Policy"),
    "/how-we-review-broadband": ("Trust & compliance", "Policy"),
    "/editorial-policy": ("Trust & compliance", "Policy"),
}


def fetch(url: str, timeout: int = 20) -> requests.Response:
    response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=timeout)
    response.raise_for_status()
    return response


def get_sitemap_urls(offline: bool) -> list[str]:
    if offline:
        return [SITE + path for path in STATIC_SITEMAP_URLS]
    try:
        response = fetch(SITEMAP_URL)
        root = ET.fromstring(response.content)
        ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
        urls = [el.text.strip() for el in root.findall(".//sm:loc", ns) if el.text]
        if urls:
            return urls
    except Exception as exc:
        print(f"Warning: could not fetch live sitemap ({exc}); using static snapshot")
    return [SITE + path for path in STATIC_SITEMAP_URLS]


def audit_page(url: str) -> dict[str, Any]:
    """Light-touch crawl: title, H1 and word count, used to sanity-check mapping."""
    row = {"url": url, "status": "", "title": "", "h1": "", "words": 0,
           "platform": "", "notes": ""}
    try:
        response = fetch(url)
        tree = html.fromstring(response.content)
        title = " ".join(tree.xpath("//title/text()")).strip()
        h1s = [re.sub(r"\s+", " ", t).strip() for t in tree.xpath("//h1//text()")]
        text_content = " ".join(tree.xpath("//body//text()[normalize-space()]"))
        word_count = len(re.findall(r"\b[\w'-]+\b", text_content))
        row.update({
            "status": response.status_code, "title": title,
            "h1": " ".join(h1s)[:160], "words": word_count,
            "platform": "Vercel" if response.headers.get("x-vercel-id")
            or "vercel" in response.headers.get("server", "").lower() else "Live host",
        })
    except Exception as exc:
        row["notes"] = str(exc)[:200]
    return row


def route_family(path: str) -> tuple[str, str]:
    """Best-guess (cluster, page type) for a URL not explicitly listed above."""
    if path in CLUSTER_BY_URL:
        return CLUSTER_BY_URL[path]
    if path.startswith("/providers/compare/"):
        return "Provider vs comparison", "Comparison page"
    if path.startswith("/providers/"):
        return "Provider brand", "Provider page"
    if path.startswith("/postcode/"):
        return "Postcode & location", "Postcode prefix page"
    if path.startswith("/guides/"):
        return "Editorial guide", "Guide"
    if path.startswith("/research/"):
        return "Trust & research", "Research page"
    return "Other", "Page"


# ---------------------------------------------------------------------------
# 2. Curated UK broadband keyword dataset — the "SEO intelligence" layer
# ---------------------------------------------------------------------------
# Each row is an editorial assignment: keyword, directional UK monthly search
# volume, difficulty (0-100), CPC estimate in GBP, search intent, funnel stage,
# topic cluster, and either the current URL it should be mapped to, or a
# recommended future page (marked GAP:) with a suggested slug and title.
# Volume/difficulty/CPC are a dated (2026-08) directional snapshot in the same
# style as docs/uk-broadband-seo-geo-plan.xlsx — not Search Console/Ads data.

KeywordRow = dict[str, Any]


def kw(
    keyword: str, volume: int, difficulty: int, cpc: float, intent: str,
    funnel: str, cluster: str, page_type: str,
    mapped_url: str | None = None,
    gap_slug: str | None = None, gap_title: str | None = None,
) -> KeywordRow:
    return {
        "keyword": keyword, "volume": volume, "difficulty": difficulty, "cpc": cpc,
        "intent": intent, "funnel": funnel, "cluster": cluster, "page_type": page_type,
        "mapped_url": mapped_url, "gap_slug": gap_slug, "gap_title": gap_title,
    }


def build_keyword_dataset() -> list[KeywordRow]:
    rows: list[KeywordRow] = []

    # --- Core commercial: home, deals, compare ---------------------------------
    rows += [
        kw("broadband deals", 165000, 71, 8.32, "Commercial", "BOFU", "Core commercial", "Deals hub", "/deals"),
        kw("best broadband deals", 40500, 70, 7.02, "Commercial", "BOFU", "Core commercial", "Deals hub", "/deals"),
        kw("cheap broadband deals", 18100, 45, 4.58, "Commercial", "BOFU", "Core commercial", "Deals hub", "/deals"),
        kw("uk broadband deals", 14800, 71, 6.93, "Commercial", "BOFU", "Core commercial", "Deals hub", "/deals"),
        kw("compare broadband deals", 9900, 62, 6.40, "Commercial", "BOFU", "Core commercial", "Comparison tool", "/compare"),
        kw("compare broadband providers", 6600, 58, 6.10, "Commercial", "BOFU", "Core commercial", "Comparison tool", "/compare"),
        kw("broadband comparison uk", 4400, 55, 5.90, "Commercial", "BOFU", "Core commercial", "Comparison tool", "/compare"),
        kw("compare broadband providers uk", 2900, 54, 5.75, "Commercial", "BOFU", "Core commercial", "Comparison tool", "/compare"),
        kw("compare broadband prices", 2400, 57, 6.20, "Commercial", "BOFU", "Core commercial", "Comparison tool", "/compare"),
        kw("broadband provider comparison", 1900, 51, 5.40, "Commercial", "MOFU", "Core commercial", "Comparison tool", "/compare"),
        kw("best broadband provider for me", 880, 42, 4.60, "Commercial investigation", "MOFU", "Core commercial", "Comparison tool", "/compare"),
        kw("broadband speed comparison", 720, 43, 3.80, "Commercial investigation", "MOFU", "Core commercial", "Comparison tool", "/compare"),
        kw("which broadband provider should i choose", 590, 39, 4.20, "Commercial investigation", "MOFU", "Core commercial", "Comparison tool", "/compare"),
        kw("broadband providers", 6600, 69, 11.05, "Commercial", "MOFU", "Core commercial", "Provider directory", "/providers"),
        kw("best broadband providers uk", 5400, 69, 6.07, "Commercial", "MOFU", "Core commercial", "Guide", "/guides/best-broadband-providers-uk"),
        kw("broadband providers for my area", 5400, 61, 7.31, "Local commercial", "MOFU", "Postcode & location", "Availability hub", "/postcode"),
        kw("broadband", 49500, 66, 11.38, "Mixed", "TOFU", "Core commercial", "Home", "/"),
        kw("switch broadband provider", 3600, 38, 13.07, "Commercial guide", "MOFU", "Core commercial", "Guide", "/guides/how-to-switch-broadband-uk"),
        kw("cheapest broadband uk", 1300, 67, 6.89, "Commercial", "BOFU", "Core commercial", "Guide", "/guides/cheapest-broadband-uk"),
        kw("broadband without landline", 2900, 40, 3.10, "Commercial guide", "MOFU", "Core commercial", "Guide", "/guides/broadband-without-phone-line"),
    ]

    # --- Provider brand pages (all 13 current providers) ------------------------
    provider_terms = [
        ("bt", "BT", 110000, 69, 2.14, 33100, 43, 2.12),
        ("sky", "Sky", 135000, 70, 1.75, 33100, 45, 1.96),
        ("virgin-media", "Virgin Media", 49500, 65, 2.51, 12100, 67, 2.67),
        ("ee", "EE", 165000, 64, 1.44, 33100, 29, 1.77),
        ("talktalk", "TalkTalk", 40500, 40, 2.41, 8100, 44, 2.35),
        ("plusnet", "Plusnet", 27100, 46, 1.92, 6600, 41, 1.85),
        ("vodafone", "Vodafone", 90500, 37, 1.94, 8100, 33, 2.05),
        ("now-broadband", "NOW Broadband", 6600, 32, 1.60, 2400, 28, 1.55),
        ("hyperoptic", "Hyperoptic", 8100, 35, 1.70, 1900, 30, 1.65),
        ("community-fibre", "Community Fibre", 4400, 28, 1.40, 1000, 24, 1.35),
        ("zen-internet", "Zen Internet", 5400, 30, 1.55, 1300, 26, 1.50),
        ("toob", "Toob", 1900, 22, 1.20, 590, 20, 1.15),
    ]
    for slug, name, brand_vol, brand_diff, brand_cpc, deal_vol, deal_diff, deal_cpc in provider_terms:
        rows.append(kw(f"{name.lower()} broadband", brand_vol, brand_diff, brand_cpc,
                        "Brand navigation", "MOFU", "Provider brand", "Provider page", f"/providers/{slug}"))
        rows.append(kw(f"{name.lower()} broadband deals", deal_vol, deal_diff, deal_cpc,
                        "Brand commercial", "BOFU", "Provider brand", "Provider page", f"/providers/{slug}"))
        rows.append(kw(f"{name.lower()} broadband reviews", max(320, deal_vol // 8), max(20, brand_diff - 10), round(brand_cpc * 0.7, 2),
                        "Research", "MOFU", "Provider brand", "Provider page", f"/providers/{slug}"))

    # Provider brands not yet built as standalone pages — content gaps.
    for slug, name, vol, diff, cpc in [
        ("shell-energy", "Shell Energy Broadband", 3600, 27, 1.35),
        ("onestream", "Onestream", 880, 18, 0.95),
        ("youfibre", "YouFibre", 4400, 26, 1.30),
        ("brsk", "Brsk", 1900, 20, 1.05),
        ("gigaclear", "Gigaclear", 2400, 24, 1.20),
        ("cuckoo", "Cuckoo Broadband", 2900, 25, 1.15),
    ]:
        rows.append(kw(f"{name.lower()} broadband", vol, diff, cpc, "Brand navigation", "MOFU",
                        "Provider brand", "Provider page", gap_slug=f"providers/{slug}",
                        gap_title=f"{name} — plans, coverage and Awin-tracked deals"))

    # --- Provider vs comparison pages -------------------------------------------
    vs_terms = [
        ("bt-vs-sky", "BT vs Sky broadband", 1900, 34, 3.20),
        ("bt-vs-virgin-media", "BT vs Virgin Media broadband", 1600, 33, 3.05),
        ("sky-vs-vodafone", "Sky vs Vodafone broadband", 480, 22, 2.40),
        ("ee-vs-bt", "EE vs BT broadband", 720, 24, 2.60),
        ("talktalk-vs-now-broadband", "TalkTalk vs NOW broadband", 320, 18, 1.90),
        ("virgin-media-vs-vodafone", "Virgin Media vs Vodafone broadband", 260, 17, 2.10),
        ("sky-vs-virgin-media", "Sky vs Virgin Media broadband", 1300, 31, 2.95),
        ("hyperoptic-vs-community-fibre", "Hyperoptic vs Community Fibre", 210, 15, 1.60),
        ("bt-vs-vodafone", "BT vs Vodafone broadband", 590, 23, 2.35),
        ("ee-vs-sky", "EE vs Sky broadband", 480, 22, 2.30),
        ("bt-vs-talktalk", "BT vs TalkTalk broadband", 880, 26, 2.55),
        ("ee-vs-vodafone", "EE vs Vodafone broadband", 320, 19, 2.00),
        ("sky-vs-talktalk", "Sky vs TalkTalk broadband", 590, 23, 2.25),
        ("plusnet-vs-bt", "Plusnet vs BT broadband", 480, 21, 2.05),
        ("plusnet-vs-sky", "Plusnet vs Sky broadband", 260, 18, 1.85),
        ("vodafone-vs-talktalk", "Vodafone vs TalkTalk broadband", 170, 16, 1.75),
    ]
    for slug, label, vol, diff, cpc in vs_terms:
        rows.append(kw(label.lower(), vol, diff, cpc, "Commercial comparison", "BOFU",
                        "Provider vs comparison", "Comparison page", f"/providers/compare/{slug}"))

    for slug, label, vol, diff, cpc in [
        ("virgin-media-vs-sky", "Virgin Media vs Sky broadband", 880, 25, 2.60),
        ("ee-vs-talktalk", "EE vs TalkTalk broadband", 390, 20, 2.10),
        ("bt-vs-plusnet", "BT vs Plusnet broadband", 320, 19, 1.95),
        ("hyperoptic-vs-youfibre", "Hyperoptic vs YouFibre broadband", 140, 14, 1.45),
    ]:
        rows.append(kw(label.lower(), vol, diff, cpc, "Commercial comparison", "BOFU",
                        "Provider vs comparison", "Comparison page", gap_slug=f"providers/compare/{slug}",
                        gap_title=f"{label}: which is better in 2026?"))

    # --- Postcode & location ------------------------------------------------
    rows += [
        kw("broadband in my area", 8100, 55, 5.40, "Local commercial", "MOFU", "Postcode & location", "Availability hub", "/postcode"),
        kw("broadband postcode checker", 3600, 40, 4.10, "Local commercial", "MOFU", "Postcode & location", "Availability hub", "/postcode"),
        kw("broadband providers london", 2400, 28, 6.935, "Local commercial", "MOFU", "Postcode & location", "City hub", "/postcode/london"),
        kw("best broadband london", 1300, 27, 5.80, "Local commercial", "MOFU", "Postcode & location", "City hub", "/postcode/london"),
    ]
    for city, vol, diff, cpc in [
        ("manchester", 1900, 26, 5.20), ("birmingham", 1600, 25, 5.10),
        ("leeds", 1000, 23, 4.60), ("glasgow", 880, 22, 4.40),
        ("bristol", 1300, 24, 4.90), ("liverpool", 880, 22, 4.35),
        ("edinburgh", 720, 21, 4.20), ("sheffield", 590, 19, 3.90),
    ]:
        rows.append(kw(f"broadband providers {city}", vol, diff, cpc, "Local commercial", "MOFU",
                        "Postcode & location", "City hub", gap_slug=f"postcode/{city}",
                        gap_title=f"Broadband providers in {city.title()}: coverage and best deals"))

    # --- Deals & pricing guides ----------------------------------------------
    rows += [
        kw("broadband deals with no mid contract price rise", 590, 24, 4.90, "Commercial guide", "BOFU", "Deals & pricing", "Guide", "/guides/broadband-deals-with-no-mid-contract-price-rise"),
        kw("best broadband and tv deals", 12100, 32, 4.49, "Commercial", "BOFU", "Deals & pricing", "Guide", "/guides/best-broadband-and-tv-deals"),
        kw("broadband deals under 20", 1000, 30, 3.60, "Commercial", "BOFU", "Deals & pricing", "Guide", "/guides/broadband-deals-under-20"),
        kw("broadband deals with cashback", 720, 26, 3.20, "Commercial", "BOFU", "Deals & pricing", "Guide", "/guides/broadband-deals-with-cashback"),
        kw("broadband deals with no setup fee", 480, 22, 2.90, "Commercial", "BOFU", "Deals & pricing", "Guide", "/guides/broadband-deals-with-no-setup-fee"),
        kw("rolling monthly broadband no contract", 1900, 34, 3.80, "Commercial guide", "MOFU", "Deals & pricing", "Guide", "/guides/best-rolling-monthly-broadband-deals"),
        kw("black friday broadband deals", 5400, 42, 4.60, "Seasonal commercial", "BOFU", "Deals & pricing", "Guide", gap_slug="guides/black-friday-broadband-deals-uk", gap_title="Black Friday broadband deals UK: what to expect and how to compare"),
        kw("january broadband deals", 1300, 28, 3.40, "Seasonal commercial", "BOFU", "Deals & pricing", "Guide", gap_slug="guides/january-broadband-deals-uk", gap_title="January broadband deals UK: new-year switching guide"),
    ]

    # --- Switching & rights ---------------------------------------------------
    rows += [
        kw("how to switch broadband provider uk", 2400, 30, 5.20, "Commercial guide", "MOFU", "Switching & rights", "Guide", "/guides/how-to-switch-broadband-uk"),
        kw("broadband price rises 2026", 1900, 33, 2.10, "Informational", "TOFU", "Switching & rights", "Guide", "/guides/broadband-price-rises-2026"),
        kw("can i leave broadband contract early", 1300, 29, 3.40, "Commercial guide", "MOFU", "Switching & rights", "Guide", "/guides/can-i-leave-broadband-early-after-price-rise"),
        kw("one touch switching broadband explained", 480, 20, 1.60, "Informational", "TOFU", "Switching & rights", "Guide", "/guides/one-touch-switching-explained"),
        kw("broadband contract end rights", 390, 19, 1.80, "Informational", "TOFU", "Switching & rights", "Guide", "/guides/broadband-contract-end-rights"),
        kw("staying with existing broadband provider deals", 590, 22, 3.90, "Commercial guide", "MOFU", "Switching & rights", "Guide", "/guides/broadband-for-existing-customers"),
        kw("broadband complaints ombudsman", 720, 26, 1.40, "Informational", "TOFU", "Trust & research", "Guide", gap_slug="guides/broadband-complaints-and-ombudsman-uk", gap_title="Broadband complaints and the ombudsman: your UK rights"),
    ]

    # --- Technology & speeds ---------------------------------------------------
    rows += [
        kw("full fibre broadband explained", 1600, 25, 2.30, "Informational", "TOFU", "Technology & speeds", "Guide", "/guides/full-fibre-broadband-explained"),
        kw("broadband speeds explained", 1000, 22, 1.90, "Informational", "TOFU", "Technology & speeds", "Guide", "/guides/broadband-speeds-explained"),
        kw("broadband speed test", 246000, 39, 1.14, "Tool", "MOFU", "Tools", "Interactive tool", "/speed-test"),
        kw("check my broadband speed", 40500, 24, 1.145, "Tool", "MOFU", "Tools", "Interactive tool", "/speed-test"),
        kw("best 5g home broadband uk", 4400, 38, 3.10, "Commercial guide", "MOFU", "Technology & speeds", "Guide", "/guides/best-5g-home-broadband-uk"),
        kw("best full fibre broadband uk", 2900, 40, 3.40, "Commercial guide", "MOFU", "Technology & speeds", "Guide", "/guides/best-full-fibre-broadband-uk"),
        kw("fttp vs fttc explained", 880, 20, 1.30, "Informational", "TOFU", "Technology & speeds", "Guide", gap_slug="guides/fttp-vs-fttc-explained", gap_title="FTTP vs FTTC: what the difference means for your speed"),
        kw("best mesh wifi for broadband router", 1300, 33, 2.80, "Commercial guide", "MOFU", "Technology & speeds", "Guide", gap_slug="guides/best-mesh-wifi-for-your-broadband-router", gap_title="Best mesh Wi-Fi systems to pair with your broadband router"),
    ]

    # --- Use cases & lifestyle --------------------------------------------------
    rows += [
        kw("best broadband for working from home", 1900, 28, 3.60, "Commercial guide", "MOFU", "Use cases & lifestyle", "Guide", "/guides/best-broadband-for-working-from-home"),
        kw("best broadband for students", 1600, 27, 2.90, "Commercial guide", "MOFU", "Use cases & lifestyle", "Guide", "/guides/best-broadband-for-students"),
        kw("best broadband for streaming", 1300, 26, 2.70, "Commercial guide", "MOFU", "Use cases & lifestyle", "Guide", "/guides/best-broadband-for-streaming"),
        kw("best broadband for gaming uk", 2400, 34, 3.90, "Commercial guide", "MOFU", "Use cases & lifestyle", "Guide", "/guides/best-broadband-for-gaming-uk"),
        kw("best broadband for rural areas uk", 1000, 24, 2.40, "Commercial guide", "MOFU", "Use cases & lifestyle", "Guide", "/guides/best-broadband-for-rural-areas-uk"),
        kw("broadband when moving house", 2900, 30, 3.30, "Commercial guide", "MOFU", "Use cases & lifestyle", "Guide", "/guides/broadband-moving-house"),
        kw("broadband for landlords hmo", 480, 20, 2.10, "Commercial guide", "MOFU", "Use cases & lifestyle", "Guide", gap_slug="guides/broadband-for-landlords-and-hmos-uk", gap_title="Broadband for landlords and HMOs: what to set up and who pays"),
        kw("student broadband cities guide", 320, 18, 1.60, "Commercial guide", "MOFU", "Use cases & lifestyle", "Guide", gap_slug="guides/student-broadband-by-university-city", gap_title="Student broadband by university city: short-term and rolling deals"),
    ]

    # --- Affordability -----------------------------------------------------------
    rows += [
        kw("broadband social tariffs uk", 2400, 22, 1.10, "Informational", "TOFU", "Affordability", "Guide", "/guides/broadband-social-tariffs-uk"),
        kw("cheapest broadband deals uk", 18100, 45, 4.58, "Commercial", "BOFU", "Affordability", "Guide", "/guides/cheapest-broadband-uk"),
        kw("broadband deals under 20 a month", 1000, 30, 3.60, "Commercial", "BOFU", "Affordability", "Guide", "/guides/broadband-deals-under-20"),
        kw("no credit check broadband", 1900, 32, 2.60, "Commercial guide", "MOFU", "Affordability", "Guide", gap_slug="guides/no-credit-check-broadband-uk", gap_title="No credit check broadband in the UK: options and what to expect"),
        kw("broadband for universal credit", 720, 20, 1.20, "Informational", "TOFU", "Affordability", "Guide", gap_slug="guides/broadband-help-if-you-claim-benefits-uk", gap_title="Cheaper broadband if you claim benefits: social tariffs explained"),
    ]

    # --- Business broadband --------------------------------------------------
    rows += [
        kw("best business broadband providers uk", 1900, 41, 5.395, "B2B commercial", "MOFU", "Business broadband", "Guide", "/guides/best-business-broadband-providers-uk"),
        kw("business broadband providers", 1000, 49, 53.80, "B2B commercial", "MOFU", "Business broadband", "Guide", "/guides/best-business-broadband-providers-uk"),
        kw("leased line cost uk", 720, 30, 18.40, "B2B commercial", "BOFU", "Business broadband", "Guide", gap_slug="guides/leased-line-cost-uk-explained", gap_title="Leased line cost UK: what small businesses actually pay"),
        kw("static ip business broadband", 480, 24, 9.20, "B2B commercial", "MOFU", "Business broadband", "Guide", gap_slug="guides/static-ip-business-broadband-explained", gap_title="Static IP business broadband: when you need one and how much it costs"),
        kw("small office broadband setup", 390, 22, 6.10, "B2B commercial", "MOFU", "Business broadband", "Guide", gap_slug="guides/small-office-broadband-setup-uk", gap_title="Small office broadband setup: routers, backup and support checklist"),
    ]

    # --- Phone & bundles ----------------------------------------------------
    rows += [
        kw("phone and broadband deals", 14800, 57, 4.07, "Commercial", "BOFU", "Phone & bundles", "Guide", "/guides/best-phone-and-broadband-deals"),
        kw("landline broadband deals", 14800, 50, 4.23, "Commercial", "BOFU", "Phone & bundles", "Guide", "/guides/best-phone-and-broadband-deals"),
        kw("broadband without phone line", 2900, 40, 3.10, "Commercial guide", "MOFU", "Phone & bundles", "Guide", "/guides/broadband-without-phone-line"),
    ]

    # --- Trust & research -----------------------------------------------------
    rows += [
        kw("uk broadband customer satisfaction", 1600, 25, 3.41, "Research", "TOFU", "Trust & research", "Research page", "/research/uk-broadband-customer-satisfaction"),
        kw("best customer service broadband provider", 880, 27, 3.10, "Research", "MOFU", "Trust & research", "Research page", gap_slug="research/broadband-customer-service-rankings-uk", gap_title="Broadband customer service rankings: Ofcom-evidenced comparison"),
        kw("how we make money broadband picker", 40, 5, 0.00, "Brand/trust", "TOFU", "Trust & compliance", "Policy", "/how-we-make-money"),
    ]

    # --- Alternative access / satellite -------------------------------------
    rows += [
        kw("satellite broadband uk", 2400, 30, 2.90, "Commercial guide", "MOFU", "Alternative access", "Guide", "/guides/satellite-broadband-uk"),
        kw("starlink vs openreach broadband", 590, 24, 2.20, "Commercial comparison", "MOFU", "Alternative access", "Guide", gap_slug="guides/starlink-vs-fibre-broadband-uk", gap_title="Starlink vs fibre broadband: when satellite makes sense in the UK"),
    ]

    # --- Tools -----------------------------------------------------------------
    rows += [
        kw("broadband cost per mbps calculator", 210, 16, 1.10, "Tool", "MOFU", "Tools", "Interactive tool", gap_slug="tools/broadband-cost-calculator", gap_title="Broadband cost calculator: true monthly and contract-length cost"),
    ]

    # --- Informational / glossary -----------------------------------------------
    rows += [
        kw("broadband glossary terms explained", 480, 15, 0.90, "Informational", "TOFU", "Informational", "Glossary", "/broadband-glossary"),
        kw("what is fttp broadband", 1300, 21, 1.40, "Informational", "TOFU", "Informational", "Glossary", "/broadband-glossary"),
        kw("what is a good broadband speed", 2400, 23, 1.60, "Informational", "TOFU", "Informational", "Glossary", "/broadband-glossary"),
    ]

    return rows


# ---------------------------------------------------------------------------
# 3. Optional: pull extra keyword ideas out of the source Google Sheet
# ---------------------------------------------------------------------------

def read_source_sheet_keywords(sheet_id: str) -> list[str]:
    """Read every non-empty cell that looks like a keyword phrase from a shared sheet.

    Requires the sheet to be shared as at least Viewer with the service account
    used for GOOGLE_APPLICATION_CREDENTIALS / GOOGLE_SERVICE_ACCOUNT_JSON.
    """
    token = google_access_token(["https://www.googleapis.com/auth/spreadsheets.readonly"])
    base = f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}"
    metadata = sheets_request(token, "GET", base + "?fields=sheets(properties(title))")
    found: set[str] = set()
    for sheet in metadata.get("sheets", []):
        title = sheet["properties"]["title"]
        quoted = urllib.parse.quote(f"'{title}'!A1:Z2000", safe="")
        values = sheets_request(token, "GET", f"{base}/values/{quoted}").get("values", [])
        for row in values:
            for cell in row:
                text = str(cell).strip().lower()
                if 3 <= len(text) <= 70 and re.search(r"[a-z]", text) and "broadband" in text:
                    found.add(text)
    return sorted(found)


def merge_source_keywords(rows: list[KeywordRow], extra_keywords: list[str]) -> list[KeywordRow]:
    existing = {r["keyword"] for r in rows}
    added = 0
    for phrase in extra_keywords:
        if phrase in existing:
            continue
        rows.append(kw(
            phrase, 0, 0, 0.0, "Unclassified — from source sheet", "Unscored",
            "From source sheet (needs triage)", "Unassigned",
            gap_slug=None, gap_title=None,
        ))
        existing.add(phrase)
        added += 1
    if added:
        print(f"Merged {added} additional keyword(s) from the source sheet (unscored, needs triage)")
    return rows


# ---------------------------------------------------------------------------
# 4. Scoring and workbook assembly
# ---------------------------------------------------------------------------

def priority_score(volume: int, difficulty: int, cpc: float, intent: str) -> float:
    """Weighted for an affiliate site: commercial value counts for more than
    on a pure-traffic play, since CPC/intent tracks likely Awin conversion value."""
    if volume == 0 and difficulty == 0 and cpc == 0:
        return 0.0
    intent_weight = 1.3 if "commercial" in intent.lower() or "b2b" in intent.lower() else 1.0
    demand = min(100, (volume ** 0.5) / 4) if volume else 5
    opportunity = max(5, 100 - difficulty)
    value = min(100, cpc * 3)
    return round((demand * 0.30 + opportunity * 0.30 + value * 0.40) * intent_weight, 1)


def ranking_ease_score(volume: int, difficulty: int, cpc: float, intent: str) -> float:
    """Score a keyword for a newer site's publish queue, with rankability first.

    Difficulty deliberately carries most of the weight. Demand and commercial
    value prevent tiny, low-value terms from dominating merely because they are
    easy. The score is directional and should be refreshed with first-party data.
    """
    if volume == 0 and difficulty == 0 and cpc == 0:
        return 0.0
    ease = max(0, min(100, 100 - difficulty))
    demand = min(100, (volume ** 0.5) / 4) if volume else 0
    commercial = 100 if "commercial" in intent.lower() or "b2b" in intent.lower() else 55
    value = min(100, cpc * 10)
    return round(ease * 0.60 + demand * 0.20 + commercial * 0.15 + value * 0.05, 1)


def difficulty_band(difficulty: float) -> str:
    if difficulty <= 20:
        return "Very easy"
    if difficulty <= 30:
        return "Easy"
    if difficulty <= 40:
        return "Moderate"
    if difficulty <= 55:
        return "Competitive"
    return "Very competitive"


def priority_rationale(difficulty: float, volume: int, intent: str, keyword_count: int) -> str:
    demand = "strong" if volume >= 1000 else "useful" if volume >= 300 else "niche"
    commercial = "commercial intent" if "commercial" in intent.lower() or "b2b" in intent.lower() else "supporting intent"
    return (
        f"{difficulty_band(difficulty)} opportunity (difficulty {difficulty:.0f}); "
        f"{demand} combined demand ({volume:,}/mo); {commercial}; "
        f"covers {keyword_count} mapped keyword{'s' if keyword_count != 1 else ''}."
    )


def affiliate_potential(cpc: float, intent: str) -> str:
    if cpc >= 5 or "commercial" in intent.lower() or "b2b" in intent.lower():
        return "High"
    if cpc >= 2:
        return "Medium"
    if cpc > 0:
        return "Low"
    return "Unscored"


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[Any]]) -> Any:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    if rows and headers:
        ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        table = Table(displayName=re.sub(r"\W+", "", name)[:20] + "Table", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    return ws


def style_sheet(ws: Any) -> None:
    navy, pale = "0F172A", "EAF6F5"
    thin = Side(style="thin", color="D7E1E8")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)
    for column in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, column).value or "") for row in range(1, min(ws.max_row, 80) + 1)]
        width = min(48, max(10, max(len(v) for v in values) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for cell in ws[row]:
                cell.fill = PatternFill("solid", fgColor=pale)


def build_workbook(
    output: Path, urls: list[str], audit: list[dict[str, Any]],
    keywords: list[KeywordRow], source_sheet_id: str | None,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    audited_by_url = {row["url"]: row for row in audit}
    live_paths = {urllib.parse.urlparse(u).path or "/" for u in urls}
    live_check_performed = any(row.get("status") for row in audit)

    # Enrich mapped rows and score everything.
    for row in keywords:
        row["priority"] = priority_score(row["volume"], row["difficulty"], row["cpc"], row["intent"])
        row["ranking_ease"] = ranking_ease_score(row["volume"], row["difficulty"], row["cpc"], row["intent"])
        row["affiliate_potential"] = affiliate_potential(row["cpc"], row["intent"])
        if row["mapped_url"]:
            path = row["mapped_url"]
            row["status"] = "Existing page" if path in live_paths or not urls else "Existing page (mapped)"
            row["target_url"] = SITE + path
        elif row["gap_slug"]:
            target_path = f"/{row['gap_slug']}"
            if target_path in live_paths:
                row["mapped_url"] = target_path
                row["status"] = "Existing page"
            else:
                row["status"] = "Content gap — recommend new page"
            row["target_url"] = SITE + target_path
        else:
            row["status"] = "Unassigned — needs triage"
            row["target_url"] = ""

    # --- Read Me -----------------------------------------------------------
    readme = [
        ["Purpose", "Detailed keyword-to-page mapping for BroadbandPicker.co.uk, covering every live page and every recommended future page, prioritised for Awin affiliate revenue."],
        ["Generated", RUN_DATE],
        ["Site analysed", SITE],
        ["Source sheet analysed", f"https://docs.google.com/spreadsheets/d/{source_sheet_id}/edit" if source_sheet_id else "None supplied"],
        ["Evidence hierarchy", "Live sitemap crawl (page titles/H1/word count) + an editorially assigned, dated UK keyword snapshot + any keywords merged from the source sheet."],
        ["Volume/CPC caveat", "Monthly search volume and CPC are directional third-party-style estimates, not Google Search Console or Google Ads data. They are not additive across variants. Re-validate in Google Ads Keyword Planner/GSC before committing budget."],
        ["Scoring", "SEO Build Priority is rankability-first: 60% inverse keyword difficulty, 20% demand, 15% intent and 5% CPC. Revenue Priority remains available as a secondary commercial signal."],
        ["Objective", "Build a defensible, evidence-led UK broadband affiliate content plan: map demand to the right page type, close content gaps in priority order, and avoid duplicate/competing pages for the same keyword."],
        ["How to use", "The standard production command builds up to five active pages from the top of the Content Gap Roadmap, one at a time. After each page it checks the live sitemap and HTTP response, moves completed pages below the active queue, labels their deployment status and strikes them through."],
        ["Re-run", "python3 scripts/build_keyword_mapping.py --create-google-doc --share-with <email>"],
    ]
    add_sheet(wb, "Read Me", ["Field", "Detail"], readme)

    # --- Keyword Mapping (main) --------------------------------------------
    mapping_headers = [
        "Keyword", "Est. Monthly UK Volume", "Difficulty (0-100)", "Est. CPC (GBP)",
        "Search Intent", "Funnel Stage", "Topic Cluster", "Page Type",
        "Mapping Status", "Mapped / Recommended URL", "SEO Ease Score", "Revenue Priority Score",
        "Affiliate Revenue Potential",
    ]
    mapping_rows = []
    for row in sorted(keywords, key=lambda r: r["priority"], reverse=True):
        mapping_rows.append([
            row["keyword"], row["volume"] or "", row["difficulty"] or "", row["cpc"] or "",
            row["intent"], row["funnel"], row["cluster"], row["page_type"],
            row["status"], row["target_url"], row["ranking_ease"], row["priority"], row["affiliate_potential"],
        ])
    add_sheet(wb, "Keyword Mapping", mapping_headers, mapping_rows)

    # --- Current Page Inventory --------------------------------------------
    inventory_headers = ["Path", "Topic Cluster", "Page Type", "Live Title", "H1", "Word Count", "Mapped Keywords", "Notes"]
    inventory_rows = []
    keyword_count_by_path: dict[str, list[str]] = {}
    for row in keywords:
        if row["mapped_url"]:
            keyword_count_by_path.setdefault(row["mapped_url"], []).append(row["keyword"])
    for url in urls:
        path = urllib.parse.urlparse(url).path or "/"
        cluster, page_type = route_family(path)
        crawled = audited_by_url.get(url, {})
        mapped_terms = keyword_count_by_path.get(path, [])
        inventory_rows.append([
            path, cluster, page_type, crawled.get("title", ""), crawled.get("h1", ""),
            crawled.get("words", ""), ", ".join(mapped_terms[:6]) + (f" (+{len(mapped_terms)-6} more)" if len(mapped_terms) > 6 else ""),
            crawled.get("notes", ""),
        ])
    add_sheet(wb, "Current Page Inventory", inventory_headers, inventory_rows)

    # --- Content Gap Roadmap -------------------------------------------------
    gap_headers = [
        "Next Build Priority", "Current Status", "Last Live Check", "Recommended URL", "Page Title", "Topic Cluster", "Page Type",
        "Primary Keywords", "Combined Est. Volume", "Weighted Difficulty", "Ranking Difficulty Band",
        "SEO Build Priority Score", "Revenue Priority Score", "Why This Position",
    ]
    gap_groups: dict[str, dict[str, Any]] = {}
    for row in keywords:
        if not row.get("gap_slug"):
            continue
        group = gap_groups.setdefault(row["gap_slug"], {
            "title": row.get("gap_title") or row["keyword"], "cluster": row["cluster"],
            "page_type": row["page_type"], "keywords": [], "volume": 0, "priority": 0.0,
            "weighted_difficulty": 0.0, "difficulty_weight": 0, "ranking_ease": 0.0,
            "commercial_intent": row["intent"], "max_cpc": 0.0,
        })
        group["keywords"].append(row["keyword"])
        group["volume"] += row["volume"]
        group["priority"] = max(group["priority"], row["priority"])
        weight = max(row["volume"], 1)
        group["weighted_difficulty"] += row["difficulty"] * weight
        group["difficulty_weight"] += weight
        group["max_cpc"] = max(group["max_cpc"], row["cpc"])
        if "commercial" in row["intent"].lower() or "b2b" in row["intent"].lower():
            group["commercial_intent"] = row["intent"]
    for slug, group in gap_groups.items():
        group["weighted_difficulty"] = round(
            group["weighted_difficulty"] / group["difficulty_weight"], 1
        ) if group["difficulty_weight"] else 100.0
        group["ranking_ease"] = ranking_ease_score(
            group["volume"], round(group["weighted_difficulty"]),
            group["max_cpc"], group["commercial_intent"],
        )
        target_url = f"{SITE}/{slug}"
        target_path = f"/{slug}"
        page_audit = audited_by_url.get(target_url, {})
        http_status = page_audit.get("status")
        group["complete"] = bool(
            target_path in live_paths and http_status and 200 <= int(http_status) < 400
        )
        if group["complete"]:
            platform = page_audit.get("platform") or "Live host"
            group["status"] = "Built, live and deployed on Vercel" if platform == "Vercel" else "Built and live"
        elif target_url in audited_by_url and http_status:
            group["status"] = f"Built route detected; HTTP {http_status}"
        elif not live_check_performed:
            group["status"] = "Planned — live check unavailable"
        else:
            group["status"] = "Planned — not live"
        group["last_live_check"] = RUN_DATE if live_check_performed else ""
    gap_rows_sorted = sorted(
        gap_groups.items(),
        key=lambda kv: (
            kv[1]["complete"], -kv[1]["ranking_ease"], kv[1]["weighted_difficulty"],
            -kv[1]["volume"], -kv[1]["priority"],
        ),
    )
    gap_rows_out = []
    next_priority = 0
    completed_row_numbers = []
    for slug, group in gap_rows_sorted:
        if not group["complete"]:
            next_priority += 1
        else:
            completed_row_numbers.append(len(gap_rows_out) + 2)
        gap_rows_out.append([
            next_priority if not group["complete"] else "", group["status"], group["last_live_check"],
            f"{SITE}/{slug}", group["title"], group["cluster"], group["page_type"],
            "; ".join(group["keywords"]), group["volume"], group["weighted_difficulty"],
            difficulty_band(group["weighted_difficulty"]), group["ranking_ease"], group["priority"],
            priority_rationale(group["weighted_difficulty"], group["volume"], group["commercial_intent"], len(group["keywords"])),
        ])
    roadmap_ws = add_sheet(wb, "Content Gap Roadmap", gap_headers, gap_rows_out)
    for row_number in completed_row_numbers:
        for cell in roadmap_ws[row_number]:
            cell.font = Font(name=cell.font.name, size=cell.font.sz, bold=cell.font.bold,
                             italic=cell.font.italic, color=cell.font.color, strike=True)

    # --- Cluster Summary -----------------------------------------------------
    cluster_headers = ["Topic Cluster", "Keyword Count", "Total Est. Volume", "Existing-Page Keywords", "Content-Gap Keywords", "Avg Priority Score"]
    cluster_stats: dict[str, dict[str, Any]] = {}
    for row in keywords:
        stats = cluster_stats.setdefault(row["cluster"], {"count": 0, "volume": 0, "existing": 0, "gap": 0, "priority_sum": 0.0})
        stats["count"] += 1
        stats["volume"] += row["volume"]
        stats["priority_sum"] += row["priority"]
        if row["mapped_url"]:
            stats["existing"] += 1
        elif row["gap_slug"]:
            stats["gap"] += 1
    cluster_rows = []
    for cluster, stats in sorted(cluster_stats.items(), key=lambda kv: kv[1]["volume"], reverse=True):
        cluster_rows.append([
            cluster, stats["count"], stats["volume"], stats["existing"], stats["gap"],
            round(stats["priority_sum"] / stats["count"], 1) if stats["count"] else 0,
        ])
    add_sheet(wb, "Cluster Summary", cluster_headers, cluster_rows)

    # --- Methodology -----------------------------------------------------------
    methodology = [
        ["Site crawl", "Live sitemap.xml fetched at run time; each URL lightly audited for <title>, H1 and word count so mapped/gap status reflects the current build, not a stale snapshot."],
        ["Keyword universe", "Editorially assembled from the site's existing 13 provider pages, 15 vs-comparison pages, 30 guides, postcode hub, priority landing pages and known UK broadband search patterns, extended with high-intent gaps competitors typically rank for."],
        ["Volume/difficulty/CPC", "Directional UK estimates in the same style/scale as the existing docs/uk-broadband-seo-geo-plan.xlsx snapshot. Treat as ranking/prioritisation signal, not a traffic forecast."],
        ["Mapping status", "'Existing page' = a live URL already targets this term. 'Content gap' = no live page targets it; a specific new URL, title and cluster is recommended. 'Unassigned' = pulled from the source sheet and needs manual triage."],
        ["SEO build priority", "Content gaps are ordered by a 0-100 rankability-first score: 60% inverse keyword difficulty, 20% demand, 15% intent and 5% CPC. Easier pages lead; value signals break ties."],
        ["Revenue priority score", "Secondary 0-130 commercial score combining demand, inverse difficulty and CPC, boosted for commercial/B2B intent. It no longer controls publish order."],
        ["Build/deployment status", "On online runs, a planned page is marked complete only when its exact URL appears in the live sitemap and returns HTTP 2xx/3xx. Vercel is confirmed from response headers. Completed rows move below the active queue and are struck through."],
        ["Affiliate revenue potential", "High: CPC >= £5 or commercial/B2B intent. Medium: CPC £2-5. Low: CPC < £2. Unscored: source-sheet keyword pending triage."],
        ["Update cadence", "Re-run monthly, or after publishing new pages, so 'Existing page' status stays accurate and newly closed content gaps drop off the roadmap."],
    ]
    add_sheet(wb, "Methodology", ["Aspect", "Detail"], methodology)

    return wb


def validate(wb: Workbook) -> None:
    required = {"Read Me", "Keyword Mapping", "Current Page Inventory", "Content Gap Roadmap", "Cluster Summary", "Methodology"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Workbook missing expected sheets: {sorted(missing)}")


# ---------------------------------------------------------------------------
# 5. Google auth + Sheets/Drive helpers (service account, same pattern as
#    scripts/build_uk_broadband_seo_plan.py)
# ---------------------------------------------------------------------------

def google_access_token(scopes: list[str]) -> str:
    try:
        from google.oauth2 import service_account
        from google.auth.transport.requests import Request
    except ImportError as exc:
        raise RuntimeError("Google sync requires: pip install -r requirements-seo.txt") from exc

    raw_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    credentials_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    if raw_json:
        credentials = service_account.Credentials.from_service_account_info(json.loads(raw_json), scopes=scopes)
    elif credentials_path:
        credentials = service_account.Credentials.from_service_account_file(credentials_path, scopes=scopes)
    else:
        raise RuntimeError(
            "Set GOOGLE_SERVICE_ACCOUNT_JSON or GOOGLE_APPLICATION_CREDENTIALS. "
            "Share the source sheet (Viewer) and, for doc creation, no sharing is "
            "needed up front — the script shares the new file itself via --share-with."
        )
    credentials.refresh(Request())
    return credentials.token


def sheets_request(token: str, method: str, url: str, **kwargs: Any) -> dict[str, Any]:
    headers = kwargs.pop("headers", {})
    headers.update({"Authorization": f"Bearer {token}", "Content-Type": "application/json"})
    response = requests.request(method, url, headers=headers, timeout=60, **kwargs)
    if not response.ok:
        raise RuntimeError(f"Google API {response.status_code}: {response.text[:800]}")
    return response.json() if response.content else {}


def create_google_sheet_from_workbook(wb: Workbook, title: str, share_with: str | None) -> dict[str, Any]:
    """Create a brand-new Google Sheet on Drive containing every tab in wb."""
    token = google_access_token([
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
    ])
    create_body = {
        "properties": {"title": title},
        "sheets": [
            {"properties": {"title": name, "gridProperties": {"frozenRowCount": 1}}}
            for name in wb.sheetnames
        ],
    }
    created = sheets_request(token, "POST", "https://sheets.googleapis.com/v4/spreadsheets", json=create_body)
    spreadsheet_id = created["spreadsheetId"]
    base = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}"

    data = []
    format_requests: list[dict[str, Any]] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [[cell.value if cell.value is not None else "" for cell in row] for row in ws.iter_rows()]
        if not rows:
            continue
        last_col = get_column_letter(len(rows[0]))
        data.append({"range": f"'{sheet}'!A1:{last_col}{len(rows)}", "values": rows})
        sheet_id = next(s["properties"]["sheetId"] for s in created["sheets"] if s["properties"]["title"] == sheet)
        format_requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                "cell": {"userEnteredFormat": {
                    "backgroundColor": {"red": 0.06, "green": 0.09, "blue": 0.16},
                    "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1}, "bold": True},
                    "wrapStrategy": "WRAP",
                }},
                "fields": "userEnteredFormat(backgroundColor,textFormat,wrapStrategy)",
            }
        })
        format_requests.append({"setBasicFilter": {"filter": {"range": {"sheetId": sheet_id}}}})

    if data:
        sheets_request(token, "POST", f"{base}/values:batchUpdate", json={"valueInputOption": "RAW", "data": data})
    if format_requests:
        sheets_request(token, "POST", f"{base}:batchUpdate", json={"requests": format_requests})

    if share_with:
        drive_token = token  # drive.file scope covers files this credential created
        requests.post(
            f"https://www.googleapis.com/drive/v3/files/{spreadsheet_id}/permissions",
            headers={"Authorization": f"Bearer {drive_token}", "Content-Type": "application/json"},
            params={"sendNotificationEmail": "false"},
            json={"type": "user", "role": "writer", "emailAddress": share_with},
            timeout=30,
        ).raise_for_status()

    return {"spreadsheetId": spreadsheet_id, "spreadsheetUrl": created.get("spreadsheetUrl"), "sharedWith": share_with}


def update_google_sheet_from_workbook(wb: Workbook, spreadsheet_id: str) -> dict[str, Any]:
    """Synchronise this workbook into an existing native Google Sheet in place."""
    token = google_access_token(["https://www.googleapis.com/auth/spreadsheets"])
    base = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}"
    metadata = sheets_request(
        token, "GET", base,
        params={"fields": "properties(title),sheets(properties(sheetId,title,gridProperties(rowCount,columnCount)),tables(tableId,range))"},
    )
    remote = {sheet["properties"]["title"]: sheet for sheet in metadata.get("sheets", [])}
    missing = [name for name in wb.sheetnames if name not in remote]
    if missing:
        sheets_request(token, "POST", f"{base}:batchUpdate", json={
            "requests": [{"addSheet": {"properties": {"title": name}}} for name in missing]
        })
        metadata = sheets_request(
            token, "GET", base,
            params={"fields": "properties(title),sheets(properties(sheetId,title,gridProperties(rowCount,columnCount)),tables(tableId,range))"},
        )
        remote = {sheet["properties"]["title"]: sheet for sheet in metadata.get("sheets", [])}

    values_payload = []
    format_requests: list[dict[str, Any]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [[cell.value if cell.value is not None else "" for cell in row] for row in ws.iter_rows()]
        last_col = get_column_letter(ws.max_column)
        quoted_name = name.replace("'", "''")
        sheet_id = remote[name]["properties"]["sheetId"]
        grid = remote[name]["properties"].get("gridProperties", {})
        row_count = int(grid.get("rowCount", 0))
        column_count = int(grid.get("columnCount", 0))
        expansion_requests: list[dict[str, Any]] = []
        if ws.max_row > row_count:
            expansion_requests.append({
                "appendDimension": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "length": ws.max_row - row_count,
                }
            })
        if ws.max_column > column_count:
            expansion_requests.append({
                "appendDimension": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "length": ws.max_column - column_count,
                }
            })
        if expansion_requests:
            sheets_request(token, "POST", f"{base}:batchUpdate", json={"requests": expansion_requests})
        # Write only the workbook's exact used rectangle. Do not clear the
        # sheet grid first: native formulas, notes or controls outside the
        # managed table may belong to a human editor and must be preserved.
        values_payload.append({"range": f"'{quoted_name}'!A1:{last_col}{ws.max_row}", "values": rows})
        format_requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": ws.max_row,
                          "startColumnIndex": 0, "endColumnIndex": ws.max_column},
                "cell": {"userEnteredFormat": {"wrapStrategy": "WRAP", "verticalAlignment": "TOP"}},
                "fields": "userEnteredFormat(wrapStrategy,verticalAlignment)",
            }
        })
        format_requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1,
                          "startColumnIndex": 0, "endColumnIndex": ws.max_column},
                "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
                "fields": "userEnteredFormat.textFormat.bold",
            }
        })
        tables = remote[name].get("tables", [])
        if tables:
            format_requests.append({
                "updateTable": {
                    "table": {"tableId": tables[0]["tableId"], "range": {
                        "sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": ws.max_row,
                        "startColumnIndex": 0, "endColumnIndex": ws.max_column,
                    }},
                    "fields": "range",
                }
            })

    sheets_request(token, "POST", f"{base}/values:batchUpdate",
                   json={"valueInputOption": "RAW", "data": values_payload})
    if "Content Gap Roadmap" in wb.sheetnames:
        ws = wb["Content Gap Roadmap"]
        sheet_id = remote["Content Gap Roadmap"]["properties"]["sheetId"]
        for row_number in range(2, ws.max_row + 1):
            if str(ws.cell(row_number, 2).value or "").startswith("Built"):
                format_requests.append({
                    "repeatCell": {
                        "range": {"sheetId": sheet_id, "startRowIndex": row_number - 1,
                                  "endRowIndex": row_number, "startColumnIndex": 0,
                                  "endColumnIndex": ws.max_column},
                        "cell": {"userEnteredFormat": {"textFormat": {"strikethrough": True}}},
                        "fields": "userEnteredFormat.textFormat.strikethrough",
                    }
                })
    if format_requests:
        sheets_request(token, "POST", f"{base}:batchUpdate", json={"requests": format_requests})
    return {"spreadsheetId": spreadsheet_id,
            "spreadsheetUrl": f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"}


# ---------------------------------------------------------------------------
# 6. Gated page-build pipeline
# ---------------------------------------------------------------------------

# Existing GSC-tracked hub/index routes that are not a-vs-b comparison pages and
# have no curated keyword row of their own. Matched exactly, before the prefix
# table, so /providers/compare is never mis-routed to the provider template.
HUB_ROUTE_OVERRIDES: dict[str, tuple[str, str, str]] = {
    "providers/compare": (
        "data/provider-comparisons.ts",
        "app/providers/compare/page.tsx",
        "ProviderComparison",
    ),
}


def route_page_build(slug: str) -> tuple[str, str, str]:
    if slug in HUB_ROUTE_OVERRIDES:
        return HUB_ROUTE_OVERRIDES[slug]
    for prefix, target in PAGE_ROUTES.items():
        if slug.startswith(prefix):
            return target
    raise RuntimeError(f"No page-build route is configured for: {slug}")


NOISE_QUERY_TOPICAL = ("broadband", "fibre", "internet", "wifi", "wi-fi", "isp", "openreach")
NOISE_QUERY_COMPARISON = (" vs ", " v ", " versus ", " or ", "compare", "comparison", "better", "cheaper", "faster")


def is_noisy_gsc_query(query: str) -> bool:
    """Reject obviously malformed or off-topic Search Console rows before one can
    seed a primary keyword. This never invents data; it only filters."""
    q = f" {(query or '').strip().lower()} "
    if len(q.strip()) < 6:
        return True
    if len(re.sub(r"[^a-z]", "", q)) < 4:
        return True
    if not any(term in q for term in NOISE_QUERY_TOPICAL) and not any(
        term in q for term in NOISE_QUERY_COMPARISON
    ):
        return True
    if re.fullmatch(r"\s*(yes|no|ok|okay|maybe|idk|i would|yes i would)\b[\s\w]{0,20}\s*", q):
        return True
    if re.search(r"\b(job|jobs|recruit|salary|vacancy|linkedin|instagram|tiktok)\b", q):
        return True
    return False


def first_party_hub_keywords(mapped_path: str) -> tuple[list[dict[str, Any]], str, str]:
    """Keyword rows for a GSC-tracked hub that has no curated mapping.

    Preference order: (1) an explicit researched primary already recorded for this
    exact slug in current-page-research.json; otherwise (2) non-noisy first-party
    GSC queries offered as candidates only, forcing an explicit researched primary
    decision before the build. Volume, difficulty and CPC are always None.
    """
    slug = mapped_path.strip("/")
    if PAGE_RESEARCH_FILE.exists():
        try:
            research = json.loads(PAGE_RESEARCH_FILE.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            research = {}
        if str(research.get("slug") or "").strip("/") == slug and str(
            research.get("primary_keyword") or ""
        ).strip():
            rows = [{
                "keyword": str(research["primary_keyword"]).strip(),
                "volume": None, "difficulty": None, "cpc": None,
                "intent": "Researched navigational comparison",
                "role": "primary",
                "recommended_slot": "title/h1/excerpt",
                "source": "Researched primary-keyword decision (current-page-research.json)",
            }]
            for item in research.get("secondary_keywords") or []:
                text = str(item.get("keyword") or "").strip()
                if text:
                    rows.append({
                        "keyword": text, "volume": None, "difficulty": None, "cpc": None,
                        "intent": str(item.get("intent") or "Comparison"),
                        "role": "secondary",
                        "recommended_slot": str(item.get("page_slot") or "body/faq"),
                        "source": "Researched secondary-keyword decision",
                    })
            return rows, "researched", "Primary keyword taken from the reviewed research file."
    brief = weekly_performance_brief(mapped_path)
    candidates = [q for q in (brief or {}).get("top_queries", []) if not is_noisy_gsc_query(q)]
    if candidates:
        rows = [{
            "keyword": q, "volume": None, "difficulty": None, "cpc": None,
            "intent": "First-party GSC query (candidate)",
            "role": "candidate",
            "recommended_slot": "requires an explicit researched primary-keyword decision",
            "source": "First-party GSC query evidence, docs/weekly-seo-intelligence.json (candidate, not a confirmed primary)",
        } for q in candidates]
        note = (
            f"No curated keyword row and no researched primary yet for /{slug}. "
            f"{len(candidates)} non-noisy GSC candidate queries attached. Decide an explicit "
            "primary keyword in current-page-research.json (distinct from the /compare tool intent) "
            "before building; the packet primary is provisional until then."
        )
        return rows, "gsc_candidates", note
    return [], "none", ""


def provider_slugs_from_gap(slug: str) -> list[str]:
    if not slug.startswith("providers/compare/"):
        return []
    comparison_slug = slug.rsplit("/", 1)[-1]
    parts = comparison_slug.split("-vs-")
    if len(parts) != 2 or not all(parts):
        raise RuntimeError(f"Comparison slug does not follow a-vs-b: {slug}")
    return parts


def provider_exists(slug: str) -> bool:
    content = (ROOT / "data" / "providers.ts").read_text(encoding="utf-8")
    return bool(re.search(rf"\bslug:\s*['\"]{re.escape(slug)}['\"]", content))


def template_secondary_keywords(slug: str) -> list[dict[str, Any]]:
    providers = provider_slugs_from_gap(slug)
    if len(providers) != 2:
        return []
    a, b = (provider.replace("-", " ") for provider in providers)
    stems = [
        (f"{a} vs {b} broadband deals", "Commercial", "keyDifferences.pricing"),
        (f"{a} vs {b} broadband speed", "Commercial research", "keyDifferences.speed"),
        (f"{a} vs {b} broadband reviews", "Research", "faqs"),
        (f"{a} vs {b} broadband coverage", "Research", "keyDifferences.coverage"),
        (f"which is better {a} or {b} broadband", "Commercial comparison", "excerpt/verdict"),
    ]
    return [{
        "keyword": phrase, "volume": None, "difficulty": None, "cpc": None,
        "intent": intent, "role": "secondary", "recommended_slot": slot,
        "source": "Template-derived supporting query; validate in GSC/Keyword Planner",
    } for phrase, intent, slot in stems]


def classify_content_format(slug: str, page_type: str) -> dict[str, Any]:
    """Choose the reader task before prescribing copy length or UI."""
    if slug.startswith("postcode/"):
        return {
            "id": "local_availability_hub",
            "primary_task": "Check local availability, coverage and relevant deals",
            "default_decision_aid": "postcode-led availability CTA plus sourced local evidence",
        }
    if slug == "providers/compare":
        return {
            "id": "comparison_directory",
            "primary_task": "Find and open the right provider head-to-head comparison, or move to the all-provider tool when no shortlist exists yet",
            "default_decision_aid": "a crawlable matchup directory plus an accessible provider-pair finder that resolves to the nearest useful comparison",
        }
    if slug.startswith("providers/compare/"):
        return {
            "id": "provider_comparison",
            "primary_task": "Choose between two named providers",
            "default_decision_aid": "side-by-side comparison with a conditional verdict",
        }
    if slug.startswith("providers/"):
        return {
            "id": "provider_review",
            "primary_task": "Assess a provider's plans, coverage, service and value",
            "default_decision_aid": "plan table, evidence-led pros and cons, and availability CTA",
        }
    if slug.startswith("tools/") or page_type == "Interactive tool":
        return {
            "id": "interactive_tool",
            "primary_task": "Complete the named calculation or diagnostic task",
            "default_decision_aid": "working accessible tool with interpretable results",
        }
    if any(term in slug for term in ("deal", "cheap", "cashback", "setup-fee", "price")):
        return {
            "id": "commercial_deals_guide",
            "primary_task": "Compare current costs, conditions and switching value",
            "default_decision_aid": "total-cost examples, eligibility checks and deal CTA",
        }
    if any(term in slug for term in ("rights", "complaint", "ombudsman", "switch", "contract")):
        return {
            "id": "consumer_action_guide",
            "primary_task": "Understand rights and complete a safe next action",
            "default_decision_aid": "step-by-step checklist, exceptions and escalation route",
        }
    return {
        "id": "editorial_explainer",
        "primary_task": "Understand the topic and make an informed broadband decision",
        "default_decision_aid": "answer-first summary, examples and a static decision table",
    }


def weekly_performance_brief(path: str) -> dict[str, Any] | None:
    """Attach first-party GSC/GA4 evidence when the weekly report contains this URL."""
    if not WEEKLY_SEO_INTELLIGENCE_FILE.exists():
        return None
    try:
        report = json.loads(WEEKLY_SEO_INTELLIGENCE_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    match = next(
        (item for item in report.get("opportunities", []) if item.get("page") == path),
        None,
    )
    if not match:
        return None
    keep = (
        "clicks", "impressions", "ctr", "position", "impression_change_pct", "sessions",
        "engagement_rate", "affiliate_clicks", "ai_referral_visits", "score", "category",
        "recommended_action", "top_queries",
    )
    return {
        "source": str(WEEKLY_SEO_INTELLIGENCE_FILE.relative_to(ROOT)),
        "generated_at": report.get("generated_at"),
        **{key: match.get(key) for key in keep},
    }


def enrich_page_packet(packet: dict[str, Any]) -> dict[str, Any]:
    path = urllib.parse.urlparse(str(packet["url"])).path.rstrip("/") or "/"
    packet["content_format"] = classify_content_format(
        str(packet["slug"]), str(packet.get("page_type") or "")
    )
    packet["performance_brief"] = weekly_performance_brief(path)
    packet["build_strategy"] = {
        "serp_feature_classification": True,
        "information_gain_required": True,
        "ctr_candidates_required": 3,
        "internal_link_plan_required": True,
        "interaction_must_match_intent": True,
        "review_days_after_launch": [7, 28, 56, 90],
    }
    return packet


def next_page_packet(wb: Workbook, keywords: list[KeywordRow]) -> dict[str, Any]:
    ws = wb["Content Gap Roadmap"]
    headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[1])}
    required = {"Next Build Priority", "Recommended URL", "Page Title", "Topic Cluster", "Page Type"}
    if missing := required - headers.keys():
        raise RuntimeError(f"Roadmap is missing pipeline columns: {sorted(missing)}")
    selected = None
    for row_number in range(2, ws.max_row + 1):
        if ws.cell(row_number, headers["Next Build Priority"]).value == 1:
            selected = row_number
            break
    if selected is None:
        raise RuntimeError("No active page remains in the Content Gap Roadmap")
    url = str(ws.cell(selected, headers["Recommended URL"]).value)
    slug = urllib.parse.urlparse(url).path.strip("/")
    data_file, template, interface = route_page_build(slug)
    mapped = [row for row in keywords if row.get("gap_slug") == slug]
    mapped.sort(key=lambda row: (row["difficulty"], -row["volume"]))
    providers = provider_slugs_from_gap(slug)
    missing_providers = [provider for provider in providers if not provider_exists(provider)]
    detailed_keywords = [{
        "keyword": row["keyword"], "volume": row["volume"],
        "difficulty": row["difficulty"], "cpc": row["cpc"],
        "intent": row["intent"], "role": "primary" if index == 0 else "secondary",
        "recommended_slot": "title/excerpt" if index == 0 else "body/faq",
        "source": "Curated keyword dataset",
    } for index, row in enumerate(mapped)]
    known = {item["keyword"] for item in detailed_keywords}
    detailed_keywords.extend(
        item for item in template_secondary_keywords(slug) if item["keyword"] not in known
    )
    return enrich_page_packet({
        "generated": RUN_DATE,
        "slug": slug,
        "url": url,
        "title": ws.cell(selected, headers["Page Title"]).value,
        "cluster": ws.cell(selected, headers["Topic Cluster"]).value,
        "page_type": ws.cell(selected, headers["Page Type"]).value,
        "data_file": data_file,
        "template": template,
        "interface": interface,
        "keywords": detailed_keywords,
        "provider_prerequisites": providers,
        "missing_provider_prerequisites": missing_providers,
        "brief": str(PIPELINE_BRIEF),
    })


def page_packet_for_url(url: str, keywords: list[KeywordRow]) -> dict[str, Any]:
    """Build a research packet for an existing or explicitly selected site URL."""
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in {"http", "https"} or parsed.netloc not in {
        "broadbandpicker.co.uk", "www.broadbandpicker.co.uk"
    }:
        raise RuntimeError("--build-page-url must be a BroadbandPicker HTTPS URL")
    slug = parsed.path.strip("/")
    if not slug:
        raise RuntimeError("--build-page-url must identify a page below the site root")
    data_file, template, interface = route_page_build(slug)
    mapped_path = f"/{slug}"
    # Explicit rebuilds must also accept a curated gap target. City pages live in
    # the nested postcode sitemap, so the top-level crawl may leave their
    # keyword row represented by ``gap_slug`` even when the route already
    # exists. Treat both representations as the same selected URL.
    mapped = [
        row for row in keywords
        if row.get("mapped_url") == mapped_path
        or f"/{str(row.get('gap_slug') or '').lstrip('/')}" == mapped_path
    ]
    mapped.sort(key=lambda row: (row["difficulty"], -row["volume"]))
    if not mapped and slug.startswith("postcode/"):
        weekly_report = WEEKLY_SEO_INTELLIGENCE_FILE
        if weekly_report.exists():
            report = json.loads(weekly_report.read_text(encoding="utf-8"))
            opportunity = next(
                (item for item in report.get("opportunities", []) if item.get("page") == mapped_path),
                None,
            )
            if opportunity and opportunity.get("top_queries"):
                mapped = [{
                    "keyword": query,
                    "volume": None,
                    "difficulty": None,
                    "cpc": None,
                    "intent": "Local commercial",
                    "cluster": "Postcode & location",
                    "page_type": "Postcode prefix page",
                } for query in opportunity["top_queries"]]
    hub_note: str | None = None
    hub_mode: str | None = None
    if not mapped:
        hub_rows, hub_mode, hub_note = first_party_hub_keywords(mapped_path)
        if hub_rows:
            mapped = hub_rows
    if not mapped:
        raise RuntimeError(
            f"No curated keyword mapping and no first-party GSC evidence for {mapped_path}. "
            "Add a curated keyword row, or ensure the URL appears in docs/weekly-seo-intelligence.json."
        )
    detailed_keywords = [{
        "keyword": row["keyword"],
        "volume": row.get("volume"),
        "difficulty": row.get("difficulty"),
        "cpc": row.get("cpc"),
        "intent": row["intent"],
        "role": row.get("role") or ("primary" if index == 0 else "secondary"),
        "recommended_slot": row.get("recommended_slot") or ("title/excerpt" if index == 0 else "body/faq"),
        "source": row.get("source") or "Curated keyword dataset",
    } for index, row in enumerate(mapped)]
    providers = provider_slugs_from_gap(slug)
    is_hub = hub_note is not None or slug in HUB_ROUTE_OVERRIDES
    packet: dict[str, Any] = {
        "generated": RUN_DATE,
        "slug": slug,
        "url": f"https://broadbandpicker.co.uk/{slug}",
        "title": slug.rsplit("/", 1)[-1].replace("-", " ").title(),
        "cluster": "Provider vs comparison" if is_hub else mapped[0]["cluster"],
        "page_type": "Comparison directory" if is_hub else mapped[0]["page_type"],
        "data_file": data_file,
        "template": template,
        "interface": interface,
        "keywords": detailed_keywords,
        "provider_prerequisites": providers,
        "missing_provider_prerequisites": [provider for provider in providers if not provider_exists(provider)],
        "brief": str(PIPELINE_BRIEF),
        "build_mode": "complete_existing_page",
    }
    if hub_note:
        packet["keyword_mapping_note"] = hub_note
        packet["primary_keyword_mode"] = hub_mode
    return enrich_page_packet(packet)


def ux_page_type_for_packet(packet: dict[str, Any]) -> str:
    """Use the same route taxonomy as analyze_page_type_ux.py."""
    slug = str(packet.get("slug") or "").strip("/")
    if not slug:
        return "homepage"
    if slug == "deals" or slug.startswith("deals/"):
        return "deals_hub"
    if slug == "providers/compare" or slug.startswith("providers/compare/"):
        return "provider_vs"
    if re.fullmatch(r"providers/[^/]+/deals", slug):
        return "provider_deals"
    if slug.startswith("providers/") and slug != "providers":
        return "provider_review"
    if slug == "providers" or slug == "compare":
        return "compare"
    if slug == "postcode" or slug.startswith("postcode/"):
        return "postcode"
    if slug == "guides" or slug == "broadband-glossary" or slug.startswith("guides/"):
        return "guide"
    if slug.startswith("research/"):
        return "research"
    if slug == "speed-test" or slug.startswith("tools/"):
        return "tool"
    if slug in {
        "about", "how-we-make-money", "how-we-review-broadband", "editorial-policy",
        "contact", "privacy-policy", "terms", "cookie-policy",
    }:
        return "trust"
    return "other"


def page_type_recommended_actions(packet: dict[str, Any]) -> list[dict[str, Any]]:
    """Return the latest recommendations for every build, not just the canonical winner."""
    try:
        scan = json.loads(PAGE_TYPE_UX_SCAN_FILE.read_text(encoding="utf-8"))
    except (FileNotFoundError, OSError, json.JSONDecodeError):
        return []
    kind = ux_page_type_for_packet(packet)
    return [
        item for item in scan.get("recommendations", [])
        if item.get("page_type") == kind
    ]


def write_page_build_packet(packet: dict[str, Any]) -> tuple[Path, Path]:
    PIPELINE_DIR.mkdir(parents=True, exist_ok=True)
    packet["benchmark_strategy"] = {
        "uk_seo_leaders_required": 3,
        "ai_cited_pages_required_when_ai_overview_present": 1,
        "international_innovators_required": 2,
        "purpose": (
            "Synthesize same-topic UK search and GEO conventions with useful same-topic "
            "international UX patterns, then implement an original BroadbandPicker layout."
        ),
        "constraints": [
            "Do not copy wording, branding, visual identity or a distinctive page layout.",
            "Do not infer that a UX pattern caused a ranking or AI citation.",
            "Adapt international patterns to UK terminology, regulation, evidence and reader needs.",
            "Use BroadbandPicker's existing design system and accessible components.",
        ],
    }
    packet["orchestrated_ux_geo_requirements"] = page_type_recommended_actions(packet)
    try:
        decision = json.loads(CANONICAL_PRIORITY_FILE.read_text(encoding="utf-8"))
        selected = decision.get("next_priority", {})
        if selected.get("url", "").rstrip("/") == packet.get("url", "").rstrip("/"):
            packet["priority_evidence"] = {
                "selected_by": selected.get("source"),
                "score": selected.get("score"),
                "reason": selected.get("reason"),
                "metrics": selected.get("metrics", {}),
                "top_queries": selected.get("top_queries", []),
            }
            canonical_requirements = selected.get("ux_geo_requirements", [])
            if canonical_requirements:
                packet["orchestrated_ux_geo_requirements"] = canonical_requirements
    except (FileNotFoundError, OSError, json.JSONDecodeError):
        pass
    packet["mandatory_recommended_actions"] = [
        item for item in packet["orchestrated_ux_geo_requirements"]
        if item.get("priority") in {"P0", "P1"}
    ]
    packet_path = PIPELINE_DIR / "next-page.json"
    prompt_path = PIPELINE_DIR / "next-page-prompt.md"
    packet_path.write_text(json.dumps(packet, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    prompt = f"""# BroadbandPicker page-build task

Read `{PIPELINE_BRIEF.relative_to(ROOT)}` completely and follow it as the controlling specification.
Read `{packet_path.relative_to(ROOT)}` for the exact page, detailed keyword mapping, template route and prerequisites.
If `performance_brief` is present in the packet, treat its GSC queries, impressions, CTR and
position as the first-party optimisation brief. Do not replace those observed queries with generic
keyword-tool guesses. If it is absent, record that this is a new/no-data page and establish a
baseline during the scheduled post-publication reviews.
If `priority_evidence` and `orchestrated_ux_geo_requirements` are present, they came from the
combined GSC, GA4, content-depth and page-type SERP/UX workflow. Treat every P0/P1 requirement as
mandatory unless current live evidence proves it is already satisfied; record that validation in
the research packet. P2 requirements are evidence-backed enhancements to implement when they help
the reader's task. Do not discard these requirements during fresh page-level research.

The packet also contains `mandatory_recommended_actions`, the build gate derived from those P0/P1
requirements. Implement every action. For each one, add a matching entry to
`implemented_recommended_actions` in the research JSON with `priority`, `feature`, `disposition`
(`implemented` or `already_satisfied`), `implementation`, `evidence`, and `validation_terms`.
`validation_terms` must be a non-empty list of stable visible words, labels or facts that the
rendered-page validator can find in this page's `<main>`. An `already_satisfied` disposition still
needs current page-specific evidence and visible validation terms. Never mark an action satisfied
merely because a shared template or unrelated page contains something similar.

## 1. Keyword research and live SERP scraping (required every build, not optional)

The packet's keyword list is a starting point, not the finished research. Before writing:
- Search the primary keyword and scrape/read at least three currently ranking UK pages. Note the
  common structural pattern among the top results: roughly how long they are, how they're
  organised (heading structure, tables, FAQs), and what specific facts or numbers they lead with.
- Check whether an AI Overview / AI answer currently appears for the primary keyword. If it does,
  read what it cites and from where — that tells you what a citable answer for this specific
  keyword looks like, more reliably than any general rule.
- Build three separate same-topic benchmark sets: at least three UK organic leaders, every
  accessible page cited by the observed AI Overview (at least one is required when an AI Overview
  is present), and at least two strong non-UK pages. Select international pages for genuinely useful
  UX, information architecture, tools or citation design that the UK set does not commonly use,
  not merely because they rank in another country. Record the country and selection evidence.
- Inspect rendered page structure and responsive behaviour where access permits, not just titles
  and snippets. Compare section order, above-the-fold answer, navigation, tables/cards, decision
  support, source presentation, mobile behaviour and accessibility. If access is blocked, record
  the limitation and do not invent observations.
- Identify at least four distinct secondary/supporting queries beyond the primary term. Use close
  variants, commercial modifiers, People Also Ask questions, entity attributes and comparison
  questions that a UK reader would plausibly search. Map every query to a specific title, intro,
  H2/H3, table, body, FAQ or verdict slot before drafting. Do not create padding variants.
- Optionally run `python3 scripts/scrape_competitor_landscape.py` for a structural read on the
  established UK broadband comparison sites (word count, schema types, tool/trust signals) as
  background context — it targets general broadband hubs, not this specific keyword, so treat it
  as supporting evidence, not a substitute for the keyword-specific research above.

## 2. Content depth (match or exceed what's actually ranking, not a fixed word count)

Do not write a thin page and call it done, and do not pad a narrow topic to hit an arbitrary
length either. The right depth is whatever the keyword-research step above shows is actually
competitive for this specific query — a niche long-tail term may need 600 well-organised words; a
broad commercial term may need substantially more. Justify the depth you chose in your report with
what you saw ranking.

## 3. Scrape and research current facts

Scrape/read the relevant official provider, regulator or government pages plus trustworthy neutral
corroboration. Use at least one primary/official source and one independent or regulatory source;
commercial/provider pages normally need several of each. Record every source URL, source type,
claim supported and verification date. If sources use different populations or methodologies,
report them separately and do not combine them into a score. A conditional verdict or an explicit
"no universal winner" conclusion is valid when supported by the evidence. Never guess a winner,
trust score or statistic. Stop with a clear BLOCKED report only when the page's required factual
claims cannot be supported safely.

If prerequisites are missing, build and validate those provider entries first, using the existing interface and sibling structure. Use only a labelled text-wordmark placeholder when no official logo asset is available.

## 4. Write for search engines and generative answer engines (GEO)

Write the page copy into the existing data file or route specified by the packet. Use natural British English. Do not use em dashes. Do not mention AI, generation, prompts or this pipeline in public copy. Avoid generic filler and keyword stuffing. Put the primary keyword in the title and standalone answer-first excerpt. Cover secondary keywords naturally in the fixed template slots. Keep factual claims attributable and useful for both search engines and generative answer engines:
- Every excerpt and every FAQ answer must stand alone as a complete, specific, 35-90 word answer
  that would make sense quoted out of context by an AI Overview or chat answer — no "it depends,
  read on" filler.
- Lead sections with the concrete fact or number, not a rhetorical question.
- Keep a genuine freshness signal (a real reviewed/updated date) and visible primary sources —
  both are things the keyword-research step should confirm the ranking pages also do.
- Make important entities and relationships explicit: who provides the service, network used,
  geographic scope, price period, speed type, contract term, eligibility and verification date.
- Include concise definitions, direct answers, comparison criteria, limitations and actionable
  next steps so passages remain accurate when quoted without the surrounding page.
- Use descriptive headings, short answer-first paragraphs, useful tables/lists where the template
  supports them, internal links to relevant BroadbandPicker pages and visible citations close to
  material claims. Do not repeat a keyword merely to increase frequency.

## 4a. Write like a helpful person, not a search-optimised machine

Full detail is in Stage 4a of the brief — read it. The short version, checked before this build
ships: self-test every page against Google's own bar from "Creating helpful, reliable, people-first
content" — if search engines didn't exist, would you still write this for an actual reader? Give a
complete, specific answer to the exact question the primary keyword implies, cite where a claim
comes from, and take a real position where the evidence supports one ("we'd start with Sky here
because..." beats "it depends on your needs").

Do not use, anywhere in public copy: delve, tapestry, boast, realm, elevate, unlock, unleash,
landscape (as a market metaphor), navigate (as a metaphor), game-changer, seamless, robust, leverage
(as a verb), dive in, embark, testament to, plethora, myriad, underscore, foster, cutting-edge,
ever-evolving, "in today's digital age," "whether you're a X or a Y" as an opener, "it's important
to note," "it's worth noting," "when it comes to," "at the end of the day," a closing paragraph that
just restates the opening, a rhetorical-question opener, a rule-of-three adjective list, or a
sentence that could be pasted unchanged into any other UK broadband article. Em dashes are already
banned above.

## 5. Save the research and keyword map (mandatory validation input)

Before running the build, write `{PAGE_RESEARCH_FILE.relative_to(ROOT)}` as valid JSON with:
- `slug`, `primary_keyword`, and `search_locale` (`UK`);
- `secondary_keywords`: at least four objects containing `keyword`, `intent`, `page_slot` and
  `coverage_note`;
- `serp_competitors`: at least three objects containing `url`, `title`, `observed_structure`,
  `approx_word_count`, `content_gap_to_improve`, `useful_ux_patterns`, `useful_ui_patterns`,
  `interactive_or_functional_elements`, `trust_signals` and `citation_patterns`;
- `ai_overview`: an object containing `checked`, `present`, `observation` and `cited_sources`;
- `design_benchmarks`: an object with `uk_seo_leaders` (at least three), `ai_cited_pages`,
  `international_innovators` (at least two) and `citation_access_limitation`. Every benchmark page
  must contain `url`, `title`, `country`, `selection_basis`, `seo_geo_evidence`, `observed_layout`,
  `ux_patterns`, `ui_patterns`, `transferable_pattern` and `copying_risk`. AI-cited pages must also
  identify the cited claim or passage in `citation_evidence`. When no AI Overview is present,
  `ai_cited_pages` may be empty; when one is present but its sources cannot be accessed, explain the
  specific limitation rather than fabricating a source;
- `llm_visibility_observations`: checks performed for the topic, with `platform_or_method`,
  `observation` and `limitations`. Treat these as volatile observations, not ranking guarantees;
- `layout_blueprint`: an ordered list of at least four page sections. Each needs `order`, `section`,
  `reader_task`, `component_or_pattern`, `benchmark_sources`, `mobile_behaviour`,
  `accessibility_notes` and `validation_terms`;
- `benchmark_synthesis`: `adopted_patterns` (at least three) and one `differentiated_pattern`.
  Each adopted pattern needs `pattern`, `source_urls`, `user_need`, `adaptation`, `implementation`,
  `originality_guard` and `validation_terms`. The differentiated pattern must state what is uncommon
  in the UK benchmark set, its international evidence, UK adaptation, implementation and visible
  validation terms;
- `serp_features`: an object recording whether snippets, PAA, local results, video, tools,
  comparison tables, forums and AI answers were observed, plus the page-format implication;
- `people_also_ask`: useful question strings discovered during research;
- `sources`: at least three objects containing `url`, `source_type` (`primary`, `regulator`,
  `government`, `independent` or `reviews`), `verified_date` and `claims_supported`;
- `recommended_min_words`: an integer justified by the ranking-page review, never below 900 for a
  Guide or Provider page, 800 for a Comparison page, or 600 for an Interactive tool;
- `required_sections`, `internal_links`, `schema_types`, `ux_ui_requirements`,
  `functional_requirements`, `research_summary` and `depth_rationale`.
- `content_format`: the packet format ID plus a short rationale based on intent and the SERP;
- `information_gain`: at least one original, useful asset this page adds beyond summarising the
  ranking pages. Each item needs `asset`, `evidence`, `implementation` and `validation_terms`;
- `ctr_candidates`: at least three distinct objects with `title`, `meta_description` and
  `rationale`, plus `selected_ctr_candidate` containing the chosen title and description;
- `internal_link_plan`: `inbound` and `outbound` arrays, each with at least two objects containing
  `url`, `anchor` and `reason`. Use existing, topically relevant pages and descriptive anchors;
- `schema_eligibility`: proposed schema objects containing `type`, `eligible`, `visible_evidence`
  and `reason`. Never add schema for content or offers that are not visibly present and current;
- `post_publication_review`: concrete checks for `day_7`, `day_28`, `day_56` and `day_90`, using
  GSC impressions/CTR/position and GA4 engagement, affiliate clicks and AI/LLM referrals.
- `implemented_recommended_actions`: one complete entry for every item in the packet's
  `mandatory_recommended_actions`, using the exact `priority` and `feature`, plus `disposition`,
  `implementation`, `evidence` and visible `validation_terms` as described above.

The validator will reject a missing, thin or mismatched research file and will check that the
rendered page covers the mapped secondary queries and the justified minimum depth.

## 6. Apply competitor UX, UI and functional learnings

Use the strongest useful patterns found across the UK, AI-cited and international benchmark sets,
without copying their wording, branding, visual identity or distinctive layout. Synthesize the
patterns into the documented `layout_blueprint`; do not reproduce any one source page. Implement
only patterns that improve this page's search intent and reader task,
such as answer summaries, comparison/checklist tables, eligibility flows, cost examples, decision
steps, jump navigation, warnings, source notes or an interactive control. Record every adopted
pattern in `ux_ui_requirements` or `functional_requirements`, explain which observed user need it
serves, and implement it using the site's existing design system and accessible components.

Do not claim that a competitor pattern caused a ranking or AI citation. Treat ranking, AI Overview
and LLM visibility as evidence-informed targets, never guarantees. Do not clone a competitor page,
add decorative UI without a task benefit, or invent data to populate a feature.

At least one implemented pattern must be a useful, evidence-backed idea found in the international
set but uncommon across the reviewed UK pages. Adapt it to UK terminology, regulation, factual
sources and expectations. Every adopted and differentiated pattern needs stable visible
`validation_terms`; the rendered-page gate will reject a build when those terms are absent.

## 7. Intent-led product, accessibility and measurement layer

Every future page must help the reader complete its primary task, but interactivity is not a
ranking ornament. Use the packet's content format, the observed SERP and the reader task to decide:
- If interaction materially improves the task, build a page-specific accessible control. It must
  work without an account and retain a crawlable static explanation.
- If a static table, checklist, worked example, local evidence block or answer is clearer, use
  that instead. Do not add a quiz or calculator merely to satisfy the pipeline.
- Make the complete journey responsive at mobile, tablet and desktop widths. Avoid horizontal
  page overflow; allow wide data tables to scroll inside a labelled region; keep controls usable
  on touch screens; and preserve task hierarchy when cards stack.
- Use semantic HTML and keyboard-operable controls with visible focus states, programmatic labels,
  accessible names, useful instructions and live status feedback where state changes. Never rely
  on colour alone and respect reduced-motion preferences.
- Track meaningful interaction and conversion steps through `trackEvent` from the site's analytics
  helper. Define stable snake_case GA4 event names and non-personal parameters. Interactive pages
  must track interaction start, completion and the primary CTA. Static pages need the relevant
  existing commercial CTA event and must not invent low-value events. Do not send postcodes,
  names, email addresses or other personal data to GA4.
- Route affiliate links through the shared `AffiliateCTA` component and give every CTA a stable,
  descriptive `placement`. Its Awin ClickRef taxonomy and GA4 parameters must remain aligned so
  page, placement and provider clicks can be joined to Awin transactions. Never create per-user
  ClickRefs or put postcodes, names, emails or other personal data in affiliate tracking URLs.

Add these mandatory fields to `{PAGE_RESEARCH_FILE.relative_to(ROOT)}`:
- `interaction_decision`: `required`, `rationale`, `reader_task` and `static_fallback`;
- when `interaction_decision.required` is true, `interactive_comparison` must contain non-empty
  `user_task`, `choices_compared`, `crawlable_fallback` and `completion_state` values;
- `responsive_requirements`: at least three concrete checks covering mobile, tablet and desktop;
- `accessibility_requirements`: at least five concrete keyboard, semantics, labels, focus,
  state-feedback or reduced-motion checks;
- `ga4_events`: complete objects with `name`, `trigger`, `parameters` and `conversion_role`; at
  least three for interactive pages and at least one meaningful CTA event for static pages.

The outer runner owns production release. A normal build must pass local validation, deploy with
`vercel --prod --yes`, verify the live route is served by Vercel, regenerate both trackers,
synchronise the existing Google Sheet, mark the page built/live and calculate the next priority.
Do not mark a page complete before all of those steps succeed.

Do not commit, push or deploy. Run the relevant deterministic validation, including `npm run build`, and report exactly which files changed, the keyword research findings (what's currently ranking, what depth/structure you matched and why, whether an AI Overview was present), sources used, unresolved factual questions and validation results.
"""
    prompt_path.write_text(prompt, encoding="utf-8")
    return packet_path, prompt_path


def run_codex_page_writer(prompt_path: Path) -> Path:
    codex = shutil.which("codex")
    if not codex:
        raise RuntimeError("Codex CLI is required for researched page drafting")
    result_path = PIPELINE_DIR / "last-agent-result.md"
    # Current Codex CLI versions make --approve-for-me select the
    # workspace-write sandbox automatically, so passing --sandbox alongside
    # it is rejected as a conflicting option.
    command = [
        codex, "exec", "-C", str(ROOT), "--approve-for-me", "--ephemeral",
        "--output-last-message", str(result_path), "-",
    ]
    result = subprocess.run(
        command, input=prompt_path.read_text(encoding="utf-8"), text=True,
        cwd=ROOT, check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"Codex page-writing stage failed with exit code {result.returncode}")
    report = result_path.read_text(encoding="utf-8") if result_path.exists() else ""
    if re.match(r"\s*BLOCKED\b", report, re.IGNORECASE):
        raise RuntimeError(f"Page drafting stopped at its factual-review gate. See {result_path}")
    return result_path


def run_codex_google_sheet_sync(workbook_path: Path, spreadsheet_id: str) -> Path:
    """Use the authenticated Drive connector when no service account is configured."""
    codex = shutil.which("codex")
    if not codex:
        raise RuntimeError("Codex CLI is required for authenticated Google Sheet sync")
    result_path = PIPELINE_DIR / "last-sheet-sync-result.md"
    prompt = f"""Update the existing Google Sheet with ID `{spreadsheet_id}` from the local
workbook `{workbook_path}`. Use the connected Google Drive/Google Sheets tools and follow the
Google Sheets skill. Synchronise all workbook tabs in place, preserve the native Sheet and its
table structure, apply completed-row strikethrough from the Content Gap Roadmap, and verify the
top active roadmap row plus all completed rows after writing. Do not create a new spreadsheet.
Do not edit local files. Return a concise success report with the Sheet URL. If authentication or
permissions prevent the update, return BLOCKED and the exact reason.
"""
    command = [
        codex, "exec", "-C", str(ROOT), "--approve-for-me", "--ephemeral",
        "--output-last-message", str(result_path), "-",
    ]
    result = subprocess.run(command, input=prompt, text=True, cwd=ROOT, check=False)
    if result.returncode != 0:
        raise RuntimeError(f"Authenticated Google Sheet sync failed with exit code {result.returncode}")
    report = result_path.read_text(encoding="utf-8") if result_path.exists() else ""
    if re.match(r"\s*BLOCKED\b", report, re.IGNORECASE):
        raise RuntimeError(f"Google Sheet sync was blocked. See {result_path}")
    return result_path


def validate_page_research(packet: dict[str, Any]) -> dict[str, Any]:
    if not PAGE_RESEARCH_FILE.exists():
        raise RuntimeError(f"Required page research file is missing: {PAGE_RESEARCH_FILE}")
    try:
        research = json.loads(PAGE_RESEARCH_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"Page research file is not valid JSON: {exc}") from exc
    if research.get("slug") != packet["slug"]:
        raise RuntimeError("Page research file does not match the page being built")
    primary = str(research.get("primary_keyword") or "").strip().lower()
    packet_primary = str(packet["keywords"][0]["keyword"] if packet["keywords"] else "").strip().lower()
    if primary != packet_primary:
        raise RuntimeError(f"Research primary keyword does not match packet: {primary!r} != {packet_primary!r}")
    secondary = research.get("secondary_keywords") or []
    if len(secondary) < 4 or any(not all(item.get(key) for key in ("keyword", "intent", "page_slot", "coverage_note")) for item in secondary):
        raise RuntimeError("Page research must map at least four complete secondary keywords")
    competitors = research.get("serp_competitors") or []
    competitor_fields = (
        "url", "title", "observed_structure", "approx_word_count", "content_gap_to_improve",
        "useful_ux_patterns", "useful_ui_patterns", "interactive_or_functional_elements",
        "trust_signals", "citation_patterns",
    )
    if len(competitors) < 3 or any(not all(item.get(key) is not None for key in competitor_fields) for item in competitors):
        raise RuntimeError("Page research must record complete content, UX, UI, function, trust and citation learnings for at least three ranking competitors")
    ai_overview = research.get("ai_overview") or {}
    if ai_overview.get("checked") is not True or "present" not in ai_overview or not ai_overview.get("observation"):
        raise RuntimeError("Page research must record the AI Overview check and observation")
    benchmarks = research.get("design_benchmarks") or {}
    benchmark_fields = (
        "url", "title", "country", "selection_basis", "seo_geo_evidence",
        "observed_layout", "ux_patterns", "ui_patterns", "transferable_pattern", "copying_risk",
    )
    uk_leaders = benchmarks.get("uk_seo_leaders") or []
    if len(uk_leaders) < 3 or any(
        not all(item.get(key) for key in benchmark_fields) for item in uk_leaders
    ):
        raise RuntimeError("Page research must document at least three complete same-topic UK SEO/GEO design benchmarks")
    international = benchmarks.get("international_innovators") or []
    if len(international) < 2 or any(
        not all(item.get(key) for key in benchmark_fields) for item in international
    ):
        raise RuntimeError("Page research must document at least two complete same-topic international UX innovators")
    if any(str(item.get("country") or "").strip().lower() in {"uk", "united kingdom", "great britain"} for item in international):
        raise RuntimeError("International UX innovators must be outside the UK")
    ai_cited_pages = benchmarks.get("ai_cited_pages") or []
    if any(
        not all(item.get(key) for key in benchmark_fields + ("citation_evidence",))
        for item in ai_cited_pages
    ):
        raise RuntimeError("Every AI-cited design benchmark must contain complete layout and citation evidence")
    if ai_overview.get("present") is True and not ai_cited_pages and not str(
        benchmarks.get("citation_access_limitation") or ""
    ).strip():
        raise RuntimeError("An observed AI Overview requires at least one cited-page benchmark or a specific access limitation")
    llm_observations = research.get("llm_visibility_observations") or []
    if not llm_observations or any(
        not all(item.get(key) for key in ("platform_or_method", "observation", "limitations"))
        for item in llm_observations
    ):
        raise RuntimeError("Page research must record LLM visibility checks and their limitations")
    blueprint = research.get("layout_blueprint") or []
    blueprint_fields = (
        "order", "section", "reader_task", "component_or_pattern", "benchmark_sources",
        "mobile_behaviour", "accessibility_notes", "validation_terms",
    )
    if len(blueprint) < 4 or any(
        not all(item.get(key) for key in blueprint_fields) for item in blueprint
    ):
        raise RuntimeError("Page research must define a complete ordered layout blueprint with at least four sections")
    synthesis = research.get("benchmark_synthesis") or {}
    adopted = synthesis.get("adopted_patterns") or []
    adopted_fields = (
        "pattern", "source_urls", "user_need", "adaptation", "implementation",
        "originality_guard", "validation_terms",
    )
    if len(adopted) < 3 or any(not all(item.get(key) for key in adopted_fields) for item in adopted):
        raise RuntimeError("Page research must synthesize at least three complete benchmark patterns")
    differentiated = synthesis.get("differentiated_pattern") or {}
    differentiated_fields = (
        "pattern", "uncommon_in_uk_evidence", "international_source_urls", "uk_adaptation",
        "implementation", "validation_terms",
    )
    if not all(differentiated.get(key) for key in differentiated_fields):
        raise RuntimeError("Page research must define one complete international-to-UK differentiated pattern")
    serp_features = research.get("serp_features") or {}
    if not serp_features.get("checked") or not serp_features.get("format_implication"):
        raise RuntimeError("Page research must classify observed SERP features and their format implication")
    sources = research.get("sources") or []
    source_types = {str(item.get("source_type") or "").lower() for item in sources}
    if len(sources) < 3 or not source_types.intersection({"primary", "regulator", "government"}) or not source_types.intersection({"independent", "regulator", "government"}):
        raise RuntimeError("Page research needs at least three sources with primary and neutral/regulatory evidence")
    floors = {"Guide": 900, "Provider page": 900, "Comparison page": 800, "Interactive tool": 600}
    minimum = research.get("recommended_min_words")
    floor = floors.get(str(packet.get("page_type")), 800)
    if not isinstance(minimum, int) or minimum < floor:
        raise RuntimeError(f"Research depth target must be an integer of at least {floor} words")
    for key in ("required_sections", "internal_links", "schema_types", "ux_ui_requirements", "functional_requirements"):
        if not research.get(key):
            raise RuntimeError(f"Page research is missing {key}")
    content_format = research.get("content_format") or {}
    if content_format.get("id") != packet.get("content_format", {}).get("id") or not content_format.get("rationale"):
        raise RuntimeError("Page research must use and justify the packet's content format")
    information_gain = research.get("information_gain") or []
    gain_fields = ("asset", "evidence", "implementation", "validation_terms")
    if not information_gain or any(not all(item.get(key) for key in gain_fields) for item in information_gain):
        raise RuntimeError("Page research must define at least one complete information-gain asset")
    ctr_candidates = research.get("ctr_candidates") or []
    if len(ctr_candidates) < 3 or any(
        not all(item.get(key) for key in ("title", "meta_description", "rationale"))
        for item in ctr_candidates
    ):
        raise RuntimeError("Page research must define at least three complete CTR candidates")
    if len({str(item["title"]).strip().lower() for item in ctr_candidates}) < 3:
        raise RuntimeError("CTR candidate titles must be meaningfully distinct")
    selected_ctr = research.get("selected_ctr_candidate") or {}
    if not selected_ctr.get("title") or not selected_ctr.get("meta_description"):
        raise RuntimeError("Page research must record the selected title and meta description")
    link_plan = research.get("internal_link_plan") or {}
    link_fields = ("url", "anchor", "reason")
    for direction in ("inbound", "outbound"):
        links = link_plan.get(direction) or []
        if len(links) < 2 or any(not all(item.get(key) for key in link_fields) for item in links):
            raise RuntimeError(f"Internal-link plan needs at least two complete {direction} links")
    interaction = research.get("interaction_decision") or {}
    if not all(key in interaction and interaction.get(key) is not None for key in (
        "required", "rationale", "reader_task", "static_fallback"
    )):
        raise RuntimeError("Page research must make and justify an intent-led interaction decision")
    if not isinstance(interaction.get("required"), bool):
        raise RuntimeError("interaction_decision.required must be true or false")
    if interaction["required"]:
        comparison = research.get("interactive_comparison") or {}
        comparison_fields = ("user_task", "choices_compared", "crawlable_fallback", "completion_state")
        if any(not comparison.get(key) for key in comparison_fields):
            raise RuntimeError("Interactive pages must define a complete comparison experience")
    if len(research.get("responsive_requirements") or []) < 3:
        raise RuntimeError("Page research must define mobile, tablet and desktop responsive checks")
    if len(research.get("accessibility_requirements") or []) < 5:
        raise RuntimeError("Page research must define at least five concrete accessibility checks")
    ga4_events = research.get("ga4_events") or []
    ga4_fields = ("name", "trigger", "parameters", "conversion_role")
    minimum_events = 3 if interaction["required"] else 1
    if len(ga4_events) < minimum_events or any(not all(event.get(key) is not None for key in ga4_fields) for event in ga4_events):
        raise RuntimeError(f"Page research must define at least {minimum_events} complete GA4 events")
    for event in ga4_events:
        if not re.fullmatch(r"[a-z][a-z0-9]*(?:_[a-z0-9]+)*", str(event["name"])):
            raise RuntimeError(f"GA4 event name must use lowercase snake_case: {event['name']}")
        serialised_parameters = json.dumps(event["parameters"]).lower()
        if any(term in serialised_parameters for term in ("postcode", "email", "phone", "address")):
            raise RuntimeError(f"GA4 event parameters may not contain personal information: {event['name']}")
    schema_eligibility = research.get("schema_eligibility") or []
    schema_fields = ("type", "eligible", "visible_evidence", "reason")
    if not schema_eligibility or any(
        not all(key in item and item.get(key) is not None for key in schema_fields)
        for item in schema_eligibility
    ):
        raise RuntimeError("Page research must record evidence-led schema eligibility")
    review = research.get("post_publication_review") or {}
    if any(not review.get(f"day_{day}") for day in (7, 28, 56, 90)):
        raise RuntimeError("Page research must define day 7, 28, 56 and 90 performance reviews")
    if not research.get("research_summary") or not research.get("depth_rationale"):
        raise RuntimeError("Page research must explain its findings and depth rationale")

    mandatory_actions = packet.get("mandatory_recommended_actions") or []
    implemented_actions = research.get("implemented_recommended_actions") or []
    action_fields = (
        "priority", "feature", "disposition", "implementation", "evidence", "validation_terms",
    )
    if mandatory_actions and any(
        not all(item.get(key) for key in action_fields) for item in implemented_actions
    ):
        raise RuntimeError("Every implemented recommended action must contain all required fields")
    implemented_by_key = {
        (str(item.get("priority")), str(item.get("feature"))): item
        for item in implemented_actions
    }
    for required in mandatory_actions:
        key = (str(required.get("priority")), str(required.get("feature")))
        action = implemented_by_key.get(key)
        if action is None:
            raise RuntimeError(
                f"Page research has not addressed mandatory recommended action {key[0]}: {key[1]}"
            )
        if action.get("disposition") not in {"implemented", "already_satisfied"}:
            raise RuntimeError(
                f"Mandatory recommended action has invalid disposition {key[0]}: {key[1]}"
            )
        terms = action.get("validation_terms")
        if isinstance(terms, str):
            terms = [terms]
        if not isinstance(terms, list) or not terms or any(not str(term).strip() for term in terms):
            raise RuntimeError(
                f"Mandatory recommended action needs visible validation terms {key[0]}: {key[1]}"
            )
    return research


def validate_page_build(packet: dict[str, Any], port: int = 4321) -> None:
    research = validate_page_research(packet)
    try:
        subprocess.run(["npm", "run", "build"], cwd=ROOT, check=True)
    except subprocess.CalledProcessError:
        # Turbopack's CSS worker binds a loopback port during compilation and
        # some managed runners deny that operation. Webpack is a supported
        # Next.js production builder and provides the same TypeScript and
        # static-generation gate without the worker-port requirement.
        print("Turbopack build failed; retrying with the supported webpack builder")
        subprocess.run(["npm", "run", "build", "--", "--webpack"], cwd=ROOT, check=True)
    server = subprocess.Popen(
        ["npm", "run", "start", "--", "-p", str(port)], cwd=ROOT,
        stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT,
    )
    try:
        local_url = f"http://127.0.0.1:{port}/{packet['slug']}"
        response = None
        for _ in range(30):
            try:
                response = requests.get(local_url, timeout=3)
                if response.status_code:
                    break
            except requests.RequestException:
                time.sleep(1)
        if response is None or response.status_code != 200:
            code = response.status_code if response is not None else "unreachable"
            raise RuntimeError(f"Local route validation failed for {local_url}: {code}")
        primary = packet["keywords"][0]["keyword"] if packet["keywords"] else ""
        tree = html.fromstring(response.content)
        selected_ctr = research["selected_ctr_candidate"]
        rendered_title = " ".join(tree.xpath("//head/title/text()") or []).strip()
        rendered_meta = " ".join(tree.xpath(
            "//head/meta[translate(@name, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')='description']/@content"
        ) or []).strip()
        if str(selected_ctr["title"]).strip().lower() not in rendered_title.lower():
            raise RuntimeError(
                f"Rendered title does not use the selected CTR candidate: {rendered_title!r}"
            )
        if str(selected_ctr["meta_description"]).strip().lower() != rendered_meta.lower():
            raise RuntimeError("Rendered meta description does not match the selected CTR candidate")
        # Validate the drafted page copy, excluding shared site chrome such as
        # the legacy footer which may contain punctuation outside this page's
        # editorial scope.
        if packet["page_type"] == "Guide":
            visible_nodes = tree.xpath(
                "//h1//text()[normalize-space()] | "
                "//div[contains(concat(' ', normalize-space(@class), ' '), ' prose ')]//text()[normalize-space()] | "
                "//h2[contains(normalize-space(.), 'Frequently Asked Questions')]/following-sibling::*//text()[normalize-space()]"
            )
        else:
            visible_nodes = tree.xpath("//main//text()[normalize-space()]")
        visible = " ".join(visible_nodes)
        visible_word_count = len(re.findall(r"\b[\w'-]+\b", visible))
        if visible_word_count < research["recommended_min_words"]:
            raise RuntimeError(
                f"Rendered page is too thin for its researched SERP target: "
                f"{visible_word_count} < {research['recommended_min_words']} words"
            )
        if primary:
            def keyword_tokens(value: str) -> list[str]:
                tokens = re.findall(r"[a-z0-9]+", value.lower().replace("-", " "))
                stopwords = {"a", "an", "and", "for", "in", "of", "the", "to", "uk"}
                return [token[:-1] if len(token) > 3 and token.endswith("s") else token
                        for token in tokens if token not in stopwords]

            required_tokens = keyword_tokens(primary)
            visible_tokens = set(keyword_tokens(visible))
            if not required_tokens or not all(token in visible_tokens for token in required_tokens):
                raise RuntimeError(f"Primary keyword is absent from rendered page: {primary}")
        if "—" in visible:
            raise RuntimeError("Rendered public copy contains an em dash")
        gain_implemented = False
        for item in research["information_gain"]:
            terms = item.get("validation_terms") or []
            if isinstance(terms, str):
                terms = [terms]
            if terms and all(str(term).lower() in visible.lower() for term in terms):
                gain_implemented = True
                break
        if not gain_implemented:
            raise RuntimeError("Rendered page does not contain a validated information-gain asset")
        benchmark_patterns = list(research["benchmark_synthesis"]["adopted_patterns"])
        benchmark_patterns.append(research["benchmark_synthesis"]["differentiated_pattern"])
        for pattern in benchmark_patterns:
            terms = pattern.get("validation_terms") or []
            if isinstance(terms, str):
                terms = [terms]
            missing = [str(term) for term in terms if str(term).lower() not in visible.lower()]
            if missing:
                raise RuntimeError(
                    f"Rendered page is missing benchmark-pattern validation terms for "
                    f"{pattern.get('pattern')}: {missing}"
                )
        for action in research.get("implemented_recommended_actions") or []:
            terms = action.get("validation_terms") or []
            if isinstance(terms, str):
                terms = [terms]
            missing = [str(term) for term in terms if str(term).lower() not in visible.lower()]
            if missing:
                raise RuntimeError(
                    "Rendered page is missing validation terms for recommended action "
                    f"{action.get('priority')}: {action.get('feature')}: {missing}"
                )
        rendered_hrefs = {
            urllib.parse.urlparse(str(value)).path.rstrip("/") or "/"
            for value in tree.xpath("//main//a[@href]/@href")
        }
        planned_outbound = {
            urllib.parse.urlparse(str(item["url"])).path.rstrip("/") or "/"
            for item in research["internal_link_plan"]["outbound"]
        }
        if len(rendered_hrefs.intersection(planned_outbound)) < 2:
            raise RuntimeError("Rendered page is missing at least two planned outbound internal links")
        secondary_covered = 0
        for item in research["secondary_keywords"]:
            tokens = keyword_tokens(str(item["keyword"]))
            if tokens and all(token in visible_tokens for token in tokens):
                secondary_covered += 1
        required_secondary = max(3, math.ceil(len(research["secondary_keywords"]) * 0.6))
        if secondary_covered < required_secondary:
            raise RuntimeError(
                f"Rendered page covers only {secondary_covered}/{len(research['secondary_keywords'])} "
                f"mapped secondary keywords; at least {required_secondary} are required"
            )
        if not tree.xpath("//h1"):
            raise RuntimeError("Rendered page has no H1")
        if packet["page_type"] == "Comparison page":
            h2_text = [" ".join(node.itertext()).strip() for node in tree.xpath("//h2")]
            expected = ["Quick Verdict", "At-a-Glance Comparison", "Key Differences",
                        "How We Think About This Matchup", "Final Verdict", "Frequently Asked Questions",
                        "Editorial and Source Notes"]
            positions = [next((i for i, heading in enumerate(h2_text) if label in heading), -1)
                         for label in expected]
            if -1 in positions or positions != sorted(positions):
                raise RuntimeError(f"Comparison H2 structure is incomplete or out of order: {h2_text}")
            if len(tree.xpath("//script[@type='application/ld+json']")) < 2:
                raise RuntimeError("Comparison page is missing expected JSON-LD blocks")
    finally:
        server.terminate()
        try:
            server.wait(timeout=10)
        except subprocess.TimeoutExpired:
            server.kill()


def deploy_page_to_vercel(packet: dict[str, Any]) -> None:
    vercel = shutil.which("vercel")
    if not vercel:
        raise RuntimeError("Vercel CLI is required for production deployment")
    subprocess.run([vercel, "--prod", "--yes"], cwd=ROOT, check=True)
    response = fetch(packet["url"], timeout=30)
    if response.status_code != 200:
        raise RuntimeError(f"Production route returned HTTP {response.status_code}: {packet['url']}")
    if not response.headers.get("x-vercel-id") and "vercel" not in response.headers.get("server", "").lower():
        raise RuntimeError("Production route is live but Vercel headers were not detected")


def schedule_post_publication_reviews(packet: dict[str, Any]) -> None:
    """Persist the measurement loop only after production verification succeeds."""
    launched = datetime.now(timezone.utc)
    try:
        existing = json.loads(POST_PUBLICATION_REVIEW_FILE.read_text(encoding="utf-8"))
    except (FileNotFoundError, OSError, json.JSONDecodeError):
        existing = {"updated_at": None, "reviews": []}
    reviews = [
        item for item in existing.get("reviews", [])
        if item.get("slug") != packet["slug"]
    ]
    checks = {
        "day_7": "Confirm indexing, impressions, query matching and analytics events",
        "day_28": "Compare CTR and average position; test title/snippet only when evidence supports it",
        "day_56": "Review engagement, affiliate clicks, internal links and missing query coverage",
        "day_90": "Decide whether to consolidate, expand, refresh or leave unchanged",
    }
    research_reviews: dict[str, Any] = {}
    try:
        research_reviews = json.loads(PAGE_RESEARCH_FILE.read_text(encoding="utf-8")).get(
            "post_publication_review", {}
        )
    except (OSError, json.JSONDecodeError):
        pass
    schedule = []
    for day in (7, 28, 56, 90):
        key = f"day_{day}"
        schedule.append({
            "day": day,
            "due_date": (launched + timedelta(days=day)).date().isoformat(),
            "status": "Pending",
            "check": research_reviews.get(key) or checks[key],
        })
    reviews.append({
        "slug": packet["slug"],
        "url": packet["url"],
        "launched_at": launched.isoformat(),
        "performance_baseline": packet.get("performance_brief"),
        "schedule": schedule,
    })
    payload = {"updated_at": launched.isoformat(), "reviews": reviews}
    POST_PUBLICATION_REVIEW_FILE.parent.mkdir(parents=True, exist_ok=True)
    POST_PUBLICATION_REVIEW_FILE.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )


# ---------------------------------------------------------------------------
# 6. CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--offline", action="store_true", help="Skip the live sitemap crawl and use the static page snapshot")
    parser.add_argument("--max-pages", type=int, default=140, help="Maximum live pages to audit")
    parser.add_argument("--source-sheet-id", default=os.environ.get("BROADBANDPICKER_SOURCE_SHEET_ID", DEFAULT_SOURCE_SHEET_ID),
                         help="Google Sheet ID to read extra keyword ideas from (must be shared with the service account)")
    parser.add_argument("--skip-source-sheet", action="store_true", help="Do not read the source sheet even if an ID is set")
    parser.add_argument("--create-google-doc", action="store_true", help="Create a brand-new Google Sheet on Drive with the finished mapping")
    parser.add_argument("--update-google-sheet-id", default=os.environ.get("BROADBANDPICKER_MAPPING_SHEET_ID", ""),
                        help=f"Update an existing mapping Sheet in place (for this plan use {DEFAULT_MAPPING_SHEET_ID})")
    parser.add_argument("--build-next-page", action="store_true",
                        help="Research, draft and locally validate the next active roadmap page")
    parser.add_argument("--build-page-url",
                        help="Research, completely rebuild and validate a specific existing BroadbandPicker URL")
    parser.add_argument("--build-all-priority", action="store_true",
                        help="Build every active roadmap page in priority order, stopping on the first failure")
    parser.add_argument("--max-priority-pages", type=int, default=DEFAULT_PRIORITY_BATCH_SIZE,
                        help=f"Maximum pages to process with --build-all-priority (default: {DEFAULT_PRIORITY_BATCH_SIZE}; 0 means every active page)")
    parser.add_argument("--prepare-only", action="store_true",
                        help="Write the next-page build packet and prompt without running the Codex writer")
    parser.add_argument("--deploy-production", action="store_true",
                        help="After validation, deploy the drafted page to Vercel production")
    parser.add_argument("--use-existing-draft", action="store_true",
                        help="Skip the writing agent and validate/deploy the already human-reviewed local draft")
    parser.add_argument("--approve-factual-review", action="store_true",
                        help="Confirm a human reviewed winner, trust, pricing and estimated-stat claims before deployment")
    parser.add_argument("--doc-title", default=f"BroadbandPicker Keyword Mapping — {RUN_DATE}")
    parser.add_argument("--master-tracker", type=Path,
                        default=ROOT / "docs" / "master-build-tracker.xlsx",
                        help="Master tracker to regenerate and merge into this workbook")
    parser.add_argument("--skip-master-tracker-merge", action="store_true",
                        help="Do not regenerate or merge the master and pending build tabs")
    parser.add_argument("--share-with", default=os.environ.get("BROADBANDPICKER_SHARE_EMAIL", ""),
                         help="Email to share the newly created Google Sheet with as Editor")
    args = parser.parse_args()

    urls = get_sitemap_urls(args.offline)
    print(f"Found {len(urls)} live sitemap URLs")
    audit = []
    if not args.offline:
        for index, url in enumerate(urls[:args.max_pages], 1):
            print(f"Auditing {index}/{min(len(urls), args.max_pages)}: {url}")
            audit.append(audit_page(url))
            time.sleep(0.05)

    keywords = build_keyword_dataset()
    print(f"Built {len(keywords)} curated keyword rows")

    if args.source_sheet_id and not args.skip_source_sheet:
        try:
            extra = read_source_sheet_keywords(args.source_sheet_id)
            keywords = merge_source_keywords(keywords, extra)
        except Exception as exc:
            print(f"Warning: could not read source sheet {args.source_sheet_id} ({exc}); continuing without it")

    wb = build_workbook(args.output.resolve(), urls, audit, keywords, args.source_sheet_id if not args.skip_source_sheet else None)
    validate(wb)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output.resolve())
    print(f"Saved {args.output.resolve()}")

    if not args.skip_master_tracker_merge:
        subprocess.run([
            sys.executable, str(ROOT / "scripts" / "build_master_tracker.py"),
            "--source", str(args.output.resolve()), "--output", str(args.master_tracker.resolve()),
        ], cwd=ROOT, check=True)
        master_wb = load_workbook(args.master_tracker.resolve(), data_only=False)
        for source_name, tab_name in (
            ("Master Tracker", "Master Build Tracker"),
            ("Pending Build Priority", "Pending Build Priority"),
        ):
            if tab_name in wb.sheetnames:
                del wb[tab_name]
            source_ws = master_wb[source_name]
            rows = [[cell.value for cell in row] for row in source_ws.iter_rows()]
            add_sheet(wb, tab_name, rows[0], rows[1:])
        wb.save(args.output.resolve())
        print(f"Merged master tracker and pending priority tabs from {args.master_tracker.resolve()}")

    if args.build_all_priority:
        args.build_next_page = True
    if args.max_priority_pages < 0:
        parser.error("--max-priority-pages cannot be negative")
    if args.build_page_url:
        args.build_next_page = True
    if args.deploy_production and not args.build_next_page:
        parser.error("--deploy-production requires --build-next-page, --build-page-url or --build-all-priority")
    if args.deploy_production and not args.approve_factual_review:
        parser.error("--deploy-production requires --approve-factual-review")
    if args.build_next_page and not args.prepare_only and not args.deploy_production:
        parser.error("Page builds must use --deploy-production; use --prepare-only to generate instructions without shipping")
    if args.deploy_production and not args.update_google_sheet_id:
        parser.error("Production page builds require --update-google-sheet-id so the shared tracker is updated")
    if args.build_all_priority and not args.deploy_production:
        parser.error("--build-all-priority requires --deploy-production so priority can advance from verified live pages")
    if args.build_all_priority and args.use_existing_draft:
        parser.error("--use-existing-draft cannot be combined with --build-all-priority")
    if args.build_all_priority and args.prepare_only:
        parser.error("--prepare-only cannot be combined with --build-all-priority")
    if args.build_all_priority and args.build_page_url:
        parser.error("--build-page-url cannot be combined with --build-all-priority")

    page_packet = None
    doc_result = None
    if args.build_next_page:
        if not PIPELINE_BRIEF.exists():
            raise RuntimeError(f"Missing pipeline brief: {PIPELINE_BRIEF}")
        completed_packets: list[dict[str, Any]] = []
        batch_limit = args.max_priority_pages if args.build_all_priority else 1
        while not batch_limit or len(completed_packets) < batch_limit:
            try:
                current_packet = (
                    page_packet_for_url(args.build_page_url, keywords)
                    if args.build_page_url else next_page_packet(wb, keywords)
                )
            except RuntimeError as exc:
                if args.build_all_priority and "No active page remains" in str(exc):
                    print("All active roadmap pages are built and live")
                    break
                raise
            page_packet = current_packet
            packet_path, prompt_path = write_page_build_packet(current_packet)
            position = len(completed_packets) + 1
            prefix = f"Priority batch page {position}" if args.build_all_priority else "Next page"
            print(f"Prepared page-build packet: {packet_path}")
            print(f"{prefix}: {current_packet['title']} ({current_packet['url']})")
            if args.prepare_only:
                break
            if not args.use_existing_draft:
                report_path = run_codex_page_writer(prompt_path)
                print(f"Page writer completed: {report_path}")
            validate_page_build(current_packet)
            print(f"Local page validation passed: {current_packet['url']}")
            if not args.deploy_production:
                break
            deploy_page_to_vercel(current_packet)
            print(f"Vercel production validation passed: {current_packet['url']}")
            schedule_post_publication_reviews(current_packet)
            print(f"Scheduled post-publication reviews: {POST_PUBLICATION_REVIEW_FILE}")
            change_type = (
                "content_refresh"
                if current_packet.get("build_mode") == "complete_existing_page"
                else "new_page"
            )
            subprocess.run([
                sys.executable, str(ROOT / "scripts" / "record_content_change.py"),
                "--url", current_packet["url"],
                "--title", current_packet["title"],
                "--change-type", change_type,
                "--source", "build_keyword_mapping.py",
                "--deployment", "vercel-production",
            ], cwd=ROOT, check=True)
            print(f"Recorded GA4 annotation and GSC measurement baseline: {current_packet['url']}")
            completed_packets.append(current_packet)

            refreshed_urls = get_sitemap_urls(False)
            refreshed_audit = [
                *[item for item in audit if item.get("url") != current_packet["url"]],
                audit_page(current_packet["url"]),
            ]
            # Carry every successful page's live audit into the next batch
            # iteration. Without this assignment, the following workbook was
            # rebuilt from the run's original audit plus only the latest page,
            # which could put an earlier deployment back into the active queue.
            audit = refreshed_audit
            wb = build_workbook(args.output.resolve(), refreshed_urls, refreshed_audit,
                                keywords, args.source_sheet_id if not args.skip_source_sheet else None)
            validate(wb)
            wb.save(args.output.resolve())

            # Checkpoint every successful batch deployment remotely. If a later
            # page fails its research, build or live checks, completed work is
            # still represented accurately in the shared roadmap.
            if args.build_all_priority and args.update_google_sheet_id:
                try:
                    doc_result = update_google_sheet_from_workbook(wb, args.update_google_sheet_id)
                    print(f"Checkpointed Google Sheet: {doc_result['spreadsheetUrl']}")
                except RuntimeError as exc:
                    if "Set GOOGLE_SERVICE_ACCOUNT_JSON" not in str(exc) and "pip install" not in str(exc):
                        raise
                    print(f"Direct Google API sync unavailable ({exc}); using authenticated Drive connector")
                    sync_report = run_codex_google_sheet_sync(args.output.resolve(), args.update_google_sheet_id)
                    doc_result = {
                        "spreadsheetId": args.update_google_sheet_id,
                        "spreadsheetUrl": f"https://docs.google.com/spreadsheets/d/{args.update_google_sheet_id}/edit",
                        "syncReport": str(sync_report),
                    }
                    print(f"Checkpointed Google Sheet through authenticated connector: {doc_result['spreadsheetUrl']}")

            if args.build_page_url:
                break
            try:
                following_packet = next_page_packet(wb, keywords)
            except RuntimeError as exc:
                if "No active page remains" not in str(exc):
                    raise
                print("No following active roadmap page remains")
                break
            following_path, _ = write_page_build_packet(following_packet)
            print(f"Prepared following page-build packet: {following_path}")
            print(f"Following page: {following_packet['title']} ({following_packet['url']})")
            if not args.build_all_priority:
                break

    if args.create_google_doc:
        wb_for_upload = load_workbook(args.output.resolve())
        doc_result = create_google_sheet_from_workbook(wb_for_upload, args.doc_title, args.share_with or None)
        print(f"Created Google Sheet: {doc_result['spreadsheetUrl']}")
    elif args.update_google_sheet_id and not (args.build_all_priority and doc_result):
        wb_for_upload = load_workbook(args.output.resolve())
        try:
            doc_result = update_google_sheet_from_workbook(wb_for_upload, args.update_google_sheet_id)
            print(f"Updated Google Sheet: {doc_result['spreadsheetUrl']}")
        except RuntimeError as exc:
            if "Set GOOGLE_SERVICE_ACCOUNT_JSON" not in str(exc) and "pip install" not in str(exc):
                raise
            print(f"Direct Google API sync unavailable ({exc}); using authenticated Drive connector")
            sync_report = run_codex_google_sheet_sync(args.output.resolve(), args.update_google_sheet_id)
            doc_result = {
                "spreadsheetId": args.update_google_sheet_id,
                "spreadsheetUrl": f"https://docs.google.com/spreadsheets/d/{args.update_google_sheet_id}/edit",
                "syncReport": str(sync_report),
            }
            print(f"Updated Google Sheet through authenticated connector: {doc_result['spreadsheetUrl']}")

    print(json.dumps({
        "output": str(args.output.resolve()),
        "sheets": load_workbook(args.output.resolve(), read_only=True).sheetnames,
        "site_urls": len(urls),
        "audited_pages": len(audit),
        "keyword_rows": len(keywords),
        "google_doc": doc_result,
        "page_pipeline": page_packet,
    }, indent=2))


if __name__ == "__main__":
    main()
